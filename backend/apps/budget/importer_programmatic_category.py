"""Validated, idempotent loader for the POAU 2027 category master."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction

try:
    import openpyxl
except ImportError:  # pragma: no cover - reported as a command error
    openpyxl = None

from .models import EstadoCategoria, NivelCategoria, ProgrammaticCategory


EXPECTED_HEADERS = (
    "Código",
    "PROG.",
    "PROYEC.",
    "ACTIV.",
    "CAT. PROGRAMATICA",
    "Descripción",
    "Nivel",
)
VALID_LEVELS = {
    NivelCategoria.PROGRAMA,
    NivelCategoria.ACTIVIDAD,
}
IMPORT_DATE = date(2027, 1, 1)
NORMATIVA = "Maestro de formulación POAU 2027"


@dataclass(frozen=True)
class ProgrammaticCategoryRow:
    excel_row: int
    source_code: str
    prog: str
    proyec: str
    activ: str
    codigo: str
    descripcion: str
    nivel: str

    @property
    def parent_codigo(self):
        return f"{self.prog} {self.proyec} 000"


class WorkbookValidationError(ValueError):
    """Raised when the source cannot be safely interpreted."""


def _text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _numeric_text(value, number_format):
    if isinstance(value, bool):
        return ""
    if isinstance(value, Decimal):
        decimal_value = value
    elif isinstance(value, (int, float)):
        decimal_value = Decimal(str(value))
    else:
        return _text(value)
    if decimal_value != decimal_value.to_integral_value():
        return _text(value)
    result = str(int(decimal_value))
    format_digits = len(number_format.split(".", 1)[0].replace("#", "").replace("?", ""))
    return result.zfill(format_digits) if format_digits else result


def _segment(cell):
    value = _numeric_text(cell.value, cell.number_format)
    if not value or not re.fullmatch(r"\d+", value):
        raise WorkbookValidationError(
            f"Fila {cell.row}, columna {cell.column_letter}: segmento inválido {cell.value!r}."
        )
    return value


def _source_code(cell):
    value = _text(cell.value)
    if not value:
        raise WorkbookValidationError(f"Fila {cell.row}: Código vacío.")
    return value


def _header(sheet, header_row):
    actual = tuple(_text(sheet.cell(header_row, column).value) for column in range(1, 8))
    if actual != EXPECTED_HEADERS:
        raise WorkbookValidationError(
            f"Encabezado inválido en la fila {header_row}: {actual!r}; "
            f"se esperaba {EXPECTED_HEADERS!r}."
        )


def _read_rows(sheet, value_sheet, header_row):
    rows = []
    formula_rows = sheet.iter_rows(min_row=header_row + 1, max_col=7)
    value_rows = value_sheet.iter_rows(min_row=header_row + 1, max_col=7)
    for row_number, (formula_cells, value_cells) in enumerate(
        zip(formula_rows, value_rows), header_row + 1,
    ):
        formula_values = [cell.value for cell in formula_cells]
        cached_values = [cell.value for cell in value_cells]
        if all(value in (None, "") for value in formula_values):
            continue
        if any(value in (None, "") for value in cached_values):
            raise WorkbookValidationError(f"Fila {row_number}: faltan valores en A:G.")

        source_code = _source_code(formula_cells[0])
        prog = _segment(formula_cells[1])
        proyec = _segment(formula_cells[2])
        activ = _segment(formula_cells[3])
        codigo = _text(value_cells[4].value)
        descripcion = _text(value_cells[5].value)
        nivel = _text(value_cells[6].value).upper()
        expected_codigo = f"{prog} {proyec} {activ}"
        if codigo != expected_codigo:
            raise WorkbookValidationError(
                f"Fila {row_number}: CAT. PROGRAMATICA {codigo!r} no coincide con "
                f"PROG./PROYEC./ACTIV. {expected_codigo!r}."
            )
        if not descripcion:
            raise WorkbookValidationError(f"Fila {row_number}: Descripción vacía.")
        if nivel not in VALID_LEVELS:
            raise WorkbookValidationError(f"Fila {row_number}: nivel desconocido {nivel!r}.")
        rows.append(ProgrammaticCategoryRow(
            excel_row=row_number,
            source_code=source_code,
            prog=prog,
            proyec=proyec,
            activ=activ,
            codigo=codigo,
            descripcion=descripcion,
            nivel=nivel,
        ))
    if not rows:
        raise WorkbookValidationError("El rango de datos no contiene filas.")
    return rows


def validate_rows(rows):
    """Validate uniqueness and the program-to-activity hierarchy."""
    seen = set()
    for row in rows:
        if row.codigo in seen:
            raise WorkbookValidationError(
                f"Fila {row.excel_row}: código de categoría duplicado {row.codigo!r}."
            )
        seen.add(row.codigo)

    programs = {row.codigo for row in rows if row.nivel == NivelCategoria.PROGRAMA}
    for row in rows:
        if row.nivel == NivelCategoria.ACTIVIDAD and row.parent_codigo not in programs:
            raise WorkbookValidationError(
                f"Fila {row.excel_row}: actividad {row.codigo!r} sin programa padre "
                f"{row.parent_codigo!r}."
            )
    return rows


def parse_workbook(file_path, sheet_name="CLASIFICADOR", header_row=8):
    """Read and validate the workbook without touching the database."""
    if openpyxl is None:
        raise WorkbookValidationError("openpyxl no está instalado.")
    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError(f"No existe el workbook: {path}")
    if header_row < 1:
        raise WorkbookValidationError("--header-row debe ser mayor o igual a 1.")

    # The supplied workbook contains pivot-cache relationships that make
    # openpyxl's normal worksheet loader unnecessarily expensive. Read-only
    # mode still exposes formulas, cached values, number formats, and rows A:G.
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    values_workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookValidationError(
                f'La hoja "{sheet_name}" no existe. Hojas: {workbook.sheetnames}'
            )
        sheet = workbook[sheet_name]
        value_sheet = values_workbook[sheet_name]
        _header(sheet, header_row)
        return validate_rows(_read_rows(sheet, value_sheet, header_row))
    finally:
        workbook.close()
        values_workbook.close()


def _origin(path, sheet_name):
    stem = re.sub(r"[^A-Za-z0-9]+", "-", Path(path).stem).strip("-")
    return f"{stem[:27]}:{sheet_name}"[:40]


def _observaciones(row, sheet_name):
    return (
        f"Fuente Código={row.source_code}; PROG.={row.prog}; "
        f"PROYEC.={row.proyec}; ACTIV.={row.activ}; "
        f"Hoja={sheet_name}; fila Excel={row.excel_row}"
    )


def _defaults(row, parent, path, sheet_name):
    return {
        "denominacion": row.descripcion,
        "nivel": row.nivel,
        "parent": parent,
        "vigencia_desde": IMPORT_DATE,
        "estado": EstadoCategoria.ACTIVA,
        "origen": _origin(path, sheet_name),
        "normativa": NORMATIVA,
        "observaciones": _observaciones(row, sheet_name),
    }


def _existing_state(gestion, rows):
    existing = {
        category.codigo: category
        for category in ProgrammaticCategory.objects.filter(gestion=gestion)
    }
    source_codes = {row.codigo for row in rows}
    return existing, len(set(existing) - source_codes)


def _counts(gestion, rows):
    existing, preserved = _existing_state(gestion, rows)
    counts = {"created": 0, "updated": 0, "unchanged": 0}
    programs = sum(row.nivel == NivelCategoria.PROGRAMA for row in rows)
    activities = sum(row.nivel == NivelCategoria.ACTIVIDAD for row in rows)
    for row in rows:
        parent = existing.get(row.parent_codigo) if row.nivel == NivelCategoria.ACTIVIDAD else None
        values = _defaults(row, parent, "", "")
        values.pop("origen")
        values.pop("observaciones")
        current = existing.get(row.codigo)
        if current is None:
            counts["created"] += 1
            continue
        changed = any(getattr(current, key) != value for key, value in values.items())
        counts["updated" if changed else "unchanged"] += 1
    return counts, programs, activities, preserved


@transaction.atomic
def load_programmatic_categories(gestion, rows, path, sheet_name, commit):
    """Return a report and optionally persist all validated rows atomically."""
    counts, programs, activities, preserved = _counts(gestion, rows)
    warnings = []
    if any(row.activ == "0150" for row in rows):
        warnings.append(
            "Fila con ACTIV.=0150 aceptada como texto; no se truncó el código 200 0 0150."
        )
    report = {
        "status": "committed" if commit else "dry-run",
        "file": Path(path).name,
        "sheet": sheet_name,
        "gestion": gestion.anio,
        "rows": len(rows),
        "programs": programs,
        "activities": activities,
        "counts": counts,
        "warnings": warnings,
        "existing_rows_preserved": preserved,
    }
    if not commit:
        return report

    # Programs are written first so every activity can use a real FK parent.
    by_code = {}
    for row in rows:
        if row.nivel != NivelCategoria.PROGRAMA:
            continue
        category, _ = ProgrammaticCategory.objects.update_or_create(
            gestion=gestion,
            codigo=row.codigo,
            defaults=_defaults(row, None, path, sheet_name),
        )
        by_code[row.codigo] = category
    for row in rows:
        if row.nivel != NivelCategoria.ACTIVIDAD:
            continue
        parent = by_code[row.parent_codigo]
        ProgrammaticCategory.objects.update_or_create(
            gestion=gestion,
            codigo=row.codigo,
            defaults=_defaults(row, parent, path, sheet_name),
        )
    return report


def import_programmatic_categories(file_path, sheet_name="CLASIFICADOR", header_row=8, commit=False):
    rows = parse_workbook(file_path, sheet_name=sheet_name, header_row=header_row)
    from apps.gestion.models import GestionFiscal

    try:
        gestion = GestionFiscal.objects.get(anio=2027)
    except GestionFiscal.DoesNotExist as exc:
        raise WorkbookValidationError("No existe GestionFiscal 2027.") from exc
    return load_programmatic_categories(
        gestion, rows, file_path, sheet_name, commit=commit,
    )
