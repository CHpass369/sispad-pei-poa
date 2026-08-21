"""API V2 del ciclo presupuestario SIS-POA.

Fase 1 — Gestión fiscal:
    GET/POST   /api/v2/sis-poa/budget/fiscal-years/
    GET/PATCH  /api/v2/sis-poa/budget/fiscal-years/{id}/
    POST       /api/v2/sis-poa/budget/fiscal-years/{id}/enable/   → HABILITADA
    POST       /api/v2/sis-poa/budget/fiscal-years/{id}/close/    → CERRADA

Fase 2 — Techo directivo:
    GET/POST            /api/v2/sis-poa/budget/directive-ceilings/
    GET/PATCH/DELETE    /api/v2/sis-poa/budget/directive-ceilings/{id}/
    POST                .../directive-ceilings/{id}/submit/   → EN_REVISION
    POST                .../directive-ceilings/{id}/observe/  → OBSERVADO
    POST                .../directive-ceilings/{id}/approve/  → APROBADO
    POST                .../directive-ceilings/{id}/freeze/   → FIJADO
    GET                 .../directive-ceilings/{id}/composition/
    CRUD                /api/v2/sis-poa/budget/resources/         (?version=)
    CRUD                /api/v2/sis-poa/budget/mandatory-expenses/ (?version=)
    POST (multipart)    /api/v2/sis-poa/budget/documents/          (upload)
    GET                 /api/v2/sis-poa/budget/documents/?gestion=

Permisos (ADR-003):
    create/update/delete → capacidad `sis_poa.budget.manage`
    submit/observe/approve/freeze → capacidad `sis_poa.budget.approve`
    el resto usa IsAuthenticated (default global).
"""
from django.core.exceptions import ValidationError as DjangoValidationError
from django.shortcuts import get_object_or_404
from drf_spectacular.utils import OpenApiTypes, extend_schema
from decimal import Decimal

from django.utils import timezone

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.permissions import TieneCapacidad
from apps.auditoria.models import EventoAuditoria
from apps.auditoria.services import registrar_evento
from apps.core.pagination import AuditoriaDualPagination, ImportacionDualPagination
from apps.gestion.models import GestionFiscal

from .models import (
    Apertura,
    EstadosTecho,
    NivelCategoria,
    RevisionApertura,
    DocumentoPresupuestario,
    RecursoTecho,
    TechoDirectivo,
    TechoVersion,
    DistribucionVersion,
    AsignacionObjetoGastoTecho,
    GastoObligatorio,
    CategoriaProgramaticaTecho,
    Reforma,
    Reserva,
    AsignacionTerritorial,
    DistribucionTerritorial,
)
from .serializers import (
    AperturaSerializer,
    AuditEventSerializer,
    DocumentoPresupuestarioSerializer,
    RecursoTechoSerializer,
    TechoDirectivoSerializer,
    DistribucionVersionSerializer,
    AsignacionObjetoGastoTechoSerializer,
    GestionFiscalPresupuestoSerializer,
    GastoObligatorioSerializer,
    CategoriaProgramaticaTechoSerializer,
    ReformaSerializer,
    ReservaSerializer,
    DistribucionTerritorialSerializer,
    _serializar_montos,
)
from .services import (
    ErrorDisponibilidad,
    ErrorObjetoGastoExcedido,
    ajuste_distribucion,
    aplicar_reform,
    aprobar,
    aprobar_distribucion,
    aprobar_reform,
    actualizar_allocation,
    actualizar_objeto_gasto,
    cerrar_allocation,
    cerrar_gestion,
    composicion_techo,
    crear_allocation,
    crear_reform,
    crear_reserva,
    eliminar_allocation,
    eliminar_objeto_gasto,
    enviar_a_revision,
    enviar_distribucion_a_revision,
    enviar_reform_a_revision,
    fijar_distribucion,
    fijar_techo,
    gestion_habilitada,
    habilitar_gestion,
    liberar_reserva,
    observar,
    observar_distribucion,
    observar_reform,
    programar_objeto_gasto,
    rechazar_reform,
    resumen_distribucion,
    validar_distribucion_completa,
)

CAPACIDAD_GESTION = 'sis_poa.budget.manage'
CAPACIDAD_APROBACION = 'sis_poa.budget.approve'
CAPACIDAD_AUDITORIA = 'sis_poa.budget.audit_read'

ERROR_409_INMUTABLE = {
    'error': {
        'detail': 'La versión está fijada (inmutable); no se puede modificar.',
    },
}


def _respuesta_error(exception):
    return Response(
        {'error': {'detail': exception.messages}},
        status=400,
    )


def _version_actual_de(ceiling):
    return TechoVersion.objects.get(
        ceiling=ceiling, numero=ceiling.version_actual,
    )


class GestionFiscalPresupuestoViewSet(viewsets.ModelViewSet):
    queryset = GestionFiscal.objects.all()
    serializer_class = GestionFiscalPresupuestoSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filterset_fields = ['anio', 'estado', 'activa']
    search_fields = ['anio', 'descripcion']

    def get_permissions(self):
        if self.action in ('enable', 'close'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()

    def _ejecutar_servicio(self, request, pk, servicio):
        gestion = self.get_object()
        try:
            servicio(gestion, request.user)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(gestion).data)

    @action(detail=True, methods=['post'], url_path='enable')
    def enable(self, request, pk=None):
        """Habilita la gestión para el ciclo presupuestario (HABILITADA)."""
        return self._ejecutar_servicio(request, pk, habilitar_gestion)

    @action(detail=True, methods=['post'], url_path='close')
    def close(self, request, pk=None):
        """Cierra la gestión del ciclo presupuestario (CERRADA)."""
        return self._ejecutar_servicio(request, pk, cerrar_gestion)


# ---------------------------------------------------------------------------
# Techo directivo
# ---------------------------------------------------------------------------
class TechoDirectivoViewSet(viewsets.ModelViewSet):
    queryset = TechoDirectivo.objects.select_related('gestion').all()
    serializer_class = TechoDirectivoSerializer
    filterset_fields = ['gestion', 'estado']
    search_fields = ['gestion__anio']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        if self.action in ('submit', 'observe', 'approve', 'freeze'):
            return [TieneCapacidad(CAPACIDAD_APROBACION)]
        return super().get_permissions()

    def _ejecutar_servicio(self, request, pk, servicio, *args, **kwargs):
        ceiling = self.get_object()
        try:
            servicio(_version_actual_de(ceiling), request.user, *args, **kwargs)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(ceiling).data)

    @action(detail=True, methods=['get'], url_path='presupuesto-recursos')
    def presupuesto_recursos(self, request, pk=None):
        """Presupuesto General de Recursos, en la forma del reporte oficial.

        Devuelve los rubros agrupadores con sus componentes anidados y el
        total general. Los porcentajes se calculan aqui y nunca se guardan:
        la planilla de origen mostraba #DIV/0! cuando el divisor era cero, y
        aqui esos casos viajan como null para que la pantalla los muestre
        como un guion.
        """
        techo = self.get_object()
        version = _version_actual_de(techo)

        def porcentaje(parte, total):
            if not total:
                return None
            return round(parte * Decimal(100) / total, 2)

        def fila(recurso, total_padre=None):
            return {
                'id': recurso.id,
                'concepto': recurso.concepto,
                'origen': recurso.origen,
                'fuente': recurso.fuente.codigo if recurso.fuente else '',
                'organismo': recurso.organismo.codigo if recurso.organismo else '',
                'ff_of': (
                    f'{recurso.fuente.codigo}/{recurso.organismo.codigo}'
                    if recurso.fuente and recurso.organismo else ''
                ),
                'monto': recurso.monto,
                # Un componente se mide contra su grupo; un grupo, contra si mismo.
                'porcentaje': porcentaje(recurso.monto, total_padre)
                if total_padre is not None else Decimal('100.00'),
                'monto_corriente': recurso.monto_corriente,
                'porcentaje_corriente': recurso.porcentaje_corriente,
                'monto_inversion': recurso.monto_inversion,
                'porcentaje_inversion': recurso.porcentaje_inversion,
                'orden': recurso.orden,
            }

        rubros = (
            version.recursos.filter(padre__isnull=True)
            .select_related('fuente', 'organismo')
            .prefetch_related('componentes__fuente', 'componentes__organismo')
            .order_by('orden', 'concepto')
        )

        filas, total, total_corriente, total_inversion = [], Decimal('0'), Decimal('0'), Decimal('0')
        for rubro in rubros:
            datos = fila(rubro)
            datos['componentes'] = [
                fila(c, rubro.monto)
                for c in sorted(rubro.componentes.all(), key=lambda c: (c.orden, c.concepto))
            ]
            filas.append(datos)
            total += rubro.monto
            total_corriente += rubro.monto_corriente or Decimal('0')
            total_inversion += rubro.monto_inversion or Decimal('0')

        # El resumen por FF/OF se calcula: varios rubros pueden compartir el
        # mismo par (41/113 aparece en Coparticipacion y en sus saldos) y hay
        # que verlos sumados, no fila por fila.
        por_par: dict[str, dict] = {}
        for rubro in rubros:
            for r in [rubro, *rubro.componentes.all()]:
                if not (r.fuente and r.organismo):
                    continue
                clave = f'{r.fuente.codigo}/{r.organismo.codigo}'
                acumulado = por_par.setdefault(clave, {
                    'ff_of': clave,
                    'fuente': r.fuente.denominacion,
                    'organismo': r.organismo.denominacion,
                    'monto': Decimal('0'),
                })
                # Solo los rubros suman: los componentes ya estan dentro del
                # total de su grupo y contarlos otra vez duplicaria.
                if r.padre_id is None:
                    acumulado['monto'] += r.monto

        resumen = sorted(por_par.values(), key=lambda x: x['ff_of'])
        for linea in resumen:
            linea['porcentaje'] = porcentaje(linea['monto'], total)

        return Response({
            'gestion': techo.gestion.anio,
            'estado': techo.estado,
            'por_fuente': resumen,
            'version_id': version.id,
            'version': version.numero if hasattr(version, 'numero') else None,
            'editable': techo.estado in (EstadosTecho.BORRADOR, EstadosTecho.OBSERVADO),
            'rubros': filas,
            'total': {
                'monto': total,
                'porcentaje': Decimal('100.00') if total else None,
                'monto_corriente': total_corriente,
                'porcentaje_corriente': porcentaje(total_corriente, total),
                'monto_inversion': total_inversion,
                'porcentaje_inversion': porcentaje(total_inversion, total),
            },
        })

    @action(detail=True, methods=['post'], url_path='submit')
    def submit(self, request, pk=None):
        """BORRADOR|OBSERVADO → EN_REVISION."""
        return self._ejecutar_servicio(request, pk, enviar_a_revision)

    @action(detail=True, methods=['post'], url_path='observe')
    def observe(self, request, pk=None):
        """EN_REVISION → OBSERVADO. Body: {'observaciones': 'motivo'}."""
        motivo = (
            request.data.get('observaciones') or request.data.get('motivo') or ''
        )
        if not motivo.strip():
            return Response(
                {'error': {'detail': ['Debe indicar el motivo de la observación.']}},
                status=400,
            )
        return self._ejecutar_servicio(request, pk, observar, motivo)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """EN_REVISION → APROBADO."""
        return self._ejecutar_servicio(request, pk, aprobar)

    @action(detail=True, methods=['post'], url_path='freeze')
    def freeze(self, request, pk=None):
        """APROBADO → FIJADO (valida §24, congela con checksum)."""
        observaciones = request.data.get('observaciones') or ''
        return self._ejecutar_servicio(
            request, pk, fijar_techo, observaciones,
        )


@extend_schema(
    responses={200: OpenApiTypes.OBJECT},
    description='Composición del techo directivo (§22): montos por origen, '
                'obligatorios, techo bruto y distribuible, y por fuente.',
)
class CompositionView(APIView):
    """GET /directive-ceilings/{id}/composition/ → composición del techo."""

    def get(self, request, pk):
        ceiling = get_object_or_404(TechoDirectivo, pk=pk)
        return Response(_serializar_montos(composicion_techo(ceiling)))


class _VersionMutableMixin:
    """Rechaza create/update/delete sobre versiones fijadas (409)."""

    def _rechazo_inmutable(self, version):
        if version is None or not version.inmutable:
            return None
        return Response(ERROR_409_INMUTABLE, status=409)

    def _version_desde_datos(self, request):
        version_id = request.data.get('version')
        if not version_id:
            return None
        return TechoVersion.objects.filter(pk=version_id).first()

    def create(self, request, *args, **kwargs):
        rechazo = self._rechazo_inmutable(self._version_desde_datos(request))
        if rechazo:
            return rechazo
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        rechazo = self._rechazo_inmutable(self.get_object().version)
        if rechazo:
            return rechazo
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        rechazo = self._rechazo_inmutable(self.get_object().version)
        if rechazo:
            return rechazo
        return super().destroy(request, *args, **kwargs)


class RecursoTechoViewSet(_VersionMutableMixin, viewsets.ModelViewSet):
    queryset = RecursoTecho.objects.select_related(
        'version', 'rubro', 'fuente', 'organismo', 'entidad_otorgante',
        'documento',
    ).all()
    serializer_class = RecursoTechoSerializer
    filterset_fields = ['version', 'origen', 'fuente']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()


class GastoObligatorioViewSet(_VersionMutableMixin, viewsets.ModelViewSet):
    queryset = GastoObligatorio.objects.select_related(
        'version', 'da', 'ue', 'fuente', 'organismo', 'objeto_gasto',
        'documento',
    ).all()
    serializer_class = GastoObligatorioSerializer
    filterset_fields = ['version', 'fuente', 'programa']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()


class DocumentoPresupuestarioViewSet(viewsets.ModelViewSet):
    queryset = DocumentoPresupuestario.objects.select_related('gestion').all()
    serializer_class = DocumentoPresupuestarioSerializer
    filterset_fields = ['gestion', 'tipo']
    search_fields = ['nombre']

    def get_permissions(self):
        if self.action in ('create', 'destroy'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()

    def perform_create(self, serializer):
        serializer.save()


# ---------------------------------------------------------------------------
# Fase 3 - CategorAas programAticas + catAAlogos para formularios
# ---------------------------------------------------------------------------
class CategoriaProgramaticaTechoViewSet(viewsets.ModelViewSet):
    """CRUD de categorAas programAticas del ciclo (por gestiA3n)."""

    queryset = CategoriaProgramaticaTecho.objects.select_related('parent').all()
    serializer_class = CategoriaProgramaticaTechoSerializer
    http_method_names = ['get', 'post', 'patch', 'delete']

    def get_queryset(self):
        qs = super().get_queryset()
        gestion = self.request.query_params.get('gestion')
        if gestion:
            qs = qs.filter(gestion_id=gestion)
        nivel = self.request.query_params.get('nivel')
        if nivel:
            qs = qs.filter(nivel=nivel)
        return qs

    def perform_create(self, serializer):
        gestion = serializer.validated_data.get('gestion')
        if gestion and not gestion_habilitada(gestion):
            raise ValidationError(
                'No se pueden crear categorAas para una gestiA3n no habilitada.'
            )
        serializer.save()

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """AArbol de categorAas por gestiA3n (parametro ?gestion= obligatorio)."""
        gestion = request.query_params.get('gestion')
        if not gestion:
            return Response(
                {'error': 'El parAametro ?gestion= es obligatorio.'},
                status=400,
            )
        categorias = CategoriaProgramaticaTecho.objects.filter(
            gestion_id=gestion, parent__isnull=True,
        ).order_by('nivel', 'codigo')

        def _nodo(cat):
            return {
                'id': str(cat.id),
                'codigo': cat.codigo,
                'denominacion': cat.denominacion,
                'nivel': cat.nivel,
                'estado': cat.estado,
                'hijos': [_nodo(h) for h in
                          cat.hijos.order_by('nivel', 'codigo')],
            }

        return Response([_nodo(c) for c in categorias])

    @action(detail=True, methods=['post'])
    def duplicar_a_gestion(self, request, pk=None):
        """Copia la categorAa (y su subArbol) a otra gestiA3n."""
        destino = request.data.get('gestion_destino')
        if not destino:
            return Response({'error': 'gestion_destino es obligatorio.'}, status=400)
        try:
            destino_obj = GestionFiscal.objects.get(pk=destino)
        except GestionFiscal.DoesNotExist:
            return Response({'error': 'GestiA3n destino no existe.'}, status=400)
        origen = self.get_object()
        copias = {}
        for cat in [origen, *origen.hijos.order_by('nivel', 'codigo')]:
            nuevo = CategoriaProgramaticaTecho.objects.create(
                gestion=destino_obj,
                codigo=cat.codigo,
                denominacion=cat.denominacion,
                nivel=cat.nivel,
                parent=copias.get(cat.parent_id),
                vigencia_desde=cat.vigencia_desde,
                vigencia_hasta=cat.vigencia_hasta,
                estado=cat.estado,
                origen=cat.origen or 'duplicado',
                normativa=cat.normativa,
            )
            copias[cat.id] = nuevo
        return Response({'detail': f'CategorAa y {len(copias)-1} hijos duplicados.'}, status=201)


class CatalogOptionsView(APIView):
    """CAtAlogos corporativos para poblar selects del ciclo presupuestario.

    GET /api/v2/sis-poa/budget/catalogs/ -> {fuentes, organismos, rubros,
    objetos_gasto, distritos, da, ue, unidades}
    """

    def get(self, request):
        from apps.catalogos.models import (
            EntidadTransferencia,
            FuenteFinanciamiento,
            ObjetoGasto,
            OrganismoFinanciador,
            RubroRecurso,
        )
        from apps.organizacion.models import (
            DireccionAdministrativa,
            UnidadEjecutora,
            UnidadOrganizacional,
        )
        from apps.territorio.models import Distrito

        gestion = request.query_params.get('gestion')

        def _opts(qs, codigo='codigo', nombre='denominacion'):
            return [
                {'id': str(o.pk), 'codigo': getattr(o, codigo), 'nombre': getattr(o, nombre)}
                for o in qs[:500]
            ]

        data = {
            'fuentes': _opts(FuenteFinanciamiento.objects.all()),
            'organismos': _opts(OrganismoFinanciador.objects.all()),
            'rubros': _opts(RubroRecurso.objects.all()),
            'objetos_gasto': _opts(ObjetoGasto.objects.all()),
            'entidades_transferencia': _opts(EntidadTransferencia.objects.all()),
            'distritos': _opts(Distrito.objects.all(), codigo='codigo', nombre='nombre'),
            'direcciones': _opts(DireccionAdministrativa.objects.all(), nombre='nombre'),
            'unidades_ejecutoras': _opts(UnidadEjecutora.objects.all(), nombre='nombre'),
            'unidades_organizacionales': _opts(
                UnidadOrganizacional.objects.all(), nombre='nombre'),
        }
        return Response(data)


# ---------------------------------------------------------------------------
# Fase 4 - Distribución presupuestaria (versiones, aperturas, reservas)
# ---------------------------------------------------------------------------

def _respuesta_exceso(exc):
    """400 BUDGET_EXCEEDED: {error: {detail}, code, details}."""
    return Response(
        {
            'error': {'detail': exc.messages},
            'code': exc.code,
            'details': exc.details,
        },
        status=400,
    )


def _respuesta_exceso_409(exc):
    """409 BUDGET_EXCEEDED (§91): {error: {detail}, code, details}."""
    return Response(
        {
            'error': {'detail': exc.messages},
            'code': exc.code,
            'details': exc.details,
        },
        status=409,
    )


class PresupuestoGastosViewSet(viewsets.ViewSet):
    """Presupuesto General de Gastos: el arbol Programa → Subprograma → Actividad.

    Reproduce la hoja `gastos` de la planilla oficial: cada actividad lleva su
    unidad ejecutora, direccion administrativa y distrito, con el monto abierto
    por fuente de financiamiento. Los totales de subprograma y programa se
    calculan aqui; en la planilla son formulas que hay que mantener a mano.
    """

    @staticmethod
    def _aportes_de_priorizacion(anio):
        """Qué actas aportaron a cada fila de gasto, por apertura.

        Import local a propósito: `priorizacion` importa `budget`, y hacerlo al
        revés a nivel de módulo cierra el círculo.
        """
        from apps.priorizacion.models import ProyectoPriorizado

        proyectos = (
            ProyectoPriorizado.objects
            .filter(acta__gestion=int(anio))
            .exclude(apertura_fuente__isnull=True)
            .select_related('acta__distrito', 'apertura_fuente',
                            'fuente', 'organismo')
        )
        por_apertura: dict = {}
        for p in proyectos:
            por_apertura.setdefault(
                p.apertura_fuente.allocation_id, []).append({
                    'acta': str(p.acta_id),
                    'otb': p.acta.otb,
                    'distrito': p.acta.distrito.nombre if p.acta.distrito else '',
                    'estado_acta': p.acta.estado,
                    'proyecto': p.nombre,
                    'par': p.par_financiamiento,
                    'monto': float(p.monto_materializado or 0),
                })
        return por_apertura

    def list(self, request):
        anio = request.query_params.get('gestion')
        if not anio:
            return Response(
                {'error': 'Indique la gestión: ?gestion=2027'}, status=400)

        aperturas = (
            Apertura.objects
            .filter(gestion__anio=int(anio))
            .select_related('categoria', 'categoria__parent', 'da', 'ue', 'distrito')
            .prefetch_related('fuentes__fuente', 'fuentes__organismo')
            .order_by('categoria__codigo', 'actividad_codigo')
        )

        # De dónde salió cada monto: sin esto, lo que aporta un acta de
        # priorización se funde dentro de una fila que ya existía y no hay
        # forma de ver que se adjuntó.
        aportes = self._aportes_de_priorizacion(anio)

        # Las fuentes viven como filas; la planilla las lee como columnas.
        pares: dict[str, str] = {}

        def montos_por_fuente(apertura):
            fila = {}
            for f in apertura.fuentes.all():
                if not (f.fuente and f.organismo):
                    continue
                clave = f'{f.fuente.codigo}/{f.organismo.codigo}'
                pares.setdefault(clave, f.organismo.denominacion)
                fila[clave] = fila.get(clave, Decimal('0')) + f.monto
            return fila

        # Arbol por programa y subprograma, deducido de la categoria.
        programas: dict[str, dict] = {}
        for apertura in aperturas:
            categoria = apertura.categoria
            if categoria is None:
                continue
            subprograma = categoria if categoria.nivel == NivelCategoria.SUBPROGRAMA \
                else categoria.parent
            programa = subprograma.parent if subprograma else None
            cod_prog = programa.codigo if programa else 'SIN PROGRAMA'
            cod_sub = subprograma.codigo if subprograma else 'SIN SUBPROGRAMA'

            rango = programa.rango_directriz if programa else None
            p_nodo = programas.setdefault(cod_prog, {
                'codigo': cod_prog,
                'denominacion': programa.denominacion if programa else 'Sin programa',
                'rango': rango,
                'subprogramas': {},
            })
            s_nodo = p_nodo['subprogramas'].setdefault(cod_sub, {
                'codigo': cod_sub,
                'denominacion': subprograma.denominacion if subprograma else 'Sin subprograma',
                'actividades': [],
            })
            s_nodo['actividades'].append({
                'id': apertura.id,
                'categoria': categoria.codigo,
                'denominacion': apertura.denominacion,
                'unidad_ejecutora': apertura.ue.codigo if apertura.ue else '',
                'direccion_administrativa': apertura.da.codigo if apertura.da else '',
                'distrito': apertura.distrito.codigo if apertura.distrito else '',
                'codigo_sisin': apertura.codigo_sisin,
                'actividad': apertura.actividad_codigo,
                'estado_revision': apertura.estado_revision,
                'observacion': apertura.observacion,
                'montos': montos_por_fuente(apertura),
                'priorizaciones': aportes.get(apertura.id, []),
                'monto_priorizado': sum(
                    a['monto'] for a in aportes.get(apertura.id, [])),
            })

        def sumar(filas):
            total: dict[str, Decimal] = {}
            for f in filas:
                for clave, monto in f['montos'].items():
                    total[clave] = total.get(clave, Decimal('0')) + monto
            return total

        # Orden secuencial por categoría programática, explícito y en los tres
        # niveles: depender del orden de inserción hace que baste con que una
        # consulta cambie para que la lista salga desordenada.
        def por_codigo(nodo):
            return nodo['codigo']

        salida, total_general = [], {}
        for p_nodo in sorted(programas.values(), key=por_codigo):
            subs = []
            for s_nodo in sorted(p_nodo['subprogramas'].values(), key=por_codigo):
                s_nodo['actividades'].sort(key=lambda a: a['categoria'])
                s_nodo['total'] = sumar(s_nodo['actividades'])
                subs.append(s_nodo)
            p_nodo['subprogramas'] = subs
            p_nodo['total'] = sumar(
                [{'montos': s['total']} for s in subs])
            for clave, monto in p_nodo['total'].items():
                total_general[clave] = total_general.get(clave, Decimal('0')) + monto
            salida.append(p_nodo)

        # El gasto se agrupa por el rango del Anexo VI, que es como lo lee el
        # Ministerio: la materia (`170-179` infraestructura urbana y rural)
        # antes que el programa suelto.
        rangos: dict[str, dict] = {}
        for p_nodo in salida:
            rango = p_nodo.pop('rango', None)
            clave = rango.codigo if rango else 'SIN RANGO'
            r_nodo = rangos.setdefault(clave, {
                'codigo': clave,
                'denominacion': (rango.denominacion if rango
                                 else 'Sin rango en la directriz'),
                'finalidad_funcion': rango.finalidad_funcion if rango else '',
                'sector_economico': rango.sector_economico if rango else '',
                'desde': rango.desde if rango else 10 ** 6,
                'programas': [],
            })
            r_nodo['programas'].append(p_nodo)

        arbol_rangos = sorted(rangos.values(), key=lambda r: r['desde'])
        for r_nodo in arbol_rangos:
            r_nodo.pop('desde')
            r_nodo['total'] = sumar(
                [{'montos': p['total']} for p in r_nodo['programas']])

        # Techo por FF/OF: sale del Presupuesto General de Recursos. Es el
        # limite contra el que se contrasta el gasto mientras se elabora el
        # POA, y la diferencia es lo que queda por asignar de cada fuente.
        techos: dict[str, Decimal] = {}
        techo = TechoDirectivo.objects.filter(gestion__anio=int(anio)).first()
        if techo:
            version_techo = TechoVersion.objects.filter(
                ceiling=techo, numero=techo.version_actual).first()
            if version_techo:
                for recurso in version_techo.recursos.select_related(
                        'fuente', 'organismo').filter(padre__isnull=True):
                    if not (recurso.fuente and recurso.organismo):
                        continue
                    clave = f'{recurso.fuente.codigo}/{recurso.organismo.codigo}'
                    techos[clave] = techos.get(clave, Decimal('0')) + recurso.monto
                    pares.setdefault(clave, recurso.organismo.denominacion)

        diferencia = {
            clave: techos.get(clave, Decimal('0')) - total_general.get(clave, Decimal('0'))
            for clave in set(techos) | set(total_general)
        }

        return Response({
            # El gasto se entrega agrupado por rango de la directriz. `programas`
            # queda como forma plana para lo que todavía no migró.
            'rangos': arbol_rangos,
            'gestion': int(anio),
            'gestion_id': GestionFiscal.objects.filter(anio=int(anio)).values_list('id', flat=True).first(),
            'techos': techos,
            'diferencia': diferencia,
            'columnas': [
                {'ff_of': k, 'denominacion': v} for k, v in sorted(pares.items())
            ],
            'programas': salida,
            'total': total_general,
        })


class DistribucionVersionViewSet(viewsets.ModelViewSet):
    """Versiones de distribución: CRUD liviano + ciclo de fijación (Fase 7).

    Acciones del ciclo (§51):
        POST .../{id}/submit/   → EN_REVISION
        POST .../{id}/observe/  → OBSERVADO  (body: observaciones|motivo)
        POST .../{id}/approve/  → APROBADO
        POST .../{id}/freeze/   → FIJADO (valida Σfuente = techo; congela)
        GET  .../{id}/validate/ → validar_distribucion_completa (diferencias)
        POST .../{id}/ajuste/   → versión siguiente (BORRADOR, contenedor)

    Permisos (ADR-003): submit/observe/approve/freeze → `sis_poa.budget.
    approve`; create/update/destroy/ajuste → `sis_poa.budget.manage`;
    validate → IsAuthenticated (default global).
    """

    queryset = DistribucionVersion.objects.select_related(
        'gestion', 'fijado_por',
    ).all()
    serializer_class = DistribucionVersionSerializer
    filterset_fields = ['gestion', 'estado', 'inmutable']
    search_fields = ['gestion__anio']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy',
                           'ajuste'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        if self.action in ('submit', 'observe', 'approve', 'freeze'):
            return [TieneCapacidad(CAPACIDAD_APROBACION)]
        return super().get_permissions()

    def perform_create(self, serializer):
        request = self.context.get('request')
        usuario = request.user if request and request.user.is_authenticated else None
        serializer.save(created_by=usuario, updated_by=usuario)

    def destroy(self, request, *args, **kwargs):
        version = self.get_object()
        if version.inmutable:
            return Response(ERROR_409_INMUTABLE, status=409)
        return super().destroy(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        version = self.get_object()
        if version.inmutable:
            return Response(ERROR_409_INMUTABLE, status=409)
        return super().update(request, *args, **kwargs)

    def _ejecutar(self, request, pk, servicio, *args, **kwargs):
        """Ejecuta un servicio de transición sobre la versión (400 si no)."""
        version = self.get_object()
        try:
            resultado = servicio(version, request.user, *args, **kwargs)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(resultado).data)

    @action(detail=False, methods=['get'], url_path='versions')
    def versions(self, request):
        """Lista las versiones de distribución por gestión (?gestion=)."""
        gestion = request.query_params.get('gestion')
        if not gestion:
            return Response(
                {'error': {'detail': ['El parámetro ?gestion= es obligatorio.']}},
                status=400,
            )
        qs = self.get_queryset().filter(gestion_id=gestion).order_by('-numero')
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='submit')
    def submit(self, request, pk=None):
        """BORRADOR|OBSERVADO → EN_REVISION."""
        return self._ejecutar(request, pk, enviar_distribucion_a_revision)

    @action(detail=True, methods=['post'], url_path='observe')
    def observe(self, request, pk=None):
        """EN_REVISION → OBSERVADO. Body: {'observaciones': 'motivo'}."""
        motivo = (
            request.data.get('observaciones') or request.data.get('motivo') or ''
        )
        if not motivo.strip():
            return Response(
                {'error': {'detail': ['Debe indicar el motivo de la observación.']}},
                status=400,
            )
        return self._ejecutar(request, pk, observar_distribucion, motivo)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """EN_REVISION → APROBADO."""
        return self._ejecutar(request, pk, aprobar_distribucion)

    @action(detail=True, methods=['post'], url_path='freeze')
    def freeze(self, request, pk=None):
        """APROBADO → FIJADO (valida Σfuente y congela con checksum)."""
        observaciones = request.data.get('observaciones') or ''
        return self._ejecutar(request, pk, fijar_distribucion, observaciones)

    @action(detail=True, methods=['get'], url_path='validate')
    def validate(self, request, pk=None):
        """Validación §49-52: {valida, diferencias[{fuente, techo, ...}]}."""
        version = self.get_object()
        return Response(
            _serializar_montos(validar_distribucion_completa(version.gestion))
        )

    @action(detail=True, methods=['post'], url_path='ajuste')
    def ajuste(self, request, pk=None):
        """Crea la versión siguiente (BORRADOR) desde la fijada."""
        return self._ejecutar(request, pk, ajuste_distribucion)


class AperturaViewSet(viewsets.ModelViewSet):
    """Aperturas programáticas: CRUD con fuentes anidadas + cerrar.

    create/update aceptan `fuentes`: [{fuente, organismo, monto}]. Errores
    de disponibilidad → 400 {error: {detail}, code: 'BUDGET_EXCEEDED',
    details: {requested, available, difference}}.
    """

    queryset = Apertura.objects.select_related(
        'gestion', 'version', 'unidad_organizacional', 'distrito',
        'da', 'ue', 'categoria',
    ).prefetch_related('fuentes__fuente', 'fuentes__organismo').all()
    serializer_class = AperturaSerializer
    filterset_fields = ['gestion', 'version', 'distrito', 'categoria', 'estado']
    search_fields = ['denominacion', 'codigo_sisin', 'proyecto_codigo',
                     'actividad_codigo']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy',
                           'cerrar'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()

    def _validar_y_servicio(self, request, serializer, allocation=None):
        """Valida el serializer y ejecuta el servicio de dominio mapeando
        errores: BUDGET_EXCEEDED → 400 con code/details; resto → 400 genérico."""
        gestion = serializer.validated_data.get('gestion')
        if gestion is None and allocation is not None:
            gestion = allocation.gestion
        if gestion is None:
            return Response(
                {'error': {'detail': ['La gestión es obligatoria.']}},
                status=400,
            )
        datos = {
            k: v for k, v in serializer.validated_data.items() if k != 'gestion'
        }
        try:
            if allocation is None:
                resultado = crear_allocation(gestion, request.user, datos)
            else:
                resultado = actualizar_allocation(
                    allocation, request.user, datos,
                )
        except ErrorDisponibilidad as exc:
            return _respuesta_exceso(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(
            self.get_serializer(resultado).data,
            status=201 if allocation is None else 200,
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return self._validar_y_servicio(request, serializer)

    def update(self, request, *args, **kwargs):
        allocation = self.get_object()
        partial = kwargs.get('partial', False)
        serializer = self.get_serializer(
            allocation, data=request.data, partial=partial,
        )
        serializer.is_valid(raise_exception=True)
        return self._validar_y_servicio(request, serializer, allocation)

    def destroy(self, request, *args, **kwargs):
        allocation = self.get_object()
        try:
            eliminar_allocation(allocation, request.user)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(status=204)

    @action(detail=True, methods=['post'], url_path='cerrar')
    def cerrar(self, request, pk=None):
        """Cierra la apertura (solo si no excede el disponible)."""
        allocation = self.get_object()
        try:
            cerrar_allocation(allocation, request.user)
        except ErrorDisponibilidad as exc:
            return _respuesta_exceso(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(allocation).data)

    # --- Revisión por categoría programática ---------------------------------
    #
    # Cada apertura se valida y aprueba por separado: las unidades presentan su
    # gasto en momentos distintos y la jefatura no espera a tenerlas todas.

    def _transicion(self, request, destino, exige_estado=None, con_motivo=False):
        apertura = self.get_object()
        if exige_estado and apertura.estado_revision != exige_estado:
            return Response(
                {'error': {'detail': [
                    f'La categoría está en {apertura.get_estado_revision_display()}: '
                    f'no admite esta acción.']}},
                status=400,
            )
        if apertura.estado_revision == RevisionApertura.APROBADO and not con_motivo:
            return Response(
                {'error': {'detail': ['Una categoría aprobada ya no se modifica.']}},
                status=400,
            )

        campos = ['estado_revision']
        apertura.estado_revision = destino
        ahora = timezone.now()
        if destino == RevisionApertura.VALIDADO:
            apertura.validado_por, apertura.validado_en = request.user, ahora
            campos += ['validado_por', 'validado_en']
        elif destino == RevisionApertura.APROBADO:
            apertura.aprobado_por, apertura.aprobado_en = request.user, ahora
            campos += ['aprobado_por', 'aprobado_en']
        elif destino == RevisionApertura.OBSERVADO:
            motivo = (request.data.get('observacion') or '').strip()
            if not motivo:
                return Response(
                    {'error': {'detail': ['Detalle qué debe corregirse.']}}, status=400)
            apertura.observacion = motivo
            apertura.observado_por, apertura.observado_en = request.user, ahora
            campos += ['observacion', 'observado_por', 'observado_en']
        apertura.save(update_fields=campos)
        return Response(self.get_serializer(apertura).data)

    @action(detail=True, methods=['post'])
    def validar(self, request, pk=None):
        """BORRADOR|OBSERVADO → VALIDADO."""
        return self._transicion(request, RevisionApertura.VALIDADO)

    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """VALIDADO → APROBADO. No se aprueba lo que nadie validó."""
        return self._transicion(
            request, RevisionApertura.APROBADO,
            exige_estado=RevisionApertura.VALIDADO)

    @action(detail=True, methods=['post'])
    def observar(self, request, pk=None):
        """Devuelve la categoría a la unidad con el motivo."""
        return self._transicion(
            request, RevisionApertura.OBSERVADO, con_motivo=True)


class ReservaViewSet(viewsets.ModelViewSet):
    """Reservas presupuestarias: CRUD + acción liberar."""

    queryset = Reserva.objects.select_related(
        'gestion', 'version', 'fuente', 'organismo',
    ).all()
    serializer_class = ReservaSerializer
    filterset_fields = ['gestion', 'version', 'estado', 'tipo', 'fuente']
    search_fields = ['motivo']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy',
                           'liberar'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()

    def update(self, request, *args, **kwargs):
        reserva = self.get_object()
        if reserva.version is not None and reserva.version.inmutable:
            return Response(ERROR_409_INMUTABLE, status=409)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        reserva = self.get_object()
        if reserva.version is not None and reserva.version.inmutable:
            return Response(ERROR_409_INMUTABLE, status=409)
        monto = reserva.monto
        gestion_anio = reserva.gestion.anio
        reserva.delete()
        registrar_evento(
            request.user,
            EventoAuditoria.Accion.ANULAR,
            'Reserva',
            reserva.id,
            resumen=(
                f'Reserva {reserva.get_tipo_display()} de {monto} eliminada '
                f'(gestión {gestion_anio})'
            ),
            datos_previos={
                'monto': str(monto), 'tipo': reserva.tipo,
                'estado': reserva.estado,
            },
            gestion=gestion_anio,
        )
        return Response(status=204)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        gestion = serializer.validated_data.get('gestion')
        if gestion is None:
            return Response(
                {'error': {'detail': ['La gestión es obligatoria.']}},
                status=400,
            )
        datos = {
            k: v for k, v in serializer.validated_data.items() if k != 'gestion'
        }
        try:
            reserva = crear_reserva(gestion, request.user, datos)
        except ErrorDisponibilidad as exc:
            return _respuesta_exceso(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(reserva).data, status=201)

    @action(detail=True, methods=['post'], url_path='liberar')
    def liberar(self, request, pk=None):
        """Libera la reserva (devuelve el disponible a la fuente)."""
        reserva = self.get_object()
        try:
            liberar_reserva(reserva, request.user)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(reserva).data)


# ---------------------------------------------------------------------------
# Fase 9 - Objetos del gasto (programación por apertura, §90-91)
# ---------------------------------------------------------------------------
class ExpenseObjectViewSet(viewsets.ModelViewSet):
    """Objetos del gasto programados por apertura (Fase 9, §90-91).

    create:  {allocation, objeto_gasto, monto} → `programar_objeto_gasto`
             (UPSERT: si la fila existe la actualiza). Requiere la versión
             de distribución FIJADA y monto <= disponible de la apertura;
             si excede → HTTP 409 {error: {detail}, code:
             'BUDGET_EXCEEDED', details: {requested, available,
             difference}} (§91).
    update / partial_update: {monto} → `actualizar_objeto_gasto`
             (recalcula contra los demás objetos, excluyendo la fila).
    destroy: `eliminar_objeto_gasto` (libera el disponible).

    Permisos (ADR-003): escritura → `sis_poa.budget.manage`; filtros
    ?allocation=.
    """

    queryset = AsignacionObjetoGastoTecho.objects.select_related(
        'allocation', 'objeto_gasto',
    ).all()
    serializer_class = AsignacionObjetoGastoTechoSerializer
    filterset_fields = ['allocation']
    search_fields = ['objeto_gasto__codigo', 'objeto_gasto__denominacion']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            fila = programar_objeto_gasto(
                serializer.validated_data['allocation'],
                serializer.validated_data['objeto_gasto'],
                serializer.validated_data['monto'],
                request.user,
            )
        except ErrorObjetoGastoExcedido as exc:
            return _respuesta_exceso_409(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(fila).data, status=201)

    def update(self, request, *args, **kwargs):
        fila = self.get_object()
        serializer = self.get_serializer(fila, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        monto = serializer.validated_data.get('monto')
        if monto is None:
            return Response(
                {'error': {'detail': ['El monto es obligatorio.']}},
                status=400,
            )
        try:
            actualizar_objeto_gasto(fila, monto, request.user)
        except ErrorObjetoGastoExcedido as exc:
            return _respuesta_exceso_409(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        fila.refresh_from_db()
        return Response(self.get_serializer(fila).data)

    def destroy(self, request, *args, **kwargs):
        fila = self.get_object()
        try:
            eliminar_objeto_gasto(fila, request.user)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(status=204)


@extend_schema(
    responses={200: OpenApiTypes.OBJECT},
    description='Resumen de la distribución (§48): cards + tabla por fuente.',
)
class DistributionDashboardView(APIView):
    """GET /distributions/dashboard/?gestion= → resumen_distribucion."""

    def get(self, request):
        gestion = request.query_params.get('gestion')
        if not gestion:
            return Response(
                {'error': {'detail': ['El parámetro ?gestion= es obligatorio.']}},
                status=400,
            )
        gestion_obj = get_object_or_404(GestionFiscal, pk=gestion)
        return Response(_serializar_montos(resumen_distribucion(gestion_obj)))


# ---------------------------------------------------------------------------
# Fase 8 - Control presupuestario central (BudgetControlService + endpoint)
# ---------------------------------------------------------------------------
from .control import BudgetControlService  # noqa: E402


@extend_schema(
    responses={200: OpenApiTypes.OBJECT},
    description='Control presupuestario central (Fase 8): '
                'GET → resumen consolidado por fuente; '
                'POST → validación pedida (distribution | expense-object | '
                'allocation) con {valido, errores}.',
)
class BudgetControlView(APIView):
    """Control presupuestario central (`BudgetControlService`, Fase 8).

    GET  /control/summary/?gestion= → get_summary (techo_bruto,
    techo_distribuible, distribuido, reservado, disponible, porcentaje,
    por_fuente).
    POST /control/validate/         → body {tipo, ...}: ejecuta la
    validación pedida y devuelve {valido, errores}:
        - distribution:   {gestion} → validate_distribution (diferencias).
        - expense-object: {allocation, objeto_gasto, monto} → Fase 9:
          apertura ACTIVA + versión de distribución FIJADA + objeto del
          gasto existente + monto <= disponible (BUDGET_EXCEEDED).
        - allocation:     {allocation} → exista y ACTIVA + {techo,
          programado, disponible} de la apertura.
    Permisos: IsAuthenticated (default global; lectura para cualquier
    usuario autenticado, como los demás endpoints de consulta).
    """

    def _gestion_obligatoria(self, request):
        gestion = request.query_params.get('gestion')
        if not gestion:
            return None, Response(
                {'error': {'detail': ['El parámetro ?gestion= es obligatorio.']}},
                status=400,
            )
        return get_object_or_404(GestionFiscal, pk=gestion), None

    def _allocation_desde(self, request):
        allocation = request.data.get('allocation')
        if not allocation:
            raise DjangoValidationError(
                'Debe indicar la apertura (allocation).'
            )
        return Apertura.objects.filter(pk=allocation).first()

    def get(self, request):
        gestion, error = self._gestion_obligatoria(request)
        if error:
            return error
        return Response(
            _serializar_montos(BudgetControlService.get_summary(gestion))
        )

    def post(self, request):
        tipo = request.data.get('tipo')
        if tipo not in ('distribution', 'expense-object', 'allocation'):
            return Response(
                {'error': {'detail': [
                    'tipo debe ser distribution, expense-object o allocation.',
                ]}},
                status=400,
            )
        try:
            if tipo == 'distribution':
                gestion = request.data.get('gestion')
                if not gestion:
                    raise DjangoValidationError(
                        'Debe indicar la gestión (gestion).'
                    )
                gestion_obj = get_object_or_404(GestionFiscal, pk=gestion)
                resultado = BudgetControlService.validate_distribution(
                    gestion_obj,
                )
                return Response({
                    'valido': resultado['valida'],
                    'errores': _serializar_montos(resultado['diferencias']),
                })
            if tipo == 'expense-object':
                allocation = self._allocation_desde(request)
                BudgetControlService.validate_expense_object(
                    allocation,
                    request.data.get('objeto_gasto')
                    or request.data.get('objeto_gasto_id'),
                    request.data.get('monto'),
                )
                return Response({'valido': True, 'errores': []})
            if tipo == 'allocation':
                allocation = self._allocation_desde(request)
                BudgetControlService.validate_expense_object(
                    allocation, None, None,
                )
                return Response({
                    'valido': True,
                    'errores': [],
                    'techo': str(
                        BudgetControlService.get_allocation_ceiling(allocation)
                    ),
                    'programado': str(
                        BudgetControlService
                        .get_allocated_to_expense_objects(allocation)
                    ),
                    'disponible': str(
                        BudgetControlService.get_allocation_available(allocation)
                    ),
                })
        except DjangoValidationError as exc:
            return Response(
                {'valido': False, 'errores': exc.messages},
                status=400,
            )
        return Response({'valido': False, 'errores': ['Validación no soportada.']})


# ---------------------------------------------------------------------------
# Fase 5 - Importador Excel (staging + validación + aplicación)
# ---------------------------------------------------------------------------
from .importer import (  # noqa: E402
    aplicar_importacion,
    parsear_libro,
    validar_importacion,
)
from .models import Importacion, ImportacionError  # noqa: E402
from .serializers import (  # noqa: E402
    ImportacionSerializer,
    ImportErrorSerializer,
)

CAPACIDAD_IMPORTACION = 'sis_poa.budget.import'


class ImportacionViewSet(viewsets.ModelViewSet):
    """Importaciones de planillas GASTOS (wizard: upload → map → validate → apply).

    Permisos: create/apply/map/validate → `sis_poa.budget.import`; el resto
    (listar/ver/hojas/errores) usa IsAuthenticated (default global).
    """

    queryset = Importacion.objects.select_related('gestion').all()
    serializer_class = ImportacionSerializer
    pagination_class = ImportacionDualPagination
    http_method_names = ['get', 'post']
    filterset_fields = ['gestion', 'estado', 'perfil']
    search_fields = ['filename']

    def get_permissions(self):
        if self.action in ('create', 'map', 'validate', 'apply'):
            return [TieneCapacidad(CAPACIDAD_IMPORTACION)]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        gestion = self.request.query_params.get('gestion')
        if gestion:
            qs = qs.filter(gestion_id=gestion)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        importacion = serializer.save()
        try:
            self._parsear(importacion, request.user)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        except Exception:
            importacion.estado = 'RECHAZADO'
            importacion.save(update_fields=['estado', 'updated_at'])
            return Response(
                {'error': {'detail': [
                    'No se pudo leer el archivo como planilla Excel/CSV válida.',
                ]}},
                status=400,
            )
        return Response(self.get_serializer(importacion).data, status=201)

    @staticmethod
    def _parsear(importacion, usuario=None):
        """Carga el libro con openpyxl y construye los ImportacionDetalle."""
        import openpyxl
        ruta = importacion.archivo.path
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        try:
            parsear_libro(importacion, wb, usuario=usuario)
        finally:
            wb.close()

    @action(detail=True, methods=['get'], url_path='hojas')
    def hojas(self, request, pk=None):
        """Lista las hojas del libro de la importación."""
        import openpyxl
        importacion = self.get_object()
        try:
            wb = openpyxl.load_workbook(
                importacion.archivo.path, read_only=True,
            )
        except Exception:
            return Response(
                {'error': {'detail': ['No se pudo leer el archivo de la importación.']}},
                status=400,
            )
        try:
            return Response({'hojas': wb.sheetnames})
        finally:
            wb.close()

    @action(detail=True, methods=['post'], url_path='map')
    def map(self, request, pk=None):
        """Configura hoja + mapeo (columnas y fuentes) y re-parsea."""
        importacion = self.get_object()
        hoja = request.data.get('hoja') or importacion.hoja_seleccionada
        mapeo = request.data.get('mapeo') or {}
        if not hoja:
            return Response(
                {'error': {'detail': ['Debe indicar la hoja a mapear.']}},
                status=400,
            )
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                importacion.archivo.path, read_only=True, data_only=True,
            )
            try:
                parsear_libro(
                    importacion, wb, hoja=hoja, mapeo=mapeo,
                    usuario=request.user,
                )
            finally:
                wb.close()
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(importacion).data)

    @action(detail=True, methods=['post'], url_path='validate')
    def validate(self, request, pk=None):
        """Ejecuta la validación (severidades) y actualiza el estado."""
        importacion = self.get_object()
        validar_importacion(importacion, request.user)
        return Response(self.get_serializer(importacion).data)

    @action(detail=True, methods=['post'], url_path='apply')
    def apply(self, request, pk=None):
        """Aplica la importación: crea aperturas BORRADOR (400 si hay CRITICAL)."""
        importacion = self.get_object()
        try:
            resultado = aplicar_importacion(importacion, request.user)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response({
            **self.get_serializer(importacion).data,
            'resultado': {
                'aperturas_creadas': resultado['aperturas_creadas'],
                'total_importado': str(resultado['total_importado']),
            },
        })

    @action(detail=True, methods=['get'], url_path='errors')
    def errors(self, request, pk=None):
        """Lista los errores/hallazgos de la validación (con severidad)."""
        importacion = self.get_object()
        qs = (
            ImportacionError.objects
            .filter(importacion=importacion)
            .select_related('detalle')
            .order_by('severidad', 'fila')
        )
        severidad = request.query_params.get('severidad')
        if severidad:
            qs = qs.filter(severidad=severidad)
        return Response(ImportErrorSerializer(qs, many=True).data)


# ---------------------------------------------------------------------------
# Fase 6 - Distribución territorial (reparto por distrito → reservas DISTRITALES)
# ---------------------------------------------------------------------------
from .territorial import (  # noqa: E402
    aplicar_reparto,
    calcular_reparto,
    liberar_reparto,
    recalcular_reparto,
)
from .services import validar_gestion_para_distribucion  # noqa: E402


class DistribucionTerritorialViewSet(viewsets.ModelViewSet):
    """Distribuciones territoriales: CRUD + calcular/aplicar/liberar.

    create acepta `distritos`: [{distrito, poblacion?, porcentaje?, monto?}].
    POST .../calcular/  → calcula el reparto (body opcional {distritos} para
    recalcular con datos actualizados; rechaza APLICADA).
    POST .../aplicar/   → materializa reservas DISTRITALES (BUDGET_EXCEEDED
    si la bolsa excede el disponible de la fuente; rollback total).
    POST .../liberar/   → solo APLICADA: libera las reservas y vuelve a
    CALCULADA.
    Escritura (incluidas las acciones) → capacidad `sis_poa.budget.manage`.
    """

    queryset = DistribucionTerritorial.objects.select_related(
        'gestion', 'version', 'fuente', 'organismo',
    ).prefetch_related('asignaciones__distrito').all()
    serializer_class = DistribucionTerritorialSerializer
    filterset_fields = ['gestion', 'estado', 'metodo', 'fuente']
    search_fields = ['observaciones', 'gestion__anio']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy',
                           'calcular', 'aplicar', 'liberar'):
            return [TieneCapacidad(CAPACIDAD_GESTION)]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        gestion = self.request.query_params.get('gestion')
        if gestion:
            qs = qs.filter(gestion_id=gestion)
        return qs

    def _ejecutar(self, servicio, distribucion, *args):
        """Ejecuta el servicio sobre la distribución y serializa la respuesta."""
        try:
            servicio(distribucion, *args)
        except ErrorDisponibilidad as exc:
            return _respuesta_exceso(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        self._descartar_cache_asignaciones(distribucion)
        return Response(self.get_serializer(distribucion).data)

    @staticmethod
    def _descartar_cache_asignaciones(distribucion):
        """Descarta el prefetch de `asignaciones` tras mutar (re-lectura)."""
        cache = getattr(distribucion, '_prefetched_objects_cache', None)
        if cache:
            cache.pop('asignaciones', None)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validos = serializer.validated_data
        gestion = validos.get('gestion')
        try:
            validar_gestion_para_distribucion(gestion)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)

        usuario = request.user
        distribucion = DistribucionTerritorial.objects.create(
            gestion=gestion,
            version=validos.get('version'),
            fuente=validos.get('fuente'),
            organismo=validos.get('organismo'),
            metodo=validos.get('metodo') or 'MANUAL',
            bolsa_total=validos.get('bolsa_total'),
            observaciones=validos.get('observaciones') or '',
            created_by=usuario,
            updated_by=usuario,
        )
        distritos = validos.get('distritos')
        if distritos:
            for fila in distritos:
                AsignacionTerritorial.objects.create(
                    distribucion=distribucion,
                    distrito_id=fila['distrito'],
                    poblacion=fila.get('poblacion'),
                    porcentaje=fila.get('porcentaje'),
                    monto_calculado=fila.get('monto') or 0,
                    created_by=usuario,
                    updated_by=usuario,
                )
        registrar_evento(
            usuario,
            EventoAuditoria.Accion.CREAR,
            'DistribucionTerritorial',
            distribucion.id,
            resumen=(
                f'Distribución territorial creada '
                f'({distribucion.get_metodo_display()}, bolsa '
                f'{distribucion.bolsa_total}) — gestión {gestion.anio}'
            ),
            datos_posteriores={
                'metodo': distribucion.metodo,
                'bolsa_total': str(distribucion.bolsa_total),
                'distritos': len(distritos or []),
                'estado': distribucion.estado,
            },
            gestion=gestion.anio,
        )
        return Response(
            self.get_serializer(distribucion).data, status=201,
        )

    def update(self, request, *args, **kwargs):
        distribucion = self.get_object()
        if distribucion.estado == 'APLICADA':
            return Response(
                {'error': {'detail': [
                    'No se puede modificar una distribución territorial '
                    'aplicada.',
                ]}},
                status=400,
            )
        return super().update(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='calcular')
    def calcular(self, request, pk=None):
        """Calcula el reparto; body opcional {distritos} para recalcular."""
        distribucion = self.get_object()
        distritos = request.data.get('distritos')
        if distritos is not None:
            try:
                recalcular_reparto(distribucion, distritos, request.user)
            except ErrorDisponibilidad as exc:
                return _respuesta_exceso(exc)
            except DjangoValidationError as exc:
                return _respuesta_error(exc)
            self._descartar_cache_asignaciones(distribucion)
            return Response(self.get_serializer(distribucion).data)
        return self._ejecutar(calcular_reparto, distribucion, request.user)

    @action(detail=True, methods=['post'], url_path='aplicar')
    def aplicar(self, request, pk=None):
        """Materializa el reparto como reservas DISTRITALES."""
        return self._ejecutar(aplicar_reparto, self.get_object(), request.user)

    @action(detail=True, methods=['post'], url_path='liberar')
    def liberar(self, request, pk=None):
        """Libera las reservas DISTRITALES (solo APLICADA → CALCULADA)."""
        return self._ejecutar(liberar_reparto, self.get_object(), request.user)


# ---------------------------------------------------------------------------
# Fase 10 - Reformulaciones (tipos + workflow + movimientos atómicos, §92-97)
# ---------------------------------------------------------------------------
CAPACIDAD_REFORM = 'sis_poa.budget.reform'


class ReformaViewSet(viewsets.ModelViewSet):
    """Reformulaciones presupuestarias (Fase 10, §92-97).

    CRUD: create acepta `movimientos_input`: [{tipo, apertura_origen?,
    apertura_destino?, fuente?, organismo?, monto, motivo?}] y delega en
    `services.crear_reform` (BORRADOR). update/patch/destroy solo en
    BORRADOR (el documento en flujo es inmodificable).

    Acciones del workflow:
        POST .../submit/  → EN_REVISION
        POST .../observe/ → OBSERVADA (body: observaciones|motivo)
        POST .../approve/ → APROBADA
        POST .../reject/  → RECHAZADA (body: motivo, obligatorio)
        POST .../apply/   → APLICADA (ATÓMICO: mueve saldos en una
                            transacción; BUDGET_EXCEEDED → 400 {code,
                            details}; fallo → rollback completo)

    Permisos (ADR-003): create/update/destroy/submit → `sis_poa.budget.
    reform`; observe/approve/reject/apply → `sis_poa.budget.approve`.
    """

    queryset = Reforma.objects.select_related(
        'gestion', 'documento', 'version_origen', 'version_resultante',
        'solicitada_por', 'aprobada_por',
    ).prefetch_related(
        'movimientos__apertura_origen', 'movimientos__apertura_destino',
        'movimientos__fuente', 'movimientos__organismo',
    ).all()
    serializer_class = ReformaSerializer
    filterset_fields = ['gestion', 'estado', 'tipo']
    search_fields = ['motivo', 'resolucion']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy',
                           'submit'):
            return [TieneCapacidad(CAPACIDAD_REFORM)]
        if self.action in ('observe', 'approve', 'reject', 'apply'):
            return [TieneCapacidad(CAPACIDAD_APROBACION)]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        gestion = self.request.query_params.get('gestion')
        if gestion:
            qs = qs.filter(gestion_id=gestion)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            reform = crear_reform(
                gestion=serializer.validated_data['gestion'],
                tipo=serializer.validated_data['tipo'],
                motivo=serializer.validated_data.get('motivo') or '',
                usuario=request.user,
                movimientos=serializer.validated_data.get('movimientos')
                or [],
            )
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        return Response(self.get_serializer(reform).data, status=201)

    def update(self, request, *args, **kwargs):
        reform = self.get_object()
        if reform.estado != 'BORRADOR':
            return Response(
                {'error': {'detail': [
                    'Solo se puede editar una reformulación en BORRADOR.',
                ]}},
                status=400,
            )
        datos_previos = {
            'estado': reform.estado,
            'tipo': reform.tipo,
            'motivo': reform.motivo,
        }
        respuesta = super().update(request, *args, **kwargs)
        if respuesta.status_code == 200:
            reform.refresh_from_db()
            registrar_evento(
                request.user,
                EventoAuditoria.Accion.MODIFICAR,
                'Reforma',
                reform.id,
                resumen=(
                    f'Reformulación {reform.get_tipo_display()} modificada '
                    f'(gestión {reform.gestion.anio})'
                ),
                datos_previos=datos_previos,
                datos_posteriores={
                    'estado': reform.estado,
                    'tipo': reform.tipo,
                    'motivo': reform.motivo,
                },
                gestion=reform.gestion.anio,
            )
        return respuesta

    def destroy(self, request, *args, **kwargs):
        reform = self.get_object()
        if reform.estado != 'BORRADOR':
            return Response(
                {'error': {'detail': [
                    'Solo se puede eliminar una reformulación en BORRADOR.',
                ]}},
                status=400,
            )
        reform_id = reform.id
        tipo = reform.tipo
        gestion_anio = reform.gestion.anio
        reform.delete()
        registrar_evento(
            request.user,
            EventoAuditoria.Accion.ANULAR,
            'Reforma',
            reform_id,
            resumen=(
                f'Reformulación {tipo} eliminada (gestión {gestion_anio})'
            ),
            datos_previos={'estado': 'BORRADOR', 'tipo': tipo},
            gestion=gestion_anio,
        )
        return Response(status=204)

    def _ejecutar(self, request, pk, servicio, *args, **kwargs):
        """Ejecuta un servicio de transición/aplicación y serializa."""
        reform = self.get_object()
        try:
            resultado = servicio(reform, request.user, *args, **kwargs)
        except ErrorDisponibilidad as exc:
            return _respuesta_exceso(exc)
        except DjangoValidationError as exc:
            return _respuesta_error(exc)
        # Refresca movimientos (saldos registrados al aplicar).
        cache = getattr(reform, '_prefetched_objects_cache', None)
        if cache:
            cache.pop('movimientos', None)
        return Response(self.get_serializer(resultado).data)

    @action(detail=True, methods=['post'], url_path='submit')
    def submit(self, request, pk=None):
        """BORRADOR|OBSERVADA → EN_REVISION."""
        return self._ejecutar(request, pk, enviar_reform_a_revision)

    @action(detail=True, methods=['post'], url_path='observe')
    def observe(self, request, pk=None):
        """EN_REVISION → OBSERVADA. Body: {'observaciones': 'motivo'}."""
        motivo = (
            request.data.get('observaciones') or request.data.get('motivo') or ''
        )
        if not motivo.strip():
            return Response(
                {'error': {'detail': [
                    'Debe indicar el motivo de la observación.',
                ]}},
                status=400,
            )
        return self._ejecutar(request, pk, observar_reform, motivo)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """EN_REVISION → APROBADA."""
        return self._ejecutar(request, pk, aprobar_reform)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """EN_REVISION → RECHAZADA. Body: {'motivo': '...'} (obligatorio)."""
        motivo = request.data.get('motivo') or ''
        if not motivo.strip():
            return Response(
                {'error': {'detail': ['Debe indicar el motivo del rechazo.']}},
                status=400,
            )
        return self._ejecutar(request, pk, rechazar_reform, motivo)

    @action(detail=True, methods=['post'], url_path='apply')
    def apply(self, request, pk=None):
        """APROBADA → APLICADA: movimientos atómicos con saldos (§97)."""
        return self._ejecutar(request, pk, aplicar_reform)


# ---------------------------------------------------------------------------
# Fase 11 - Auditoría de trazabilidad (consulta de EventoAuditoria del ciclo)
# ---------------------------------------------------------------------------

# Entidades del ciclo auditables por slug de la API (`?entidad=`) → nombre de
# modelo con el que `EventoAuditoria.entidad` identifica los registros.
ENTIDADES_AUDITORIA = {
    'allocation': 'Apertura',
    'reserve': 'Reserva',
    'directive-ceiling': 'TechoVersion',
    'distribution': 'DistribucionVersion',
    'expense-object': 'AsignacionObjetoGastoTecho',
    'reform': 'Reforma',
    'import': 'Importacion',
    'territorial': 'DistribucionTerritorial',
    'fiscal-year': 'GestionFiscal',
}


class TieneCapacidadAuditoria(TieneCapacidad):
    """`TieneCapacidad` sin argumentos para `permission_classes` (DRF
    instancia las clases): exige `sis_poa.budget.audit_read`."""

    def __init__(self):
        super().__init__(CAPACIDAD_AUDITORIA)


@extend_schema(
    responses={200: AuditEventSerializer(many=True)},
    description='Registro de auditoría del ciclo presupuestario (Fase 11): '
                'filtros ?gestion=&entidad=&registro_id=&usuario=&accion='
                '&desde=&hasta=, paginado (DRF). Capacidad '
                'sis_poa.budget.audit_read.',
)
class AuditLogView(APIView):
    """GET /budget/audit/ → EventoAuditoria del ciclo, paginado y filtrable.

    `entidad` acepta los slugs de `ENTIDADES_AUDITORIA` (allocation, reserve,
    directive-ceiling, distribution, expense-object, reform, import,
    territorial, fiscal-year) o el nombre de modelo directo; `accion` acepta
    los códigos del catálogo (crear/modificar/anular/...) o las acciones
    semánticas de `services.ACCIONES_AUDITORIA` (CREATE/UPDATE/DELETE/...).
    """

    permission_classes = [TieneCapacidadAuditoria]

    def get(self, request):
        from .services import ACCIONES_AUDITORIA

        qs = EventoAuditoria.objects.select_related('usuario')

        gestion = request.query_params.get('gestion')
        if gestion:
            # PIP-DB-008: ?gestion=<año> filtra por anio de GestionFiscal;
            # también acepta el UUID de la gestión.
            if gestion.isdigit():
                qs = qs.filter(gestion__anio=gestion)
            else:
                qs = qs.filter(gestion_id=gestion)
        entidad = request.query_params.get('entidad')
        if entidad:
            qs = qs.filter(entidad=ENTIDADES_AUDITORIA.get(entidad, entidad))
        registro_id = request.query_params.get('registro_id')
        if registro_id:
            qs = qs.filter(entidad_id=str(registro_id))
        usuario = request.query_params.get('usuario')
        if usuario:
            qs = qs.filter(usuario_id=usuario)
        accion = request.query_params.get('accion')
        if accion:
            qs = qs.filter(accion=ACCIONES_AUDITORIA.get(accion, accion))
        desde = request.query_params.get('desde')
        if desde:
            qs = qs.filter(creado_en__date__gte=desde)
        hasta = request.query_params.get('hasta')
        if hasta:
            qs = qs.filter(creado_en__date__lte=hasta)

        qs = qs.order_by('-creado_en')
        paginator = AuditoriaDualPagination()
        pagina = paginator.paginate_queryset(qs, request)
        return paginator.get_paginated_response(
            AuditEventSerializer(pagina, many=True).data
        )
