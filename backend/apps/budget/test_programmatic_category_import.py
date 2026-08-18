import json
from io import StringIO

import openpyxl
import pytest
from django.core import management
from django.core.management.base import CommandError

from apps.budget.importer_programmatic_category import (
    WorkbookValidationError,
    parse_workbook,
)
from apps.budget.models import ProgrammaticCategory
from apps.gestion.models import GestionFiscal


HEADERS = ["Código", "PROG.", "PROYEC.", "ACTIV.", "CAT. PROGRAMATICA", "Descripción", "Nivel"]


def make_workbook(path, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CLASIFICADOR"
    for column, value in enumerate(HEADERS, 1):
        sheet.cell(8, column, value)
    for row_number, values in enumerate(rows, 9):
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
    workbook.save(path)


def sample_rows():
    return [
        ["0.0.0", "000", "0", "000", "000 0 000", "Funcionamiento", "PROGRAMA"],
        ["0.0.1", "000", "0", "001", "000 0 001", "Alcaldía", "ACTIVIDAD"],
        ["97.0.0", "097", "0", "000", "097 0 000", "Salud", "PROGRAMA"],
        ["200.0.0", "200", "0", "000", "200 0 000", "Coronavirus", "PROGRAMA"],
        ["200.0.150", "200", "0", "0150", "200 0 0150", "Actividad especial", "ACTIVIDAD"],
    ]


def test_parser_preserves_leading_zeroes_and_accepts_0150(tmp_path):
    path = tmp_path / "master.xlsx"
    make_workbook(path, sample_rows())

    rows = parse_workbook(path)

    assert rows[1].codigo == "000 0 001"
    assert rows[-1].activ == "0150"
    assert rows[-1].codigo == "200 0 0150"


@pytest.mark.parametrize("change", ["duplicate", "unknown_level", "missing_parent"])
def test_parser_rejects_invalid_master(tmp_path, change):
    rows = sample_rows()
    if change == "duplicate":
        rows.append(["x", "000", "0", "001", "000 0 001", "Duplicada", "ACTIVIDAD"])
    elif change == "unknown_level":
        rows[0][-1] = "SUBPROGRAMA"
    else:
        rows[-1][1] = "351"
        rows[-1][4] = "351 0 0150"
    path = tmp_path / f"{change}.xlsx"
    make_workbook(path, rows)

    with pytest.raises(WorkbookValidationError):
        parse_workbook(path)


def test_parser_rejects_missing_sheet_and_malformed_row(tmp_path):
    path = tmp_path / "invalid.xlsx"
    make_workbook(path, sample_rows())
    workbook = openpyxl.load_workbook(path)
    workbook["CLASIFICADOR"]["F10"] = None
    workbook.save(path)
    with pytest.raises(WorkbookValidationError):
        parse_workbook(path)

    with pytest.raises(WorkbookValidationError):
        parse_workbook(path, sheet_name="MISSING")
    with pytest.raises(FileNotFoundError):
        parse_workbook(tmp_path / "absent.xlsx")


@pytest.mark.django_db
def test_dry_run_has_no_writes_and_commit_is_idempotent(tmp_path):
    path = tmp_path / "master.xlsx"
    make_workbook(path, sample_rows())
    gestion, _ = GestionFiscal.objects.get_or_create(anio=2027)
    ProgrammaticCategory.objects.filter(gestion=gestion).delete()
    historical_gestion, _ = GestionFiscal.objects.get_or_create(anio=2026)
    historical = ProgrammaticCategory.objects.create(
        gestion=historical_gestion,
        codigo="000 0 000",
        denominacion="Historical program",
        nivel="PROGRAMA",
    )
    preserved = ProgrammaticCategory.objects.create(
        gestion=gestion,
        codigo="999 0 000",
        denominacion="Existing outside source",
        nivel="PROGRAMA",
    )

    output = StringIO()
    management.call_command(
        "importar_catalogo_programatico_2027", file=path, dry_run=True, stdout=output,
    )
    assert ProgrammaticCategory.objects.filter(gestion=gestion).count() == 1
    assert ProgrammaticCategory.objects.filter(pk=preserved.pk).exists()
    dry_report = json.loads(output.getvalue())
    assert dry_report["status"] == "dry-run"
    assert dry_report["existing_rows_preserved"] == 1

    output = StringIO()
    management.call_command("importar_catalogo_programatico_2027", file=path, commit=True, stdout=output)
    assert ProgrammaticCategory.objects.filter(gestion=gestion).count() == 6
    first = list(ProgrammaticCategory.objects.filter(gestion=gestion).values_list("codigo", "updated_at"))
    assert json.loads(output.getvalue())["counts"]["created"] == 5

    output = StringIO()
    management.call_command("importar_catalogo_programatico_2027", file=path, commit=True, stdout=output)
    assert ProgrammaticCategory.objects.filter(gestion=gestion).count() == 6
    second = list(ProgrammaticCategory.objects.filter(gestion=gestion).values_list("codigo", "updated_at"))
    assert [code for code, _ in first] == [code for code, _ in second]
    assert json.loads(output.getvalue())["counts"]["unchanged"] == 5
    assert ProgrammaticCategory.objects.get(gestion=gestion, codigo="000 0 001").parent.codigo == "000 0 000"
    assert ProgrammaticCategory.objects.get(gestion=gestion, codigo="200 0 0150").observaciones.find("ACTIV.=0150") >= 0
    assert ProgrammaticCategory.objects.filter(pk=preserved.pk).exists()
    assert ProgrammaticCategory.objects.get(pk=historical.pk).denominacion == "Historical program"


@pytest.mark.django_db
def test_missing_gestion_fails_closed(tmp_path):
    path = tmp_path / "master.xlsx"
    make_workbook(path, sample_rows())
    GestionFiscal.objects.filter(anio=2027).delete()
    with pytest.raises(CommandError):
        management.call_command("importar_catalogo_programatico_2027", file=path)
