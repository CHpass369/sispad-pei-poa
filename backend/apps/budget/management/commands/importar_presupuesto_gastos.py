"""Importa la hoja `gastos` de la planilla de distribución de prioridades.

    python manage.py importar_presupuesto_gastos --archivo ruta.xlsx --gestion 2027

La planilla es un árbol: filas de Programa (P), de Subprograma (SP), de
actividad, y de totales (T/TS) que aquí NO se importan porque se calculan.
Las fuentes son columnas fijas en el Excel y aquí pasan a ser filas de
AperturaFuente, que es lo que permite agregar una fuente nueva sin rehacer
la estructura.
"""
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.budget.models import (
    Apertura, AperturaFuente, CategoriaProgramaticaTecho, EstadoCategoria,
    NivelCategoria, RevisionApertura,
)
from apps.catalogos.models import FuenteFinanciamiento, OrganismoFinanciador
from apps.gestion.models import GestionFiscal

# Columna del Excel -> par (fuente, organismo) del clasificador.
COLUMNAS = {
    11: ('41', '113'),   # CT   Coparticipación Tributaria
    12: ('20', '210'),   # RE   Recursos Específicos
    13: ('20', '230'),   # ORE  Otros Recursos Específicos
    14: ('41', '119'),   # IDH  Impuestos Directos a Hidrocarburos
    15: ('41', '111'),   # TGN  Tesoro General de la Nación
}
FILA_INICIAL = 9


class Command(BaseCommand):
    help = 'Importa el Presupuesto General de Gastos desde la planilla oficial.'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', required=True)
        parser.add_argument('--gestion', type=int, default=2027)
        parser.add_argument('--dry-run', action='store_true')

    @transaction.atomic
    def handle(self, *args, **opciones):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Falta openpyxl para leer la planilla.')

        try:
            gestion = GestionFiscal.objects.get(anio=opciones['gestion'])
        except GestionFiscal.DoesNotExist:
            raise CommandError(f'No existe la gestión {opciones["gestion"]}.')

        hoja = openpyxl.load_workbook(opciones['archivo'], data_only=True)['gastos']

        fuentes = {f.codigo: f for f in FuenteFinanciamiento.objects.all()}
        organismos = {o.codigo: o for o in OrganismoFinanciador.objects.all()}

        def texto(fila, col):
            v = hoja.cell(fila, col).value
            return '' if v in (None, '') else str(v).strip()

        seco = opciones['dry_run']

        def categoria(codigo, denominacion, nivel, padre=None):
            if seco:
                # En seco NO se escribe: se devuelve un marcador para que el
                # recorrido siga y solo se informe el conteo.
                vistas.add(codigo)
                return None
            obj, _ = CategoriaProgramaticaTecho.objects.get_or_create(
                gestion=gestion, codigo=codigo,
                defaults={
                    'denominacion': denominacion[:300] or codigo,
                    'nivel': nivel, 'parent': padre,
                    'estado': EstadoCategoria.ACTIVA,
                },
            )
            return obj

        programa = subprograma = None
        creadas = montos = 0
        vistas: set[str] = set()

        for i in range(FILA_INICIAL, hoja.max_row + 1):
            col_a, col_e = texto(i, 1), texto(i, 5)
            prog, act = texto(i, 6), texto(i, 8)
            denominacion = texto(i, 9)

            if not any(texto(i, c) for c in range(1, 16)):
                continue
            # Los totales son formula: se recalculan, no se importan.
            if col_a == 'T' or col_e == 'TS' or 'TOTAL' in denominacion.upper():
                continue

            if col_a == 'P' or col_e == 'P':
                programa = categoria(prog[:20], denominacion, NivelCategoria.PROGRAMA)
                subprograma = None
                continue
            if col_e == 'SP':
                subprograma = categoria(
                    f'{prog}.SP'[:20], denominacion, NivelCategoria.SUBPROGRAMA, programa)
                continue

            if not act:
                continue

            # El campo admite 20 caracteres y la planilla trae rangos
            # ("100-109") y denominaciones pegadas al codigo.
            codigo_cat = f'{prog} {texto(i, 7) or "0"} {act}'[:20]
            cat = categoria(
                codigo_cat, denominacion, NivelCategoria.ACTIVIDAD,
                subprograma or programa)

            if opciones['dry_run']:
                creadas += 1
                continue

            apertura = Apertura.objects.create(
                gestion=gestion, categoria=cat,
                denominacion=denominacion[:300] or codigo_cat,
                codigo_sisin=texto(i, 7), actividad_codigo=act,
                estado_revision=RevisionApertura.BORRADOR,
            )
            creadas += 1

            for columna, (ff, of) in COLUMNAS.items():
                valor = hoja.cell(i, columna).value
                if not isinstance(valor, (int, float)) or not valor:
                    continue
                AperturaFuente.objects.create(
                    allocation=apertura, fuente=fuentes.get(ff),
                    organismo=organismos.get(of), monto=Decimal(str(round(valor, 2))),
                )
                montos += 1

        etiqueta = '[dry-run] ' if opciones['dry_run'] else ''
        self.stdout.write(self.style.SUCCESS(
            f'{etiqueta}Gestión {gestion.anio}: {creadas} actividades, '
            f'{montos} montos por fuente, '
            f'{len(vistas) if seco else CategoriaProgramaticaTecho.objects.filter(gestion=gestion).count()}'
            f' categorías.'
        ))
