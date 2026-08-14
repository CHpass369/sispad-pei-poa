import io
import logging
from celery import shared_task

logger = logging.getLogger(__name__)


@shared_task(bind=True, soft_time_limit=300, time_limit=600)
def generar_reporte_presupuestario_async(self, params: dict):
    """Genera un reporte presupuestario de forma asíncrona"""
    from apps.reportes.services import (
        generar_poa_consolidado_xlsx,
        generar_poa_unidad_xlsx,
    )
    try:
        gestion = params.get('gestion')
        tipo = params.get('tipo', 'consolidado')

        if tipo == 'consolidado':
            output, filename = generar_poa_consolidado_xlsx(gestion)
        elif tipo == 'unidad':
            unidad_id = params.get('unidad_id')
            output, filename = generar_poa_unidad_xlsx(gestion, unidad_id)
        else:
            return {"status": "error", "error": f"Tipo de reporte no soportado: {tipo}"}

        logger.info(f"Reporte {tipo} generado: {filename}")
        return {
            "status": "ok",
            "filename": filename,
            "gestion": gestion,
            "tipo": tipo,
            "tamanio_bytes": output.tell(),
        }
    except Exception as e:
        logger.error(f"Error generando reporte: {e}")
        raise self.retry(exc=e, countdown=60)


def _generar_lineas_presupuestarias_xlsx(gestion):
    """Genera Excel de líneas presupuestarias."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    from apps.presupuesto.models import LineaPresupuestaria

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Líneas {gestion}'

    header_fill = PatternFill(start_color='1B5E3B', end_color='1B5E3B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)

    ws.merge_cells('A1:H1')
    ws['A1'] = f'LÍNEAS PRESUPUESTARIAS - GESTIÓN {gestion}'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    headers = ['Programa', 'Proyecto', 'Actividad', 'Objeto Gasto',
               'Fuente', 'Importe', 'Importe G.Anterior', 'Plurianual']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    lineas = LineaPresupuestaria.objects.filter(
        gestion=gestion, activo=True
    ).select_related('programa', 'proyecto', 'actividad', 'objeto_gasto', 'fuente')

    row = 4
    for lp in lineas:
        datos = [
            str(lp.programa),
            str(lp.proyecto) if lp.proyecto else '',
            str(lp.actividad) if lp.actividad else '',
            str(lp.objeto_gasto),
            str(lp.fuente),
            float(lp.importe),
            float(lp.importe_gestion_anterior) if lp.importe_gestion_anterior else 0,
            float(lp.importe_plurianual) if lp.importe_plurianual else 0,
        ]
        for col, val in enumerate(datos, 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, f'lineas_presupuestarias_{gestion}.xlsx'
