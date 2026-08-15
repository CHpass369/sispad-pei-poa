import os
from io import BytesIO
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

# Adaptado a main: el seed legacy de la rama (scripts.seed) no existe en main;
# los usuarios y datos demo se crean en el propio test.
TEST_DEMO_PASSWORDS = {
    'PIP_DEMO_ADMIN_PASSWORD': 'test-only-admin-credential',
    'PIP_DEMO_AUDITOR_PASSWORD': 'test-only-auditor-credential',
}


@patch.dict(os.environ, TEST_DEMO_PASSWORDS, clear=False)
class MatrizCompletaXlsxIntegrationTest(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_superuser(
            email='admin@demo.pip.local', password='real-password'
        )
        self.auditor = user_model.objects.create_user(
            email='auditor@demo.pip.local', password='real-password'
        )
        # Seed mínimo: un plan con un nodo para que la matriz tenga filas.
        from datetime import date
        from apps.planificacion.models import Plan, NodoPlanificacion

        self.plan = Plan.objects.create(
            codigo='PEI-DEMO', tipo='pei', nombre='PEI demo',
            gestion_inicio=2026, gestion_fin=2030,
            fecha_vigencia_desde=date(2026, 1, 1),
        )
        NodoPlanificacion.objects.create(
            plan=self.plan, nivel='accion_mediano', codigo='AMP-1',
            nombre='Acción demo', gestion=2026,
        )

    def test_authenticated_endpoint_returns_readable_workbook(self):
        client = APIClient()
        client.force_authenticate(user=self.admin)

        response = client.get(
            '/api/v1/reportes/matriz_completa_xlsx/',
            {'gestion': 2026},
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook['Matriz Articulación Completa']
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=4, max_row=4))]
        self.assertEqual(
            headers,
            [
                'Código Completo',
                'Nivel',
                'Nombre',
                'Plan',
                'Código Padre',
                'Nodos Vinculados',
            ],
        )
        self.assertGreater(worksheet.max_row, 4)
        workbook.close()

        client.force_authenticate(user=self.auditor)
        denied_response = client.get(
            '/api/v1/reportes/matriz_completa_xlsx/',
            {'gestion': 2026},
            HTTP_HOST='localhost',
        )
        self.assertEqual(denied_response.status_code, 403)
