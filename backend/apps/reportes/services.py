"""
Servicios de generación de reportes del POA.
Soporta PDF (reportlab), XLSX (openpyxl), CSV y GeoJSON.
"""
import csv, io
from decimal import Decimal
from datetime import datetime
from django.db.models import Sum, Count, DecimalField, IntegerField
from django.db.models.functions import Coalesce

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from apps.presupuesto.models import ProgramaPresupuestario, LineaPresupuestaria
from apps.planificacion.models import AccionCortoPlazo
from apps.techos.models import DistribucionTecho
from apps.techos.services import budget_service
from apps.inversion.models import ProyectoInversion
from apps.workflow.models import Observacion, Aprobacion
from apps.articulacion.models import (
    ArticulacionPADPEI, ProductoPAD, ProductoPEI, ResultadoPAD, ResultadoPEI,
    AccionPOA, OperacionPOAU, ActividadPOAU, TareaPOAU,
    AsignacionObjetoGasto, SeguimientoPresupuesto,
)


HEADER_FILL = PatternFill(start_color='1B5E3B', end_color='1B5E3B', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _build_response_filename(tipo: str, formato: str, gestion: int) -> str:
    return f'poa_{tipo}_{gestion}_{datetime.now().strftime("%Y%m%d")}.{formato}'


# ===== REPORTE POA POR UNIDAD =====


def generar_poa_unidad_xlsx(gestion: int, unidad_id: str = None) -> tuple:
    """Genera XLSX del POA por unidad organizacional."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'POA Unidad {gestion}'

    # Encabezado
    ws.merge_cells('A1:H1')
    ws['A1'] = f'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA - POA {gestion}'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    # Headers
    headers = ['Código', 'Acción', 'Indicador', 'Meta', 'Unidad', 'Operaciones', 'Presupuesto (Bs)', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    acciones = AccionCortoPlazo.objects.filter(gestion=gestion)
    if unidad_id:
        acciones = acciones.filter(unidad_responsable_id=unidad_id)

    row = 4
    for acp in acciones:
        datos = [
            acp.codigo, acp.nombre, '',
            '', '',
            acp.operaciones.count(),
            0,
            'Formulado' if acp.activo else 'Inactivo',
        ]
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER_THIN
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    filename = _build_response_filename('unidad', 'xlsx', gestion)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== REPORTE POA CONSOLIDADO =====


def generar_poa_consolidado_xlsx(gestion: int) -> tuple:
    """Genera XLSX del POA institucional consolidado."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Consolidado {gestion}'

    ws.merge_cells('A1:G1')
    ws['A1'] = f'POA CONSOLIDADO {gestion} - GAM SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    headers = ['Programa', 'Acciones', 'Presupuesto (Bs)', 'Techo (Bs)', 'Saldo (Bs)', 'Avance %', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    programas = ProgramaPresupuestario.objects.filter(gestion=gestion).annotate(
        total_presupuesto=Coalesce(
            Sum('lineas__importe'), 0,
            output_field=DecimalField(max_digits=20, decimal_places=2)
        ),
        total_acciones=Coalesce(
            Count('proyectos_inversion'), 0,
            output_field=IntegerField()
        ),
    )

    row = 4
    for prog in programas:
        techo_total = budget_service.get_distribuido_por_programa(prog)

        # Decimal en todo el cálculo (nunca float, D5): el saldo y el avance
        # se derivan del motor único de saldos (D11).
        presupuesto = prog.total_presupuesto
        techo = techo_total
        saldo = techo - presupuesto
        avance = (presupuesto / techo * 100) if techo > 0 else 0

        datos = [prog.codigo, prog.total_acciones, round(presupuesto, 2),
                 round(techo, 2), round(saldo, 2), round(avance, 1),
                 'Completo' if saldo >= 0 else 'Sobregiro']
        for col, val in enumerate(datos, 1):
            ws.cell(row=row, column=col, value=val).border = BORDER_THIN
        row += 1

    filename = _build_response_filename('consolidado', 'xlsx', gestion)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== REPORTE DE PROYECTOS DE INVERSIÓN =====


def generar_proyectos_xlsx(gestion: int) -> tuple:
    """Genera XLSX de proyectos de inversión."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Proyectos {gestion}'

    ws.merge_cells('A1:I1')
    ws['A1'] = f'PROYECTOS DE INVERSIÓN {gestion} - GAM SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    headers = ['Código', 'Nombre', 'SISIN', 'Prioridad', 'Etapa', 'Costo Total',
               'Ejecutado', 'Fuente', 'Organismo']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    proyectos = ProyectoInversion.objects.filter(gestion_inicio__lte=gestion) | \
                ProyectoInversion.objects.filter(gestion_fin__gte=gestion)
    proyectos = proyectos.filter(activo=True).distinct()

    row = 4
    for p in proyectos:
        datos = [p.codigo_interno, p.nombre, p.codigo_sisin or '—',
                 p.get_prioridad_display(), p.get_etapa_display(),
                 float(p.costo_total), float(p.ejecucion_acumulada),
                 p.fuente.denominacion if p.fuente else '—',
                 p.organismo.denominacion if p.organismo else '—']
        for col, val in enumerate(datos, 1):
            ws.cell(row=row, column=col, value=val).border = BORDER_THIN
        row += 1

    filename = _build_response_filename('proyectos', 'xlsx', gestion)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== REPORTE DE OBSERVACIONES =====


def generar_observaciones_csv(gestion: int) -> tuple:
    """Genera CSV de observaciones."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Código', 'Tipo', 'Severidad', 'Estado', 'Módulo',
                     'Texto', 'Creado', 'Responsable'])

    observaciones = Observacion.objects.filter(gestion=gestion)
    for obs in observaciones:
        writer.writerow([
            obs.codigo, obs.get_tipo_display(), obs.get_severidad_display(),
            obs.get_estado_display(), obs.modulo, obs.texto,
            obs.created_at.isoformat() if hasattr(obs, 'created_at') else '',
            str(obs.responsable_subsanacion) if obs.responsable_subsanacion else '',
        ])

    filename = _build_response_filename('observaciones', 'csv', gestion)
    return io.BytesIO(output.getvalue().encode('utf-8-sig')), filename



def generar_acta_aprobacion_pdf(gestion: int) -> tuple:
    """Genera PDF de acta de aprobación."""
    if not HAS_REPORTLAB:
        raise RuntimeError('reportlab no está instalado')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=inch, leftMargin=inch,
        topMargin=inch, bottomMargin=inch
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='ActaTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1B5E3B'),
        spaceAfter=20,
    ))

    elements = []
    elements.append(Paragraph(f'ACTA DE APROBACIÓN POA {gestion}', styles['ActaTitle']))
    elements.append(Paragraph(
        f'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA<br/>'
        f'<br/>'
        f'<b>Gestión:</b> {gestion}<br/>'
        f'<b>Fecha de emisión:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}<br/>'
        f'<b>Tipo:</b> Acta de aprobación institucional del POA',
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))

    # Tabla de aprobaciones
    aprobaciones = Aprobacion.objects.filter(gestion=gestion)
    if aprobaciones.exists():
        data = [['Tipo', 'Estado', 'Aprobado por', 'Fecha']]
        for ap in aprobaciones:
            data.append([
                ap.get_tipo_display(),
                ap.get_estado_display(),
                str(ap.aprobado_por) if ap.aprobado_por else '—',
                ap.created_at.strftime('%d/%m/%Y') if hasattr(ap, 'created_at') and ap.created_at else '—',
            ])

        table = Table(data, colWidths=[2*inch, 1*inch, 2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B5E3B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    filename = _build_response_filename('acta_aprobacion', 'pdf', gestion)
    return buffer, filename


# ===== AUXILIAR PLURI =====

GRUPOS_GASTO = {
    '1': 'SERVICIOS PERSONALES',
    '2': 'SERVICIOS NO PERSONALES',
    '3': 'MATERIALES Y SUMINISTROS',
    '4': 'ACTIVOS REALES',
    '5': 'FINANCIEROS',
    '6': 'TRANSFERENCIAS',
    '7': 'IMPUESTOS Y TASAS',
    '8': 'DEUDA PÚBLICA',
    '9': 'OTROS',
}


def _obtener_grupo_gasto(codigo_objeto: str) -> tuple:
    """Deriva el grupo de gasto del primer dígito del código."""
    if not codigo_objeto:
        return ('9', 'OTROS')
    primer_digito = codigo_objeto[0]
    nombre = GRUPOS_GASTO.get(primer_digito, 'OTROS')
    return (primer_digito, nombre)


def generar_auxiliar_pluri_xlsx(gestion: int) -> tuple:
    """Genera XLSX del Auxiliar Pluri (presupuesto plurianual por objeto de gasto y FF/OF)."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Auxiliar Pluri {gestion}'

    # Encabezado institucional
    ws.merge_cells('A1:G1')
    ws['A1'] = f'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    ws.merge_cells('A2:G2')
    ws['A2'] = f'AUXILIAR PLURI - Gestión {gestion} (Expresado en Bolivianos)'
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    # Headers
    headers = ['GRUPO DE GASTO', 'OBJETO DE GASTO', 'DESCRIPCIÓN',
               'FF/OF', 'IMPORTE G. ANTERIOR', 'IMPORTE', 'IMPORTE PLURIANUAL']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Consultar líneas presupuestarias
    lineas = LineaPresupuestaria.objects.filter(
        gestion=gestion, activo=True
    ).select_related(
        'objeto_gasto', 'fuente', 'organismo'
    ).order_by('objeto_gasto__codigo', 'fuente__codigo')

    row = 5
    grupo_actual = None
    subtotal_grupo_ant = 0
    subtotal_grupo_imp = 0
    subtotal_grupo_pluri = 0
    total_gral_ant = 0
    total_gral_imp = 0
    total_gral_pluri = 0

    for lp in lineas:
        grupo_cod, grupo_nom = _obtener_grupo_gasto(lp.objeto_gasto.codigo)

        # Subtotales al cambiar de grupo
        if grupo_actual and grupo_actual != grupo_cod:
            _write_subtotal(ws, row, grupo_actual, subtotal_grupo_ant, subtotal_grupo_imp, subtotal_grupo_pluri)
            row += 1
            subtotal_grupo_ant = 0
            subtotal_grupo_imp = 0
            subtotal_grupo_pluri = 0

        grupo_actual = grupo_cod

        ff_of = f'{lp.fuente.codigo}/{lp.organismo.codigo if lp.organismo else "—"}'
        imp_ant = float(lp.importe_gestion_anterior or 0)
        imp = float(lp.importe or 0)
        imp_pluri = float(lp.importe_plurianual or 0)

        ws.cell(row=row, column=1, value=grupo_nom)
        ws.cell(row=row, column=2, value=lp.objeto_gasto.codigo)
        ws.cell(row=row, column=3, value=lp.objeto_gasto.denominacion)
        ws.cell(row=row, column=4, value=ff_of)
        ws.cell(row=row, column=5, value=round(imp_ant, 2)).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=round(imp, 2)).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=round(imp_pluri, 2)).number_format = '#,##0.00'

        for c in range(1, 8):
            ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=c).font = Font(size=9)

        subtotal_grupo_ant += imp_ant
        subtotal_grupo_imp += imp
        subtotal_grupo_pluri += imp_pluri
        total_gral_ant += imp_ant
        total_gral_imp += imp
        total_gral_pluri += imp_pluri
        row += 1

    # Último subtotal
    if grupo_actual:
        _write_subtotal(ws, row, grupo_actual, subtotal_grupo_ant, subtotal_grupo_imp, subtotal_grupo_pluri)
        row += 2

    # Total general
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value='TOTAL GENERAL')
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color='1B5E3B')
    ws.cell(row=row, column=5, value=round(total_gral_ant, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_gral_imp, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_gral_pluri, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='double'), bottom=Side(style='double')
        )
        ws.cell(row=row, column=c).fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = _build_response_filename('auxiliar_pluri', 'xlsx', gestion)
    return buffer, filename


def _write_subtotal(ws, row, grupo_cod, sub_ant, sub_imp, sub_pluri):
    """Escribe fila de subtotal para un grupo de gasto."""
    grupo_nom = GRUPOS_GASTO.get(str(grupo_cod), 'OTROS')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=f'Subtotal {grupo_cod} - {grupo_nom}')
    ws.cell(row=row, column=1).font = Font(bold=True, italic=True, size=9, color='1B5E3B')
    ws.cell(row=row, column=5, value=round(sub_ant, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(sub_imp, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(sub_pluri, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=c).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')


# ===== EVALUACIÓN — CUADRO N°1 =====


def generar_evaluacion_cuadro1_xlsx(gestion: int) -> tuple:
    """Cuadro N°1: Comparación de programación presupuestaria PTDI vs PEI vs POA por fuente."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Eval Cuadro1 {gestion}'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    ws.merge_cells('A2:F2')
    ws['A2'] = f'CUADRO N°1 — Evaluación de Programación de Recursos (Gestión {gestion})'
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    headers = ['FUENTE DE INGRESOS', 'PTDI/PGTC (Bs)', 'PEI (Bs)',
               'POA (Bs)', 'DIFERENCIA (Bs)', 'DIFERENCIA (%)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Datos de LineaPresupuestaria agrupados por fuente
    from django.db.models import Sum
    lineas = LineaPresupuestaria.objects.filter(gestion=gestion, activo=True)
    fuentes_data = lineas.values('fuente__codigo', 'fuente__denominacion').annotate(
        total_poa=Sum('importe')
    ).order_by('fuente__codigo')

    row = 5
    total_poa = 0
    for f in fuentes_data:
        poa = float(f['total_poa'] or 0)
        # PTDI y PEI son estimativos sin datos en BD actual — se muestran como 0
        ws.cell(row=row, column=1, value=f'{f["fuente__codigo"]} — {f["fuente__denominacion"]}')
        ws.cell(row=row, column=2, value=0).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=0).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=round(poa, 2)).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=round(poa, 2)).number_format = '#,##0.00'  # dif = POA - PTDI
        ws.cell(row=row, column=6, value='100%' if poa > 0 else '0%')
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=c).font = Font(size=9)
        total_poa += poa
        row += 1

    # Total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True, size=10, color='1B5E3B')
    ws.cell(row=row, column=4, value=round(total_poa, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_poa, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=5).font = Font(bold=True)
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                     top=Side(style='double'), bottom=Side(style='double'))

    for col_letter, width in [('A', 40), ('B', 18), ('C', 18), ('D', 18), ('E', 18), ('F', 18)]:
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, _build_response_filename('evaluacion_cuadro1', 'xlsx', gestion)


# ===== EVALUACIÓN — CUADRO N°2 =====


def generar_evaluacion_cuadro2_xlsx(gestion: int) -> tuple:
    """Cuadro N°2: Vinculación de acciones PTDI/PGTC ↔ PEI ↔ POA."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Eval Cuadro2 {gestion}'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')
    ws.merge_cells('A2:F2')
    ws['A2'] = f'CUADRO N°2 — Vinculación de Acciones PTDI/PGTC ↔ PEI ↔ POA (Gestión {gestion})'
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    headers = ['PTDI/PGTC (Acción ETA)', 'PEI (AMP)', 'POA (ACP)',
               'CÓD. PROG.', '¿VINCULADO?', 'ARTICULACIÓN']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Trazar relaciones reales: ACP → AMP → Nodo (vía ArticulacionPlanificacion)
    from apps.planificacion.models import AccionCortoPlazo, AccionMedianoPlazo, ArticulacionPlanificacion

    acps = AccionCortoPlazo.objects.filter(gestion=gestion).select_related(
        'accion_mediano_plazo'
    ).prefetch_related('accion_mediano_plazo__nodo_planificacion')

    row = 5
    for acp in acps:
        amp = acp.accion_mediano_plazo
        amp_nombre = amp.nombre[:100] if amp else '—'
        ptdi_accion = amp_nombre if amp else '—'

        # Verificar articulación
        articulado = 'SÍ' if amp and amp.nodo_planificacion else 'NO'
        nivel_art = amp.nodo_planificacion.nivel if amp and amp.nodo_planificacion else '—'

        ws.cell(row=row, column=1, value=ptdi_accion)
        ws.cell(row=row, column=2, value=amp_nombre)
        ws.cell(row=row, column=3, value=acp.nombre[:100])
        ws.cell(row=row, column=4, value=acp.codigo)
        ws.cell(row=row, column=5, value=articulado)
        ws.cell(row=row, column=6, value=nivel_art)

        fill_color = 'E8F5E9' if articulado == 'SÍ' else 'FFEBEE'
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=c).font = Font(size=9)
            ws.cell(row=row, column=c).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        row += 1

    for col_letter, width in [('A', 40), ('B', 40), ('C', 40), ('D', 15), ('E', 15), ('F', 20)]:
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, _build_response_filename('evaluacion_cuadro2', 'xlsx', gestion)


def reporte_presupuesto_por_linea(linea_id=None):
    """Detalle de movimientos de una línea presupuestaria."""
    lineas = LineaPresupuestaria.objects.filter(activo=True).select_related(
        'programa', 'objeto_gasto', 'fuente', 'organismo', 'ue'
    )
    if linea_id:
        lineas = lineas.filter(id=linea_id)

    resultados = []
    for lp in lineas:
        movimientos = []
        techos_asociados = DistribucionTecho.objects.filter(
            programa=lp.programa, activo=True
        )
        for dist in techos_asociados:
            movs = MovimientoTecho.objects.filter(techo=dist.techo).order_by('-date')
            for mv in movs:
                movimientos.append({
                    'tipo': mv.get_movement_type_display(),
                    'monto': float(mv.amount),
                    'fecha': mv.date.isoformat() if mv.date else '',
                    'justificacion': mv.justification,
                })

        resultados.append({
            'linea_id': str(lp.id),
            'gestion': lp.gestion,
            'programa': lp.programa.codigo,
            'objeto_gasto': lp.objeto_gasto.denominacion,
            'fuente': lp.fuente.denominacion,
            'ue': str(lp.ue),
            'importe': float(lp.importe),
            'importe_anterior': float(lp.importe_gestion_anterior or 0),
            'importe_plurianual': float(lp.importe_plurianual or 0),
            'movimientos': movimientos,
        })

    return resultados


def reporte_acciones_correctivas_pendientes():
    """Acciones correctivas pendientes con fechas límite y responsable."""
    from apps.acciones_correctivas.models import AccionCorrectiva

    pendientes = AccionCorrectiva.objects.filter(
        status__in=['pendiente', 'en_ejecucion']
    ).select_related('responsible', 'responsible_unit')

    resultados = []
    for ac in pendientes:
        vencida = ac.esta_vencida
        resultados.append({
            'id': str(ac.id),
            'descripcion': ac.description,
            'causa': ac.cause,
            'responsable': str(ac.responsible) if ac.responsible else '',
            'unidad_responsable': str(ac.responsible_unit) if ac.responsible_unit else '',
            'fecha_inicio': ac.start_date.isoformat() if ac.start_date else '',
            'fecha_limite': ac.due_date.isoformat() if ac.due_date else '',
            'resultado_esperado': ac.expected_result,
            'estado': ac.get_status_display(),
            'vencida': vencida,
            'porcentaje_cumplimiento': ac.porcentaje_cumplimiento,
            'gestion': ac.gestion,
        })

    return resultados


def reporte_productos_por_pad(pad_id):
    """Productos de un PAD con resultados y líneas presupuestarias."""
    from apps.pad.models import ProductoTerritorial, ResultadoTerritorial

    resultado = ResultadoTerritorial.objects.filter(id=pad_id).first()
    if not resultado:
        resultado = ResultadoTerritorial.objects.filter(lineamiento__politica__id=pad_id).first()

    if not resultado:
        return []

    productos = ProductoTerritorial.objects.filter(
        resultado=resultado
    ).select_related('resultado')

    resultados = []
    for prod in productos:
        presupuesto = LineaPresupuestaria.objects.filter(
            activo=True
        ).aggregate(
            total=Coalesce(Sum('importe'), 0, output_field=DecimalField(max_digits=20, decimal_places=2))
        )

        resultados.append({
            'producto_id': str(prod.id),
            'producto_codigo': prod.codigo,
            'producto_nombre': prod.nombre,
            'resultado_codigo': resultado.codigo,
            'resultado_nombre': resultado.nombre,
            'indicador': prod.indicador,
            'linea_base': float(prod.linea_base or 0),
            'meta_2030': float(prod.meta_2030 or 0),
            'cuenta_financiamiento': prod.cuenta_con_financiamiento,
            'presupuesto_total_pad': float(prod.presupuesto_total_pad or 0),
            'presupuesto_general': float(presupuesto['total']),
            'gestion': prod.gestion,
        })

    return resultados


def reporte_historial_aprobaciones(gestion):
    """Historial de aprobaciones con fechas, usuarios y acciones."""
    aprobaciones = Aprobacion.objects.filter(
        gestion=gestion
    ).select_related('aprobado_por', 'documento').order_by('-created_at')

    resultados = []
    for ap in aprobaciones:
        resultados.append({
            'aprobacion_id': str(ap.id),
            'tipo': ap.get_tipo_display(),
            'tipo_codigo': ap.tipo,
            'estado': ap.get_estado_display(),
            'estado_codigo': ap.estado,
            'aprobado_por': str(ap.aprobado_por) if ap.aprobado_por else '',
            'comentario': ap.comentario,
            'version': ap.version,
            'huella_documento': ap.huella_documento,
            'es_reapertura': ap.es_reapertura,
            'motivo_reapertura': ap.motivo_reapertura,
            'fecha': ap.created_at.isoformat() if hasattr(ap, 'created_at') and ap.created_at else '',
        })

    return resultados


# ===== MATRIZ 1 — ARTICULACIÓN PAD → PEI =====


def generar_matriz_pad_pei_xlsx(gestion=None) -> tuple:
    """Genera XLSX de la Matriz 1 (PAD→PEI) con estilo institucional."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matriz PAD-PEI'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    ws.merge_cells('A2:I2')
    titulo_gestion = f'MATRIZ 1 — Articulación PAD → PEI (Gestión {gestion})' if gestion else 'MATRIZ 1 — Articulación PAD → PEI'
    ws['A2'] = titulo_gestion
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    headers = ['Código Resultado PAD', 'Resultado PAD', 'Código Producto PAD',
               'Producto PAD', 'Código Resultado PEI', 'Resultado PEI',
               'Código Producto PEI', 'Producto PEI', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    articulaciones = ArticulacionPADPEI.objects.select_related(
        'producto_pad__resultado_pad', 'producto_pei__resultado_pei'
    ).all()
    if gestion:
        pass  # PAD/PEI no tienen gestión directa, se filtran por vigencia si aplica

    row = 5
    for art in articulaciones:
        prod_pad = art.producto_pad
        prod_pei = art.producto_pei
        res_pad = prod_pad.resultado_pad if prod_pad else None
        res_pei = prod_pei.resultado_pei if prod_pei else None

        datos = [
            res_pad.codigo_resultado if res_pad else '—',
            (res_pad.denominacion or '—')[:120] if res_pad else '—',
            prod_pad.codigo_producto if prod_pad else '—',
            (prod_pad.denominacion or '—')[:120] if prod_pad else '—',
            res_pei.codigo_resultado if res_pei else '—',
            (res_pei.denominacion or '—')[:120] if res_pei else '—',
            prod_pei.codigo_producto if prod_pei else '—',
            (prod_pei.denominacion or '—')[:120] if prod_pei else '—',
            art.estado or '—',
        ]
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER_THIN
            cell.font = Font(size=9)
            if col in (1, 3, 5, 7):
                cell.font = Font(size=9, bold=True)
        row += 1

    for col_letter, width in [('A', 22), ('B', 35), ('C', 22), ('D', 35),
                               ('E', 22), ('F', 35), ('G', 22), ('H', 35), ('I', 15)]:
        ws.column_dimensions[col_letter].width = width

    filename = _build_response_filename('matriz_pad_pei', 'xlsx', gestion or 0)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== MATRIZ 2 — ARTICULACIÓN PEI → POA =====


def generar_matriz_pei_poa_xlsx(gestion=None) -> tuple:
    """Genera XLSX de la Matriz 2 (PEI→POA) con estilo institucional."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matriz PEI-POA'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    ws.merge_cells('A2:H2')
    titulo_gestion = f'MATRIZ 2 — Articulación PEI → POA (Gestión {gestion})' if gestion else 'MATRIZ 2 — Articulación PEI → POA'
    ws['A2'] = titulo_gestion
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    headers = ['Código Acción POA', 'Acción POA', 'Producto PEI', 'Indicador',
               'Unidad Medida', 'Meta Gestión', 'Presupuesto (Bs)', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    acciones = AccionPOA.objects.select_related('producto_pei').all()
    if gestion:
        acciones = acciones.filter(gestion=gestion)

    row = 5
    for act in acciones:
        datos = [
            act.codigo_accion,
            act.denominacion[:150],
            act.producto_pei.denominacion[:120] if act.producto_pei else '—',
            act.indicador or '—',
            act.unidad_medida or '—',
            float(act.meta_gestion) if act.meta_gestion is not None else '—',
            float(act.presupuesto_programado) if act.presupuesto_programado is not None else 0,
            act.estado or '—',
        ]
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER_THIN
            cell.font = Font(size=9)
            if col in (7,):
                cell.number_format = '#,##0.00'
        row += 1

    for col_letter, width in [('A', 22), ('B', 40), ('C', 35), ('D', 30),
                               ('E', 14), ('F', 16), ('G', 18), ('H', 15)]:
        ws.column_dimensions[col_letter].width = width

    filename = _build_response_filename('matriz_pei_poa', 'xlsx', gestion or 0)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== MATRIZ 4 — PRESUPUESTO Y SEGUIMIENTO =====


def generar_matriz_presupuesto_seguimiento_xlsx(gestion=None) -> tuple:
    """Genera XLSX de la Matriz 4 (Presupuesto y Seguimiento)."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Presupuesto y Seguimiento'

    headers = [
        'ID Cadena', 'Acción POA', 'Operación', 'Actividad',
        'Categoría Programática', 'DA', 'UE', 'Programa',
        'Presupuesto Inicial', 'Modificaciones', 'Presupuesto Vigente',
        'Ejecución Financiera', '% Ejecución Financiera', 'Meta Física',
        'Ejecución Física', '% Ejecución Física', 'Eficacia',
    ]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'] = 'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws['A2'] = (
        f'MATRIZ 4 — Presupuesto y Seguimiento (Gestión {gestion})'
        if gestion else 'MATRIZ 4 — Presupuesto y Seguimiento'
    )
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    rows = SeguimientoPresupuesto.objects.select_related(
        'accion_poa', 'operacion', 'actividad',
    )
    if gestion:
        rows = rows.filter(gestion=gestion)

    for row_number, item in enumerate(rows.order_by('id_cadena'), start=5):
        values = [
            item.id_cadena,
            item.accion_poa.denominacion,
            item.operacion.denominacion,
            item.actividad.denominacion,
            item.categoria_programatica,
            item.da,
            item.ue,
            item.programa,
            item.presupuesto_inicial,
            item.modificaciones,
            item.presupuesto_vigente,
            item.ejecutado_total,
            item.porcentaje_ejecucion_financiera,
            item.meta_fisica,
            item.ejecucion_fisica,
            item.porcentaje_ejecucion_fisica,
            item.eficacia,
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.border = BORDER_THIN
            cell.font = Font(size=9)
            if column >= 9:
                cell.number_format = '#,##0.00'

    for column in range(1, len(headers) + 1):
        letter = openpyxl.utils.get_column_letter(column)
        ws.column_dimensions[letter].width = 24 if column in (2, 3, 4, 8) else 18

    filename = _build_response_filename(
        'matriz_presupuesto_seguimiento', 'xlsx', gestion or 0,
    )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== MATRIZ 5 — OBJETOS DE GASTO =====


def generar_matriz_objetos_gasto_xlsx(gestion=None) -> tuple:
    """Genera XLSX de la Matriz 5 (Objetos de Gasto) con estilo institucional."""
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Objetos de Gasto'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    ws.merge_cells('A2:G2')
    titulo_gestion = f'MATRIZ 5 — Objetos de Gasto (Gestión {gestion})' if gestion else 'MATRIZ 5 — Objetos de Gasto'
    ws['A2'] = titulo_gestion
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    headers = ['Código', 'Objeto de Gasto', 'Grupo', 'Tipo Gasto', 'FF/OF',
               'Monto Programado (Bs)', 'Monto Vigente (Bs)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    asignaciones = AsignacionObjetoGasto.objects.all()
    if gestion:
        asignaciones = asignaciones.filter(gestion=gestion)

    row = 5
    total_programado = 0
    total_vigente = 0
    for asig in asignaciones:
        ff_of = f'{asig.fuente_financiamiento}/{asig.organismo_financiador}'
        monto_prog = float(asig.monto_programado or 0)
        monto_vig = float(asig.monto_vigente or 0)
        datos = [
            asig.codigo_asignacion,
            asig.descripcion_objeto[:120],
            asig.grupo_gasto or '—',
            asig.tipo_gasto or '—',
            ff_of,
            monto_prog,
            monto_vig,
        ]
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER_THIN
            cell.font = Font(size=9)
            if col in (6, 7):
                cell.number_format = '#,##0.00'
        total_programado += monto_prog
        total_vigente += monto_vig
        row += 1

    # Total row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='TOTAL GENERAL').font = Font(bold=True, size=10, color='1B5E3B')
    ws.cell(row=row, column=6, value=round(total_programado, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_vigente, 2)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='double'), bottom=Side(style='double')
        )
        ws.cell(row=row, column=c).fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

    for col_letter, width in [('A', 18), ('B', 40), ('C', 16), ('D', 16),
                               ('E', 18), ('F', 22), ('G', 22)]:
        ws.column_dimensions[col_letter].width = width

    filename = _build_response_filename('matriz_objetos_gasto', 'xlsx', gestion or 0)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


# ===== MATRIZ ARTICULACIÓN COMPLETA =====


def generar_matriz_completa_xlsx(gestion=None) -> tuple:
    """Genera XLSX de la Matriz de Articulación Completa (PGDESA→PDESA→PAD→PEI→POA).

    Columnas: Código Completo, Nivel, Nombre, Plan, Código Padre, Nodos Vinculados
    """
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matriz Articulación Completa'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'GOBIERNO AUTÓNOMO MUNICIPAL DE SACABA'
    ws['A1'].font = Font(bold=True, size=14, color='1B5E3B')

    ws.merge_cells('A2:F2')
    titulo_gestion = f'MATRIZ ARTICULACIÓN COMPLETA (Gestión {gestion})' if gestion else 'MATRIZ ARTICULACIÓN COMPLETA'
    ws['A2'] = titulo_gestion
    ws['A2'].font = Font(bold=True, size=11, color='1B5E3B')

    headers = ['Código Completo', 'Nivel', 'Nombre', 'Plan', 'Código Padre', 'Nodos Vinculados']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    from apps.planificacion.models import NodoPlanificacion, ArticulacionPlanificacion
    from apps.articulacion.models import ResultadoPAD

    nodos = NodoPlanificacion.objects.select_related('plan', 'padre').all()
    if gestion:
        nodos = nodos.filter(gestion=gestion)

    row = 5
    for nodo in nodos.order_by('plan__tipo', 'nivel', 'codigo'):
        # Reconstruir código completo
        codes = []
        n = nodo
        while n:
            codes.append(n.codigo)
            n = n.padre
        codigo_completo = '.'.join(reversed(codes))

        # Nodos vinculados (ArticulacionPlanificacion + ResultadoPAD bridge)
        vinculados = []
        arts = ArticulacionPlanificacion.objects.filter(nodo_origen=nodo).select_related('nodo_destino')
        for a in arts:
            vinculados.append(f'[{a.nodo_destino.codigo}] {a.nodo_destino.nombre[:60]}')
        if nodo.plan and nodo.plan.tipo == 'pdesa' and nodo.nivel == 'accion':
            for rp in ResultadoPAD.objects.filter(nodo_pdesa=nodo):
                vinculados.append(f'[RP:{rp.codigo_resultado}] {rp.denominacion[:60]}')

        datos = [
            codigo_completo,
            nodo.get_nivel_display(),
            nodo.nombre[:200],
            nodo.plan.get_tipo_display() if nodo.plan else '—',
            nodo.padre.codigo if nodo.padre else '—',
            '; '.join(vinculados) if vinculados else '—',
        ]
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER_THIN
            cell.font = Font(size=9)
            if col == 1:
                cell.font = Font(size=9, bold=True)
        row += 1

    for col_letter, width in [('A', 25), ('B', 18), ('C', 45), ('D', 15), ('E', 15), ('F', 55)]:
        ws.column_dimensions[col_letter].width = width

    filename = _build_response_filename('matriz_completa', 'xlsx', gestion or 0)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename
