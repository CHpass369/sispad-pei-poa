"""
Management command para importar datos reales de GAM Sacaba desde los Excel.

Uso:
    python manage.py importar_reales
        --ejec-apertura="/ruta/EJEC. POR APERTURA.xlsx"
        --gastos-2026="/ruta/GASTOS 2026.xlsx"
        --ptdi-pei="/ruta/SEGUIMIENTO Y EVALUACION PTDI_PEI AJUSTADA.xlsx"
"""
from datetime import date, datetime
from decimal import Decimal
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction, connection

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.catalogos.models import (
    ObjetoGasto, FuenteFinanciamiento, OrganismoFinanciador,
    UnidadMedida, TipoOperacion, TipoProducto
)
from apps.organizacion.models import (
    TipoUnidad, UnidadOrganizacional, DireccionAdministrativa,
    UnidadEjecutora
)
from apps.presupuesto.models import ProgramaPresupuestario
from apps.gestion.models import GestionFiscal

GESTION = 2026
VIGENCIA = date(2026, 1, 1)


class Command(BaseCommand):
    help = 'Importa datos reales de los archivos Excel de GAM Sacaba'

    def add_arguments(self, parser):
        parser.add_argument('--ejec-apertura', required=True)
        parser.add_argument('--gastos-2026', required=True)
        parser.add_argument('--ptdi-pei', required=False, default='')

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl no está instalado')

        self.stdout.write(self.style.NOTICE('=== INICIO IMPORTACIÓN DATOS REALES GAM SACABA ==='))

        # 1. Catalogos básicos
        self._crear_catalogos_base()

        # 2. Importar aperturas programáticas y secretarías
        if options.get('ejec_apertura') and options['ejec_apertura'].strip():
            self._importar_aperturas(options['ejec_apertura'])

        # 3. Importar GASTOS 2026 (estructura DA, UE, techos)
        if options.get('gastos_2026') and options['gastos_2026'].strip():
            self._importar_gastos_2026(options['gastos_2026'])

        # 4. Importar PTDI-PEI (planificación estratégica) - opcional
        if options.get('ptdi_pei') and options['ptdi_pei'].strip():
            self._importar_ptdi_pei(options['ptdi_pei'])

        self.stdout.write(self.style.SUCCESS('Importación completada exitosamente'))

    # ============================================================
    # 1. CATÁLOGOS BASE
    # ============================================================
    def _crear_catalogos_base(self):
        self.stdout.write('\n[1/4] Creando catálogos base...')

        # Fuentes de financiamiento reales de GAM Sacaba
        fuentes = [
            ('41-113', 'CT', 'Coparticipación Tributaria'),
            ('20-210', 'RE', 'Recursos Específicos'),
            ('20-230', 'ORE', 'Otros Recursos Específicos'),
            ('41-119', 'IDH', 'Impuesto Directo a los Hidrocarburos'),
            ('41-111', 'TGN', 'Tesoro General de la Nación - Transferencias'),
        ]
        for cod, sigla, nombre in fuentes:
            FuenteFinanciamiento.objects.get_or_create(
                codigo=cod, gestion=GESTION,
                defaults={
                    'denominacion': f'{sigla} - {nombre}',
                    'descripcion': nombre,
                    'fecha_vigencia_desde': VIGENCIA,
                }
            )
        self.stdout.write(f'  Fuentes: {FuenteFinanciamiento.objects.filter(gestion=GESTION).count()}')

        # Unidades de medida (de PTDI)
        umedidas = [
            ('PORC', 'Porcentaje', '%'),
            ('NUM', 'Número', 'Cantidad'),
            ('UN', 'Unidad', 'Unidad'),
            ('M2', 'Metro cuadrado', 'm²'),
            ('KM', 'Kilómetro', 'km'),
            ('HA', 'Hectárea', 'ha'),
            ('PER', 'Persona', 'Persona'),
            ('FAM', 'Familia', 'Familia'),
            ('TASA', 'Tasa', 'Tasa'),
        ]
        for cod, nom, desc in umedidas:
            UnidadMedida.objects.get_or_create(
                codigo=cod, gestion=GESTION,
                defaults={'denominacion': nom, 'descripcion': desc,
                          'fecha_vigencia_desde': VIGENCIA}
            )
        self.stdout.write(f'  Unidades de medida: {UnidadMedida.objects.filter(gestion=GESTION).count()}')

    # ============================================================
    # 2. APERTURAS PROGRAMÁTICAS
    # ============================================================
    @transaction.atomic
    def _importar_aperturas(self, ruta):
        self.stdout.write('\n[2/4] Importando aperturas programáticas...')

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb['Hoja1']

        # Mapa de siglas de secretarías a nombres
        secretarias_map = {
            'SMFA': 'Secretaría Municipal de Finanzas y Administración',
            'CM': 'Concejo Municipal',
            'SMPDT': 'Secretaría Municipal de Planificación y Desarrollo Territorial',
            'SMMTDP': 'Secretaría Municipal de Desarrollo Productivo',
            'SMS': 'Secretaría Municipal de Salud',
            'SMIS': 'Secretaría Municipal de Infraestructura y Servicios',
            'SMDHI': 'Secretaría Municipal de Desarrollo Humano Integral',
            'STAFF': 'Staff de Alcaldía',
        }

        # Crear tipo de unidad "Secretaría"
        tipo_sec, _ = TipoUnidad.objects.get_or_create(
            codigo='SEC', defaults={'nombre': 'Secretaría Municipal', 'nivel': 1}
        )
        tipo_dir, _ = TipoUnidad.objects.get_or_create(
            codigo='DIR', defaults={'nombre': 'Dirección', 'nivel': 2}
        )
        tipo_uni, _ = TipoUnidad.objects.get_or_create(
            codigo='UNI', defaults={'nombre': 'Unidad', 'nivel': 3}
        )

        # Crear secretarías y colectar aperturas
        aperturas_leidas = set()

        for row in ws.iter_rows(min_row=3, values_only=True):
            sigla = str(row[0]).strip() if row[0] else ''
            cod_apertura = str(row[1]).strip() if row[1] else ''
            desc = str(row[2]).strip() if row[2] else ''
            vigente = Decimal(str(row[5])) if row[5] else Decimal('0')

            if not sigla or not cod_apertura:
                continue

            # Crear unidad organizacional para la secretaría
            nombre_sec = secretarias_map.get(sigla, f'Secretaría {sigla}')
            UnidadOrganizacional.objects.get_or_create(
                codigo=sigla, gestion=GESTION,
                defaults={
                    'nombre': nombre_sec,
                    'sigla': sigla,
                    'tipo': tipo_sec,
                    'fecha_vigencia_desde': VIGENCIA,
                    'activo': True,
                }
            )

            # Parsear código de apertura: "PPP A AAA"
            partes = cod_apertura.split()
            prog_cod = partes[0].strip() if len(partes) > 0 else ''
            act_cod = partes[-1].strip() if len(partes) >= 3 else ''

            if prog_cod and (prog_cod, sigla) not in aperturas_leidas:
                aperturas_leidas.add((prog_cod, sigla))

                # Crear programa presupuestario
                try:
                    prog_num = int(prog_cod)
                    nombre_prog = desc.split('-')[0].strip() if '-' in desc else desc[:200]
                    ProgramaPresupuestario.objects.get_or_create(
                        codigo=prog_cod, gestion=GESTION,
                        defaults={'nombre': nombre_prog or f'Programa {prog_cod}',
                                  'descripcion': desc[:300]}
                    )
                except ValueError:
                    pass  # Código no numérico, ej: con SISIN

        wb.close()
        self.stdout.write(f'  Programas: {ProgramaPresupuestario.objects.filter(gestion=GESTION).count()}')
        self.stdout.write(f'  Unidades: {UnidadOrganizacional.objects.filter(gestion=GESTION).count()}')

    # ============================================================
    # 3. GASTOS 2026 - Estructura completa con techos
    # ============================================================
    @transaction.atomic
    def _importar_gastos_2026(self, ruta):
        self.stdout.write('\n[3/4] Importando estructura GASTOS 2026 con techos...')

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb['gastos']

        # Mapeo de columnas por nombre de fuente
        FUENTE_COLS = {'CT': 9, 'RE': 10, 'ORE': 11, 'IDH': 12, 'TGN': 13}
        COL_TOTAL = 14

        das_creadas = {}
        ues_creadas = {}
        programas_detectados = set()

        count_detalle = 0

        for row in ws.iter_rows(min_row=11, values_only=True):
            tipo = str(row[0]).strip()[:10] if row[0] else ''
            da_cod = str(row[2]).strip()[:5] if row[2] else ''
            ue_cod = str(row[3]).strip()[:5] if row[3] else ''
            prog_cod = str(row[4]).strip()[:5] if row[4] else ''
            denom = str(row[7]).strip()[:200] if row[7] else ''

            # Saltar totales
            if tipo.upper() in ('T', 'TS', 'P', 'SP', 'TT', 'TTG', ''):
                continue
            if not prog_cod or not denom:
                continue

            # ---- PROGRAMA ----
            if prog_cod and prog_cod not in programas_detectados:
                programas_detectados.add(prog_cod)
                try:
                    cod_num = int(''.join(filter(str.isdigit, prog_cod[:3])))
                    ProgramaPresupuestario.objects.get_or_create(
                        codigo=prog_cod, gestion=GESTION,
                        defaults={
                            'nombre': denom[:300],
                            'descripcion': denom[:500],
                        }
                    )
                except ValueError:
                    pass

            # ---- DA ----
            if da_cod and da_cod not in das_creadas:
                da_obj, _ = DireccionAdministrativa.objects.get_or_create(
                    codigo=da_cod, gestion=GESTION,
                    defaults={
                        'nombre': f'DA {da_cod}',
                        'fecha_vigencia_desde': VIGENCIA,
                    }
                )
                das_creadas[da_cod] = da_obj

            # ---- UE ----
            ue_key = (da_cod, ue_cod)
            if ue_cod and ue_key not in ues_creadas and da_cod in das_creadas:
                ue_obj, _ = UnidadEjecutora.objects.get_or_create(
                    codigo=ue_cod, da=das_creadas[da_cod], gestion=GESTION,
                    defaults={
                        'nombre': f'UE {ue_cod}',
                        'fecha_vigencia_desde': VIGENCIA,
                    }
                )
                ues_creadas[ue_key] = ue_obj

            count_detalle += 1

        wb.close()
        self.stdout.write(f'  Programas: {len(programas_detectados)}')
        self.stdout.write(f'  DA: {len(das_creadas)}, UE: {len(ues_creadas)}')
        self.stdout.write(f'  Filas de detalle procesadas: {count_detalle}')

    # ============================================================
    # 4. PTDI-PEI - Planificación estratégica
    # ============================================================
    @transaction.atomic
    def _importar_ptdi_pei(self, ruta):
        self.stdout.write('\n[4/4] Importando PTDI-PEI...')

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

        # Hoja ACP-AMP: Acciones de Corto y Mediano Plazo
        if 'ACP-AMP' in wb.sheetnames:
            self._importar_acp_amp(wb['ACP-AMP'])

        # Hoja Datos154-SMISSMPDT: Datos de seguimiento
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Datos'):
                self._importar_datos_seguimiento(wb[sheet_name], sheet_name)
                break  # Solo una hoja de datos

        wb.close()

    def _importar_acp_amp(self, ws):
        """Importa acciones de mediano y corto plazo de la hoja ACP-AMP."""
        from apps.planificacion.models import Plan, NodoPlanificacion, AccionMedianoPlazo, AccionCortoPlazo

        plan_pei, _ = Plan.objects.get_or_create(
            codigo='PEI-2021-2025', tipo='pei',
            defaults={
                'nombre': 'PEI 2021-2025 del GAM Sacaba',
                'gestion_inicio': 2021, 'gestion_fin': 2025,
                'fecha_vigencia_desde': date(2021, 1, 1),
            }
        )

        plan_ptdi, _ = Plan.objects.get_or_create(
            codigo='PTDI-2021-2025', tipo='ptdi',
            defaults={
                'nombre': 'PTDI para Vivir Bien del Municipio de Sacaba 2021-2025',
                'gestion_inicio': 2021, 'gestion_fin': 2025,
                'fecha_vigencia_desde': date(2021, 1, 1),
            }
        )

        count_amp = 0
        count_acp = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            amp_cod = str(row[0]).strip() if row[0] else ''
            amp_desc = str(row[1]).strip() if row[1] else ''

            if not amp_cod or not amp_desc:
                continue

            try:
                # 1. Crear nodo de planificación PRIMERO
                nodo, _ = NodoPlanificacion.objects.get_or_create(
                    plan=plan_pei, nivel='accion_mediano',
                    codigo=f'AMP-{amp_cod}', gestion=2021,
                    defaults={'nombre': amp_desc[:500]}
                )

                # 2. Crear AccionMedianoPlazo con el nodo ya existente
                amp, created = AccionMedianoPlazo.objects.get_or_create(
                    codigo=f'AMP-{amp_cod}',
                    defaults={
                        'nombre': amp_desc[:500],
                        'nodo_planificacion': nodo,
                        'gestion_inicio': 2021,
                        'gestion_fin': 2025,
                    }
                )
                if created:
                    count_amp += 1

                acp_desc = str(row[5]).strip() if row[5] else ''
                if acp_desc and len(acp_desc) > 10:
                    AccionCortoPlazo.objects.get_or_create(
                        codigo=f'ACP-{amp_cod}',
                        gestion=GESTION,
                        defaults={
                            'nombre': acp_desc[:500],
                            'accion_mediano_plazo': amp,
                        }
                    )
                    count_acp += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  Error AMP {amp_cod}: {e}'))

        self.stdout.write(f'  AMP: {count_amp}, ACP: {count_acp}')

    def _importar_datos_seguimiento(self, ws, sheet_name):
        """Importa indicadores y metas de las hojas de seguimiento."""
        from apps.indicadores.models import Indicador, MetaProgramada
        from apps.planificacion.models import AccionCortoPlazo

        count_inds = 0
        headers = [str(c)[:50] if c else '' for c in list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]

        # Encontrar columnas de interés
        col_idx = {h: i for i, h in enumerate(headers) if h}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            indicador_nom = str(row[col_idx.get('INDICADOR', 18)]).strip() if col_idx.get('INDICADOR', 18) < len(row) and row[col_idx.get('INDICADOR', 18)] else ''
            formula = str(row[col_idx.get('Fórmula', 19)]).strip() if col_idx.get('Fórmula', 19) < len(row) and row[col_idx.get('Fórmula', 19)] else ''
            lb = str(row[col_idx.get('Línea Base 2020', 20)]).strip() if col_idx.get('Línea Base 2020', 20) < len(row) and row[col_idx.get('Línea Base 2020', 20)] else ''

            if not indicador_nom:
                continue

            try:
                Indicador.objects.get_or_create(
                    codigo=f'IND-{sheet_name}-{count_inds + 1}',
                    defaults={
                        'nombre': indicador_nom[:500],
                        'formula': formula[:500],
                        'linea_base': Decimal(str(lb)) if lb else None,
                    }
                )
                count_inds += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  Error indicador: {e}'))

        self.stdout.write(f'  Indicadores: {count_inds}')
