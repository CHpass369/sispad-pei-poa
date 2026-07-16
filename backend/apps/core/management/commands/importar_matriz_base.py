"""
Management command para importar la hoja MATRIZ BASE del archivo
SEGUIMIENTO Y EVALUACION PTDI_PEI AJUSTADA.xlsx de GAM Sacaba.

Crea/actualiza la jerarquía completa de planificación estratégica:
  PDES  → Pilar → Eje → Meta
  PTDI  → Resultado (vinculado a Meta)
  PEI   → Acción PDES (vinculado a Resultado) + AccionMedianoPlazo
  POA   → AccionCortoPlazo por año (2021–2025)

Además importa Sectores, Unidades Organizacionales e Indicadores.

Uso:
    python manage.py importar_matriz_base /ruta/al/archivo.xlsx

Idempotente: se puede ejecutar múltiples veces sin duplicar datos.
"""
import re
from datetime import date, datetime
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.planificacion.models import (
    Plan,
    NodoPlanificacion,
    Sector,
    AccionMedianoPlazo,
    AccionCortoPlazo,
)
from apps.organizacion.models import TipoUnidad, UnidadOrganizacional
from apps.indicadores.models import Indicador

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
ANIOS_POA = [2021, 2022, 2023, 2024, 2025]

COL_ACCION_POA = {2021: 15, 2022: 17, 2023: 19, 2024: 21, 2025: 23}
COL_PROD_POAU = {2021: 16, 2022: 18, 2023: 20, 2024: 22, 2025: 24}

COL_PROG_FISICA = {
    2021: (29, 30, 31),
    2022: (32, 33, 34),
    2023: (35, 36, 37),
    2024: (38, 39, 40),
    2025: (41, 42, 43),
}

COL_CAT_PROG = {2021: 47, 2022: 47, 2023: 48, 2024: 49, 2025: 50}

COL_FINANCIERA = {
    2021: (51, 52, 53),
    2022: (54, 55, 56),
    2023: (57, 58, 59),
    2024: (60, 61, 62),
    2025: (63, 64, 65),
}


# ---------------------------------------------------------------------------
# Helpers de limpieza
# ---------------------------------------------------------------------------
def _clean(val):
    """Limpia un valor a string; retorna '' si es None o vacío."""
    if val is None:
        return ""
    s = str(val).strip()
    return s


def _codigo(val):
    """Extrae el código (primera palabra antes de espacio o fecha)."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        # Excel interpretó "3.3" como 3-mar — extraemos mes.día
        return f"{val.month}.{val.day}"
    if isinstance(val, (int, float)):
        # Expresado como número: 2.0 → "2",  3.0 → "3"
        if val == int(val):
            return str(int(val))
        return str(val)
    s = str(val).strip()
    return s.split(" ", 1)[0].strip()


def _nombre(val):
    """Extrae el nombre descriptivo después del código."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return ""
    s = str(val).strip()
    partes = s.split(" ", 1)
    if len(partes) > 1 and partes[1].strip():
        return re.sub(r"\s+", " ", partes[1].strip())
    return ""


def _decimal(val):
    """Convierte a Decimal o None."""
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _orden_pilar(codigo):
    """Extrae el orden numérico de un código de pilar."""
    m = re.match(r"(\d+)", codigo)
    return int(m.group(1)) if m else 0


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------
class Command(BaseCommand):
    help = "Importa la hoja MATRIZ BASE del archivo SEGUIMIENTO Y EVALUACION PTDI_PEI AJUSTADA.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Ruta al archivo Excel")

    # ================================================================
    # handle
    # ================================================================
    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl no está instalado. Ejecute: pip install openpyxl")

        file_path = options["file_path"]
        self.stdout.write(self.style.NOTICE(f"📂 Leyendo archivo: {file_path}"))

        # --- Cargar todas las filas ---
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if "MATRIZ BASE" not in wb.sheetnames:
            wb.close()
            raise CommandError('La hoja "MATRIZ BASE" no existe en el archivo.')

        ws = wb["MATRIZ BASE"]
        filas = list(ws.iter_rows(min_row=6, values_only=True))
        wb.close()
        self.stdout.write(f"  Filas cargadas: {len(filas)}")

        # ---- 1. Planes ----
        self._crear_planes()

        # ---- 2. Catálogos base ----
        self._crear_catalogos()

        # ---- 3. Recolectar nombres de jerarquía ----
        nombres = self._recolectar_nombres(filas)

        # ---- 4. Crear jerarquía de planificación ----
        nodos_creados = self._crear_jerarquia(filas, nombres)

        # ---- 5. Crear sectores ----
        sectores = self._crear_sectores(filas)

        # ---- 6. Acciones PEI (AccionMedianoPlazo) ----
        amp_count = self._crear_acciones_pei(filas, nodos_creados)

        # ---- 7. Acciones POA (AccionCortoPlazo) ----
        acp_count = self._crear_acciones_poa(filas)

        # ---- 8. Indicadores ----
        ind_count = self._crear_indicadores(filas)

        # ---- Resumen ----
        self._resumen(nombres, sectores, amp_count, acp_count, ind_count)

    # ================================================================
    # 1. Planes
    # ================================================================
    def _crear_planes(self):
        self.stdout.write("\n[1/8] Creando/verificando planes…")
        self.plan_pdes, _ = Plan.objects.get_or_create(
            codigo="PDES",
            tipo="pdes",
            defaults={
                "nombre": "Plan de Desarrollo Económico y Social 2021-2025",
                "gestion_inicio": 2021,
                "gestion_fin": 2025,
                "descripcion": "PDES del Estado Plurinacional de Bolivia 2021-2025",
                "fecha_vigencia_desde": date(2021, 1, 1),
            },
        )
        self.plan_ptdi, _ = Plan.objects.get_or_create(
            codigo="PTDI",
            tipo="ptdi",
            defaults={
                "nombre": "Plan Territorial de Desarrollo Integral 2021-2025",
                "gestion_inicio": 2021,
                "gestion_fin": 2025,
                "descripcion": "PTDI del Municipio de Sacaba 2021-2025",
                "fecha_vigencia_desde": date(2021, 1, 1),
            },
        )
        self.plan_pei, _ = Plan.objects.get_or_create(
            codigo="PEI",
            tipo="pei",
            defaults={
                "nombre": "Plan Estratégico Institucional 2021-2025",
                "gestion_inicio": 2021,
                "gestion_fin": 2025,
                "descripcion": "PEI del GAM Sacaba 2021-2025",
                "fecha_vigencia_desde": date(2021, 1, 1),
            },
        )
        self.stdout.write("  ✓ PDES, PTDI y PEI listos")

    # ================================================================
    # 2. Catálogos base (TipoUnidad)
    # ================================================================
    def _crear_catalogos(self):
        self.stdout.write("\n[2/8] Creando catálogos base…")
        self.tipo_area, _ = TipoUnidad.objects.get_or_create(
            codigo="AREA", defaults={"nombre": "Área Organizacional", "nivel": 3}
        )
        self.tipo_sec, _ = TipoUnidad.objects.get_or_create(
            codigo="SEC", defaults={"nombre": "Secretaría Municipal", "nivel": 1}
        )
        self.stdout.write("  ✓ TipoUnidad listo")

    # ================================================================
    # 3. Recolectar nombres de jerarquía
    # ================================================================
    def _recolectar_nombres(self, filas):
        self.stdout.write("\n[3/8] Recolectando nombres de jerarquía…")

        pilar_names = {}
        pilar_names_col0 = {}
        eje_names = {}
        eje_names_col1 = {}
        meta_names = {}
        resultado_names = {}
        resultado_desc = {}
        accion_names = {}

        for row in filas:
            p_cod = _codigo(row[2])
            e_cod = _codigo(row[3])
            m_cod = _codigo(row[4])
            r_cod = _codigo(row[5])
            a_cod = _codigo(row[6])

            if not a_cod:
                continue

            # Pilar — nombre desde col 2 (si viene "2 NOMBRE") o desde col 0
            nombre_p = _nombre(row[2])
            if p_cod and nombre_p:
                pilar_names[p_cod] = nombre_p
            if p_cod and _clean(row[0]):
                pilar_names_col0[p_cod] = _clean(row[0])

            # Eje
            nombre_e = _nombre(row[3])
            if e_cod and nombre_e:
                eje_names[e_cod] = nombre_e
            if e_cod and _clean(row[1]):
                eje_names_col1[e_cod] = _clean(row[1])

            # Meta
            nombre_m = _nombre(row[4])
            if m_cod and nombre_m:
                meta_names[m_cod] = nombre_m

            # Resultado
            nombre_r = _nombre(row[5])
            if r_cod and nombre_r:
                resultado_names[r_cod] = nombre_r
            desc_r = _clean(row[12])
            if r_cod and desc_r and (r_cod not in resultado_desc or len(desc_r) > len(resultado_desc.get(r_cod, ""))):
                resultado_desc[r_cod] = desc_r

            # Acción
            nombre_a = _nombre(row[6])
            pei_desc = _clean(row[14])
            if a_cod:
                if nombre_a:
                    accion_names[a_cod] = nombre_a
                elif pei_desc and a_cod not in accion_names:
                    accion_names[a_cod] = pei_desc

        # Fallback: si un pilar no tiene nombre desde col 2, usa col 0
        for cod in list(pilar_names_col0.keys()):
            if cod not in pilar_names:
                pilar_names[cod] = pilar_names_col0[cod]

        for cod in list(eje_names_col1.keys()):
            if cod not in eje_names:
                eje_names[cod] = eje_names_col1[cod]

        self.stdout.write(f"  Pilares: {len(pilar_names)}")
        self.stdout.write(f"  Ejes:    {len(eje_names)}")
        self.stdout.write(f"  Metas:   {len(meta_names)}")
        self.stdout.write(f"  Rdos:    {len(resultado_names)}")
        self.stdout.write(f"  Accs:    {len(accion_names)}")

        return {
            "pilar": pilar_names,
            "eje": eje_names,
            "meta": meta_names,
            "resultado": resultado_names,
            "resultado_desc": resultado_desc,
            "accion": accion_names,
        }

    # ================================================================
    # 4. Jerarquía de planificación
    # ================================================================
    @transaction.atomic
    def _crear_jerarquia(self, filas, nombres):
        self.stdout.write("\n[4/8] Creando jerarquía de planificación…")

        cache = {}  # (plan_id, nivel, codigo) → NodoPlanificacion
        contadores = {"pilar": set(), "eje": set(), "meta": set(), "resultado": set(), "accion": set()}

        total = len(filas)

        def _nodo(plan, nivel, codigo, defaults):
            key = (plan.pk, nivel, codigo)
            if key in cache:
                return cache[key]
            obj, created = NodoPlanificacion.objects.get_or_create(
                plan=plan,
                nivel=nivel,
                codigo=codigo,
                defaults=defaults,
            )
            # Si ya existía y no tenía nombre, actualizarlo
            if not created and defaults.get("nombre") and not obj.nombre:
                obj.nombre = defaults["nombre"]
                obj.save(update_fields=["nombre"])
            cache[key] = obj
            return obj

        for idx, row in enumerate(filas):
            p_cod = _codigo(row[2])
            e_cod = _codigo(row[3])
            m_cod = _codigo(row[4])
            r_cod = _codigo(row[5])
            a_cod = _codigo(row[6])

            if not a_cod:
                continue

            # --- PILAR ---
            if p_cod:
                nombre = nombres["pilar"].get(p_cod, p_cod)
                pilar = _nodo(
                    self.plan_pdes,
                    "pilar",
                    p_cod,
                    {"nombre": nombre, "gestion": 2021, "orden": _orden_pilar(p_cod)},
                )
                contadores["pilar"].add(p_cod)

            # --- EJE (padre = pilar) ---
            if e_cod and p_cod:
                nombre = nombres["eje"].get(e_cod, e_cod)
                eje = _nodo(
                    self.plan_pdes,
                    "eje",
                    e_cod,
                    {"nombre": nombre, "padre": pilar, "gestion": 2021},
                )
                contadores["eje"].add(e_cod)

            # --- META (padre = eje) ---
            if m_cod and e_cod:
                nombre = nombres["meta"].get(m_cod, m_cod)
                meta = _nodo(
                    self.plan_pdes,
                    "meta",
                    m_cod,
                    {"nombre": nombre, "padre": eje, "gestion": 2021},
                )
                contadores["meta"].add(m_cod)

            # --- RESULTADO (padre = meta, plan = PTDI) ---
            if r_cod and m_cod:
                nombre = nombres["resultado"].get(r_cod, r_cod)
                desc = nombres["resultado_desc"].get(r_cod, "")
                resultado = _nodo(
                    self.plan_ptdi,
                    "resultado",
                    r_cod,
                    {"nombre": nombre, "descripcion": desc, "padre": meta, "gestion": 2021},
                )
                contadores["resultado"].add(r_cod)

            # --- ACCIÓN PDES (padre = resultado, plan = PEI) ---
            if a_cod and r_cod:
                nombre = nombres["accion"].get(a_cod, a_cod)
                pei_desc = _clean(row[14]) or _clean(row[13]) or ""
                _nodo(
                    self.plan_pei,
                    "accion_pdes",
                    a_cod,
                    {"nombre": nombre, "descripcion": pei_desc, "padre": resultado, "gestion": 2021},
                )
                contadores["accion"].add(a_cod)

            if (idx + 1) % 500 == 0:
                self.stdout.write(f"  … fila {idx + 1}/{total}")

        for k, v in contadores.items():
            self.stdout.write(f"  {k}: {len(v)}")

        return cache  # para fases posteriores

    # ================================================================
    # 5. Sectores
    # ================================================================
    def _crear_sectores(self, filas):
        self.stdout.write("\n[5/8] Creando sectores…")
        creados = set()
        for row in filas:
            nombre = _clean(row[7]).upper()
            if nombre:
                cod = nombre[:50]
                Sector.objects.get_or_create(codigo=cod, defaults={"nombre": nombre})
                creados.add(cod)
        self.stdout.write(f"  Sectores: {len(creados)}")
        return creados

    # ================================================================
    # 6. Acciones PEI → AccionMedianoPlazo
    # ================================================================
    @transaction.atomic
    def _crear_acciones_pei(self, filas, cache_nodos):
        self.stdout.write("\n[6/8] Creando acciones PEI (AccionMedianoPlazo)…")
        count = 0
        total = len(filas)

        for idx, row in enumerate(filas):
            a_cod = _codigo(row[6])
            pei_desc = _clean(row[14])
            if not a_cod or not pei_desc:
                continue

            key = (self.plan_pei.pk, "accion_pdes", a_cod)
            nodo = cache_nodos.get(key)
            if not nodo:
                continue

            # Grabar descripción en el nodo si está vacía
            if not nodo.descripcion:
                nodo.descripcion = pei_desc
                nodo.save(update_fields=["descripcion"])

            _, created = AccionMedianoPlazo.objects.get_or_create(
                codigo=f"AMP-{a_cod}",
                defaults={
                    "nombre": pei_desc,
                    "nodo_planificacion": nodo,
                    "gestion_inicio": 2021,
                    "gestion_fin": 2025,
                },
            )
            if created:
                count += 1

            if (idx + 1) % 500 == 0:
                self.stdout.write(f"  … fila {idx + 1}/{total}")

        self.stdout.write(f"  Creadas: {count}")
        return count

    # ================================================================
    # 7. Acciones POA → AccionCortoPlazo
    # ================================================================
    @transaction.atomic
    def _crear_acciones_poa(self, filas):
        self.stdout.write("\n[7/8] Creando acciones POA (AccionCortoPlazo)…")
        cache_uos = {}
        count = 0
        total = len(filas)

        def _uo(nombre):
            if not nombre:
                return None
            key = nombre.upper().strip()
            if key in cache_uos:
                return cache_uos[key]
            cod = "".join(c for c in key if c.isalnum() or c in "_-")[:20]
            if not cod:
                cod = "GEN"
            obj, _ = UnidadOrganizacional.objects.get_or_create(
                codigo=cod,
                gestion=2021,
                defaults={
                    "nombre": nombre[:300],
                    "sigla": key[:30],
                    "tipo": self.tipo_area,
                    "fecha_vigencia_desde": date(2021, 1, 1),
                },
            )
            cache_uos[key] = obj
            return obj

        for idx, row in enumerate(filas):
            a_cod = _codigo(row[6])
            if not a_cod:
                continue

            try:
                amp = AccionMedianoPlazo.objects.get(codigo=f"AMP-{a_cod}")
            except AccionMedianoPlazo.DoesNotExist:
                continue

            unidad = _uo(_clean(row[45]) or _clean(row[9]))
            if not unidad:
                continue

            for anio in ANIOS_POA:
                col = COL_ACCION_POA[anio]
                accion_poa = _clean(row[col]) if col < len(row) else ""
                if not accion_poa:
                    continue

                cod_acp = f"ACP-{a_cod}-{anio}"
                _, created = AccionCortoPlazo.objects.get_or_create(
                    codigo=cod_acp,
                    gestion=anio,
                    defaults={
                        "nombre": accion_poa,
                        "descripcion": accion_poa,
                        "accion_mediano_plazo": amp,
                        "unidad_responsable": unidad,
                    },
                )
                if created:
                    count += 1

            if (idx + 1) % 500 == 0:
                self.stdout.write(f"  … fila {idx + 1}/{total}")

        self.stdout.write(f"  Creadas: {count}")
        return count

    # ================================================================
    # 8. Indicadores
    # ================================================================
    def _crear_indicadores(self, filas):
        self.stdout.write("\n[8/8] Creando indicadores…")
        count = 0
        total = len(filas)

        for idx, row in enumerate(filas):
            a_cod = _codigo(row[6])
            nombre_ind = _clean(row[25])
            if not a_cod or not nombre_ind:
                continue

            formula = _clean(row[26])
            lb = _decimal(row[27])
            meta25 = _decimal(row[28])

            _, created = Indicador.objects.get_or_create(
                codigo=f"IND-{a_cod}",
                defaults={
                    "nombre": nombre_ind,
                    "formula": formula,
                    "linea_base": lb,
                    "meta_anual": meta25,
                },
            )
            if created:
                count += 1

            if (idx + 1) % 500 == 0:
                self.stdout.write(f"  … fila {idx + 1}/{total}")

        self.stdout.write(f"  Creados: {count}")
        return count

    # ================================================================
    # Resumen
    # ================================================================
    def _resumen(self, nombres, sectores, amp_count, acp_count, ind_count):
        self.stdout.write(self.style.SUCCESS("\n" + "=" * 60))
        self.stdout.write(self.style.SUCCESS("  ✅ IMPORTACIÓN COMPLETADA"))
        self.stdout.write(self.style.SUCCESS("=" * 60))
        self.stdout.write(f"  Pilares:          {len(nombres['pilar'])}")
        self.stdout.write(f"  Ejes:             {len(nombres['eje'])}")
        self.stdout.write(f"  Metas:            {len(nombres['meta'])}")
        self.stdout.write(f"  Resultados:       {len(nombres['resultado'])}")
        self.stdout.write(f"  Acciones PDES:    {len(nombres['accion'])}")
        self.stdout.write(f"  Sectores:         {len(sectores)}")
        self.stdout.write(f"  Acciones PEI:     {amp_count}")
        self.stdout.write(f"  Acciones POA:     {acp_count}")
        self.stdout.write(f"  Indicadores:      {ind_count}")
        self.stdout.write("=" * 60)
        self.stdout.write(
            self.style.WARNING(
                "⚠  No olvide ejecutar:  python manage.py makemigrations && python manage.py migrate"
            )
        )
