import json
from datetime import datetime
from io import StringIO
from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import SimpleTestCase, TestCase
from rest_framework.test import APIClient

from apps.articulacion.models import (
    AccionPOA,
    ActividadPOAU,
    OperacionPOAU,
    ProductoPEI,
    ResultadoPEI,
    TareaPOAU,
)
from apps.planificacion.models import AccionCortoPlazo
from apps.poau.models import EjecucionFisica, POAU, POAUActividad


class FormulacionPOAU2027CommandTest(SimpleTestCase):
    def test_source_file_argument_is_required(self):
        from apps.core.management.commands.importar_formulacion_poau_2027 import Command

        parser = Command().create_parser('manage.py', 'importar_formulacion_poau_2027')
        file_action = next(action for action in parser._actions if action.dest == 'file')

        self.assertTrue(file_action.required)


class FormulacionPOAU2027ImportTest(TestCase):
    def setUp(self):
        self.workbook_path = Path(self._create_workbook())

    def tearDown(self):
        self.workbook_path.unlink(missing_ok=True)

    def _create_workbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Base"
        catalog = workbook.create_sheet("CLASIFICADOR cat. progra")
        catalog["E9"] = "000 0 001"
        catalog["F9"] = "FUNCIONAMIENTO ALCALDIA MUNICIPAL"
        catalog["G9"] = "ACTIVIDAD"

        sheet["A5"] = "EM-DJR-01"
        sheet["B5"] = "STAFF MAE"
        sheet["C5"] = "STAFF MAE"
        sheet["D5"] = "DIRECCIÓN JURÍDICA"
        sheet["E5"] = "JURÍDICA"
        sheet["F5"] = "EM-DJR-01"
        for column in "BCDEF":
            sheet.merge_cells(f"{column}5:{column}166")
        sheet["A6"] = "ACP"
        sheet["S7"] = "Acción de corto plazo simulada"

        sheet["U8"] = "000 0 001"
        sheet["V8"] = "=#REF!"
        sheet["W8"] = "Operación simulada"
        sheet["Z8"] = "Producto intermedio simulado"
        sheet["AA8"] = "Indicador de operación"
        sheet["AB8"] = "Meta ejecutada / meta programada"
        sheet["AC8"] = "Número"
        sheet["AH8"] = 1
        sheet["BJ8"] = 12
        sheet["AI8"] = datetime(2027, 1, 1)
        sheet["AJ8"] = datetime(2027, 6, 30)
        for index, column in enumerate(
            ("AL", "AN", "AP", "AR", "AT", "AV", "AX", "AZ", "BB", "BD", "BF", "BH"),
            1,
        ):
            sheet[f"{column}8"] = index

        sheet["X9"] = "Actividad simulada"
        sheet["Z9"] = "Resultado de actividad simulado"
        sheet["AA9"] = "Indicador de actividad"
        sheet["AC9"] = "Número"
        sheet["AH9"] = 2
        sheet["BJ9"] = 2
        sheet["AI9"] = datetime(2027, 1, 1)
        sheet["AJ9"] = datetime(2027, 6, 30)
        for month, column in enumerate(
            ("AL", "AN", "AP", "AR", "AT", "AV", "AX", "AZ", "BB", "BD", "BF", "BH"),
            1,
        ):
            sheet[f"{column}9"] = 1 if month <= 2 else 0
        sheet["Y10"] = "Tarea simulada"
        sheet["AA10"] = "Indicador de tarea"
        sheet["AC10"] = "Número"
        sheet["AH10"] = 2
        sheet["BJ10"] = 2
        sheet["AI10"] = datetime(2027, 1, 1)
        sheet["AJ10"] = datetime(2027, 6, 30)
        sheet["Y167"] = "No debe importarse"
        workbook.save(self._temporary_path())
        workbook.close()
        return self._temporary_path()

    def _temporary_path(self):
        return "/tmp/formulacion-poau-2027-test.xlsx"

    def _run(self, *extra):
        output = StringIO()
        call_command(
            "importar_formulacion_poau_2027",
            "--file",
            str(self.workbook_path),
            "--sheet",
            "Base",
            "--max-row",
            "166",
            *extra,
            stdout=output,
        )
        return json.loads(output.getvalue())

    def test_dry_run_does_not_write_and_excludes_formula_errors(self):
        report = self._run("--dry-run")

        self.assertEqual(report["rows_seen"], 166)
        self.assertEqual(report["candidate_rows"], 4)
        self.assertEqual(report["imported"], 0)
        self.assertEqual(report["inherited_context"]["A"], "EM-DJR-01")
        self.assertEqual(report["inherited_context"]["F"], "EM-DJR-01")
        self.assertEqual(report["warnings_count"], len(report["warnings"]))
        self.assertEqual(AccionCortoPlazo.objects.count(), 0)
        self.assertEqual(OperacionPOAU.objects.count(), 0)
        self.assertTrue(any("#REF!" in warning for warning in report["warnings"]))

    def test_commit_creates_simulated_chain(self):
        report = self._run("--commit")

        self.assertEqual(report["operational_imported"], 4)
        self.assertEqual(AccionCortoPlazo.objects.count(), 1)
        self.assertEqual(OperacionPOAU.objects.count(), 1)
        self.assertEqual(ActividadPOAU.objects.count(), 1)
        self.assertEqual(TareaPOAU.objects.count(), 1)
        self.assertEqual(POAU.objects.count(), 1)
        self.assertEqual(ResultadoPEI.objects.count(), 1)
        self.assertEqual(ProductoPEI.objects.count(), 1)
        self.assertEqual(AccionPOA.objects.count(), 1)
        self.assertTrue(
            AccionPOA.objects.get().codigo_accion.startswith("SIM-2027-POA-")
        )

    def test_commit_populates_native_poau_activity_and_physical_schedule(self):
        report = self._run("--commit")

        native_activity = POAUActividad.objects.get()
        self.assertEqual(report["native_imported"], 13)
        self.assertEqual(native_activity.codigo, "SIM-2027-ACT-EM-DJR-01-01")
        self.assertEqual(native_activity.poau, POAU.objects.get())
        self.assertEqual(native_activity.accion_corto_plazo, AccionCortoPlazo.objects.get())
        self.assertEqual(native_activity.meta_fisica_anual, 2)
        self.assertEqual(native_activity.meta_q1, 2)
        self.assertIsNone(native_activity.presupuesto_anual)
        self.assertIsNone(native_activity.objeto_gasto)
        executions = EjecucionFisica.objects.filter(actividad=native_activity)
        self.assertEqual(executions.count(), 12)
        self.assertEqual(
            list(executions.order_by("periodo").values_list("periodo", "programado")[:2]),
            [("2027-01", 1), ("2027-02", 1)],
        )
        self.assertEqual(report["native_limitations"], [
            "TareaPOAU no tiene equivalente en apps.poau; se conserva en apps.articulacion.",
        ])

    def test_native_mapping_keeps_unsynchronized_source_schedule_without_budget(self):
        workbook = openpyxl.load_workbook(self.workbook_path)
        workbook["Base"]["BJ9"] = 3
        workbook["Base"]["AP9"] = 1
        workbook.save(self.workbook_path)
        workbook.close()

        report = self._run("--commit")

        native_activity = POAUActividad.objects.get()
        self.assertIsNone(native_activity.meta_q1)
        self.assertIsNone(native_activity.meta_q2)
        self.assertEqual(
            list(
                EjecucionFisica.objects.filter(actividad=native_activity)
                .order_by("periodo")
                .values_list("periodo", "programado")[:2]
            ),
            [("2027-01", 1), ("2027-02", 1)],
        )
        self.assertIsNone(native_activity.presupuesto_anual)
        self.assertTrue(any("meta actual" in warning for warning in report["warnings"]))

    def test_native_activity_is_visible_through_poau_api(self):
        self._run("--commit")
        client = APIClient()
        client.force_authenticate(
            user=get_user_model().objects.create_user(
                email="poau-import-test@example.test",
                password="test-password",
            )
        )

        response = client.get(
            "/api/v1/poau/actividades/",
            {"poau": POAU.objects.get().id},
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(
            response.data["results"][0]["codigo"],
            "SIM-2027-ACT-EM-DJR-01-01",
        )

    def test_commit_is_idempotent_and_does_not_overwrite(self):
        first = self._run("--commit")
        action = AccionPOA.objects.get()
        action_name = action.denominacion
        second = self._run("--commit")

        self.assertGreater(first["imported"], 0)
        self.assertEqual(second["imported"], 0)
        self.assertGreater(second["existing_total"], 0)
        self.assertEqual(AccionPOA.objects.count(), 1)
        self.assertEqual(AccionPOA.objects.get().denominacion, action_name)
        self.assertEqual(POAUActividad.objects.count(), 1)
        self.assertEqual(EjecucionFisica.objects.count(), 12)

    def test_max_row_166_excludes_later_source_rows(self):
        report = self._run("--dry-run")

        self.assertEqual(report["max_row"], 166)
        self.assertEqual(report["rows_seen"], 166)
        self.assertNotIn("No debe importarse", str(report))
