"""Merge-aware parser and importer for the simulated 2027 POAU workbook.

The source does not contain complete official PAD/PEI/POA/POAU keys. This module
therefore creates a clearly namespaced ``SIM-2027`` chain and only updates rows
in that namespace. Monetary fields are intentionally not mapped from this
workbook.
"""

import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction

try:
    import openpyxl
except ImportError:  # pragma: no cover - surfaced by the command
    openpyxl = None


GESTION = 2027
MONTH_COLUMNS = ("AL", "AN", "AP", "AR", "AT", "AV", "AX", "AZ", "BB", "BD", "BF", "BH")
ERROR_MARKERS = ("#REF!", "#N/A")
LEVEL_COLUMNS = {"acp": "S", "operation": "W", "activity": "X", "task": "Y"}


def _text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _decimal(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    value = _text(value).replace(",", "")
    if not value:
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


def _json_number(value):
    if value is None or not isinstance(value, Decimal):
        return value
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _slug(value):
    value = re.sub(r"[^A-Za-z0-9]+", "-", _text(value)).strip("-").upper()
    return value or "SIN-CODIGO"


def _error_value(value):
    return isinstance(value, str) and any(marker in value.upper() for marker in ERROR_MARKERS)


class _WorkbookReader:
    def __init__(self, file_path, sheet_name, max_row):
        if openpyxl is None:
            raise ValueError("openpyxl no está instalado.")
        self.path = Path(file_path)
        if not self.path.exists():
            raise FileNotFoundError(f"No existe el workbook: {self.path}")
        self.sheet_name = sheet_name
        self.max_row = max_row
        self.wb_formula = openpyxl.load_workbook(self.path, data_only=False, read_only=False)
        self.wb_value = openpyxl.load_workbook(self.path, data_only=True, read_only=False)
        if sheet_name not in self.wb_formula.sheetnames:
            self.close()
            raise ValueError(f'La hoja "{sheet_name}" no existe. Hojas: {self.wb_formula.sheetnames}')
        self.formula_sheet = self.wb_formula[sheet_name]
        self.value_sheet = self.wb_value[sheet_name]
        self.warnings = []
        self._warning_keys = set()
        self._merge_anchors = self._build_merge_anchors(self.formula_sheet)

    @staticmethod
    def _build_merge_anchors(sheet):
        anchors = {}
        for merged_range in sheet.merged_cells.ranges:
            anchor = merged_range.start_cell.coordinate
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_col, merged_range.max_col + 1):
                    anchors[(row, column)] = anchor
        return anchors

    def close(self):
        self.wb_formula.close()
        self.wb_value.close()

    def _warn(self, key, message):
        if key not in self._warning_keys:
            self._warning_keys.add(key)
            self.warnings.append(message)

    def value(self, row, column):
        coordinate = f"{column}{row}"
        anchor = self._merge_anchors.get((row, self.formula_sheet[coordinate].column), coordinate)
        formula_value = self.formula_sheet[anchor].value
        cached_value = self.value_sheet[anchor].value
        if isinstance(formula_value, str) and formula_value.startswith("="):
            if _error_value(formula_value) or _error_value(cached_value):
                self._warn(
                    f"formula:{anchor}",
                    f"{anchor}: fórmula con #REF!/#N/A excluida del valor importado.",
                )
                return None
            if cached_value is None:
                self._warn(
                    f"formula-cache:{anchor}",
                    f"{anchor}: fórmula sin valor calculado excluida del valor importado.",
                )
                return None
            return cached_value
        value = cached_value
        if _error_value(value):
            self._warn(
                f"value:{anchor}",
                f"{anchor}: valor #REF!/#N/A excluido del valor importado.",
            )
            return None
        return value

    def read_catalog(self):
        sheet_name = "CLASIFICADOR cat. progra"
        if sheet_name not in self.wb_value.sheetnames:
            self._warn("catalog-sheet", f"No existe la hoja de catálogo {sheet_name}; se conserva solo el código.")
            return {}
        sheet = self.wb_value[sheet_name]
        catalog = {}
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if len(row) < 6:
                continue
            code = _text(row[4])
            if code and code.upper() not in ERROR_MARKERS:
                catalog[code] = {"description": _text(row[5]), "level": _text(row[6])}
        return catalog


def _contextual_rows(reader):
    context = {
        "organization": {column: "" for column in "ABCDEF"},
        "acp": "",
        "category": "",
        "category_description": "",
        "operation": "",
        "activity": "",
    }
    catalog = reader.read_catalog()
    rows = []
    columns = (
        "A", "B", "C", "D", "E", "F", "S", "U", "V", "W", "X", "Y", "Z",
        "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "BJ",
    )
    for row_number in range(1, reader.max_row + 1):
        values = {column: reader.value(row_number, column) for column in columns}
        if row_number < 5:
            continue
        for column in "ABCDEF":
            value = _text(values[column])
            if column == "A" and row_number > 5 and value.upper() in {"ACP", "OPE", "P"}:
                continue
            if value:
                context["organization"][column] = value
        for key, column in (("acp", "S"), ("category", "U"), ("operation", "W"), ("activity", "X")):
            value = _text(values[column])
            if value:
                context[key] = value
        category = context["category"]
        if category in catalog:
            context["category_description"] = catalog[category]["description"]
        elif category:
            reader._warn(f"category:{category}", f"Código de categoría {category} no fue resuelto en el catálogo.")

        level = None
        if _text(values["S"]):
            level = "acp"
        elif _text(values["W"]):
            level = "operation"
        elif _text(values["X"]):
            level = "activity"
        elif _text(values["Y"]):
            level = "task"
        if not level:
            continue

        record = {
            "source_row": row_number,
            "level": level,
            "organization": dict(context["organization"]),
            "acp": context["acp"],
            "category": context["category"],
            "category_description": context["category_description"],
            "operation": context["operation"],
            "activity": context["activity"],
            "name": _text(values[LEVEL_COLUMNS[level]]),
            "intermediate_result": _text(values["Z"]),
            "indicator": _text(values["AA"]),
            "formula": _text(values["AB"]),
            "unit": _text(values["AC"]),
            "baseline_2026": _decimal(values["AE"]),
            "target": _decimal(values["AF"]),
            "target_modification": _decimal(values["AG"]),
            "current_target": _decimal(values["AH"]),
            "date_start": _date(values["AI"]),
            "date_end": _date(values["AJ"]),
            "annual_total": _decimal(values["BJ"]),
            "monthly": {
                f"{GESTION}-{month:02d}": _json_number(_decimal(reader.value(row_number, column)))
                for month, column in enumerate(MONTH_COLUMNS, 1)
            },
            "source_key": f"{reader.path.name}:{reader.sheet_name}:{row_number}",
        }
        if not record["name"]:
            reader._warn(f"row:{row_number}:name", f"Fila {row_number}: registro {level} sin descripción; omitido.")
            continue
        rows.append(record)
        if record["current_target"] is not None and record["annual_total"] is not None and record["current_target"] != record["annual_total"]:
            reader._warn(
                f"discrepancy:{row_number}",
                f"Fila {row_number}: meta actual ({record['current_target']}) != total anual programado ({record['annual_total']}); se conservan ambos como warning.",
            )
    return rows


def _report_base(reader, rows, max_row):
    operational = {"acp": 0, "operation": 0, "activity": 0, "task": 0}
    for row in rows:
        operational[row["level"]] += 1
    native = {
        "POAUActividad": operational["activity"],
        "EjecucionFisica": sum(
            value is not None
            for row in rows
            if row["level"] == "activity"
            for value in row["monthly"].values()
        ),
        "EjecucionFinanciera": 0,
    }
    if reader.formula_sheet.max_row > max_row:
        reader._warn("row-limit", f"Se aplicó --max-row={max_row}; filas posteriores no fueron revisadas.")
    reader._warn("simulation", "Los registros se namespacean como SIM-2027; no representan códigos oficiales ni se mezclan con datos legacy.")
    reader._warn("budget", "No se importaron montos presupuestarios: la fuente no contiene presupuesto monetario confiable en estas filas.")
    reader._warn("official-codes", "No se importaron códigos oficiales PAD/PEI/POA/POAU completos; se generó únicamente la cadena SIM-2027.")
    limitations = []
    if operational["task"]:
        limitations.append(
            "TareaPOAU no tiene equivalente en apps.poau; se conserva en apps.articulacion."
        )
    return {
        "file": str(reader.path),
        "sheet": reader.sheet_name,
        "max_row": max_row,
        "rows_seen": max_row,
        "candidate_rows": len(rows),
        "skipped": max_row - len(rows),
        "errors": [],
        "warnings": list(reader.warnings),
        "inherited_context": rows[0]["organization"] if rows else {},
        "operational_candidates": operational,
        "planned_support_records": 9 if rows else 0,
        "native_candidates": native,
        "planned_total": len(rows) + (9 if rows else 0) + sum(native.values()),
        "native_imported": 0,
        "native_existing": 0,
        "native_limitations": limitations,
        "imported": 0,
        "operational_imported": 0,
        "support_imported": 0,
        "existing_total": 0,
        "record_counts": {},
    }


def _provenance(row, reader):
    return f"[SIM-2027] Fuente simulada: {reader.path.name}, hoja {reader.sheet_name}, fila {row['source_row']}."


def _get_or_create(model, lookup, defaults, report, label):
    obj, created = model.objects.get_or_create(defaults=defaults, **lookup)
    counts = report["record_counts"].setdefault(label, {"created": 0, "existing": 0})
    counts["created" if created else "existing"] += 1
    if created:
        report["imported"] += 1
    else:
        report["existing_total"] += 1
    return obj, created


def _get_or_update_simulated(model, lookup, defaults, report, label):
    """Upsert a SIM-2027 row without touching official or legacy records."""
    obj, created = model.objects.get_or_create(defaults=defaults, **lookup)
    counts = report["record_counts"].setdefault(label, {"created": 0, "existing": 0})
    counts["created" if created else "existing"] += 1
    if created:
        report["imported"] += 1
    else:
        report["existing_total"] += 1
        changed = False
        for field, value in defaults.items():
            if getattr(obj, field) != value:
                setattr(obj, field, value)
                changed = True
        if changed:
            obj.save()
    return obj, created


def _quarterly_targets(row):
    """Return quarterly physical targets only when they reconcile to the source."""
    annual_target = _decimal(row["current_target"])
    monthly = [
        _decimal(row["monthly"].get(f"{GESTION}-{month:02d}"))
        for month in range(1, 13)
    ]
    if annual_target is None or any(value is None for value in monthly):
        return [None, None, None, None]

    quarterly = [
        sum(monthly[index:index + 3], Decimal("0"))
        for index in range(0, 12, 3)
    ]
    if sum(quarterly, Decimal("0")) != annual_target:
        return [None, None, None, None]
    return quarterly


def _native_imported_counts(report):
    labels = {"POAUActividad", "EjecucionFisica", "EjecucionFinanciera"}
    report["native_imported"] = sum(
        report["record_counts"].get(label, {}).get("created", 0)
        for label in labels
    )
    report["native_existing"] = sum(
        report["record_counts"].get(label, {}).get("existing", 0)
        for label in labels
    )


@transaction.atomic
def _write_records(rows, reader, report):
    from apps.articulacion.models import AccionPOA, ActividadPOAU, OperacionPOAU, ProductoPEI, ResultadoPEI, TareaPOAU
    from apps.organizacion.models import TipoUnidad, UnidadOrganizacional
    from apps.planificacion.models import AccionCortoPlazo, AccionMedianoPlazo, NodoPlanificacion, Plan
    from apps.poau.models import EjecucionFisica, POAU, POAUActividad

    first = rows[0]
    organization_code = _slug(first["organization"].get("F") or "EM-DJR-01")
    base_code = f"SIM-2027-{organization_code}"
    provenance = _provenance(first, reader)
    start = first["date_start"] or date(GESTION, 1, 1)
    end = first["date_end"] or date(GESTION, 12, 31)

    area_type, _ = _get_or_create(TipoUnidad, {"codigo": "SIM-2027-AREA"}, {"nombre": "Área organizacional simulada SIM-2027", "nivel": 3}, report, "TipoUnidad")
    unit, _ = _get_or_create(
        UnidadOrganizacional,
        {"codigo": base_code, "gestion": GESTION},
        {"nombre": f"{first['organization'].get('E') or 'JURÍDICA'} — SIM-2027", "sigla": organization_code, "tipo": area_type, "gestion": GESTION, "fecha_vigencia_desde": start, "fecha_vigencia_hasta": end},
        report,
        "UnidadOrganizacional",
    )
    plan, _ = _get_or_create(
        Plan,
        {"codigo": f"{base_code}-PLAN-PEI", "tipo": "pei"},
        {"nombre": "Plan Estratégico Institucional simulado 2027", "gestion_inicio": GESTION, "gestion_fin": GESTION, "descripcion": provenance, "fecha_vigencia_desde": start},
        report,
        "Plan",
    )
    node, _ = _get_or_create(
        NodoPlanificacion,
        {"plan": plan, "nivel": "accion_mediano", "codigo": f"{base_code}-AMP"},
        {"nombre": "Acción de mediano plazo simulada", "descripcion": provenance, "gestion": GESTION, "orden": 1},
        report,
        "NodoPlanificacion",
    )
    amp, _ = _get_or_create(
        AccionMedianoPlazo,
        {"codigo": f"{base_code}-AMP"},
        {"nombre": "Acción de mediano plazo simulada", "descripcion": provenance, "nodo_planificacion": node, "gestion_inicio": GESTION, "gestion_fin": GESTION},
        report,
        "AccionMedianoPlazo",
    )
    result_pei, _ = _get_or_create(
        ResultadoPEI,
        {"codigo_resultado": f"SIM-2027-PEI-{organization_code}", "vigencia_desde": GESTION},
        {"denominacion": "Resultado PEI simulado de servicios jurídicos", "cod_entidad": "SIM-2027", "entidad": "Gobierno Autónomo Municipal de Sacaba — dato simulado", "cod_oei": "SIM-2027", "vigencia_hasta": GESTION},
        report,
        "ResultadoPEI",
    )
    product_pei, _ = _get_or_create(
        ProductoPEI,
        {"codigo_producto": f"SIM-2027-PEI-{organization_code}-01", "resultado_pei": result_pei},
        {"denominacion": first["acp"], "cod_programa_presup": "SIM-2027", "programa_presup": "Producto PEI simulado; sin presupuesto monetario importado."},
        report,
        "ProductoPEI",
    )
    action_row = next(row for row in rows if row["level"] == "operation")
    action, _ = _get_or_create(
        AccionPOA,
        {"codigo_accion": f"SIM-2027-POA-{organization_code}-01"},
        {"denominacion": first["acp"], "resultado_esperado": action_row["intermediate_result"], "producto_pei": product_pei, "indicador": action_row["indicator"], "formula": action_row["formula"], "unidad_medida": action_row["unit"], "fecha_inicio": action_row["date_start"] or start, "fecha_fin": action_row["date_end"] or end, "tipo_operacion": "SIMULADA", "categoria_programatica": action_row["category"], "programa": action_row["category_description"], "medio_verificacion": provenance, "riesgo": "Registro simulado; requiere validación de códigos y fuente oficial.", "gestion": GESTION, "unidad_responsable": unit, "estado": "REFERENCIAL"},
        report,
        "AccionPOA",
    )
    poau, _ = _get_or_create(
        POAU,
        {"codigo": f"SIM-2027-POAU-{organization_code}-01"},
        {"unidad": unit, "gestion": GESTION, "nombre": f"POAU simulado — {first['organization'].get('E') or 'JURÍDICA'}", "descripcion": provenance, "estado": "borrador"},
        report,
        "POAU",
    )
    acp, _ = _get_or_create(
        AccionCortoPlazo,
        {"codigo": f"SIM-2027-ACP-{organization_code}-01", "gestion": GESTION},
        {"nombre": first["acp"], "descripcion": provenance, "justificacion": "Acción simulada para visualizar la formulación 2027; no es un código oficial.", "accion_mediano_plazo": amp, "unidad_responsable": unit, "gestion": GESTION, "fecha_inicio": start, "fecha_fin": end},
        report,
        "AccionCortoPlazo",
    )
    operation, _ = _get_or_create(
        OperacionPOAU,
        {"codigo_operacion": f"SIM-2027-OPE-{organization_code}-01"},
        {"denominacion": action_row["name"], "tipo_operacion": "SIMULADA", "producto_entregable": action_row["intermediate_result"], "accion_poa": action, "unidad_ejecutora": first["organization"].get("E") or "JURÍDICA", "codigo_unidad_ejecutora": organization_code, "responsable": first["organization"].get("E") or "JURÍDICA", "codigo_responsable": organization_code, "meta_anual": action_row["current_target"], "indicador": action_row["indicator"], "formula": action_row["formula"], "unidad_medida": action_row["unit"], "fecha_inicio": action_row["date_start"] or start, "fecha_fin": action_row["date_end"] or end, "programacion_mensual": action_row["monthly"], "total_programado": action_row["annual_total"], "medio_verificacion": provenance, "requerimientos": "Registro simulado; no se importaron montos presupuestarios.", "riesgo": "Meta actual y total anual presentan discrepancias en la fuente cuando corresponde.", "estado": "REFERENCIAL"},
        report,
        "OperacionPOAU",
    )

    activity_number = 0
    task_number_by_activity = defaultdict(int)
    activities = {}
    for row in rows:
        if row["level"] == "activity":
            activity_number += 1
            activity_code = f"SIM-2027-ACT-{organization_code}-{activity_number:02d}"
            activity, _ = _get_or_create(
                ActividadPOAU,
                {"codigo_actividad": activity_code},
                {"denominacion": row["name"], "operacion": operation, "producto_entregable": row["intermediate_result"], "meta_anual": row["current_target"], "indicador": row["indicator"], "formula": row["formula"], "unidad_medida": row["unit"], "fecha_inicio": row["date_start"] or start, "fecha_fin": row["date_end"] or end, "programacion_mensual": row["monthly"], "total_programado": row["annual_total"], "medio_verificacion": provenance, "requerimientos": "Registro simulado; sin detalle presupuestario confiable.", "riesgo": "Verificar meta física frente al total anual programado.", "estado": "REFERENCIAL"},
                report,
                "ActividadPOAU",
            )
            quarterly_targets = _quarterly_targets(row)
            native_activity, _ = _get_or_update_simulated(
                POAUActividad,
                {"poau": poau, "codigo": activity_code},
                {
                    "nombre": row["name"],
                    "objeto_gasto": None,
                    "meta_fisica_anual": row["current_target"],
                    "presupuesto_anual": None,
                    "meta_q1": quarterly_targets[0],
                    "meta_q2": quarterly_targets[1],
                    "meta_q3": quarterly_targets[2],
                    "meta_q4": quarterly_targets[3],
                    "accion_corto_plazo": acp,
                },
                report,
                "POAUActividad",
            )
            for month in range(1, 13):
                planned = _decimal(row["monthly"].get(f"{GESTION}-{month:02d}"))
                if planned is None:
                    continue
                _get_or_update_simulated(
                    EjecucionFisica,
                    {
                        "actividad": native_activity,
                        "periodo": f"{GESTION}-{month:02d}",
                    },
                    {
                        "tipo_periodo": "mensual",
                        "programado": planned,
                        "ejecutado": Decimal("0"),
                        "observaciones": (
                            f"{provenance} No se reportó ejecución mensual en la fuente; "
                            "se conserva en cero."
                        ),
                    },
                    report,
                    "EjecucionFisica",
                )
            activities[activity_number] = activity
        elif row["level"] == "task":
            if not activities:
                report["errors"].append(f"Fila {row['source_row']}: tarea sin actividad padre.")
                continue
            task_number_by_activity[activity_number] += 1
            task_code = f"SIM-2027-TAR-{organization_code}-{activity_number:02d}-{task_number_by_activity[activity_number]:02d}"
            _get_or_create(
                TareaPOAU,
                {"codigo_tarea": task_code},
                {"denominacion": row["name"], "actividad": activities[activity_number], "responsable": first["organization"].get("E") or "JURÍDICA", "fecha_inicio": row["date_start"] or start, "fecha_fin": row["date_end"] or end, "metas": row["current_target"], "programacion_mensual": row["monthly"], "requerimientos": "Registro simulado; no se importaron montos.", "evidencia": provenance, "estado": "REFERENCIAL"},
                report,
                "TareaPOAU",
            )

    report["operational_imported"] = sum(
        counts.get("created", 0)
        for label, counts in report["record_counts"].items()
        if label in {"AccionCortoPlazo", "OperacionPOAU", "ActividadPOAU", "TareaPOAU"}
    )
    _native_imported_counts(report)
    report["support_imported"] = (
        report["imported"]
        - report["operational_imported"]
        - report["native_imported"]
    )
    return report


def import_formulacion_poau_2027(file_path, sheet_name="Base", max_row=166, commit=False):
    """Parse and optionally persist the selected workbook rows."""
    reader = _WorkbookReader(file_path, sheet_name, max_row)
    try:
        rows = _contextual_rows(reader)
        report = _report_base(reader, rows, max_row)
        if not rows:
            report["errors"].append("No se encontraron filas operativas válidas en el rango seleccionado.")
        elif commit:
            report = _write_records(rows, reader, report)
        else:
            report["record_counts"] = {
                "AccionCortoPlazo": {"planned": report["operational_candidates"]["acp"]},
                "OperacionPOAU": {"planned": report["operational_candidates"]["operation"]},
                "ActividadPOAU": {"planned": report["operational_candidates"]["activity"]},
                "TareaPOAU": {"planned": report["operational_candidates"]["task"]},
                "POAUActividad": {"planned": report["native_candidates"]["POAUActividad"]},
                "EjecucionFisica": {"planned": report["native_candidates"]["EjecucionFisica"]},
                "EjecucionFinanciera": {"planned": report["native_candidates"]["EjecucionFinanciera"]},
            }
        report["errors_count"] = len(report["errors"])
        report["warnings_count"] = len(report["warnings"])
        report["mode"] = "commit" if commit else "dry-run"
        return report
    finally:
        reader.close()
