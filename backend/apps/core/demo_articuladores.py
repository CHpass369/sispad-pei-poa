"""Safe refresh of the provisional 2027 articulation demonstration dataset."""

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction

try:
    import openpyxl
except ImportError:  # pragma: no cover - handled with a clear runtime error
    openpyxl = None

from apps.articulacion.models import (
    AccionPOA,
    ActividadPOAU,
    ArticulacionPADPEI,
    AsignacionObjetoGasto,
    IndicadorCadena,
    LineamientoPAD as LineamientoPADLegacy,
    OperacionPOAU,
    ProductoPAD,
    ProductoPEI,
    ResultadoPAD,
    ResultadoPEI,
    SeguimientoPresupuesto,
    TareaPOAU,
)
from apps.catalogos.models import (
    ClasificadorInstitucional,
    FinalidadFuncion,
    FuenteFinanciamiento,
    ObjetoGasto,
    OrganismoFinanciador,
    VersionClasificador,
)
from apps.codificacion.models import (
    ComponentePDESA,
    EjePGDESA,
    EntidadCodificadora,
    EntidadTerritorialCGEO,
    LineamientoPAD,
    ResultadoSectorial,
    SectorEconomico,
    VersionCatalogoPlan,
)
from apps.codificacion.services.codificador import CodificadorService
from apps.core.models import DemoDatasetManifest
from apps.gestion.models import GestionFiscal
from apps.organizacion.models import (
    DireccionAdministrativa,
    TipoUnidad,
    UnidadEjecutora,
    UnidadOrganizacional,
)
from apps.planificacion.models import ArticulacionPlanificacion, NodoPlanificacion, Plan
from apps.poau.models import EjecucionFinanciera, EjecucionFisica, POAU, POAUActividad
from apps.presupuesto.models import (
    ActividadPresupuestaria,
    AsignacionPresupuestariaUnidad,
    CategoriaProgramatica,
    LineaPresupuestaria,
    ProgramaPresupuestario,
    ProyectoPresupuestario,
)
from apps.seguimiento.models import EntradaSeguimiento, ReporteSeguimiento
from apps.techos.models import DistribucionTecho, TechoPresupuestario


EXPECTED_SOURCE_COUNTS = {
    'acciones': 1,
    'operaciones': 1,
    'actividades': 19,
    'tareas': 139,
    'programaciones_fisicas': 228,
}


def _text(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().split())


def _decimal(value, default='0'):
    if value in (None, ''):
        return Decimal(default)
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _date(value, default):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return default


@dataclass
class SourceTask:
    name: str
    deliverable: str
    indicator: str
    formula: str
    unit: str
    meta: Decimal
    start: date
    end: date


@dataclass
class SourceActivity:
    name: str
    deliverable: str
    indicator: str
    formula: str
    unit: str
    meta: Decimal
    start: date
    end: date
    programmed: list[Decimal]
    executed: list[Decimal]
    source_row: int
    tasks: list[SourceTask] = field(default_factory=list)


@dataclass
class SourceDataset:
    unit_code: str
    unit_name: str
    action_name: str
    operation_name: str
    category_code: str
    operation_deliverable: str
    operation_indicator: str
    operation_formula: str
    operation_unit: str
    operation_meta: Decimal
    operation_start: date
    operation_end: date
    activities: list[SourceActivity]
    file_hash: str

    @property
    def counts(self):
        return {
            'acciones': 1,
            'operaciones': 1,
            'actividades': len(self.activities),
            'tareas': sum(len(activity.tasks) for activity in self.activities),
            'programaciones_fisicas': len(self.activities) * 12,
        }


def load_source_dataset(source_file, gestion):
    """Read the first complete organizational block from the 2027 workbook."""
    if openpyxl is None:
        raise ValidationError('openpyxl no está instalado.')
    source_path = Path(source_file).expanduser().resolve()
    if not source_path.is_file():
        raise ValidationError(f'No existe el archivo fuente: {source_path}')

    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    if 'Base' not in workbook.sheetnames:
        workbook.close()
        raise ValidationError('El archivo fuente no contiene la hoja Base.')

    rows = list(workbook['Base'].iter_rows(min_row=5, values_only=True))
    workbook.close()
    first_unit_index = next(
        (index for index, row in enumerate(rows) if _text(row[0]).startswith('EM-')),
        None,
    )
    if first_unit_index is None:
        raise ValidationError('No se encontró una unidad EM-* en la hoja Base.')

    unit_code = _text(rows[first_unit_index][0])
    block = []
    for row in rows[first_unit_index:]:
        row_unit = _text(row[0])
        if block and row_unit.startswith('EM-') and row_unit != unit_code:
            break
        block.append(row)

    start_default = date(gestion, 1, 1)
    end_default = date(gestion, 12, 31)
    first = lambda column: next((_text(row[column]) for row in block if _text(row[column])), '')
    first_raw = lambda column: next((row[column] for row in block if row[column] not in (None, '')), None)
    activities = []
    current = None
    for offset, row in enumerate(block, start=first_unit_index + 5):
        activity_name = _text(row[23])
        task_name = _text(row[24])
        if activity_name:
            current = SourceActivity(
                name=activity_name,
                deliverable=_text(row[25]),
                indicator=_text(row[26]),
                formula=_text(row[27]),
                unit=_text(row[28]) or 'Número',
                meta=_decimal(row[33] if row[33] not in (None, '') else row[31]),
                start=_date(row[34], start_default),
                end=_date(row[35], end_default),
                programmed=[_decimal(row[column]) for column in range(37, 61, 2)],
                executed=[_decimal(row[column]) for column in range(38, 62, 2)],
                source_row=offset,
            )
            activities.append(current)
        elif task_name and current is not None:
            current.tasks.append(SourceTask(
                name=task_name,
                deliverable=_text(row[25]),
                indicator=_text(row[26]),
                formula=_text(row[27]),
                unit=_text(row[28]) or current.unit,
                meta=_decimal(row[33] if row[33] not in (None, '') else row[31]),
                start=_date(row[34], current.start),
                end=_date(row[35], current.end),
            ))

    operation_row = next((row for row in block if _text(row[22])), None)
    if operation_row is None:
        raise ValidationError('No se encontró la operación de la unidad seleccionada.')
    source = SourceDataset(
        unit_code=unit_code,
        unit_name=_text(rows[first_unit_index][4]) or 'Dirección Jurídica',
        action_name=first(18),
        operation_name=_text(operation_row[22]),
        category_code=_text(operation_row[20]) or '000 0 001',
        operation_deliverable=_text(operation_row[25]),
        operation_indicator=_text(operation_row[26]),
        operation_formula=_text(operation_row[27]),
        operation_unit=_text(operation_row[28]) or 'Porcentaje',
        operation_meta=_decimal(
            operation_row[33]
            if operation_row[33] not in (None, '')
            else operation_row[31],
            '100',
        ),
        operation_start=_date(operation_row[34], start_default),
        operation_end=_date(operation_row[35], end_default),
        activities=activities,
        file_hash=sha256(source_path.read_bytes()).hexdigest(),
    )
    if source.counts != EXPECTED_SOURCE_COUNTS:
        raise ValidationError(
            f'La forma del Excel no coincide con el contrato esperado: {source.counts}.'
        )
    return source


class DemoArticuladoresSeeder:
    NAMESPACE = 'demo-articuladores-numericos-v2'
    PREVIOUS_NAMESPACE = 'demo-articuladores-numericos-v1'
    PREVIOUS_MARKER = 'Demostración provisional articuladores'
    GESTION = 2027
    REFERENCE_NOTE = (
        'Dato demostrativo PROVISIONAL y referencial; no constituye fuente oficial.'
    )

    def __init__(self, *, source_file, gestion=GESTION, refresh=False):
        if int(gestion) != self.GESTION:
            raise ValidationError(
                'Este conjunto está respaldado por el Excel 2027 y solo admite gestión 2027.'
            )
        self.gestion = int(gestion)
        self.source_file = str(Path(source_file).expanduser().resolve())
        self.refresh = refresh
        self.source = load_source_dataset(self.source_file, self.gestion)
        self.created = Counter()
        self.reused = Counter()
        self.ids = {}
        self.bridges = {}
        self.ownership = {
            'owned': defaultdict(list),
            'reused': defaultdict(list),
        }
        self.retired_namespaces = []
        self._known_owned_ids = self._load_known_owned_ids()

    def _load_known_owned_ids(self):
        owned = defaultdict(set)
        manifest = DemoDatasetManifest.objects.filter(namespace=self.NAMESPACE).first()
        if manifest:
            for model_label, values in (
                manifest.payload.get('ownership', {}).get('owned', {}).items()
            ):
                owned[model_label].update(str(value) for value in values)
        return owned

    def _is_known_owned(self, obj):
        return str(obj.pk) in self._known_owned_ids[obj._meta.label]

    def _track(self, obj, created, label, *, owned=False):
        model_label = obj._meta.label
        obj_id = str(obj.pk)
        bucket = self.created if created else self.reused
        bucket[model_label] += 1
        ownership_key = 'owned' if created or owned or self._is_known_owned(obj) else 'reused'
        if obj_id not in self.ownership[ownership_key][model_label]:
            self.ownership[ownership_key][model_label].append(obj_id)
        self.ids[label] = obj_id
        return obj

    @staticmethod
    def _value_changed(obj, field_name, value):
        field = obj._meta.get_field(field_name)
        if field.is_relation:
            expected = getattr(value, 'pk', value)
            return getattr(obj, field.attname) != expected
        return getattr(obj, field_name) != value

    def _sync(self, obj, values):
        changed = False
        for field_name, value in values.items():
            if self._value_changed(obj, field_name, value):
                setattr(obj, field_name, value)
                changed = True
        if changed:
            obj.save()
        return obj

    def _get(self, model, label, defaults=None, *, sync=True, owned=False, **lookup):
        obj, created = model.objects.get_or_create(defaults=defaults or {}, **lookup)
        self._track(obj, created, label, owned=owned)
        if not created and sync and self.refresh and defaults:
            self._sync(obj, defaults)
        return obj

    def _owned_plan(self, label, *, codigo, tipo, defaults):
        existing = Plan.objects.filter(codigo=codigo, tipo=tipo).first()
        if existing is not None:
            is_previous_demo = self.PREVIOUS_MARKER in existing.nombre
            if not self._is_known_owned(existing) and not is_previous_demo:
                raise ValidationError(
                    f'El plan {tipo}/{codigo} ya existe y no pertenece al demo; no se modificó.'
                )
            self._track(existing, False, label, owned=True)
            if self.refresh:
                self._sync(existing, defaults)
            return existing
        obj = Plan.objects.create(codigo=codigo, tipo=tipo, **defaults)
        return self._track(obj, True, label)

    @staticmethod
    def _assert_relation(obj, field, expected):
        if getattr(obj, f'{field}_id') != expected.pk:
            raise ValidationError({field: 'La relación existente no coincide con la cadena 2027.'})

    def _classifier_version(self, tipo, label):
        return self._get(
            VersionClasificador,
            label,
            tipo=tipo,
            gestion=self.gestion,
            vigente=False,
            defaults={
                'codigo_fuente': '2026',
                'procedencia_normativa': (
                    'Clasificador 2026 reutilizado como referencia provisional para 2027.'
                ),
                'clasificacion_fuente': VersionClasificador.FUENTE_INCIERTA,
            },
        )

    def _code(self, instance):
        generated = CodificadorService.generar_codigo_completo(instance)
        if not instance.articulacion_incompleta:
            CodificadorService.validar_codigo(generated)
        instance.save(update_fields=[
            'codigo_completo_articulacion',
            'articulacion_incompleta',
            'updated_at',
        ])

    def run(self, *, commit=False):
        if not commit:
            return self._preview()
        return self._commit()

    def _preview(self):
        manifest = DemoDatasetManifest.objects.filter(namespace=self.NAMESPACE).first()
        return {
            'mode': 'dry-run',
            'namespace': self.NAMESPACE,
            'gestion': self.gestion,
            'source_file': self.source_file,
            'source_hash': self.source.file_hash,
            'source_counts': self.source.counts,
            'would_refresh': bool(manifest or self.refresh),
            'would_create_total': 0 if manifest else None,
            'manifest_id': str(manifest.pk) if manifest else None,
        }

    @transaction.atomic
    def _commit(self):
        self._seed_strategy()
        source_objects = self._seed_operational()
        native_objects = self._seed_native(source_objects)
        self._seed_budget(source_objects, native_objects)
        self._seed_tracking(native_objects)

        manifest, created = DemoDatasetManifest.objects.get_or_create(
            namespace=self.NAMESPACE,
            defaults={'gestion': self.gestion, 'payload': {}},
        )
        self._track(manifest, created, 'manifest', owned=True)
        payload = {
            'namespace': self.NAMESPACE,
            'gestion': self.gestion,
            'source': {
                'file': self.source_file,
                'sha256': self.source.file_hash,
                'sheet': 'Base',
                'status': 'PROVISIONAL',
            },
            'source_counts': self.source.counts,
            'ids': self.ids,
            'bridges': self.bridges,
            'ownership': {
                key: dict(sorted(values.items()))
                for key, values in self.ownership.items()
            },
            'financial_rule': (
                'monto_ejecutado <= monto_vigente <= monto_formulado; '
                'clasificadores 2026 usados solo como referencia provisional 2027.'
            ),
        }
        manifest.gestion = self.gestion
        manifest.payload = payload
        manifest.save(update_fields=['gestion', 'payload', 'updated_at'])

        previous = DemoDatasetManifest.objects.filter(
            namespace=self.PREVIOUS_NAMESPACE,
        ).first()
        if previous:
            previous.delete()
            self.retired_namespaces.append(self.PREVIOUS_NAMESPACE)

        return {
            'mode': 'commit',
            'namespace': self.NAMESPACE,
            'gestion': self.gestion,
            'source_counts': self.source.counts,
            'created_total': sum(self.created.values()),
            'reused_total': sum(self.reused.values()),
            'created_by_model': dict(sorted(self.created.items())),
            'reused_by_model': dict(sorted(self.reused.items())),
            'manifest_id': str(manifest.pk),
            'retired_namespaces': self.retired_namespaces,
        }

    def _seed_strategy(self):
        start = date(self.gestion, 1, 1)
        self._get(
            GestionFiscal,
            'gestion_fiscal',
            anio=self.gestion,
            defaults={
                'estado': GestionFiscal.Estado.ABIERTA,
                'descripcion': 'Gestión demostrativa 2027 abierta para formulación provisional.',
                'anio_inicio_plurianual': 2026,
                'anio_fin_plurianual': 2030,
            },
        )
        plan_pgdesa = self._owned_plan(
            'plan_pgdesa', codigo='04', tipo='pgdesa',
            defaults={
                'nombre': 'PGDESA referencial — eje institucional 04',
                'gestion_inicio': 2026,
                'gestion_fin': 2050,
                'fecha_vigencia_desde': start,
                'descripcion': self.REFERENCE_NOTE,
            },
        )
        plan_pdesa = self._owned_plan(
            'plan_pdesa', codigo='04.02', tipo='pdesa',
            defaults={
                'nombre': 'PDESA referencial — gestión pública eficaz',
                'gestion_inicio': 2026,
                'gestion_fin': 2030,
                'fecha_vigencia_desde': start,
                'descripcion': self.REFERENCE_NOTE,
            },
        )
        plan_pei = self._owned_plan(
            'plan_pei', codigo='1312.03', tipo='pei',
            defaults={
                'nombre': 'PEI referencial GAM Sacaba — fortalecimiento institucional',
                'gestion_inicio': 2026,
                'gestion_fin': 2030,
                'fecha_vigencia_desde': start,
                'descripcion': self.REFERENCE_NOTE,
            },
        )

        pg_eje = self._get(
            NodoPlanificacion, 'nodo_pgdesa_eje', plan=plan_pgdesa, codigo='04', nivel='eje',
            defaults={
                'nombre': 'Gestión pública transparente, eficaz y al servicio de la población',
                'gestion': self.gestion, 'orden': 1,
            }, owned=True,
        )
        pg_meta = self._get(
            NodoPlanificacion, 'nodo_pgdesa_meta', plan=plan_pgdesa, codigo='01', nivel='meta',
            defaults={
                'nombre': 'Fortalecer las capacidades institucionales de las entidades territoriales',
                'gestion': self.gestion, 'orden': 1, 'padre': pg_eje,
            }, owned=True,
        )
        self._assert_relation(pg_meta, 'padre', pg_eje)
        pg_result = self._get(
            NodoPlanificacion, 'nodo_pgdesa_resultado', plan=plan_pgdesa,
            codigo='01', nivel='resultado',
            defaults={
                'nombre': 'Entidades públicas con servicios oportunos y gestión administrativa fortalecida',
                'gestion': self.gestion, 'orden': 1, 'padre': pg_meta,
            }, owned=True,
        )
        self._assert_relation(pg_result, 'padre', pg_meta)
        pd_component = self._get(
            NodoPlanificacion, 'nodo_pdesa_componente', plan=plan_pdesa,
            codigo='02', nivel='componente',
            defaults={
                'nombre': 'Fortalecimiento institucional y seguridad jurídica',
                'gestion': self.gestion, 'orden': 1,
            }, owned=True,
        )
        pd_action = self._get(
            NodoPlanificacion, 'nodo_pdesa_accion', plan=plan_pdesa,
            codigo='01', nivel='accion',
            defaults={
                'nombre': 'Mejorar la gestión jurídica y la defensa de los intereses institucionales',
                'gestion': self.gestion, 'orden': 1, 'padre': pd_component,
            }, owned=True,
        )
        self._assert_relation(pd_action, 'padre', pd_component)
        bridge = self._get(
            ArticulacionPlanificacion, 'bridge_pgdesa_pdesa',
            nodo_origen=pg_result, nodo_destino=pd_component, gestion=self.gestion,
            defaults={'es_principal': True},
        )
        pei_objective = self._get(
            NodoPlanificacion, 'nodo_objetivo_pei', plan=plan_pei,
            codigo='03', nivel='accion_mediano',
            defaults={
                'nombre': 'Fortalecer la gestión institucional y la seguridad jurídica municipal',
                'gestion': self.gestion, 'orden': 1,
            }, owned=True,
        )

        catalog_version = self._get(
            VersionCatalogoPlan, 'version_catalogo_plan', plan=plan_pgdesa,
            gestion=self.gestion,
            defaults={
                'estado': VersionCatalogoPlan.ESTADO_BORRADOR,
                'clasificacion_fuente': VersionCatalogoPlan.FUENTE_INCIERTA,
                'procedencia_fuente': self.REFERENCE_NOTE,
            },
        )
        eje = self._get(
            EjePGDESA, 'catalogo_eje_pgdesa', version_catalogo=catalog_version,
            codigo='04', defaults={'denominacion': pg_eje.nombre},
        )
        component = self._get(
            ComponentePDESA, 'catalogo_componente_pdesa',
            version_catalogo=catalog_version, eje=eje, codigo='02',
            defaults={'denominacion': pd_component.nombre},
        )
        sector = self._get(
            SectorEconomico, 'catalogo_sector', version_catalogo=catalog_version,
            componente=component, codigo='14',
            defaults={'denominacion': 'Administración pública y servicios institucionales'},
        )
        sector_result = self._get(
            ResultadoSectorial, 'catalogo_resultado_sectorial',
            version_catalogo=catalog_version, sector=sector, codigo='01',
            defaults={'denominacion': pg_result.nombre},
        )
        department = self._get(
            EntidadTerritorialCGEO, 'cgeo_departamento', codigo='03',
            defaults={
                'nombre': 'Cochabamba',
                'nivel': EntidadTerritorialCGEO.NIVEL_DEPARTAMENTO,
                'estado': EntidadTerritorialCGEO.ESTADO_PROVISIONAL,
            }, sync=False,
        )
        province = self._get(
            EntidadTerritorialCGEO, 'cgeo_provincia', codigo='0310',
            defaults={
                'nombre': 'Chapare', 'nivel': EntidadTerritorialCGEO.NIVEL_PROVINCIA,
                'padre': department, 'estado': EntidadTerritorialCGEO.ESTADO_PROVISIONAL,
            }, sync=False,
        )
        municipality = self._get(
            EntidadTerritorialCGEO, 'cgeo_municipio', codigo='031001',
            defaults={
                'nombre': 'Sacaba', 'nivel': EntidadTerritorialCGEO.NIVEL_MUNICIPIO,
                'padre': province, 'estado': EntidadTerritorialCGEO.ESTADO_PROVISIONAL,
            }, sync=False,
        )
        entity = self._get(
            EntidadCodificadora, 'entidad_codificadora', codigo='1312',
            defaults={'denominacion': 'Gobierno Autónomo Municipal de Sacaba'},
            sync=False,
        )
        line_catalog = self._get(
            LineamientoPAD, 'catalogo_lineamiento_pad',
            version_catalogo=catalog_version, entidad_territorial=municipality, codigo='02',
            defaults={'denominacion': 'Gestión institucional y seguridad jurídica municipal'},
        )
        legacy_line = self._get(
            LineamientoPADLegacy, 'lineamiento_pad', codigo='02',
            gestion_desde=2026, gestion_hasta=2030,
            defaults={'denominacion': line_catalog.denominacion},
        )

        result_pad = self._get(
            ResultadoPAD, 'resultado_pad', codigo_resultado='01',
            vigencia_desde=self.gestion,
            defaults={
                'id_cadena': '2027.04.02.14.01',
                'denominacion': 'Gestión jurídica municipal fortalecida y servicios legales oportunos',
                'lineamiento_pad': legacy_line.codigo,
                'vigencia_hasta': 2030,
                'cod_geografico': municipality.codigo,
                'eta': 'Gobierno Autónomo Municipal de Sacaba',
                'resultado_sectorial_catalogo': sector_result,
                'entidad_territorial_cgeo': municipality,
                'lineamiento_pad_catalogo': line_catalog,
                'nodo_pdesa': pd_action,
                'cod_eje_pgdesa': '04',
                'objetivo_impacto': pg_result.nombre,
                'cod_componente_pdesa': '02',
                'objetivo_efecto': pd_action.nombre,
                'cod_sector': '14',
                'sector': sector.denominacion,
                'cod_resultado_pds': '01',
                'resultado_pds': sector_result.denominacion,
                'estado': 'PROVISIONAL',
                'correlativo': 1, 'segmento': '01',
                'codigo_normalizado': '01', 'codigo_fuente': 'REF-2027-PAD-01',
            },
        )
        product_pad = self._get(
            ProductoPAD, 'producto_pad', codigo_producto='01', resultado_pad=result_pad,
            defaults={
                'denominacion': 'Servicios jurídicos y patrocinio legal institucional prestados',
                'territorializacion': municipality.codigo,
                'responsable': 'Dirección Jurídica',
                'correlativo': 1, 'segmento': '01',
                'codigo_normalizado': '01', 'codigo_fuente': 'REF-2027-PAD-PROD-01',
            },
        )
        result_pei = self._get(
            ResultadoPEI, 'resultado_pei', codigo_resultado='01',
            vigencia_desde=self.gestion,
            defaults={
                'denominacion': 'Capacidad institucional jurídica fortalecida',
                'cod_entidad': '1312',
                'entidad': 'Gobierno Autónomo Municipal de Sacaba',
                'entidad_codificadora': entity,
                'cod_oei': pei_objective.codigo,
                'vigencia_hasta': 2030,
                'correlativo': 1, 'segmento': '01',
                'codigo_normalizado': '01', 'codigo_fuente': 'REF-2027-PEI-RES-01',
            },
        )
        product_pei = self._get(
            ProductoPEI, 'producto_pei', codigo_producto='01', resultado_pei=result_pei,
            defaults={
                'denominacion': 'Servicios de asesoramiento jurídico y defensa legal institucional',
                'cod_programa_presup': '000',
                'programa_presup': 'Gestión administrativa institucional',
                'correlativo': 1, 'segmento': '01',
                'codigo_normalizado': '01', 'codigo_fuente': 'REF-2027-PEI-PROD-01',
            },
        )
        pad_pei = self._get(
            ArticulacionPADPEI, 'bridge_pad_pei',
            producto_pad=product_pad, producto_pei=product_pei,
            defaults={
                'tipo_contribucion': 'directa',
                'ponderacion': Decimal('100.00'),
                'justificacion': 'Contribución directa, provisional y referencial para el demo 2027.',
                'estado': 'PROVISIONAL',
            },
        )
        self._get(
            IndicadorCadena, 'indicador_cadena',
            nivel_indicador='producto_pei', producto_pad=product_pad, producto_pei=product_pei,
            defaults={
                'indicador': 'Porcentaje de servicios jurídicos atendidos dentro del plazo programado',
                'tipo_indicador': 'eficacia',
                'unidad_medida': 'Porcentaje',
                'formula': '(Servicios atendidos / servicios programados) x 100',
                'linea_base': Decimal('78.0000'),
                'meta_2030': Decimal('95.0000'),
                'programacion_fisica': {'2027': '85.00'},
                'presupuesto_corriente_total': Decimal('285000.00'),
                'corriente_2027': Decimal('285000.00'),
                'fuente_dato': self.REFERENCE_NOTE,
            },
        )
        for coded in (result_pad, product_pad, result_pei, product_pei):
            self._code(coded)
        self.bridges.update({
            'catalog_to_node': {
                'eje': [str(eje.pk), str(pg_eje.pk)],
                'componente': [str(component.pk), str(pd_component.pk)],
                'resultado_sectorial': [str(sector_result.pk), str(pg_result.pk)],
            },
            'pgdesa_pdesa': str(bridge.pk),
            'pdesa_pad': str(result_pad.pk),
            'pad_pei': str(pad_pei.pk),
            'pei_objective_result': [str(pei_objective.pk), str(result_pei.pk)],
        })
        self.strategy = {
            'product_pei': product_pei,
            'product_pad': product_pad,
            'pei_objective': pei_objective,
        }

    def _segmented_source(self, model, label, legacy_field, source_code, numeric_code, defaults):
        obj = model.objects.filter(codigo_fuente=source_code).first()
        legacy_source = obj is not None
        if obj is None:
            obj = model.objects.filter(**{legacy_field: source_code}).first()
        if obj is None:
            obj = model.objects.filter(**{legacy_field: numeric_code}).first()
        if obj is None:
            obj = model.objects.create(
                **{legacy_field: numeric_code},
                codigo_fuente=source_code,
                **defaults,
            )
            self._track(obj, True, label)
        else:
            if not legacy_source and not self._is_known_owned(obj):
                raise ValidationError(
                    f'El registro {model._meta.label}/{numeric_code} ya existe y '
                    'no pertenece al demo; no se modificó.'
                )
            conflict = model.objects.filter(
                **{legacy_field: numeric_code}
            ).exclude(pk=obj.pk).first()
            if conflict is not None:
                raise ValidationError(
                    f'El código {numeric_code} ya pertenece a otro registro; no se modificó.'
                )
            self._track(obj, False, label, owned=True)
            values = {legacy_field: numeric_code, 'codigo_fuente': source_code, **defaults}
            self._sync(obj, values)
        return obj

    def _seed_operational(self):
        source = self.source
        start = date(self.gestion, 1, 1)
        action_source = f'SIM-2027-POA-{source.unit_code}-01'
        action = self._segmented_source(
            AccionPOA, 'accion_poa', 'codigo_accion', action_source,
            '2027.1312.001',
            {
                'denominacion': source.action_name,
                'resultado_esperado': source.operation_deliverable,
                'producto_pei': self.strategy['product_pei'],
                'indicador': source.operation_indicator,
                'formula': source.operation_formula,
                'unidad_medida': source.operation_unit,
                'linea_base': Decimal('0.0000'),
                'meta_gestion': Decimal('100.0000'),
                'fecha_inicio': start,
                'fecha_fin': date(self.gestion, 12, 31),
                'presupuesto_programado': Decimal('285000.00'),
                'fuente_financiamiento': '20',
                'organismo_financiador': '210',
                'estado': 'PROVISIONAL',
                'gestion': self.gestion,
                'correlativo': 1, 'segmento': '001',
                'codigo_normalizado': '001', 'estado_codigo': 'provisional',
            },
        )
        operation_source = f'SIM-2027-OPE-{source.unit_code}-01'
        operation = self._segmented_source(
            OperacionPOAU, 'operacion_poau', 'codigo_operacion', operation_source,
            '2027.1312.001.001',
            {
                'denominacion': source.operation_name,
                'tipo_operacion': 'funcionamiento',
                'producto_entregable': source.operation_deliverable,
                'accion_poa': action,
                'unidad_ejecutora': source.unit_name,
                'codigo_unidad_ejecutora': source.unit_code,
                'responsable': 'Dirección Jurídica',
                'meta_anual': source.operation_meta,
                'indicador': source.operation_indicator,
                'formula': source.operation_formula,
                'unidad_medida': source.operation_unit,
                'fecha_inicio': source.operation_start,
                'fecha_fin': source.operation_end,
                'estado': 'PROVISIONAL',
                'correlativo': 1, 'segmento': '001',
                'codigo_normalizado': '001', 'estado_codigo': 'provisional',
            },
        )
        activities = []
        tasks = []
        for activity_index, source_activity in enumerate(source.activities, start=1):
            activity_source = f'SIM-2027-ACT-{source.unit_code}-{activity_index:02d}'
            activity = self._segmented_source(
                ActividadPOAU, f'actividad_poau_{activity_index:02d}',
                'codigo_actividad', activity_source,
                f'2027.1312.001.001.{activity_index:03d}',
                {
                    'denominacion': source_activity.name,
                    'operacion': operation,
                    'producto_entregable': source_activity.deliverable,
                    'meta_anual': source_activity.meta,
                    'indicador': source_activity.indicator,
                    'formula': source_activity.formula,
                    'unidad_medida': source_activity.unit,
                    'fecha_inicio': source_activity.start,
                    'fecha_fin': source_activity.end,
                    'programacion_mensual': {
                        f'{month:02d}': str(value)
                        for month, value in enumerate(source_activity.programmed, start=1)
                    },
                    'total_programado': sum(source_activity.programmed, Decimal('0')),
                    'medio_verificacion': source_activity.deliverable,
                    'estado': 'PROVISIONAL',
                    'correlativo': activity_index,
                    'segmento': f'{activity_index:03d}',
                    'codigo_normalizado': f'{activity_index:03d}',
                    'estado_codigo': 'provisional',
                },
            )
            activities.append(activity)
            activity_tasks = []
            for task_index, source_task in enumerate(source_activity.tasks, start=1):
                task_source = (
                    f'SIM-2027-TAR-{source.unit_code}-{activity_index:02d}-{task_index:02d}'
                )
                task = self._segmented_source(
                    TareaPOAU,
                    f'tarea_poau_{activity_index:02d}_{task_index:02d}',
                    'codigo_tarea', task_source,
                    f'2027.1312.001.001.{activity_index:03d}.{task_index:03d}',
                    {
                        'denominacion': source_task.name,
                        'actividad': activity,
                        'responsable': source.unit_name,
                        'fecha_inicio': source_task.start,
                        'fecha_fin': source_task.end,
                        'metas': source_task.meta,
                        'requerimientos': source_task.deliverable,
                        'evidencia': source_task.indicator,
                        'estado': 'PROVISIONAL',
                        'correlativo': task_index,
                        'segmento': f'{task_index:03d}',
                        'codigo_normalizado': f'{task_index:03d}',
                        'estado_codigo': 'provisional',
                    },
                )
                activity_tasks.append(task)
                tasks.append(task)
            for task in activity_tasks:
                self._code(task)
            self._code(activity)
        self._code(action)
        self._code(operation)
        self.bridges['pei_poa'] = [
            str(self.strategy['product_pei'].pk), str(action.pk),
        ]
        return {'action': action, 'operation': operation, 'activities': activities, 'tasks': tasks}

    def _source_unit(self):
        unit_code = f'SIM-2027-{self.source.unit_code}'
        unit = UnidadOrganizacional.objects.filter(
            codigo=unit_code, gestion=self.gestion,
        ).first()
        if unit:
            self._track(unit, False, 'unidad_fuente')
            return unit
        unit_type = self._get(
            TipoUnidad, 'tipo_unidad', codigo='DIR',
            defaults={'nombre': 'Dirección municipal', 'nivel': 2}, sync=False,
        )
        unit = UnidadOrganizacional.objects.create(
            codigo=unit_code,
            gestion=self.gestion,
            nombre=self.source.unit_name,
            sigla='DJR',
            tipo=unit_type,
            fecha_vigencia_desde=date(self.gestion, 1, 1),
        )
        return self._track(unit, True, 'unidad_fuente')

    def _seed_native(self, source_objects):
        unit = self._source_unit()
        source_poau_code = f'SIM-2027-POAU-{self.source.unit_code}-01'
        numeric_poau_code = '2027.1312.001'
        poau = POAU.objects.filter(
            gestion=self.gestion, codigo=source_poau_code,
        ).first()
        legacy_poau = poau is not None
        if poau is None:
            poau = POAU.objects.filter(
                gestion=self.gestion, codigo=numeric_poau_code,
            ).first()
        poau_defaults = {
            'unidad': unit,
            'gestion': self.gestion,
            'nombre': 'POAU 2027 — Dirección Jurídica',
            'descripcion': self.REFERENCE_NOTE,
            'estado': 'borrador',
        }
        legacy_poau = legacy_poau or bool(
            poau
            and poau.nombre == poau_defaults['nombre']
            and poau.descripcion == self.REFERENCE_NOTE
            and poau.unidad.codigo.startswith('SIM-2027-')
        )
        if poau:
            if not legacy_poau and not self._is_known_owned(poau):
                raise ValidationError(
                    f'El POAU {numeric_poau_code} ya existe y no pertenece al demo; '
                    'no se modificó.'
                )
            self._track(poau, False, 'poau_nativo', owned=True)
            self._sync(poau, {'codigo': numeric_poau_code, **poau_defaults})
        else:
            poau = POAU.objects.create(codigo=numeric_poau_code, **poau_defaults)
            self._track(poau, True, 'poau_nativo')

        native_activities = []
        for index, (activity, source_activity) in enumerate(
            zip(source_objects['activities'], self.source.activities), start=1,
        ):
            source_code = f'SIM-2027-ACT-{self.source.unit_code}-{index:02d}'
            numeric_code = f'{index:03d}'
            native = POAUActividad.objects.filter(
                poau=poau, codigo=source_code,
            ).first()
            legacy_native = native is not None
            if native is None:
                native = POAUActividad.objects.filter(
                    poau=poau, codigo=numeric_code,
                ).first()
            legacy_native = legacy_native or bool(
                legacy_poau
                and native
                and native.nombre == source_activity.name
            )
            quarter = (source_activity.meta / Decimal('4')).quantize(Decimal('0.0001'))
            quarters = [quarter, quarter, quarter, source_activity.meta - quarter * 3]
            values = {
                'nombre': source_activity.name,
                'meta_fisica_anual': source_activity.meta,
                'presupuesto_anual': Decimal('15000.00'),
                'meta_q1': quarters[0], 'meta_q2': quarters[1],
                'meta_q3': quarters[2], 'meta_q4': quarters[3],
            }
            if native:
                if not legacy_native and not self._is_known_owned(native):
                    raise ValidationError(
                        f'La actividad POAU {numeric_code} ya existe y no pertenece '
                        'al demo; no se modificó.'
                    )
                self._track(
                    native, False, f'poau_actividad_{index:02d}', owned=True,
                )
                self._sync(native, {'codigo': numeric_code, **values})
            else:
                native = POAUActividad.objects.create(
                    poau=poau, codigo=numeric_code, **values,
                )
                self._track(native, True, f'poau_actividad_{index:02d}')
            native_activities.append(native)
            for month, programmed_source in enumerate(source_activity.programmed, start=1):
                ratio = (Decimal('0.45'), Decimal('0.65'), Decimal('0.85'))[
                    (index + month) % 3
                ]
                executed_source = source_activity.executed[month - 1]
                executed = executed_source if executed_source > 0 else (
                    programmed_source * ratio
                ).quantize(Decimal('0.0001'))
                self._get(
                    EjecucionFisica,
                    f'ejecucion_fisica_{index:02d}_{month:02d}',
                    actividad=native, periodo=f'{self.gestion}-{month:02d}',
                    defaults={
                        'tipo_periodo': 'mensual',
                        'programado': programmed_source,
                        'ejecutado': min(executed, programmed_source),
                        'observaciones': (
                            f'[PROVISIONAL 2027] Excel hoja Base, fila {source_activity.source_row}; '
                            'si la fuente no informó ejecución, se aplicó una razón referencial.'
                        ),
                    },
                )
        self.bridges['poau_native_segmented'] = [
            {
                'poau_actividad': str(native.pk),
                'actividad_poau': str(segmented.pk),
            }
            for native, segmented in zip(native_activities, source_objects['activities'])
        ]
        return {'unit': unit, 'poau': poau, 'activities': native_activities}

    def _seed_budget(self, source_objects, native_objects):
        start = date(self.gestion, 1, 1)
        unit = native_objects['unit']
        da = self._get(
            DireccionAdministrativa, 'direccion_administrativa',
            codigo='91', gestion=self.gestion,
            defaults={
                'nombre': 'Dirección Administrativa 91 — Gestión institucional',
                'fecha_vigencia_desde': start,
            },
        )
        ue = self._get(
            UnidadEjecutora, 'unidad_ejecutora', codigo='91', da=da, gestion=self.gestion,
            defaults={
                'nombre': 'Unidad Ejecutora 91 — Dirección Jurídica',
                'unidad_organizacional': unit,
                'fecha_vigencia_desde': start,
            },
        )
        category_version = self._classifier_version(
            VersionClasificador.TIPO_CATEGORIA_PROGRAMATICA, 'version_categoria',
        )
        object_version = self._classifier_version(
            VersionClasificador.TIPO_OBJETO_GASTO, 'version_objeto_gasto',
        )
        source_version = self._classifier_version(
            VersionClasificador.TIPO_FUENTE_FINANCIAMIENTO, 'version_fuente',
        )
        funding_version = self._classifier_version(
            VersionClasificador.TIPO_ORGANISMO_FINANCIADOR, 'version_organismo',
        )
        institution = self._get(
            ClasificadorInstitucional, 'clasificador_institucional',
            codigo='1312', gestion=self.gestion,
            defaults={
                'denominacion': 'Gobierno Autónomo Municipal de Sacaba',
                'fecha_vigencia_desde': start,
                'fuente_normativa': self.REFERENCE_NOTE,
            },
        )
        program = self._get(
            ProgramaPresupuestario, 'programa_presupuestario',
            codigo='000', gestion=self.gestion,
            defaults={
                'nombre': 'Gestión administrativa institucional',
                'descripcion': 'Programa referencial para servicios jurídicos municipales.',
                'ue_responsable': ue,
            },
        )
        project = self._get(
            ProyectoPresupuestario, 'proyecto_presupuestario',
            codigo='0', programa=program, gestion=self.gestion,
            defaults={'nombre': 'Administración central'},
        )
        budget_activity = self._get(
            ActividadPresupuestaria, 'actividad_presupuestaria',
            codigo='001', proyecto=project, gestion=self.gestion,
            defaults={'nombre': 'Servicios jurídicos institucionales'},
        )
        category = self._get(
            CategoriaProgramatica, 'categoria_programatica',
            version_clasificador=category_version,
            entidad=institution, da=da, ue=ue, programa=program,
            proyecto=project, actividad=budget_activity,
            defaults={
                'codigo_fuente': self.source.category_code,
                'procedencia_normativa': self.REFERENCE_NOTE,
            },
        )
        source = self._get(
            FuenteFinanciamiento, 'fuente_financiamiento',
            codigo='20', gestion=self.gestion,
            defaults={
                'denominacion': 'Recursos específicos — referencia provisional',
                'descripcion': self.REFERENCE_NOTE,
                'fecha_vigencia_desde': start,
                'fuente_normativa': self.REFERENCE_NOTE,
                'version_clasificador': source_version,
            },
        )
        funding = self._get(
            OrganismoFinanciador, 'organismo_financiador',
            codigo='210', gestion=self.gestion,
            defaults={
                'denominacion': 'Tesoro municipal — referencia provisional',
                'descripcion': self.REFERENCE_NOTE,
                'fecha_vigencia_desde': start,
                'fuente_normativa': self.REFERENCE_NOTE,
                'version_clasificador': funding_version,
            },
        )
        purpose = self._get(
            FinalidadFuncion, 'finalidad_funcion', codigo='01110', gestion=self.gestion,
            defaults={
                'denominacion': 'Administración general y servicios jurídicos',
                'descripcion': self.REFERENCE_NOTE,
                'fecha_vigencia_desde': start,
                'fuente_normativa': self.REFERENCE_NOTE,
            },
        )
        object_rows = [
            ('21100', 'Servicios básicos y comunicaciones'),
            ('22110', 'Pasajes al interior del país'),
            ('25220', 'Consultores individuales de línea'),
            ('25600', 'Servicios de imprenta y reproducción'),
            ('31110', 'Gastos de oficina y material de escritorio'),
            ('32200', 'Productos de artes gráficas y papel'),
        ]
        amounts = [
            ('95000.00', '92000.00', '52000.00'),
            ('65000.00', '64000.00', '38000.00'),
            ('50000.00', '49000.00', '29000.00'),
            ('35000.00', '34000.00', '20000.00'),
            ('25000.00', '24000.00', '14000.00'),
            ('15000.00', '15000.00', '9000.00'),
        ]
        levels = [
            ('operacion', source_objects['operation']),
            ('operacion', source_objects['operation']),
            ('actividad', source_objects['activities'][0]),
            ('actividad', source_objects['activities'][1]),
            ('tarea', source_objects['tasks'][0]),
            ('tarea', source_objects['tasks'][1]),
        ]
        assignments = []
        legacy_assignments = []
        followups = []
        objects = []
        for index, ((object_code, object_name), amount_values, level_data) in enumerate(
            zip(object_rows, amounts, levels), start=1,
        ):
            formulated, current, executed = map(Decimal, amount_values)
            expense = self._get(
                ObjetoGasto, f'objeto_gasto_{index:02d}',
                codigo=object_code, gestion=self.gestion,
                defaults={
                    'denominacion': object_name,
                    'descripcion': self.REFERENCE_NOTE,
                    'fecha_vigencia_desde': start,
                    'fuente_normativa': self.REFERENCE_NOTE,
                    'version_clasificador': object_version,
                    'nivel': ObjetoGasto.NIVEL_DETALLE,
                },
            )
            objects.append(expense)
            level, operational = level_data
            relations = {'operacion': None, 'actividad': None, 'tarea': None}
            relations[level] = operational
            assignment = self._get(
                AsignacionPresupuestariaUnidad,
                f'asignacion_presupuestaria_{index:02d}',
                categoria_programatica=category,
                fuente=source, organismo=funding, objeto_gasto=expense,
                unidad=unit, gestion=self.gestion, **relations,
                defaults={
                    'monto_formulado': formulated,
                    'monto_vigente': current,
                    'monto_ejecutado': executed,
                },
            )
            assignments.append(assignment)
            self._get(
                LineaPresupuestaria, f'linea_presupuestaria_{index:02d}',
                gestion=self.gestion, entidad='1312', da=da, ue=ue,
                programa=program, proyecto=project, actividad=budget_activity,
                finalidad_funcion=purpose, fuente=source, organismo=funding,
                objeto_gasto=expense, version=1,
                defaults={'importe': formulated},
            )
            segmented_activity = source_objects['activities'][min(index - 1, 18)]
            segmented_task = source_objects['tasks'][min(index - 1, 138)]
            legacy = self._get(
                AsignacionObjetoGasto, f'asignacion_objeto_gasto_{index:02d}',
                codigo_asignacion=f'{index:02d}', gestion=self.gestion,
                defaults={
                    'accion_poa': source_objects['action'],
                    'operacion': source_objects['operation'],
                    'actividad': segmented_activity,
                    'tarea': segmented_task if level == 'tarea' else None,
                    'categoria_programatica': category.codigo_compuesto,
                    'da': da.codigo, 'ue': ue.codigo, 'programa': program.codigo,
                    'actividad_presup': budget_activity.codigo,
                    'cod_objeto_gasto': expense.codigo,
                    'descripcion_objeto': expense.denominacion,
                    'grupo_gasto': expense.codigo[0],
                    'tipo_gasto': 'funcionamiento referencial',
                    'fuente_financiamiento': source.codigo,
                    'organismo_financiador': funding.codigo,
                    'monto_programado': formulated,
                    'monto_modificado': current - formulated,
                    'monto_vigente': current,
                    'justificacion': self.REFERENCE_NOTE,
                    'memoria_calculo': 'Estimación referencial por volumen de servicios 2027.',
                    'estado': 'PROVISIONAL',
                },
            )
            legacy_assignments.append(legacy)
            physical_programmed = segmented_activity.meta_anual or Decimal('0')
            physical_executed = (physical_programmed * Decimal('0.65')).quantize(
                Decimal('0.0001')
            )
            followup = self._get(
                SeguimientoPresupuesto, f'seguimiento_presupuesto_{index:02d}',
                id_cadena=f'2027.{index:03d}', gestion=self.gestion,
                accion_poa=source_objects['action'],
                defaults={
                    'operacion': source_objects['operation'],
                    'actividad': segmented_activity,
                    'tarea': segmented_task if level == 'tarea' else None,
                    'categoria_programatica': category.codigo_compuesto,
                    'da': da.codigo, 'ue': ue.codigo, 'programa': program.codigo,
                    'actividad_presup': budget_activity.codigo,
                    'tipo_gasto': 'funcionamiento referencial',
                    'presupuesto_inicial': formulated,
                    'modificaciones': current - formulated,
                    'presupuesto_vigente': current,
                    'ejecucion_mensual': {
                        f'{month:02d}': str((executed / 12).quantize(Decimal('0.01')))
                        for month in range(1, 13)
                    },
                    'ejecutado_total': executed,
                    'porcentaje_ejecucion_financiera': (
                        executed / current * 100
                    ).quantize(Decimal('0.0001')),
                    'meta_fisica': physical_programmed,
                    'ejecucion_fisica': physical_executed,
                    'porcentaje_ejecucion_fisica': Decimal('65.0000'),
                    'eficacia': Decimal('65.0000'),
                    'eficiencia': Decimal('95.0000'),
                    'desviacion': 'Ejecución referencial dentro del rango esperado.',
                    'evidencia': self.REFERENCE_NOTE,
                    'fecha_actualizacion': date(self.gestion, 12, 31),
                    'estado': 'PROVISIONAL',
                },
            )
            followups.append(followup)

        ceiling = self._get(
            TechoPresupuestario, 'techo_presupuestario',
            gestion=self.gestion, fuente=source, organismo=funding, version=1,
            defaults={
                'monto_total': Decimal('300000.00'),
                'descripcion': self.REFERENCE_NOTE,
                'activo': True,
            },
        )
        self._get(
            DistribucionTecho, 'distribucion_techo',
            techo=ceiling, da=da, ue=ue, unidad=unit, programa=program, version=1,
            defaults={
                'monto_asignado': Decimal('300000.00'),
                'monto_reserva': Decimal('15000.00'),
                'activo': True,
            },
        )
        for native, expense in zip(native_objects['activities'], objects * 4):
            if native.objeto_gasto_id != expense.pk:
                native.objeto_gasto = expense
                native.save()
        self.bridges['budget'] = {
            'canonical_assignments': [str(row.pk) for row in assignments],
            'legacy_assignments': [str(row.pk) for row in legacy_assignments],
            'followups': [str(row.pk) for row in followups],
        }

    def _seed_tracking(self, native_objects):
        for month in range(1, 13):
            period = f'{self.gestion}-{month:02d}'
            report = self._get(
                ReporteSeguimiento, f'reporte_seguimiento_{month:02d}',
                gestion=self.gestion, periodo=period,
                unidad_organizacional=native_objects['unit'],
                defaults={'estado': 'borrador'},
            )
            for index, activity in enumerate(native_objects['activities'], start=1):
                physical = EjecucionFisica.objects.get(actividad=activity, periodo=period)
                ratio = (Decimal('0.45'), Decimal('0.65'), Decimal('0.85'))[
                    (index + month) % 3
                ]
                financial_programmed = Decimal('1250.00')
                financial_executed = (financial_programmed * ratio).quantize(Decimal('0.01'))
                financial = self._get(
                    EjecucionFinanciera,
                    f'ejecucion_financiera_{index:02d}_{month:02d}',
                    actividad=activity, periodo=period,
                    defaults={
                        'tipo_periodo': 'mensual',
                        'programado': financial_programmed,
                        'ejecutado': financial_executed,
                        'observaciones': self.REFERENCE_NOTE,
                    },
                )
                physical_pct = (
                    physical.ejecutado / physical.programado * 100
                    if physical.programado else Decimal('0')
                ).quantize(Decimal('0.01'))
                financial_pct = (
                    financial.ejecutado / financial.programado * 100
                    if financial.programado else Decimal('0')
                ).quantize(Decimal('0.01'))
                self._get(
                    EntradaSeguimiento,
                    f'entrada_seguimiento_{index:02d}_{month:02d}',
                    reporte=report, actividad=activity,
                    defaults={
                        'programado_fisico': physical.programado,
                        'ejecutado_fisico': physical.ejecutado,
                        'porcentaje_avance_fisico': physical_pct,
                        'presupuesto_inicial': Decimal('15000.00'),
                        'presupuesto_actual': Decimal('15000.00'),
                        'programado_financiero': financial.programado,
                        'ejecutado_financiero': financial.ejecutado,
                        'porcentaje_avance_financiero': financial_pct,
                        'desviacion': physical_pct - Decimal('100'),
                        'causa_desviacion': (
                            'Avance referencial sujeto a validación con evidencia mensual.'
                        ),
                        'proyeccion_cierre': 'Cumplimiento progresivo previsto al cierre 2027.',
                        'evidencia': self.REFERENCE_NOTE,
                    },
                )
