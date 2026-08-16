"""Importa las metas NDC (Contribución Nacional Determinada) al catálogo.

Lee la hoja ``13_NDC_PUBLICADA`` del catálogo maestro
(Plataforma_Integral_Catalogo_Maestro_Completo_v1_0.xlsx) o la hoja
``09_NDC_TABLA1`` del XLSX SISPE y crea/actualiza
``articulacion.AcuerdoInternacional`` con ``tipo_acuerdo='NDC'``.

Idempotente: ``update_or_create`` por (tipo_acuerdo, codigo) con
``codigo = nro_publicado``.

Uso:
    python manage.py importar_metas_ndc --archivo <ruta.xlsx>
"""
import sys

from django.core.management.base import BaseCommand, CommandError

from apps.articulacion.models import AcuerdoInternacional

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

HOJA_NOMBRES = ('13_NDC_PUBLICADA', '09_NDC_TABLA1')


def _texto(val):
    if val is None:
        return ''
    return str(val).strip()


class Command(BaseCommand):
    help = 'Importa las metas NDC publicadas como AcuerdoInternacional (idempotente).'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            required=True,
            help='XLSX del catálogo maestro (hoja 13_NDC_PUBLICADA) o SISPE (09_NDC_TABLA1).',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Solo informa qué metas se importarían, sin persistir.',
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl no está instalado')
        archivo = options['archivo']
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError(f'No se pudo abrir el archivo: {exc}') from exc

        hoja = next((h for h in HOJA_NOMBRES if h in wb.sheetnames), None)
        if hoja is None:
            raise CommandError(
                f'El archivo no tiene ninguna hoja NDC ({", ".join(HOJA_NOMBRES)}); '
                f'hojas: {", ".join(wb.sheetnames[:12])}...'
            )

        filas = list(wb[hoja].iter_rows(values_only=True))
        if not filas:
            raise CommandError(f'La hoja {hoja} está vacía')
        header = [str(c).strip().lower() if c else '' for c in filas[0]]
        idx_nro = header.index('nro_publicado')
        idx_texto = header.index('texto_meta_publicado')
        idx_enfoque = header.index('enfoque') if 'enfoque' in header else None
        idx_cond = header.index('condicionalidad') if 'condicionalidad' in header else None

        creados = actualizados = omitidos = 0
        for fila in filas[1:]:
            nro = _texto(fila[idx_nro])
            texto = _texto(fila[idx_texto])
            if not nro or not texto:
                omitidos += 1
                continue
            rango = _texto(fila[idx_enfoque]) if idx_enfoque is not None else ''
            cond = _texto(fila[idx_cond]) if idx_cond is not None else ''
            if cond and cond != rango:
                rango = f'{rango} · {cond}'.strip(' ·')
            self.stdout.write(
                f'  NDC [{nro}] ({len(texto)} chars)'
            )
            if not options['dry_run']:
                _, created = AcuerdoInternacional.objects.update_or_create(
                    tipo_acuerdo='NDC',
                    codigo=nro,
                    defaults={
                        'denominacion': texto,
                        'rango_valido': rango[:100],
                        'es_codigo_oficial': True,
                        'activo': True,
                    },
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1

        modo = 'DRY-RUN (sin persistir)' if options['dry_run'] else 'persistido'
        self.stdout.write(
            self.style.SUCCESS(
                f'NDC {modo}: {creados} creados, {actualizados} actualizados, '
                f'{omitidos} omitidos (sin nro/texto).'
            )
        )
