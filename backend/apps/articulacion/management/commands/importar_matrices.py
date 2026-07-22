"""
Management command para importar datos desde el archivo Excel de matrices
de articulación (PAD - PEI - POA - POAU) a los modelos de la app articulacion.

Uso:
    python manage.py importar_matrices /ruta/al/archivo.xlsx
    python manage.py importar_matrices /ruta/al/archivo.xlsx --clear

Idempotente: usa get_or_create en todos los modelos.
Transaccional: todo el proceso dentro de transaction.atomic().
"""
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction, IntegrityError

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.articulacion.models import (
    AcuerdoInternacional,
    CodigoNivel,
    Normativa,
    ResultadoPAD,
    ProductoPAD,
    ResultadoPEI,
    ProductoPEI,
    ArticulacionPADPEI,
    IndicadorCadena,
    AccionPOA,
    OperacionPOAU,
    ActividadPOAU,
    TareaPOAU,
    ActividadNormativa,
    TareaNormativa,
    SeguimientoPresupuesto,
    AsignacionObjetoGasto,
)
from apps.organizacion.models import UnidadOrganizacional, TipoUnidad

# ---------------------------------------------------------------------------
# Constantes de mapeo
# ---------------------------------------------------------------------------
TIPO_ACUERDO_MAP = {
    "30/30": "COMPROMISO_3030",
    "ODS": "ODS",
    "NDC": "NDC",
    "NDT": "NDT",
}

MESES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
         "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean(val):
    """Retorna string limpio o vacío."""
    if val is None:
        return ""
    return str(val).strip()


def _decimal(val):
    """Convierte a Decimal o None."""
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _int(val):
    """Convierte a int o None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, InvalidOperation):
        return None


def _parse_date(val):
    """Convierte a date o None. Soporta datetime, string ISO, y serial Excel."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        from datetime import timedelta
        try:
            return date(1899, 12, 30) + timedelta(days=int(val))
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    if not s:
        return None
    # intentar ISO
    try:
        return datetime.fromisoformat(s).date()
    except (ValueError, TypeError):
        pass
    # formatos comunes
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _build_mensual(row, start_col, n=12):
    """Construye dict {'ENE': val, ...} desde columnas consecutivas."""
    result = {}
    for i in range(min(n, len(row) - start_col)):
        val = _decimal(row[start_col + i])
        if val is not None and i < len(MESES):
            result[MESES[i]] = str(val)
    return result if result else None


def _si_no(val):
    """Mapea 'Sí'/'No' a booleano."""
    if val is None:
        return True  # default
    s = str(val).strip().upper()
    if s in ("NO", "N", "FALSE", "0"):
        return False
    if s in ("SÍ", "SI", "S", "TRUE", "1"):
        return True
    return True


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa datos del archivo Excel de matrices de articulación"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Ruta al archivo Excel")
        parser.add_argument(
            "--clear", action="store_true",
            help="Eliminar datos existentes antes de importar"
        )

    # ================================================================
    # handle
    # ================================================================
    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl no está instalado. Ejecute: pip install openpyxl")

        file_path = options["file_path"]
        clear = options["clear"]

        self.stdout.write(self.style.NOTICE(f"\n📂 Leyendo archivo: {file_path}"))

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheets_disponibles = wb.sheetnames
        self.stdout.write(f"   Hojas disponibles: {sheets_disponibles}")

        # --clear: eliminar datos existentes en orden inverso de dependencia
        if clear:
            self._clear_data()

        stats = {}

        with transaction.atomic():
            # 1 – CAT_ACUERDOS
            if "CAT_ACUERDOS" in sheets_disponibles:
                ws = wb["CAT_ACUERDOS"]
                stats["CAT_ACUERDOS (AcuerdoInternacional)"] = self._importar_cat_acuerdos(ws)
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja CAT_ACUERDOS no encontrada"))

            # 2 – 01_CODIFICACION
            if "01_CODIFICACION" in sheets_disponibles:
                ws = wb["01_CODIFICACION"]
                stats["01_CODIFICACION (CodigoNivel)"] = self._importar_codificacion(ws)
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja 01_CODIFICACION no encontrada"))

            # 3 – CAT_NORMATIVA
            if "CAT_NORMATIVA" in sheets_disponibles:
                ws = wb["CAT_NORMATIVA"]
                stats["CAT_NORMATIVA (Normativa)"] = self._importar_normativa(ws)
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja CAT_NORMATIVA no encontrada"))

            # 4 – M1_PAD_PEI
            if "M1_PAD_PEI" in sheets_disponibles:
                ws = wb["M1_PAD_PEI"]
                stats.update(self._importar_m1_pad_pei(ws))
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja M1_PAD_PEI no encontrada"))

            # 5 – M2_PEI_POA
            if "M2_PEI_POA" in sheets_disponibles:
                ws = wb["M2_PEI_POA"]
                stats.update(self._importar_m2_pei_poa(ws))
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja M2_PEI_POA no encontrada"))

            # 6 – M3_POA_POAU
            if "M3_POA_POAU" in sheets_disponibles:
                ws = wb["M3_POA_POAU"]
                stats.update(self._importar_m3_poa_poau(ws))
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja M3_POA_POAU no encontrada"))

            # 7 – M4_PRESUP_SEGUIM
            if "M4_PRESUP_SEGUIM" in sheets_disponibles:
                ws = wb["M4_PRESUP_SEGUIM"]
                stats.update(self._importar_m4_presup_seguim(ws))
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja M4_PRESUP_SEGUIM no encontrada"))

            # 8 – M5_OBJETOS_GASTO
            if "M5_OBJETOS_GASTO" in sheets_disponibles:
                ws = wb["M5_OBJETOS_GASTO"]
                stats.update(self._importar_m5_objetos_gasto(ws))
            else:
                self.stdout.write(self.style.WARNING("  ⚠ Hoja M5_OBJETOS_GASTO no encontrada"))

        wb.close()
        self._resumen(stats)

    # ================================================================
    # Clear
    # ================================================================

    def _clear_data(self):
        """Elimina datos existentes en orden inverso de dependencia."""
        self.stdout.write(self.style.WARNING("\n🧹 Eliminando datos existentes…"))
        orden_borrado = [
            AsignacionObjetoGasto,
            SeguimientoPresupuesto,
            TareaNormativa,
            TareaPOAU,
            ActividadNormativa,
            ActividadPOAU,
            OperacionPOAU,
            AccionPOA,
            IndicadorCadena,
            ArticulacionPADPEI,
            ProductoPEI,
            ResultadoPEI,
            ProductoPAD,
            ResultadoPAD,
            Normativa,
            CodigoNivel,
            AcuerdoInternacional,
        ]
        for modelo in orden_borrado:
            cnt = modelo.objects.all().delete()[0]
            if cnt:
                self.stdout.write(f"  ✓ {modelo.__name__}: {cnt} eliminados")
        self.stdout.write(self.style.SUCCESS("  ✅ Datos eliminados correctamente\n"))

    # ================================================================
    # Cache de UnidadOrganizacional
    # ================================================================

    def _get_or_create_uo(self, codigo, nombre, gestion):
        """Busca o crea una UnidadOrganizacional."""
        if not codigo:
            return None
        try:
            return UnidadOrganizacional.objects.get(codigo=codigo, gestion=gestion)
        except UnidadOrganizacional.DoesNotExist:
            from datetime import date
            tipo, _ = TipoUnidad.objects.get_or_create(
                codigo="AREA",
                defaults={"nombre": "Área Organizacional", "nivel": 3},
            )
            uo, _ = UnidadOrganizacional.objects.get_or_create(
                codigo=codigo,
                gestion=gestion,
                defaults={
                    "nombre": nombre or f"Unidad {codigo}",
                    "sigla": codigo[:30],
                    "tipo": tipo,
                    "orden": 0,
                    "fecha_vigencia_desde": date(gestion, 1, 1),
                },
            )
            return uo

    # ================================================================
    # 1 – CAT_ACUERDOS → AcuerdoInternacional
    # ================================================================

    def _importar_cat_acuerdos(self, ws):
        self.stdout.write("\n[1] CAT_ACUERDOS → AcuerdoInternacional")
        rows = list(ws.iter_rows(min_row=5, values_only=True))
        creados = 0
        existentes = 0

        for idx, row in enumerate(rows):
            try:
                tipo_raw = _clean(row[0])
                tipo = TIPO_ACUERDO_MAP.get(tipo_raw, tipo_raw)
                codigo = _clean(row[1])
                denominacion = _clean(row[2]) if row[2] else ""
                rango_valido = _clean(row[3]) if len(row) > 3 and row[3] else ""
                es_oficial = _si_no(row[5]) if len(row) > 5 else True

                if not tipo or not codigo:
                    continue

                _, created = AcuerdoInternacional.objects.get_or_create(
                    tipo_acuerdo=tipo,
                    codigo=codigo,
                    defaults={
                        "denominacion": denominacion,
                        "rango_valido": rango_valido,
                        "es_codigo_oficial": es_oficial,
                    },
                )
                if created:
                    creados += 1
                else:
                    existentes += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error fila {idx + 5}: {e}"
                ))

        total = creados + existentes
        self.stdout.write(f"  → {creados} creados, {existentes} existentes (total {total})")
        return {"creados": creados, "existentes": existentes, "total": total}

    # ================================================================
    # 2 – 01_CODIFICACION → CodigoNivel
    # ================================================================

    def _importar_codificacion(self, ws):
        self.stdout.write("\n[2] 01_CODIFICACION → CodigoNivel")
        rows = list(ws.iter_rows(min_row=5, values_only=True))
        creados = 0
        existentes = 0

        # cache de niveles para resolver codigo_padre
        cache_niveles = {}

        for idx, row in enumerate(rows):
            try:
                nivel = _clean(row[0])
                codigo_nivel = _clean(row[1])
                segmentos = _clean(row[2]) if row[2] else ""
                longitud = _clean(row[3]) if row[3] else ""
                codigo_padre_raw = _clean(row[4]) if row[4] else ""
                ejemplo = _clean(row[5]) if row[5] else ""
                regla = _clean(row[7]) if len(row) > 7 and row[7] else ""
                editable = not _si_no(row[8]) if len(row) > 8 else False
                vigencia = _clean(row[9]) if len(row) > 9 and row[9] else ""

                if not nivel or not codigo_nivel:
                    continue

                # resolver código padre
                codigo_padre_obj = None
                if codigo_padre_raw and codigo_padre_raw in cache_niveles:
                    codigo_padre_obj = cache_niveles[codigo_padre_raw]

                obj, created = CodigoNivel.objects.get_or_create(
                    nivel=nivel,
                    defaults={
                        "codigo_nivel": codigo_nivel,
                        "segmentos": segmentos,
                        "longitud": longitud,
                        "codigo_padre": codigo_padre_obj,
                        "ejemplo": ejemplo,
                        "regla_generacion": regla,
                        "editable": editable,
                        "vigencia": vigencia,
                    },
                )
                # si ya existe, actualizar codigo_padre si cambió
                if not created and codigo_padre_obj and obj.codigo_padre != codigo_padre_obj:
                    obj.codigo_padre = codigo_padre_obj
                    obj.save(update_fields=["codigo_padre"])

                cache_niveles[nivel] = obj

                if created:
                    creados += 1
                else:
                    existentes += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error fila {idx + 5}: {e}"
                ))

        total = creados + existentes
        self.stdout.write(f"  → {creados} creados, {existentes} existentes (total {total})")
        return {"creados": creados, "existentes": existentes, "total": total}

    # ================================================================
    # 3 – CAT_NORMATIVA → Normativa
    # ================================================================

    def _importar_normativa(self, ws):
        self.stdout.write("\n[3] CAT_NORMATIVA → Normativa")
        rows = list(ws.iter_rows(min_row=5, values_only=True))
        creados = 0
        existentes = 0

        for idx, row in enumerate(rows):
            try:
                codigo_norma = _clean(row[0])
                nivel = _clean(row[1])
                tipo_norma = _clean(row[2])
                numero_id = _clean(row[3])
                denominacion = _clean(row[4])
                ambito = _clean(row[5]) if len(row) > 5 and row[5] else ""
                vigencia = _clean(row[6]) if len(row) > 6 and row[6] else ""
                estado = _clean(row[7]) if len(row) > 7 and row[7] else "VALIDAR"
                fuente = _clean(row[8]) if len(row) > 8 and row[8] else ""
                observacion = _clean(row[9]) if len(row) > 9 and row[9] else ""

                if not codigo_norma:
                    continue

                _, created = Normativa.objects.get_or_create(
                    codigo_norma=codigo_norma,
                    defaults={
                        "nivel": nivel,
                        "tipo_norma": tipo_norma,
                        "numero_identificador": numero_id,
                        "denominacion": denominacion,
                        "ambito_aplicacion": ambito,
                        "vigencia": vigencia,
                        "estado": estado,
                        "fuente": fuente,
                        "observacion": observacion,
                    },
                )
                if created:
                    creados += 1
                else:
                    existentes += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error fila {idx + 5}: {e}"
                ))

        total = creados + existentes
        self.stdout.write(f"  → {creados} creados, {existentes} existentes (total {total})")
        return {"creados": creados, "existentes": existentes, "total": total}

    # ================================================================
    # 4 – M1_PAD_PEI → PAD/PEI + Indicadores
    # ================================================================

    def _importar_m1_pad_pei(self, ws):
        self.stdout.write("\n[M1] PAD → PEI")
        rows = list(ws.iter_rows(min_row=6, values_only=True))

        acuerdos_cache = {}  # (tipo, codigo) → AcuerdoInternacional
        stats = {
            "M1_ResultadoPAD": {"creados": 0, "existentes": 0},
            "M1_ProductoPAD": {"creados": 0, "existentes": 0},
            "M1_ResultadoPEI": {"creados": 0, "existentes": 0},
            "M1_ProductoPEI": {"creados": 0, "existentes": 0},
            "M1_ArticulacionPADPEI": {"creados": 0, "existentes": 0},
            "M1_IndicadorCadena": {"creados": 0, "existentes": 0},
        }

        def _get_acuerdo(tipo, codigo):
            if not codigo or codigo in ("00", "0", ""):
                return None
            key = (tipo, codigo)
            if key in acuerdos_cache:
                return acuerdos_cache[key]
            try:
                obj = AcuerdoInternacional.objects.get(tipo_acuerdo=tipo, codigo=codigo)
                acuerdos_cache[key] = obj
                return obj
            except AcuerdoInternacional.DoesNotExist:
                acuerdos_cache[key] = None
                return None

        for idx, row in enumerate(rows):
            try:
                id_cadena = _clean(row[0])
                vigencia_desde = _int(row[1])
                vigencia_hasta = _int(row[2])
                if not id_cadena:
                    continue

                # --- ResultadoPAD ---
                cod_resultado_pad = _clean(row[23])
                nombre_resultado_pad = _clean(row[24]) if row[24] else ""
                lineamiento_pad = _clean(row[21])
                territorializacion = _clean(row[27]) if row[27] else ""
                responsable_pad = _clean(row[28]) if row[28] else ""
                cod_geografico = _clean(row[19]) if row[19] else ""
                eta = _clean(row[20]) if row[20] else ""

                resultado_pad = None
                if cod_resultado_pad:
                    resultado_pad, created = ResultadoPAD.objects.get_or_create(
                        codigo_resultado=cod_resultado_pad,
                        vigencia_desde=vigencia_desde or 2026,
                        defaults={
                            "id_cadena": id_cadena,
                            "denominacion": nombre_resultado_pad,
                            "lineamiento_pad": lineamiento_pad,
                            "territorializacion": territorializacion,
                            "responsable_pad": responsable_pad,
                            "vigencia_desde": vigencia_desde or 2026,
                            "vigencia_hasta": vigencia_hasta or 2030,
                            "cod_geografico": cod_geografico,
                            "eta": eta,
                            "cod_eje_pgdesa": _clean(row[3]) if row[3] else "",
                            "objetivo_impacto": _clean(row[4]) if row[4] else "",
                            "cod_componente_pdesa": _clean(row[5]) if row[5] else "",
                            "objetivo_efecto": _clean(row[6]) if row[6] else "",
                            "cod_sector": _clean(row[15]) if row[15] else "",
                            "sector": _clean(row[16]) if row[16] else "",
                            "cod_resultado_pds": _clean(row[17]) if row[17] else "",
                            "resultado_pds": _clean(row[18]) if row[18] else "",
                        },
                    )
                    key = "creados" if created else "existentes"
                    stats["M1_ResultadoPAD"][key] += 1

                    # Acuerdos M2M
                    ods = _get_acuerdo("ODS", _clean(row[7]))
                    if ods and not resultado_pad.acuerdo_ods.filter(pk=ods.pk).exists():
                        resultado_pad.acuerdo_ods.add(ods)
                    ndc = _get_acuerdo("NDC", _clean(row[9]))
                    if ndc and not resultado_pad.acuerdo_ndc.filter(pk=ndc.pk).exists():
                        resultado_pad.acuerdo_ndc.add(ndc)
                    ndt = _get_acuerdo("NDT", _clean(row[11]))
                    if ndt and not resultado_pad.acuerdo_ndt.filter(pk=ndt.pk).exists():
                        resultado_pad.acuerdo_ndt.add(ndt)
                    c3030 = _get_acuerdo("COMPROMISO_3030", _clean(row[13]))
                    if c3030 and not resultado_pad.acuerdo_3030.filter(pk=c3030.pk).exists():
                        resultado_pad.acuerdo_3030.add(c3030)

                # --- ProductoPAD ---
                cod_producto_pad = _clean(row[25])
                nombre_producto_pad = _clean(row[26]) if row[26] else ""
                producto_pad = None
                if cod_producto_pad and resultado_pad:
                    producto_pad, created = ProductoPAD.objects.get_or_create(
                        codigo_producto=cod_producto_pad,
                        resultado_pad=resultado_pad,
                        defaults={
                            "denominacion": nombre_producto_pad,
                            "territorializacion": territorializacion,
                            "responsable": responsable_pad,
                        },
                    )
                    key = "creados" if created else "existentes"
                    stats["M1_ProductoPAD"][key] += 1

                # --- ResultadoPEI ---
                cod_resultado_pei = _clean(row[32])
                nombre_resultado_pei = _clean(row[33]) if row[33] else ""
                cod_entidad = _clean(row[29]) if row[29] else ""
                entidad = _clean(row[30]) if row[30] else ""
                cod_oei = _clean(row[31]) if row[31] else ""
                resultado_pei = None
                if cod_resultado_pei:
                    resultado_pei, created = ResultadoPEI.objects.get_or_create(
                        codigo_resultado=cod_resultado_pei,
                        vigencia_desde=vigencia_desde or 2026,
                        defaults={
                            "denominacion": nombre_resultado_pei,
                            "cod_entidad": cod_entidad,
                            "entidad": entidad,
                            "cod_oei": cod_oei,
                            "vigencia_desde": vigencia_desde or 2026,
                            "vigencia_hasta": vigencia_hasta or 2030,
                        },
                    )
                    key = "creados" if created else "existentes"
                    stats["M1_ResultadoPEI"][key] += 1

                # --- ProductoPEI ---
                cod_producto_pei = _clean(row[36])
                nombre_producto_pei = _clean(row[37]) if row[37] else ""
                cod_programa_presup = _clean(row[34]) if row[34] else ""
                programa_presup = _clean(row[35]) if row[35] else ""
                producto_pei = None
                if cod_producto_pei and resultado_pei:
                    producto_pei, created = ProductoPEI.objects.get_or_create(
                        codigo_producto=cod_producto_pei,
                        resultado_pei=resultado_pei,
                        defaults={
                            "denominacion": nombre_producto_pei,
                            "cod_programa_presup": cod_programa_presup,
                            "programa_presup": programa_presup,
                        },
                    )
                    key = "creados" if created else "existentes"
                    stats["M1_ProductoPEI"][key] += 1

                # --- ArticulacionPADPEI ---
                if producto_pad and producto_pei:
                    _, created = ArticulacionPADPEI.objects.get_or_create(
                        producto_pad=producto_pad,
                        producto_pei=producto_pei,
                        defaults={"estado": "REFERENCIAL"},
                    )
                    key = "creados" if created else "existentes"
                    stats["M1_ArticulacionPADPEI"][key] += 1

                # --- IndicadorCadena ---
                nivel_indicador = _clean(row[38]) if row[38] else ""
                indicador = _clean(row[39]) if row[39] else ""
                if indicador:
                    tipo_indicador = _clean(row[40]) if row[40] else ""
                    unidad_medida = _clean(row[41]) if row[41] else ""
                    formula = _clean(row[42]) if row[42] else ""
                    linea_base = _decimal(row[43])
                    meta_2030 = _decimal(row[44])

                    # programación física PF_2026..PF_2030 (cols 45-49)
                    prog_fisica = {}
                    for i, yr in enumerate(["2026", "2027", "2028", "2029", "2030"]):
                        val = _decimal(row[45 + i])
                        if val is not None:
                            prog_fisica[yr] = str(val)

                    inv_total = _decimal(row[50])
                    inv_2026 = _decimal(row[51])
                    inv_2027 = _decimal(row[52])
                    inv_2028 = _decimal(row[53])
                    inv_2029 = _decimal(row[54])
                    inv_2030 = _decimal(row[55])
                    corr_total = _decimal(row[56])
                    corr_2026 = _decimal(row[57])
                    corr_2027 = _decimal(row[58])
                    corr_2028 = _decimal(row[59])
                    corr_2029 = _decimal(row[60])
                    corr_2030 = _decimal(row[61])
                    fuente_dato = _clean(row[62]) if row[62] else ""

                    # lookup por la combinación más específica disponible
                    lookup = {}
                    if producto_pad:
                        lookup["producto_pad"] = producto_pad
                    if producto_pei:
                        lookup["producto_pei"] = producto_pei
                    if nivel_indicador:
                        lookup["nivel_indicador"] = nivel_indicador
                    if indicador:
                        lookup["indicador"] = indicador

                    if lookup:
                        _, created = IndicadorCadena.objects.get_or_create(
                            **lookup,
                            defaults={
                                "tipo_indicador": tipo_indicador,
                                "unidad_medida": unidad_medida,
                                "formula": formula,
                                "linea_base": linea_base,
                                "meta_2030": meta_2030,
                                "programacion_fisica": prog_fisica if prog_fisica else None,
                                "presupuesto_inversion_total": inv_total,
                                "inversion_2026": inv_2026,
                                "inversion_2027": inv_2027,
                                "inversion_2028": inv_2028,
                                "inversion_2029": inv_2029,
                                "inversion_2030": inv_2030,
                                "presupuesto_corriente_total": corr_total,
                                "corriente_2026": corr_2026,
                                "corriente_2027": corr_2027,
                                "corriente_2028": corr_2028,
                                "corriente_2029": corr_2029,
                                "corriente_2030": corr_2030,
                                "fuente_dato": fuente_dato,
                            },
                        )
                        key = "creados" if created else "existentes"
                        stats["M1_IndicadorCadena"][key] += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error M1 fila {idx + 6}: {e}"
                ))

        for k, v in stats.items():
            self.stdout.write(f"  → {k}: {v['creados']} creados, {v['existentes']} existentes")
        return stats

    # ================================================================
    # 5 – M2_PEI_POA → AccionPOA
    # ================================================================

    def _importar_m2_pei_poa(self, ws):
        self.stdout.write("\n[M2] PEI → POA")
        rows = list(ws.iter_rows(min_row=6, values_only=True))
        stats = {"M2_AccionPOA": {"creados": 0, "existentes": 0}}

        for idx, row in enumerate(rows):
            try:
                cod_accion = _clean(row[9])
                if not cod_accion:
                    continue

                cod_producto_pei = _clean(row[4])
                if not cod_producto_pei:
                    continue

                # buscar ProductoPEI
                try:
                    producto_pei = ProductoPEI.objects.get(codigo_producto=cod_producto_pei)
                except ProductoPEI.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M2 fila {idx + 6}: ProductoPEI '{cod_producto_pei}' no encontrado"
                    ))
                    continue

                gestion = _int(row[1]) or 2026

                # Unidad responsable
                cod_unidad = _clean(row[7])
                nombre_unidad = _clean(row[8]) if row[8] else ""
                unidad_resp = self._get_or_create_uo(cod_unidad, nombre_unidad, gestion)

                denominacion = _clean(row[10]) if row[10] else ""
                resultado_esperado = _clean(row[11]) if row[11] else ""
                indicador = _clean(row[12]) if row[12] else ""
                formula = _clean(row[13]) if row[13] else ""
                unidad_medida = _clean(row[14]) if row[14] else ""
                linea_base = _decimal(row[15])
                meta_gestion = _decimal(row[16])
                cod_rea = _clean(row[17]) if row[17] else ""
                cargo_resp = _clean(row[18]) if row[18] else ""
                fecha_inicio = _parse_date(row[19])
                fecha_fin = _parse_date(row[20])
                tipo_op = _clean(row[21]) if row[21] else ""
                cat_prog = _clean(row[22]) if row[22] else ""
                programa = _clean(row[23]) if row[23] else ""
                proy_sisin = _clean(row[24]) if row[24] else ""
                act_presup = _clean(row[25]) if row[25] else ""
                presup_prog = _decimal(row[26])
                fte = _clean(row[27]) if row[27] else ""
                org = _clean(row[28]) if row[28] else ""
                medio_verif = _clean(row[29]) if row[29] else ""
                riesgo = _clean(row[30]) if row[30] else ""
                estado = _clean(row[31]) if row[31] else "REFERENCIAL"

                _, created = AccionPOA.objects.get_or_create(
                    codigo_accion=cod_accion,
                    defaults={
                        "denominacion": denominacion,
                        "resultado_esperado": resultado_esperado,
                        "producto_pei": producto_pei,
                        "indicador": indicador,
                        "formula": formula,
                        "unidad_medida": unidad_medida,
                        "linea_base": linea_base,
                        "meta_gestion": meta_gestion,
                        "codigo_rea": cod_rea,
                        "cargo_responsable": cargo_resp,
                        "fecha_inicio": fecha_inicio,
                        "fecha_fin": fecha_fin,
                        "tipo_operacion": tipo_op,
                        "categoria_programatica": cat_prog,
                        "programa": programa,
                        "proyecto_sisin": proy_sisin,
                        "actividad_presupuestaria": act_presup,
                        "presupuesto_programado": presup_prog,
                        "fuente_financiamiento": fte,
                        "organismo_financiador": org,
                        "medio_verificacion": medio_verif,
                        "riesgo": riesgo,
                        "estado": estado,
                        "gestion": gestion,
                        "unidad_responsable": unidad_resp,
                    },
                )
                key = "creados" if created else "existentes"
                stats["M2_AccionPOA"][key] += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error M2 fila {idx + 6}: {e}"
                ))

        v = stats["M2_AccionPOA"]
        self.stdout.write(f"  → AccionPOA: {v['creados']} creados, {v['existentes']} existentes")
        return stats

    # ================================================================
    # 6 – M3_POA_POAU → Operaciones / Actividades / Tareas
    # ================================================================

    def _importar_m3_poa_poau(self, ws):
        self.stdout.write("\n[M3] POA → POAU")
        rows = list(ws.iter_rows(min_row=6, values_only=True))
        stats = {
            "M3_OperacionPOAU": {"creados": 0, "existentes": 0},
            "M3_ActividadPOAU": {"creados": 0, "existentes": 0},
            "M3_TareaPOAU": {"creados": 0, "existentes": 0},
            "M3_ActividadNormativa": {"creados": 0},
            "M3_TareaNormativa": {"creados": 0},
        }

        def _parse_normativas(codigos_str):
            """Parsea '001;002;005;...' → lista de objetos Normativa."""
            if not codigos_str:
                return []
            s = str(codigos_str).strip()
            if not s:
                return []
            codigos = [c.strip() for c in s.split(";") if c.strip()]
            normativas = []
            for cod in codigos:
                try:
                    n = Normativa.objects.get(codigo_norma=cod)
                    normativas.append(n)
                except Normativa.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"    ⚠ Normativa '{cod}' no encontrada"
                    ))
            return normativas

        for idx, row in enumerate(rows):
            try:
                cod_accion = _clean(row[2])
                cod_operacion = _clean(row[4])
                cod_actividad = _clean(row[8])
                cod_tarea = _clean(row[10])

                if not cod_accion or not cod_operacion or not cod_actividad or not cod_tarea:
                    continue

                gestion = _int(row[1]) or 2026

                # --- AccionPOA lookup ---
                try:
                    accion_poa = AccionPOA.objects.get(codigo_accion=cod_accion)
                except AccionPOA.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M3 fila {idx + 6}: AccionPOA '{cod_accion}' no encontrada"
                    ))
                    continue

                # --- OperacionPOAU ---
                tipo_operacion = _clean(row[6]) if row[6] else ""
                producto_entregable = _clean(row[7]) if row[7] else ""
                nombre_operacion = _clean(row[5]) if row[5] else ""
                cod_uni_ej = _clean(row[12]) if row[12] else ""
                uni_ejecutora = _clean(row[13]) if row[13] else ""
                cod_resp = _clean(row[14]) if row[14] else ""
                responsable = _clean(row[15]) if row[15] else ""

                operacion, created = OperacionPOAU.objects.get_or_create(
                    codigo_operacion=cod_operacion,
                    defaults={
                        "denominacion": nombre_operacion,
                        "tipo_operacion": tipo_operacion,
                        "producto_entregable": producto_entregable,
                        "accion_poa": accion_poa,
                        "codigo_unidad_ejecutora": cod_uni_ej,
                        "unidad_ejecutora": uni_ejecutora,
                        "codigo_responsable": cod_resp,
                        "responsable": responsable,
                    },
                )
                key = "creados" if created else "existentes"
                stats["M3_OperacionPOAU"][key] += 1

                # --- ActividadPOAU ---
                meta_anual = _decimal(row[16])
                indicador = _clean(row[17]) if row[17] else ""
                formula = _clean(row[18]) if row[18] else ""
                unidad_medida = _clean(row[19]) if row[19] else ""
                fecha_inicio = _parse_date(row[20])
                fecha_fin = _parse_date(row[21])
                prog_mensual = _build_mensual(row, 22, 12)
                total_programado = _decimal(row[34])
                medio_verif = _clean(row[37]) if len(row) > 37 and row[37] else ""
                requerimientos = _clean(row[38]) if len(row) > 38 and row[38] else ""
                riesgo = _clean(row[39]) if len(row) > 39 and row[39] else ""
                acc_correctiva = _clean(row[40]) if len(row) > 40 and row[40] else ""
                estado = _clean(row[41]) if len(row) > 41 and row[41] else "REFERENCIAL"

                cod_normativas_str = row[35] if len(row) > 35 else None

                nombre_actividad = _clean(row[9]) if row[9] else ""

                actividad, created = ActividadPOAU.objects.get_or_create(
                    codigo_actividad=cod_actividad,
                    defaults={
                        "denominacion": nombre_actividad,
                        "operacion": operacion,
                        "meta_anual": meta_anual,
                        "indicador": indicador,
                        "formula": formula,
                        "unidad_medida": unidad_medida,
                        "fecha_inicio": fecha_inicio,
                        "fecha_fin": fecha_fin,
                        "programacion_mensual": prog_mensual,
                        "total_programado": total_programado,
                        "medio_verificacion": medio_verif,
                        "requerimientos": requerimientos,
                        "riesgo": riesgo,
                        "accion_correctiva": acc_correctiva,
                        "estado": estado,
                    },
                )
                key = "creados" if created else "existentes"
                stats["M3_ActividadPOAU"][key] += 1

                # Normativas de actividad
                if cod_normativas_str:
                    for norm in _parse_normativas(cod_normativas_str):
                        _, n_created = ActividadNormativa.objects.get_or_create(
                            actividad=actividad,
                            normativa=norm,
                        )
                        if n_created:
                            stats["M3_ActividadNormativa"]["creados"] += 1

                # --- TareaPOAU ---
                nombre_tarea = _clean(row[11]) if row[11] else ""

                tarea, created = TareaPOAU.objects.get_or_create(
                    codigo_tarea=cod_tarea,
                    defaults={
                        "denominacion": nombre_tarea,
                        "actividad": actividad,
                        "responsable": responsable,
                        "fecha_inicio": fecha_inicio,
                        "fecha_fin": fecha_fin,
                        "metas": meta_anual,
                        "programacion_mensual": prog_mensual,
                        "requerimientos": requerimientos,
                        "estado": estado,
                    },
                )
                key = "creados" if created else "existentes"
                stats["M3_TareaPOAU"][key] += 1

                # Normativas de tarea
                if cod_normativas_str:
                    for norm in _parse_normativas(cod_normativas_str):
                        _, n_created = TareaNormativa.objects.get_or_create(
                            tarea=tarea,
                            normativa=norm,
                        )
                        if n_created:
                            stats["M3_TareaNormativa"]["creados"] += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error M3 fila {idx + 6}: {e}"
                ))

        for k, v in stats.items():
            if isinstance(v, dict):
                self.stdout.write(f"  → {k}: {v}")
        return stats

    # ================================================================
    # 7 – M4_PRESUP_SEGUIM → SeguimientoPresupuesto
    # ================================================================

    def _importar_m4_presup_seguim(self, ws):
        self.stdout.write("\n[M4] Seguimiento Presupuestario")
        rows = list(ws.iter_rows(min_row=6, values_only=True))
        stats = {"M4_SeguimientoPresupuesto": {"creados": 0, "existentes": 0}}

        for idx, row in enumerate(rows):
            try:
                id_cadena = _clean(row[0])
                gestion = _int(row[1]) or 2026
                cod_accion = _clean(row[2])
                cod_operacion = _clean(row[3])
                cod_actividad = _clean(row[4])
                cod_tarea = _clean(row[5]) if row[5] else ""

                if not cod_accion or not cod_operacion or not cod_actividad:
                    continue

                # resolver FK
                try:
                    accion = AccionPOA.objects.get(codigo_accion=cod_accion)
                except AccionPOA.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M4 fila {idx + 6}: AccionPOA '{cod_accion}' no encontrada"
                    ))
                    continue

                try:
                    operacion = OperacionPOAU.objects.get(codigo_operacion=cod_operacion)
                except OperacionPOAU.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M4 fila {idx + 6}: Operacion '{cod_operacion}' no encontrada"
                    ))
                    continue

                try:
                    actividad = ActividadPOAU.objects.get(codigo_actividad=cod_actividad)
                except ActividadPOAU.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M4 fila {idx + 6}: Actividad '{cod_actividad}' no encontrada"
                    ))
                    continue

                tarea = None
                if cod_tarea:
                    try:
                        tarea = TareaPOAU.objects.get(codigo_tarea=cod_tarea)
                    except TareaPOAU.DoesNotExist:
                        pass  # nullable

                cat_prog = _clean(row[6]) if row[6] else ""
                da = _clean(row[7]) if row[7] else ""
                ue = _clean(row[8]) if row[8] else ""
                programa = _clean(row[9]) if row[9] else ""
                proy_sisin = _clean(row[10]) if row[10] else ""
                act_presup = _clean(row[11]) if row[11] else ""
                tipo_gasto = _clean(row[12]) if row[12] else ""
                presup_inicial = _decimal(row[13]) or Decimal("0")
                modificaciones = _decimal(row[14]) or Decimal("0")
                presup_vigente = _decimal(row[15]) or (presup_inicial + modificaciones)

                ejecucion_mensual = _build_mensual(row, 16, 12)
                ejecutado_total = _decimal(row[28]) or Decimal("0")

                meta_fisica = _decimal(row[30])
                ejec_fisica = _decimal(row[31])
                eficacia = _decimal(row[33])
                eficiencia = _decimal(row[34])

                # calcular porcentajes
                pct_ejec_fin = None
                if presup_vigente and presup_vigente > 0:
                    pct_ejec_fin = (ejecutado_total / presup_vigente) * 100

                pct_ejec_fis = None
                if meta_fisica and meta_fisica > 0 and ejec_fisica is not None:
                    pct_ejec_fis = (ejec_fisica / meta_fisica) * 100

                desviacion = _clean(row[35]) if len(row) > 35 and row[35] else ""
                acc_correctiva = _clean(row[36]) if len(row) > 36 and row[36] else ""
                evidencia = _clean(row[37]) if len(row) > 37 and row[37] else ""
                fecha_act = _parse_date(row[38]) if len(row) > 38 else None
                estado = _clean(row[39]) if len(row) > 39 and row[39] else "REFERENCIAL"

                _, created = SeguimientoPresupuesto.objects.get_or_create(
                    id_cadena=id_cadena,
                    gestion=gestion,
                    accion_poa=accion,
                    operacion=operacion,
                    actividad=actividad,
                    tarea=tarea,
                    defaults={
                        "categoria_programatica": cat_prog,
                        "da": da,
                        "ue": ue,
                        "programa": programa,
                        "proyecto_sisin": proy_sisin,
                        "actividad_presup": act_presup,
                        "tipo_gasto": tipo_gasto,
                        "presupuesto_inicial": presup_inicial,
                        "modificaciones": modificaciones,
                        "presupuesto_vigente": presup_vigente,
                        "ejecucion_mensual": ejecucion_mensual,
                        "ejecutado_total": ejecutado_total,
                        "porcentaje_ejecucion_financiera": pct_ejec_fin,
                        "meta_fisica": meta_fisica,
                        "ejecucion_fisica": ejec_fisica,
                        "porcentaje_ejecucion_fisica": pct_ejec_fis,
                        "eficacia": eficacia,
                        "eficiencia": eficiencia,
                        "desviacion": desviacion,
                        "accion_correctiva": acc_correctiva,
                        "evidencia": evidencia,
                        "fecha_actualizacion": fecha_act,
                        "estado": estado,
                    },
                )
                key = "creados" if created else "existentes"
                stats["M4_SeguimientoPresupuesto"][key] += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error M4 fila {idx + 6}: {e}"
                ))

        v = stats["M4_SeguimientoPresupuesto"]
        self.stdout.write(f"  → SeguimientoPresupuesto: {v['creados']} creados, {v['existentes']} existentes")
        return stats

    # ================================================================
    # 8 – M5_OBJETOS_GASTO → AsignacionObjetoGasto
    # ================================================================

    def _importar_m5_objetos_gasto(self, ws):
        self.stdout.write("\n[M5] Objetos de Gasto")
        rows = list(ws.iter_rows(min_row=6, values_only=True))
        stats = {"M5_AsignacionObjetoGasto": {"creados": 0, "existentes": 0}}

        for idx, row in enumerate(rows):
            try:
                cod_asignacion = _clean(row[0])
                gestion = _int(row[2]) or 2026
                cod_accion = _clean(row[3])
                cod_operacion = _clean(row[4])
                cod_actividad = _clean(row[5])
                cod_tarea = _clean(row[6]) if row[6] else ""

                if not cod_asignacion or not cod_accion or not cod_operacion or not cod_actividad:
                    continue

                # resolver FK
                try:
                    accion = AccionPOA.objects.get(codigo_accion=cod_accion)
                except AccionPOA.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M5 fila {idx + 6}: AccionPOA '{cod_accion}' no encontrada"
                    ))
                    continue

                try:
                    operacion = OperacionPOAU.objects.get(codigo_operacion=cod_operacion)
                except OperacionPOAU.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M5 fila {idx + 6}: Operacion '{cod_operacion}' no encontrada"
                    ))
                    continue

                try:
                    actividad = ActividadPOAU.objects.get(codigo_actividad=cod_actividad)
                except ActividadPOAU.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f"  ⚠ M5 fila {idx + 6}: Actividad '{cod_actividad}' no encontrada"
                    ))
                    continue

                tarea = None
                if cod_tarea:
                    try:
                        tarea = TareaPOAU.objects.get(codigo_tarea=cod_tarea)
                    except TareaPOAU.DoesNotExist:
                        pass

                cat_prog = _clean(row[7]) if row[7] else ""
                da = _clean(row[8]) if row[8] else ""
                ue = _clean(row[9]) if row[9] else ""
                programa = _clean(row[10]) if row[10] else ""
                proy_sisin = _clean(row[11]) if row[11] else ""
                act_presup = _clean(row[12]) if row[12] else ""
                cod_og = _clean(row[13]) if row[13] else ""
                desc_og = _clean(row[14]) if row[14] else ""
                grupo_gasto = _clean(row[15]) if row[15] else ""
                tipo_gasto = _clean(row[16]) if row[16] else ""
                fte = _clean(row[17]) if row[17] else ""
                org = _clean(row[18]) if row[18] else ""
                monto_prog = _decimal(row[19]) or Decimal("0")
                justificacion = _clean(row[20]) if len(row) > 20 and row[20] else ""
                memoria_calculo = _clean(row[21]) if len(row) > 21 and row[21] else ""
                estado = _clean(row[22]) if len(row) > 22 and row[22] else "REFERENCIAL"

                _, created = AsignacionObjetoGasto.objects.get_or_create(
                    codigo_asignacion=cod_asignacion,
                    gestion=gestion,
                    defaults={
                        "accion_poa": accion,
                        "operacion": operacion,
                        "actividad": actividad,
                        "tarea": tarea,
                        "categoria_programatica": cat_prog,
                        "da": da,
                        "ue": ue,
                        "programa": programa,
                        "proyecto_sisin": proy_sisin,
                        "actividad_presup": act_presup,
                        "cod_objeto_gasto": cod_og,
                        "descripcion_objeto": desc_og,
                        "grupo_gasto": grupo_gasto,
                        "tipo_gasto": tipo_gasto,
                        "fuente_financiamiento": fte,
                        "organismo_financiador": org,
                        "monto_programado": monto_prog,
                        "monto_vigente": monto_prog,  # inicialmente = monto programado
                        "justificacion": justificacion,
                        "memoria_calculo": memoria_calculo,
                        "estado": estado,
                    },
                )
                key = "creados" if created else "existentes"
                stats["M5_AsignacionObjetoGasto"][key] += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(
                    f"  ⚠ Error M5 fila {idx + 6}: {e}"
                ))

        v = stats["M5_AsignacionObjetoGasto"]
        self.stdout.write(f"  → AsignacionObjetoGasto: {v['creados']} creados, {v['existentes']} existentes")
        return stats

    # ================================================================
    # Resumen final
    # ================================================================

    def _resumen(self, stats):
        self.stdout.write(self.style.SUCCESS("\n" + "=" * 60))
        self.stdout.write(self.style.SUCCESS("  ✅ IMPORTACIÓN COMPLETADA"))
        self.stdout.write(self.style.SUCCESS("=" * 60))

        for key, value in stats.items():
            if isinstance(value, dict):
                if "creados" in value and "existentes" in value:
                    self.stdout.write(
                        f"  {key}: {value['creados']} creados, "
                        f"{value['existentes']} existentes"
                        f"  (total {value.get('total', value['creados'] + value['existentes'])})"
                    )
                else:
                    self.stdout.write(f"  {key}: {value}")
            else:
                self.stdout.write(f"  {key}: {value}")

        self.stdout.write("=" * 60)
