"""Preview/apply ETL for the physical programming shown by Matriz POAU."""

from __future__ import annotations

import hashlib
import io
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from urllib.parse import parse_qs, urlencode, urlsplit
from urllib.request import HTTPRedirectHandler, Request, build_opener

import openpyxl
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.db import connection, transaction
from django.db.models import Max, Q
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime

from apps.accounts.services_scope import ScopeResolver
from apps.catalogos.models import TipoOperacion, UnidadMedida
from apps.gestion.mixins import gestion_del_candado
from apps.organizacion.models import UnidadOrganizacional
from apps.presupuesto.models import AsignacionPresupuestariaUnidad

from .models import (
    AccionPOA,
    ActividadPOAU,
    ImportacionProgramacionFisica,
    OperacionPOAU,
    ProductoPEI,
    ResultadoPEI,
    TareaPOAU,
    VersionImportacionPOAU,
)
from .revision_poau import EstadosPOAU

MESES = (
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
)
MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
PREVIEW_TTL = timedelta(minutes=30)
GOOGLE_ID_RE = re.compile(r'^[A-Za-z0-9_-]{20,100}$')


class ImportacionError(ValidationError):
    """Stable validation error used by the V2 transport adapter."""


def _texto(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r'\s+', ' ', str(value)).strip()


def _clave(value):
    text = unicodedata.normalize('NFKD', _texto(value))
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return re.sub(r'[^a-z0-9]+', ' ', text.lower()).strip()


def _decimal(value, field, row, errors, required=False):
    text = _texto(value)
    if not text:
        if required:
            errors.append(_error(row, field, f'La columna {field} es obligatoria.'))
        return None
    normalized = text.replace(' ', '')
    if normalized.endswith('%'):
        normalized = normalized[:-1]
    if ',' in normalized and '.' in normalized:
        if normalized.rfind(',') > normalized.rfind('.'):
            normalized = normalized.replace('.', '').replace(',', '.')
        else:
            normalized = normalized.replace(',', '')
    elif ',' in normalized:
        normalized = normalized.replace(',', '.')
    try:
        number = Decimal(normalized)
    except InvalidOperation:
        errors.append(_error(row, field, f'«{text}» no es un número válido.'))
        return None
    if number < 0:
        errors.append(_error(row, field, 'No se admiten valores negativos.'))
        return None
    return number


def _fecha(value, field, row, gestion, errors):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        result = value.date()
    elif isinstance(value, date):
        result = value
    else:
        text = _texto(value)
        result = None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                result = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        if result is None:
            errors.append(_error(row, field, f'«{text}» no es una fecha válida.'))
            return None
    if result.year != gestion.anio:
        errors.append(_error(
            row, field,
            f'La fecha debe pertenecer a la gestión {gestion.anio}.',
        ))
        return None
    return result


def _error(row, field, message, code='invalid'):
    return {
        'fila': row, 'campo': field, 'codigo': code, 'mensaje': message,
        'severidad': 'error',
    }


def _warning(row, field, message, code='warning'):
    return {
        'fila': row, 'campo': field, 'codigo': code, 'mensaje': message,
        'severidad': 'advertencia',
    }


def _is_blocking(issue):
    return issue.get('severidad', 'error') != 'advertencia'


HEADER_ALIASES = {
    'nivel': {'nivel', 'tipo fila', 'nivel jerarquico'},
    'aie': {
        'accion institucional especifica pei',
        'accion institucional especifica', 'aie',
    },
    'accion': {
        'accion de corto plazo', 'accion corto plazo', 'accion poa',
        'accion institucional de corto plazo',
    },
    'unidad_nombre': {'unidades', 'unidad', 'unidad organizacional'},
    'unidad_codigo': {
        'codigo unidad', 'codigo uo', 'unidad codigo',
        'codigo codificacion unidades',
    },
    'accion_codigo': {
        'codigo accion', 'codigo accion de corto plazo',
        'cod accion corto plazo', 'accion codigo',
    },
    'categoria_programatica': {
        'categoria programatica', 'categoria programatica poa',
    },
    'operacion_codigo': {'codigo operacion', 'cod operacion'},
    'operacion': {'operacion', 'operaciones', 'operaciones producto intermedio'},
    'actividad_codigo': {'codigo actividad', 'cod actividad'},
    'actividad': {'actividad', 'actividades'},
    'tarea_codigo': {'codigo tarea', 'cod tarea'},
    'tarea': {'tarea', 'tareas', 'tareas especificas'},
    'tipo_operacion': {'tipo operacion', 'tipo de operacion'},
    'indicador': {'indicador'},
    'formula': {'formula'},
    'unidad_medida': {'unidad de medida', 'unidad medida'},
    'linea_base': {'linea base'},
    'meta': {'meta', 'meta anual'},
    'meta_actual': {
        'meta actual', 'estimacion linea base', 'estimacion de linea base',
    },
    'ponderacion': {'ponderacion', 'porcentaje ponderacion'},
    'fecha_inicio': {'fecha inicio', 'inicio'},
    'fecha_fin': {'fecha final', 'fecha fin', 'fin', 'final'},
    'responsable': {'responsable', 'cargo responsable'},
    'total_anual': {'total anual', 'total programado'},
}
for _mes in MESES:
    HEADER_ALIASES[_mes] = {_mes, _mes[:3]}


def _canonical_header(value):
    normalized = _clave(value)
    normalized = re.sub(r'\b(?:19|20)\d{2}\b', '', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    # La matriz oficial agrega gestión, horizonte y descripciones entre
    # paréntesis a estas dos columnas. Esos textos son parte de la etiqueta
    # visual, no cambian el significado del campo que consume el importador.
    if normalized.startswith('accion institucional especifica pei'):
        return 'aie'
    if normalized.startswith('accion de corto plazo gestion'):
        return 'accion'
    if normalized.startswith('responsable reacp'):
        return 'responsable'
    for field, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _header_map(values):
    mapping = {}
    for index, value in enumerate(values):
        normalized = _clave(value)
        field = _canonical_header(value)
        if normalized == 'codigo' and 'unidad_nombre' in mapping and 'unidad_codigo' not in mapping:
            field = 'unidad_codigo'
        if field and field not in mapping:
            mapping[field] = index
    return mapping


def _find_header(sheet, fallback_action_code='', fallback_unit_code=''):
    best = None
    for number, values in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        mapping = _header_map(values)
        score = len(mapping)
        executable = {'operacion', 'actividad', 'tarea'} & mapping.keys()
        has_action = bool({'accion', 'accion_codigo'} & mapping.keys()) or fallback_action_code
        if executable and has_action:
            if best is None or score > best[0]:
                best = (score, number, mapping)
    if best is None:
        raise ImportacionError(
            'No se encontró una cabecera con Acción de corto plazo y columnas '
            'de operación, actividad o tarea.',
        )
    missing_months = [month for month in MESES if month not in best[2]]
    if missing_months:
        raise ImportacionError(
            'Faltan columnas mensuales: ' + ', '.join(missing_months) + '.',
        )
    return best[1], best[2]


def _cell(values, mapping, field):
    index = mapping.get(field)
    return values[index] if index is not None and index < len(values) else None


def _parse_sheet(
    sheet, gestion, fallback_action_code='', fallback_unit_code='',
):
    header_row, mapping = _find_header(sheet, fallback_action_code, fallback_unit_code)
    nodes, issues = [], []
    rows_read = 0
    emitted_actions = {}
    context = {
        'aie': '', 'accion': '', 'operacion': '', 'actividad': '',
        'operacion_codigo': '', 'actividad_codigo': '',
        'categoria_programatica': '',
        'unidad_codigo': fallback_unit_code.upper(),
    }

    for row_number, values in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1,
    ):
        if not any(_texto(value) for value in values):
            continue
        if all(_clave(value) in {'', 'programado', 'ejecutado'} for value in values):
            continue
        rows_read += 1
        row_issues = []
        unit_code = _texto(_cell(values, mapping, 'unidad_codigo')).upper()
        if unit_code:
            context['unidad_codigo'] = unit_code
        aie = _texto(_cell(values, mapping, 'aie'))
        action = _texto(_cell(values, mapping, 'accion'))

        # La matriz puede contener, después del POAU propuesto,
        # una sección auxiliar/legacy denominada "INDICADORES EXISTENTE".
        # Esa sección no forma parte del árbol POAU a importar.
        if _clave(aie) in {
            'indicadores existente',
            'indicadores existentes',
        }:
            break

        if aie and _clave(aie) != _clave(context['aie']):
            context.update(
                aie=aie, accion='', operacion='', actividad='',
                categoria_programatica='',
            )
        if action and _clave(action) != _clave(context['accion']):
            context.update(
                accion=action, operacion='', actividad='',
                categoria_programatica='',
            )
        category = _texto(_cell(values, mapping, 'categoria_programatica'))
        if category:
            context['categoria_programatica'] = category

        raw = {
            level: _texto(_cell(values, mapping, level))
            for level in ('operacion', 'actividad', 'tarea')
        }
        codes = {
            level: _texto(_cell(values, mapping, f'{level}_codigo')).upper()
            for level in ('operacion', 'actividad', 'tarea')
        }
        if raw['operacion']:
            context.update(
                operacion=raw['operacion'], operacion_codigo=codes['operacion'],
                actividad='', actividad_codigo='',
            )
        if raw['actividad']:
            context.update(actividad=raw['actividad'], actividad_codigo=codes['actividad'])

        if context['accion']:
            action_key = f"{_clave(context['aie'])}|{_clave(context['accion'])}"
            # Sólo se completa si la acción todavía no tiene categoría. Antes
            # se reasignaba en cada fila que trajera una, así que con una
            # matriz de varias categorías —una por operación, que es el caso
            # real— la acción terminaba con la ÚLTIMA y las demás se perdían.
            # La categoría de cada operación viaja ahora en su propio nodo.
            if (
                action_key in emitted_actions
                and context['categoria_programatica']
                and not emitted_actions[action_key]['categoria_programatica']
            ):
                emitted_actions[action_key]['categoria_programatica'] = (
                    context['categoria_programatica']
                )
            if action_key not in emitted_actions:
                if not context['aie']:
                    row_issues.append(_error(
                        row_number, 'aie',
                        'Falta la Acción Institucional Específica (PEI).',
                        'missing_parent',
                    ))
                nodes.append({
                    'fila': row_number, 'nivel': 'accion',
                    'aie': context['aie'], 'accion': context['accion'],
                    'unidad_codigo': context['unidad_codigo'],
                    'accion_clave': action_key,
                    'categoria_programatica': context['categoria_programatica'],
                    'indicador': '', 'formula': '', 'unidad_medida': '',
                    'meta': '0', 'fecha_inicio': None, 'fecha_fin': None,
                    'responsable': '',
                    'programacion_mensual': {month: '0' for month in MESES},
                    'total_anual': '0',
                })
                emitted_actions[action_key] = nodes[-1]
        else:
            action_key = ''

        explicit = _clave(_cell(values, mapping, 'nivel'))
        explicit_level = {
            'operacion': 'operacion', 'operaciones': 'operacion',
            'actividad': 'actividad', 'actividades': 'actividad',
            'tarea': 'tarea', 'tareas': 'tarea',
        }.get(explicit)
        present = [level for level in ('operacion', 'actividad', 'tarea') if raw[level]]
        levels = [explicit_level] if explicit_level else present
        if not levels:
            issues.extend(row_issues)
            continue
        if explicit_level and not raw[explicit_level]:
            row_issues.append(_error(
                row_number, explicit_level,
                f'La denominación de {explicit_level} es obligatoria.',
            ))
            issues.extend(row_issues)
            continue

        months = {}
        missing_months = []
        for month in MESES:
            value = _decimal(_cell(values, mapping, month), month, row_number, row_issues)
            if value is None:
                missing_months.append(month)
                value = Decimal('0')
            months[month] = str(value)
        if len(missing_months) == 12:
            row_issues.append(_warning(
                row_number, 'programacion_mensual',
                'No hay programación mensual; se usarán ceros.',
                'empty_months',
            ))
        elif missing_months:
            row_issues.append(_warning(
                row_number, 'programacion_mensual',
                'Los meses vacíos se completarán con cero: ' + ', '.join(missing_months) + '.',
                'missing_months',
            ))
        total = sum(Decimal(value) for value in months.values())
        declared_total = _decimal(
            _cell(values, mapping, 'total_anual'), 'total_anual', row_number, row_issues,
        )
        if declared_total is not None and abs(declared_total - total) > Decimal('0.0001'):
            row_issues.append(_warning(
                row_number, 'total_anual',
                f'El total declarado ({declared_total}) no coincide con los meses ({total}).',
                'total_mismatch',
            ))
        meta = _decimal(_cell(values, mapping, 'meta'), 'meta', row_number, row_issues)
        if meta is not None and abs(meta - total) > Decimal('0.0001'):
            row_issues.append(_warning(
                row_number, 'meta',
                f'La meta ({meta}) no coincide con la programación mensual ({total}).',
                'meta_mismatch',
            ))
        if meta is None:
            meta = total
            row_issues.append(_warning(
                row_number, 'meta',
                'La meta está vacía; se usará el total programado.',
                'missing_value',
            ))

        linea_base = _decimal(
            _cell(values, mapping, 'linea_base'), 'linea_base',
            row_number, row_issues,
        )
        meta_actual = _decimal(
            _cell(values, mapping, 'meta_actual'), 'meta_actual',
            row_number, row_issues,
        )
        ponderacion = _decimal(
            _cell(values, mapping, 'ponderacion'), 'ponderacion',
            row_number, row_issues,
        )
        if ponderacion is not None and ponderacion > Decimal('100'):
            row_issues.append(_warning(
                row_number, 'ponderacion',
                'La ponderación supera el 100%; se importará igual.',
                'out_of_range',
            ))

        start = _fecha(
            _cell(values, mapping, 'fecha_inicio'), 'fecha_inicio',
            row_number, gestion, row_issues,
        )
        end = _fecha(
            _cell(values, mapping, 'fecha_fin'), 'fecha_fin',
            row_number, gestion, row_issues,
        )
        if bool(start) != bool(end):
            row_issues.append(_warning(
                row_number, 'fechas',
                'La fecha de inicio o fin está vacía.', 'missing_value',
            ))
        elif not start and not end:
            row_issues.append(_warning(
                row_number, 'fechas', 'Las fechas están vacías.', 'missing_value',
            ))
        elif start > end:
            row_issues.append(_warning(
                row_number, 'fecha_fin',
                'La fecha final es anterior al inicio; se importará igual.',
                'invalid_range',
            ))

        if not context['accion']:
            row_issues.append(_error(
                row_number, 'accion', 'Falta la Acción de corto plazo.', 'missing_parent',
            ))
        for level in levels:
            if level in ('actividad', 'tarea') and not context['operacion']:
                row_issues.append(_error(
                    row_number, 'operacion', 'Falta la operación padre.', 'missing_parent',
                ))
            if level == 'tarea' and not context['actividad']:
                row_issues.append(_error(
                    row_number, 'actividad', 'Falta la actividad padre.', 'missing_parent',
                ))
            node = {
                'fila': row_number, 'nivel': level,
                'aie': context['aie'], 'accion': context['accion'],
                'unidad_codigo': context['unidad_codigo'],
                'accion_clave': action_key,
                # La categoría vigente en esta fila. La columna la trae la fila
                # de la operación, y las de actividad/tarea que cuelgan de ella
                # la heredan del contexto hasta que aparezca otra.
                'categoria_programatica': context['categoria_programatica'],
                'operacion_codigo': context['operacion_codigo'],
                'operacion': context['operacion'],
                'actividad_codigo': context['actividad_codigo'],
                'actividad': context['actividad'],
                'tarea_codigo': codes['tarea'] if level == 'tarea' else '',
                'tarea': raw['tarea'] if level == 'tarea' else '',
                'tipo_operacion': _texto(_cell(values, mapping, 'tipo_operacion')),
                'indicador': _texto(_cell(values, mapping, 'indicador')),
                'formula': _texto(_cell(values, mapping, 'formula')),
                'unidad_medida': _texto(_cell(values, mapping, 'unidad_medida')),
                'linea_base': str(linea_base) if linea_base is not None else None,
                'meta': str(meta),
                'meta_actual': str(meta_actual) if meta_actual is not None else None,
                'ponderacion': str(ponderacion) if ponderacion is not None else None,
                'fecha_inicio': start.isoformat() if start else None,
                'fecha_fin': end.isoformat() if end else None,
                'responsable': _texto(_cell(values, mapping, 'responsable')),
                'programacion_mensual': months,
                'total_anual': str(total),
            }
            nodes.append(node)
        issues.extend(row_issues)

    return nodes, issues, rows_read


def _node_key(node):
    parts = [node.get('accion_clave') or _clave(node.get('accion'))]
    if node['nivel'] == 'accion':
        return '|'.join(parts)
    for level in ('operacion', 'actividad', 'tarea'):
        if level == 'tarea' and node['nivel'] != 'tarea':
            break
        code = node[f'{level}_codigo']
        name = node[level]
        if code or name:
            parts.append(f'{level}:{_clave(code) if code else "nombre:" + _clave(name)}')
        if node['nivel'] == level:
            break
    return '|'.join(parts)


def _match_existing(node, actions, operations, activities, tasks):
    action = actions.get(node['accion_codigo'])
    if action is None:
        return None
    op = _match_child(
        operations.get(action.id, []), 'codigo_operacion',
        node['operacion_codigo'], node['operacion'],
    )
    if node['nivel'] == 'operacion' or op is None:
        return op
    activity = _match_child(
        activities.get(op.id, []), 'codigo_actividad',
        node['actividad_codigo'], node['actividad'],
    )
    if node['nivel'] == 'actividad' or activity is None:
        return activity
    return _match_child(
        tasks.get(activity.id, []), 'codigo_tarea',
        node['tarea_codigo'], node['tarea'],
    )


def _match_child(objects, code_field, code, name):
    if code:
        return next((obj for obj in objects if getattr(obj, code_field).upper() == code.upper()), None)
    matches = [obj for obj in objects if _clave(obj.denominacion) == _clave(name)]
    return matches[0] if len(matches) == 1 else None


def _catalog_value(model, gestion, value):
    if not value:
        return None, False
    key = _clave(value)
    entries = list(model.objects.filter(gestion=gestion, activo=True))
    if not entries:
        return value, False
    matches = [entry for entry in entries if key in (_clave(entry.codigo), _clave(entry.denominacion))]
    return (matches[0].denominacion, True) if len(matches) == 1 else (value, False)


def _database_errors(nodes, gestion, unidad):
    errors = []
    actions_qs = AccionPOA.objects.filter(
        gestion=gestion.anio, unidad_responsable=unidad,
    )
    actions = {action.codigo_accion.upper(): action for action in actions_qs}
    operations = defaultdict(list)
    activities = defaultdict(list)
    tasks = defaultdict(list)
    for operation in OperacionPOAU.objects.filter(accion_poa__in=actions_qs):
        operations[operation.accion_poa_id].append(operation)
    for activity in ActividadPOAU.objects.filter(operacion__accion_poa__in=actions_qs):
        activities[activity.operacion_id].append(activity)
    for task in TareaPOAU.objects.filter(actividad__operacion__accion_poa__in=actions_qs):
        tasks[task.actividad_id].append(task)

    seen = {}
    imported_keys = {
        level: {_node_key(node) for node in nodes if node['nivel'] == level}
        for level in ('operacion', 'actividad', 'tarea')
    }
    all_codes = defaultdict(dict)
    for model, code_field in (
        (OperacionPOAU, 'codigo_operacion'),
        (ActividadPOAU, 'codigo_actividad'),
        (TareaPOAU, 'codigo_tarea'),
    ):
        all_codes[model] = {
            code.upper(): pk for code, pk in model.objects.values_list(code_field, 'pk')
        }

    for node in nodes:
        row = node['fila']
        if node['nivel'] in ('actividad', 'tarea'):
            operation_key = _node_key({**node, 'nivel': 'operacion'})
            if operation_key not in imported_keys['operacion']:
                errors.append(_error(
                    row, 'operacion',
                    'La operación padre debe estar incluida en la importación.',
                    'missing_parent',
                ))
        if node['nivel'] == 'tarea':
            activity_key = _node_key({**node, 'nivel': 'actividad'})
            if activity_key not in imported_keys['actividad']:
                errors.append(_error(
                    row, 'actividad',
                    'La actividad padre debe estar incluida en la importación.',
                    'missing_parent',
                ))
        if node['unidad_codigo'] != unidad.codigo.upper():
            errors.append(_error(
                row, 'unidad_codigo',
                f'La fila pertenece a {node["unidad_codigo"]}, no a {unidad.codigo}.',
                'mixed_unit',
            ))
        action = actions.get(node['accion_codigo'])
        if action is None:
            errors.append(_error(
                row, 'accion_codigo',
                'La acción no existe para la unidad y gestión seleccionadas.',
                'foreign_key',
            ))
            continue
        key = _node_key(node)
        if key in seen:
            errors.append(_error(
                row, node['nivel'],
                f'Duplica la fila {seen[key]} dentro de la misma jerarquía.',
                'duplicate',
            ))
        else:
            seen[key] = row

        existing = _match_existing(node, actions, operations, activities, tasks)
        model, code_field = {
            'operacion': (OperacionPOAU, 'operacion_codigo'),
            'actividad': (ActividadPOAU, 'actividad_codigo'),
            'tarea': (TareaPOAU, 'tarea_codigo'),
        }[node['nivel']]
        code = node[code_field]
        if code and code.upper() in all_codes[model] and (
            existing is None or all_codes[model][code.upper()] != existing.pk
        ):
            errors.append(_error(
                row, code_field,
                f'El código {code} ya pertenece a otra jerarquía.',
                'duplicate_code',
            ))

        if node['nivel'] in ('operacion', 'actividad'):
            if not node['unidad_medida'] and existing is not None:
                node['unidad_medida'] = existing.unidad_medida
            canonical, valid = _catalog_value(
                UnidadMedida, gestion, node['unidad_medida'],
            )
            if not node['unidad_medida']:
                errors.append(_error(row, 'unidad_medida', 'La unidad de medida es obligatoria.'))
            elif not valid:
                errors.append(_error(
                    row, 'unidad_medida', 'La unidad de medida no existe en el catálogo vigente.',
                    'foreign_key',
                ))
            else:
                node['unidad_medida'] = canonical
        if node['nivel'] == 'operacion':
            if not node['tipo_operacion'] and existing is not None:
                node['tipo_operacion'] = existing.tipo_operacion
            canonical, valid = _catalog_value(
                TipoOperacion, gestion, node['tipo_operacion'],
            )
            if not node['tipo_operacion']:
                errors.append(_error(row, 'tipo_operacion', 'El tipo de operación es obligatorio.'))
            elif not valid:
                errors.append(_error(
                    row, 'tipo_operacion', 'El tipo de operación no existe en el catálogo vigente.',
                    'foreign_key',
                ))
            else:
                node['tipo_operacion'] = canonical

    return errors


def _database_errors_v2(nodes, gestion, unidad):
    issues, seen = [], {}
    products = defaultdict(list)
    for product in ProductoPEI.objects.all():
        products[_clave(product.denominacion)].append(product)
    imported = {
        level: {_node_key(node) for node in nodes if node['nivel'] == level}
        for level in ('accion', 'operacion', 'actividad', 'tarea')
    }
    for node in nodes:
        row, level, key = node['fila'], node['nivel'], _node_key(node)
        if node.get('unidad_codigo') != unidad.codigo.upper():
            issues.append(_error(
                row, 'unidad_codigo',
                f'La fila pertenece a {node.get("unidad_codigo") or "otra unidad"}, '
                f'no a {unidad.codigo}.',
                'mixed_unit',
            ))
        if key in seen:
            issues.append(_error(
                row, level,
                f'El registro está duplicado; ya aparece en la fila {seen[key]}.',
                'duplicate',
            ))
        else:
            seen[key] = row
        if level == 'accion':
            matches = products.get(_clave(node.get('aie')), [])
            if not node.get('aie'):
                issues.append(_error(row, 'aie', 'La AIE (PEI) es obligatoria.'))
            elif not matches:
                node['producto_pei_provisional'] = True
                issues.append(_warning(
                    row, 'aie',
                    'La AIE no existe todavía en la matriz PEI; se creará con '
                    'código provisional al aplicar la importación.',
                    'provisional_foreign_key',
                ))
            elif len(matches) > 1:
                issues.append(_error(
                    row, 'aie',
                    'La AIE coincide con más de un Producto PEI; debe ser inequívoca.',
                    'ambiguous_foreign_key',
                ))
            else:
                node['producto_pei_id'] = str(matches[0].id)
                node['producto_pei_codigo'] = matches[0].codigo_producto
            if not node.get('accion'):
                issues.append(_error(
                    row, 'accion', 'La Acción de corto plazo es obligatoria.',
                ))
            continue
        if node.get('accion_clave') not in imported['accion']:
            issues.append(_error(
                row, 'accion',
                'La Acción de corto plazo padre debe estar incluida.',
                'missing_parent',
            ))
        if level in ('actividad', 'tarea'):
            operation_key = _node_key({**node, 'nivel': 'operacion'})
            if operation_key not in imported['operacion']:
                issues.append(_error(
                    row, 'operacion',
                    'La operación padre debe estar incluida en la importación.',
                    'missing_parent',
                ))
        if level == 'tarea':
            activity_key = _node_key({**node, 'nivel': 'actividad'})
            if activity_key not in imported['actividad']:
                issues.append(_error(
                    row, 'actividad',
                    'La actividad padre debe estar incluida en la importación.',
                    'missing_parent',
                ))
        canonical, valid = _catalog_value(
            UnidadMedida, gestion, node['unidad_medida'],
        )
        if not node['unidad_medida']:
            issues.append(_warning(
                row,
                'unidad_medida',
                f'El registro de nivel {level} no tiene unidad de medida; '
                'se importará sin este dato.',
                'missing_value',
            ))
        elif not valid:
            issues.append(_error(
                row,
                'unidad_medida',
                'La unidad de medida no existe en el catálogo vigente.',
                'foreign_key',
            ))
        else:
            node['unidad_medida'] = canonical
        if level == 'operacion':
            canonical, valid = _catalog_value(
                TipoOperacion, gestion, node['tipo_operacion'],
            )
            if not node['tipo_operacion']:
                issues.append(_warning(
                    row, 'tipo_operacion',
                    'Seleccione el tipo de operación antes de aplicar.',
                    'missing_value',
                ))
            elif not valid:
                issues.append(_error(
                    row, 'tipo_operacion',
                    'El tipo de operación no existe en el catálogo vigente.',
                    'foreign_key',
                ))
            else:
                node['tipo_operacion'] = canonical
        for field, label in (('indicador', 'indicador'), ('formula', 'fórmula')):
            if not node.get(field):
                issues.append(_warning(
                    row, field, f'Falta {label}.', 'missing_value',
                ))
        if level in ('operacion', 'tarea') and not node.get('responsable'):
            issues.append(_warning(
                row, 'responsable', 'Falta responsable.', 'missing_value',
            ))
    return issues


def parse_workbook(content, sheet_name, gestion, unidad, fallback_action_code=''):
    if len(content) > MAX_WORKBOOK_BYTES:
        raise ImportacionError('El archivo supera el límite de 10 MiB.')
    if not content.startswith(b'PK\x03\x04'):
        raise ImportacionError('El archivo no tiene una firma XLSX válida.')
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True,
        )
    except Exception as exc:
        raise ImportacionError('No se pudo leer el libro XLSX.') from exc
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ImportacionError(f'La hoja «{sheet_name}» no existe en el libro.')
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        nodes, errors, rows_read = _parse_sheet(
            sheet, gestion, '', unidad.codigo.upper(),
        )
        errors.extend(_database_errors_v2(nodes, gestion, unidad))
        blocking = [issue for issue in errors if _is_blocking(issue)]
        warnings = [issue for issue in errors if not _is_blocking(issue)]
        invalid_rows = {error['fila'] for error in blocking if error['fila'] > 0}
        summary = {
            'filas_leidas': rows_read,
            'filas_validas': max(0, len(nodes) - len(invalid_rows)),
            'filas_rechazadas': len(invalid_rows),
            'errores': len(blocking),
            'advertencias': len(warnings),
            'registros_preview': len(nodes),
        }
        return nodes, errors, summary, sheet.title
    finally:
        workbook.close()


class _SafeGoogleRedirect(HTTPRedirectHandler):
    def redirect_request(self, request, fp, code, msg, headers, newurl):
        parts = urlsplit(newurl)
        host = (parts.hostname or '').lower()
        allowed = host == 'docs.google.com' or host.endswith('.googleusercontent.com')
        if parts.scheme != 'https' or not allowed or parts.username or parts.password:
            raise ImportacionError('Google devolvió una redirección no permitida.')
        return super().redirect_request(request, fp, code, msg, headers, newurl)


def download_google_sheet(url, sheet_name=''):
    parts = urlsplit(url)
    if (
        parts.scheme != 'https' or parts.hostname != 'docs.google.com'
        or parts.port not in (None, 443) or parts.username or parts.password
    ):
        raise ImportacionError('La URL debe pertenecer a https://docs.google.com.')
    match = re.fullmatch(r'/spreadsheets/d/([^/]+)(?:/edit)?/?', parts.path)
    if not match or not GOOGLE_ID_RE.fullmatch(match.group(1)):
        raise ImportacionError('La URL de Google Sheets no tiene un identificador válido.')
    query = parse_qs(parts.query)
    fragment = parse_qs(parts.fragment)
    gid = (query.get('gid') or fragment.get('gid') or [''])[0]
    if gid and not gid.isdigit():
        raise ImportacionError('El gid de la hoja debe ser numérico.')
    if not gid and not sheet_name:
        raise ImportacionError('Indique la hoja o use una URL que incluya gid.')
    export_query = {'format': 'xlsx'}
    if gid:
        export_query['gid'] = gid
    export_url = (
        f'https://docs.google.com/spreadsheets/d/{match.group(1)}/export?'
        f'{urlencode(export_query)}'
    )
    request = Request(export_url, headers={'User-Agent': 'PIP-POAU-Importer/1.0'})
    try:
        response = build_opener(_SafeGoogleRedirect()).open(request, timeout=10)
        try:
            declared = response.headers.get('Content-Length')
            if declared and int(declared) > MAX_WORKBOOK_BYTES:
                raise ImportacionError('La hoja supera el límite de 10 MiB.')
            content = response.read(MAX_WORKBOOK_BYTES + 1)
        finally:
            response.close()
    except ImportacionError:
        raise
    except Exception as exc:
        raise ImportacionError('No se pudo descargar la hoja de Google.') from exc
    if len(content) > MAX_WORKBOOK_BYTES:
        raise ImportacionError('La hoja supera el límite de 10 MiB.')
    if not content.startswith(b'PK\x03\x04'):
        raise ImportacionError('Google no devolvió un archivo XLSX válido.')
    return content, f'Google Sheets {match.group(1)}', gid


def create_preview(
    *, request, origin, unit_code, content, source_name, sheet_name='',
    action_code='',
):
    gestion = gestion_del_candado(request)
    try:
        unidad = UnidadOrganizacional.objects.get(
            codigo__iexact=unit_code, gestion=gestion, activo=True,
        )
    except UnidadOrganizacional.DoesNotExist as exc:
        raise ImportacionError('La unidad no existe en la gestión habilitada.') from exc
    if not request.user.is_superuser and not ScopeResolver.puede_operar(
        request.user, unidad.id, gestion.id,
    ):
        raise ImportacionError('La unidad está fuera de su alcance organizacional.')
    request_data = getattr(request, 'data', {})
    fallback_action_code = _texto(
        action_code or request_data.get('accion_codigo', ''),
    ).upper()
    if fallback_action_code:
        action = AccionPOA.objects.filter(
            codigo_accion__iexact=fallback_action_code,
            gestion=gestion.anio,
            unidad_responsable=unidad,
        ).first()
        if action is None:
            raise ImportacionError(
                'La Acción POA objetivo no pertenece a la unidad y gestión '
                'habilitada.',
            )
        fallback_action_code = action.codigo_accion.upper()
    nodes, errors, summary, selected_sheet = parse_workbook(
        content, sheet_name, gestion, unidad, fallback_action_code,
    )
    if not nodes and not errors:
        errors.append(_error(0, 'archivo', 'La hoja no contiene programación física.'))
        summary['errores'] = 1
    state = (
        ImportacionProgramacionFisica.Estado.INVALIDO
        if any(_is_blocking(issue) for issue in errors)
        else ImportacionProgramacionFisica.Estado.VALIDO
    )
    return ImportacionProgramacionFisica.objects.create(
        gestion=gestion,
        unidad=unidad,
        creado_por=request.user,
        origen=origin,
        fuente_nombre=source_name[:300],
        hoja=selected_sheet[:200],
        fuente_sha256=hashlib.sha256(content).hexdigest(),
        filas_normalizadas=nodes,
        errores=errors,
        resumen=summary,
        estado=state,
        expira_en=timezone.now() + PREVIEW_TTL,
    )


def serialize_preview(preview):
    return {
        'id': str(preview.id),
        'estado': preview.estado,
        'gestion': preview.gestion.anio,
        'unidad': {
            'id': str(preview.unidad_id),
            'codigo': preview.unidad.codigo,
            'nombre': preview.unidad.nombre,
        },
        'origen': preview.origen,
        'fuente_nombre': preview.fuente_nombre,
        'hoja': preview.hoja,
        'resumen': preview.resumen,
        'errores': preview.errores,
        # Se manda la vista previa completa. El recorte anterior a 100 filas
        # escondía las operaciones que caían más abajo, y el asistente pide el
        # tipo de cada operación en su propia fila: una operación fuera del
        # corte no tenía desplegable y no había forma de tiparla.
        'filas': preview.filas_normalizadas,
        'tipos_operacion': list(
            TipoOperacion.objects.filter(
                gestion=preview.gestion, activo=True,
            ).order_by('denominacion').values('codigo', 'denominacion')
        ),
        'expira_en': preview.expira_en.isoformat(),
        'resultado': preview.resultado,
    }


def _fields_for(node):
    common = {
        'denominacion': node[node['nivel']],
        'fecha_inicio': date.fromisoformat(node['fecha_inicio']) if node['fecha_inicio'] else None,
        'fecha_fin': date.fromisoformat(node['fecha_fin']) if node['fecha_fin'] else None,
        'programacion_mensual': node['programacion_mensual'],
        'estado': EstadosPOAU.BORRADOR,
        'observacion': '',
        'linea_base': Decimal(node['linea_base']) if node.get('linea_base') is not None else None,
        'meta_actual': Decimal(node['meta_actual']) if node.get('meta_actual') is not None else None,
        'ponderacion': Decimal(node['ponderacion']) if node.get('ponderacion') is not None else None,
    }
    if node['nivel'] == 'operacion':
        return {
            **common,
            'categoria_programatica': node.get('categoria_programatica', ''),
            'tipo_operacion': node['tipo_operacion'],
            'indicador': node['indicador'],
            'formula': node['formula'],
            'unidad_medida': node['unidad_medida'],
            'meta_anual': Decimal(node['meta']),
            'total_programado': Decimal(node['total_anual']),
            'responsable': node['responsable'],
        }
    if node['nivel'] == 'actividad':
        return {
            **common,
            'indicador': node['indicador'],
            'formula': node['formula'],
            'unidad_medida': node['unidad_medida'],
            'meta_anual': Decimal(node['meta']),
            'total_programado': Decimal(node['total_anual']),
        }
    return {
        **common,
        'metas': Decimal(node['meta']),
        'responsable': node['responsable'],
        'indicador': node['indicador'],
        'formula': node['formula'],
        'unidad_medida': node['unidad_medida'],
    }


def _next_identity(model, parent_filter, parent_code, code_field, requested_code):
    maximum = model.objects.filter(**parent_filter).aggregate(value=Max('correlativo'))['value'] or 0
    correlativo = maximum + 1
    while correlativo <= 999:
        code = requested_code or f'{parent_code}.{correlativo}'[:50]
        if not model.objects.filter(**{code_field: code}).exists():
            return code, correlativo, model.generar_segmento(correlativo)
        if requested_code:
            raise ImportacionError(f'El código {requested_code} ya existe.')
        correlativo += 1
    raise ImportacionError('No quedan correlativos disponibles para la jerarquía.')


def _has_external_dependencies(obj, ignored_models=()):
    dependencies = []
    for relation in obj._meta.related_objects:
        if relation.related_model in ignored_models:
            continue
        accessor = relation.get_accessor_name()
        if not accessor or accessor == '+':
            continue
        try:
            related = getattr(obj, accessor)
            exists = related.exists() if hasattr(related, 'exists') else related is not None
        except ObjectDoesNotExist:
            exists = False
        if exists:
            dependencies.append(relation.related_model._meta.verbose_name)
    return sorted(set(dependencies))


def _apply_values(obj, values):
    changed = []
    for field, value in values.items():
        if getattr(obj, field) != value:
            setattr(obj, field, value)
            changed.append(field)
    if changed:
        obj.save(update_fields=[*changed, 'updated_at'])
    return bool(changed)


def _json_value(value):
    if isinstance(value, (date, datetime, Decimal, uuid.UUID)):
        return value.isoformat() if hasattr(value, 'isoformat') else str(value)
    return value


def _serialize_queryset(queryset):
    return [{
        field.attname: _json_value(getattr(obj, field.attname))
        for field in obj._meta.concrete_fields
    } for obj in queryset]


def snapshot_unit_tree(gestion, unidad):
    actions = list(AccionPOA.objects.filter(
        gestion=gestion.anio, unidad_responsable=unidad,
    ).order_by('codigo_accion'))
    operations = list(OperacionPOAU.objects.filter(
        accion_poa__in=actions,
    ).order_by('codigo_operacion'))
    activities = list(ActividadPOAU.objects.filter(
        operacion__in=operations,
    ).order_by('codigo_actividad'))
    tasks = list(TareaPOAU.objects.filter(
        actividad__in=activities,
    ).order_by('codigo_tarea'))
    # Las asignaciones presupuestarias no son parte del árbol POAU, pero el
    # reemplazo las borra (es la única FK con on_delete=PROTECT sobre esta
    # jerarquía) — quedan acá para que el monto asignado nunca se pierda,
    # aunque el registro vivo ya no exista. `schema` se mantiene en 1:
    # `restore_version()` solo lee las cuatro claves de árbol y no le importa
    # esta adicional.
    asignaciones = list(AsignacionPresupuestariaUnidad.objects.filter(
        Q(operacion__in=operations) | Q(actividad__in=activities) | Q(tarea__in=tasks),
    ))
    return {
        'schema': 1,
        'acciones': _serialize_queryset(actions),
        'operaciones': _serialize_queryset(operations),
        'actividades': _serialize_queryset(activities),
        'tareas': _serialize_queryset(tasks),
        'asignaciones_presupuestarias': _serialize_queryset(asignaciones),
    }


def _locked_tree(gestion, unidad):
    actions = list(AccionPOA.objects.select_for_update().filter(
        gestion=gestion.anio, unidad_responsable=unidad,
    ))
    operations = list(OperacionPOAU.objects.select_for_update().filter(accion_poa__in=actions))
    activities = list(ActividadPOAU.objects.select_for_update().filter(operacion__in=operations))
    tasks = list(TareaPOAU.objects.select_for_update().filter(actividad__in=activities))
    return actions, operations, activities, tasks


def _assert_replaceable(actions, operations, activities, tasks):
    checks = (
        *((obj, ()) for obj in tasks),
        *((obj, (TareaPOAU,)) for obj in activities),
        *((obj, (ActividadPOAU,)) for obj in operations),
        *((obj, (OperacionPOAU,)) for obj in actions),
    )
    for obj, ignored in checks:
        dependencies = _has_external_dependencies(obj, ignored)
        if dependencies:
            raise ImportacionError(
                f'No se puede reemplazar {obj}: tiene dependencias no incluidas '
                f'en el historial ({", ".join(dependencies)}).'
            )


def _enable_rebuild_override():
    if connection.vendor == 'postgresql':
        with connection.cursor() as cursor:
            cursor.execute("SET LOCAL pip.poau_rebuild = 'on'")


def _raw_delete(objects):
    if not objects:
        return
    model = type(objects[0])
    table = connection.ops.quote_name(model._meta.db_table)
    pk_field = model._meta.pk
    column = connection.ops.quote_name(pk_field.column)
    with connection.cursor() as cursor:
        for obj in objects:
            prepared = pk_field.get_db_prep_value(obj.pk, connection, prepared=False)
            cursor.execute(f'DELETE FROM {table} WHERE {column} = %s', [prepared])


def _delete_tree(actions, operations, activities, tasks):
    _enable_rebuild_override()
    for objects in (tasks, activities, operations, actions):
        _raw_delete(objects)


def _action_values(node, gestion, unidad, product):
    return {
        'denominacion': node['accion'], 'producto_pei': product,
        'gestion': gestion.anio, 'unidad_responsable': unidad,
        'estado': EstadosPOAU.BORRADOR,
        'indicador': node.get('indicador', ''), 'formula': node.get('formula', ''),
        'unidad_medida': node.get('unidad_medida', ''),
        'meta_gestion': Decimal(node.get('meta') or '0'),
        'fecha_inicio': date.fromisoformat(node['fecha_inicio']) if node.get('fecha_inicio') else None,
        'fecha_fin': date.fromisoformat(node['fecha_fin']) if node.get('fecha_fin') else None,
        'cargo_responsable': node.get('responsable', ''),
        'categoria_programatica': node.get('categoria_programatica', ''),
    }


def _create_provisional_product(aie, gestion, unidad, result=None):
    """Create the temporary PEI references authorized for matrix-first POAUs."""
    if result is None:
        start = gestion.anio - 1
        institutional = ResultadoPEI.objects.order_by(
            'vigencia_desde', 'codigo_resultado',
        ).values('cod_entidad', 'entidad').first() or {}
        code, corr, segment = _next_identity(
            ResultadoPEI, {'vigencia_desde': start}, f'PROV-RI-{start}',
            'codigo_resultado', '',
        )
        result = ResultadoPEI.objects.create(
            codigo_resultado=code,
            correlativo=corr,
            segmento=segment,
            codigo_fuente=f'POAU-{gestion.anio}-{unidad.codigo}',
            denominacion=(
                f'Resultado PEI provisional para {unidad.codigo}, '
                f'gestión {gestion.anio}'
            ),
            cod_entidad=institutional.get('cod_entidad') or 'PROV',
            entidad=institutional.get('entidad') or 'Entidad pendiente de matriz PEI',
            vigencia_desde=start,
            vigencia_hasta=start + 4,
        )
    code, corr, segment = _next_identity(
        ProductoPEI, {'resultado_pei': result}, result.codigo_resultado,
        'codigo_producto', '',
    )
    product = ProductoPEI.objects.create(
        codigo_producto=code,
        correlativo=corr,
        segmento=segment,
        codigo_fuente=f'POAU-{gestion.anio}-{unidad.codigo}',
        denominacion=aie,
        resultado_pei=result,
        tipo_producto='INTERMEDIO',
    )
    return product, result


@transaction.atomic
def apply_preview(preview_id, user, confirmation_code='', operation_types=None):
    preview = (
        ImportacionProgramacionFisica.objects.select_for_update()
        .select_related('gestion', 'unidad').get(pk=preview_id)
    )
    if preview.creado_por_id != user.id and not user.is_superuser:
        raise ImportacionError('La vista previa pertenece a otro usuario.')
    if preview.estado == ImportacionProgramacionFisica.Estado.APLICADO:
        raise ImportacionError('La vista previa ya fue aplicada.')
    if preview.estado != ImportacionProgramacionFisica.Estado.VALIDO or any(
        _is_blocking(issue) for issue in preview.errores
    ):
        raise ImportacionError('No se puede aplicar una vista previa con errores.')
    if preview.expira_en <= timezone.now():
        raise ImportacionError('La vista previa expiró; genere una nueva.')
    if not user.is_superuser and not ScopeResolver.puede_operar(
        user, preview.unidad_id, preview.gestion_id,
    ):
        raise ImportacionError('La unidad está fuera de su alcance organizacional.')
    if _texto(confirmation_code).upper() != preview.unidad.codigo.upper():
        raise ImportacionError('Para reconstruir el POAU escriba exactamente el código de la unidad.')

    operation_types = operation_types or {}
    actions, operations, activities, tasks = _locked_tree(preview.gestion, preview.unidad)

    # Reemplaza también registros OFICIAL/APROBADO: la decisión confirmada es
    # "avisar y reconstruir todo", sin bloqueo manual previo — cubierto por
    # test_approved_record_is_replaced_and_new_tree_returns_to_draft.

    # Auditoría inmutable ANTES de tocar nada: cubre ejecución, presupuesto y
    # normativa que el reemplazo va a borrar, sin bloquear por su existencia.
    # `VersionImportacionPOAU` es append-only; `restore_version()` la usa para
    # deshacer el reemplazo.
    snapshot = snapshot_unit_tree(preview.gestion, preview.unidad)
    old_count = sum(len(items) for items in (actions, operations, activities, tasks))
    resumen = {
        'registros': old_count,
        'asignaciones_presupuestarias': len(
            snapshot.get('asignaciones_presupuestarias', []),
        ),
    }
    history = VersionImportacionPOAU.objects.create(
        gestion=preview.gestion, unidad=preview.unidad, usuario=user,
        tipo_evento=VersionImportacionPOAU.TipoEvento.REEMPLAZO,
        snapshot=snapshot, resumen=resumen,
        fuente_nombre=preview.fuente_nombre,
        fuente_sha256=preview.fuente_sha256, hoja=preview.hoja,
    )

    # Borrado por ORM, no SQL crudo: `AsignacionPresupuestariaUnidad` es la
    # única dependencia externa con on_delete=PROTECT sobre operación/
    # actividad/tarea, así que hay que borrarla primero explícitamente (ya
    # quedó en el snapshot de arriba, no se pierde). Las otras dependencias
    # externas (seguimiento de ejecución, asignación de objeto de gasto,
    # normativa) son CASCADE: el delete() del ORM las arrastra solo.
    AsignacionPresupuestariaUnidad.objects.filter(
        Q(tarea_id__in=[obj.id for obj in tasks])
        | Q(actividad_id__in=[obj.id for obj in activities])
        | Q(operacion_id__in=[obj.id for obj in operations]),
    ).delete()
    TareaPOAU.objects.filter(pk__in=[obj.id for obj in tasks]).delete()
    ActividadPOAU.objects.filter(pk__in=[obj.id for obj in activities]).delete()
    OperacionPOAU.objects.filter(pk__in=[obj.id for obj in operations]).delete()
    AccionPOA.objects.filter(pk__in=[obj.id for obj in actions]).delete()
    resolved_actions, resolved_ops, resolved_acts = {}, {}, {}
    provisional_products = {}
    provisional_result = None
    created = 0
    for node in preview.filas_normalizadas:
        key = _node_key(node)
        if node['nivel'] == 'accion':
            product_id = node.get('producto_pei_id')
            if product_id:
                try:
                    product = ProductoPEI.objects.get(pk=product_id)
                except ProductoPEI.DoesNotExist as exc:
                    raise ImportacionError('La AIE cambió desde la vista previa.') from exc
            elif node.get('producto_pei_provisional'):
                product_key = _clave(node['aie'])
                product = provisional_products.get(product_key)
                if product is None:
                    product, provisional_result = _create_provisional_product(
                        node['aie'], preview.gestion, preview.unidad,
                        provisional_result,
                    )
                    provisional_products[product_key] = product
            else:
                raise ImportacionError('La AIE cambió desde la vista previa.')
            code, corr, segment = _next_identity(
                AccionPOA,
                {'producto_pei': product, 'gestion': preview.gestion.anio},
                f'PROV-{preview.gestion.anio}-{preview.unidad.codigo}',
                'codigo_accion', '',
            )
            action = AccionPOA.objects.create(
                codigo_accion=code, correlativo=corr, segmento=segment,
                **_action_values(node, preview.gestion, preview.unidad, product),
            )
            resolved_actions[node['accion_clave']] = action
            created += 1
            continue
        action = resolved_actions.get(node.get('accion_clave'))
        if action is None:
            raise ImportacionError('La acción padre no pudo reconstruirse.')
        values = _fields_for(node)
        if node['nivel'] == 'operacion':
            selected = operation_types.get(key) or operation_types.get(str(node['fila']))
            if selected:
                canonical, valid = _catalog_value(TipoOperacion, preview.gestion, selected)
                if not valid:
                    raise ImportacionError('El tipo de operación seleccionado no es válido.')
                values['tipo_operacion'] = canonical
            code, corr, segment = _next_identity(
                OperacionPOAU, {'accion_poa': action}, action.codigo_accion,
                'codigo_operacion', '',
            )
            obj = OperacionPOAU.objects.create(
                codigo_operacion=code, correlativo=corr, segmento=segment,
                accion_poa=action, **values,
            )
            resolved_ops[key] = obj
        elif node['nivel'] == 'actividad':
            operation = resolved_ops.get(_node_key({**node, 'nivel': 'operacion'}))
            if operation is None:
                raise ImportacionError('La actividad no tiene una operación importada válida.')
            code, corr, segment = _next_identity(
                ActividadPOAU, {'operacion': operation}, operation.codigo_operacion,
                'codigo_actividad', '',
            )
            obj = ActividadPOAU.objects.create(
                codigo_actividad=code, correlativo=corr, segmento=segment,
                operacion=operation, **values,
            )
            resolved_acts[key] = obj
        else:
            activity = resolved_acts.get(_node_key({**node, 'nivel': 'actividad'}))
            if activity is None:
                raise ImportacionError('La tarea no tiene una actividad importada válida.')
            code, corr, segment = _next_identity(
                TareaPOAU, {'actividad': activity}, activity.codigo_actividad,
                'codigo_tarea', '',
            )
            TareaPOAU.objects.create(
                codigo_tarea=code, correlativo=corr, segmento=segment,
                actividad=activity, **values,
            )
        created += 1
    result = {
        'creados': created, 'actualizados': 0, 'eliminados': old_count,
        'sin_cambios': 0, 'reemplazados': old_count,
        'filas_aplicadas': len(preview.filas_normalizadas),
        'version_historial_id': str(history.id),
    }
    preview.estado = ImportacionProgramacionFisica.Estado.APLICADO
    preview.aplicado_en = timezone.now()
    preview.resultado = result
    preview.save(update_fields=['estado', 'aplicado_en', 'resultado', 'updated_at'])
    return preview


def _decode_value(field, value):
    if value is None:
        return None
    kind = field.get_internal_type()
    if kind == 'UUIDField':
        return uuid.UUID(str(value))
    if kind == 'DecimalField':
        return Decimal(str(value))
    if kind == 'DateTimeField':
        return parse_datetime(value) if isinstance(value, str) else value
    if kind == 'DateField':
        return parse_date(value) if isinstance(value, str) else value
    return value


def _restore_rows(model, rows):
    for row in rows:
        values = {
            field.attname: _decode_value(field, row[field.attname])
            for field in model._meta.concrete_fields
            if field.attname in row
        }
        obj = model(**values)
        obj.estado = EstadosPOAU.BORRADOR
        if obj.estado_codigo == obj.ESTADO_CODIGO_OFICIAL:
            obj._permitir_promocion_oficial = True
        obj.save(force_insert=True)


@transaction.atomic
def restore_version(version_id, user, confirmation_code=''):
    version = VersionImportacionPOAU.objects.select_for_update().select_related(
        'gestion', 'unidad',
    ).get(pk=version_id)
    if not user.is_superuser and not ScopeResolver.puede_operar(
        user, version.unidad_id, version.gestion_id,
    ):
        raise ImportacionError('La unidad está fuera de su alcance organizacional.')
    if _texto(confirmation_code).upper() != version.unidad.codigo.upper():
        raise ImportacionError(
            'Para restaurar escriba exactamente el código de la unidad.',
        )
    actions, operations, activities, tasks = _locked_tree(
        version.gestion, version.unidad,
    )
    _assert_replaceable(actions, operations, activities, tasks)
    current = snapshot_unit_tree(version.gestion, version.unidad)
    VersionImportacionPOAU.objects.create(
        gestion=version.gestion, unidad=version.unidad, usuario=user,
        tipo_evento=VersionImportacionPOAU.TipoEvento.RESTAURACION,
        snapshot=current,
        resumen={
            'registros': sum(
                len(items) for items in (actions, operations, activities, tasks)
            ),
        },
        motivo=f'Restauración de la versión {version.id}',
    )
    _delete_tree(actions, operations, activities, tasks)
    snapshot = version.snapshot
    if snapshot.get('schema') != 1:
        raise ImportacionError(
            'La versión histórica tiene un formato no soportado.',
        )
    _restore_rows(AccionPOA, snapshot.get('acciones', []))
    _restore_rows(OperacionPOAU, snapshot.get('operaciones', []))
    _restore_rows(ActividadPOAU, snapshot.get('actividades', []))
    _restore_rows(TareaPOAU, snapshot.get('tareas', []))
    return version


def serialize_version(version, include_snapshot=False):
    data = {
        'id': str(version.id), 'gestion': version.gestion.anio,
        'unidad_codigo': version.unidad.codigo,
        'tipo_evento': version.tipo_evento,
        'fuente_nombre': version.fuente_nombre, 'hoja': version.hoja,
        'resumen': version.resumen,
        'creado_en': version.created_at.isoformat(),
    }
    if include_snapshot:
        data['snapshot'] = version.snapshot
    return data
