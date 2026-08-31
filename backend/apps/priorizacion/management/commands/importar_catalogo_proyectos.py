"""Arma el catálogo maestro de nombres de proyecto.

**La categoría programática es condición de entrada.** Una fila sin categoría no
sirve para lo único que el catálogo hace —copiarla al proyecto priorizado del
acta—, así que no se crea. Lo que llega sin categoría se cuenta y se informa,
nunca se guarda a medias.

Tres fuentes, y hacen falta las tres:

`--sigep`
    Reportes planos: una fila por proyecto, con SISIN y categoría.

`--ejecucion`
    Reportes de ejecución por objeto de gasto. La categoría va como encabezado
    de bloque, no como columna, y trae el SISIN embebido en su propio código.
    Es la única fuente de las categorías `<programa> 0 <actividad>`, que son las
    que no cuelgan de un proyecto de inversión.

`--actas`
    El libro de actas. No crea filas: solo cuenta cuántas veces se priorizó cada
    nombre, que es lo que ordena el buscador.
"""
import difflib
import glob
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.priorizacion.catalogo_sigep import (distrito_de, leer_ejecucion,
                                              leer_plano)
from apps.priorizacion.models import (OrigenProyecto, ProyectoCatalogo,
                                      normalizar)

COL_ACTAS = [7, 9, 11, 13, 15, 17, 19]  # PROYECTO 1..7

UMBRAL_EMPAREJADO = 0.85


class Command(BaseCommand):
    help = ('Importa el catálogo de proyectos desde reportes del SIGEP. '
            'Solo entran las filas que traen categoría programática.')

    def add_arguments(self, parser):
        parser.add_argument('--sigep', help='Carpeta o glob de reportes planos')
        parser.add_argument('--ejecucion',
                            help='Carpeta o glob de reportes de ejecución')
        parser.add_argument('--actas', help='Libro de actas (.xlsx)')
        parser.add_argument('--hoja', default='REGISTRO ACTAS')
        parser.add_argument(
            '--solo-altas', action='store_true',
            help='Solo da de alta lo que no existe. No toca ninguna fila ya '
                 'cargada, ni siquiera para completarle la categoría.')
        parser.add_argument(
            '--solo-obra', action='store_true',
            help='Descarta lo que no sea obra priorizable por una OTB '
                 '(funcionamiento, servicios y transferencias).')
        parser.add_argument(
            '--emparejar-distrito', nargs='?', type=float,
            const=UMBRAL_EMPAREJADO, metavar='UMBRAL',
            help='Completa la categoría de las filas que no la tienen usando '
                 f'el nombre más parecido del MISMO distrito (umbral '
                 f'{UMBRAL_EMPAREJADO} por omisión).')
        parser.add_argument(
            '--purgar-sin-categoria', action='store_true',
            help='Borra del catálogo las filas que quedaron sin categoría. '
                 'Las actas ya emitidas no se tocan: guardan su propia copia.')
        parser.add_argument('--dry-run', action='store_true',
                            help='Informa sin escribir nada.')

    def handle(self, *args, **opciones):
        self.seco = opciones['dry_run']
        registros = {}

        if opciones['sigep']:
            self._leer_reportes(opciones['sigep'], registros, leer_plano,
                                'SIGEP', opciones['solo_obra'])
        if opciones['ejecucion']:
            self._leer_reportes(opciones['ejecucion'], registros, leer_ejecucion,
                                'Ejecución', opciones['solo_obra'])
        if opciones['actas']:
            self._leer_actas(opciones['actas'], opciones['hoja'], registros)

        if registros:
            self._guardar(registros, opciones['solo_altas'])
        elif not (opciones['emparejar_distrito'] or
                  opciones['purgar_sin_categoria']):
            raise CommandError(
                'Nada que importar: indique --sigep, --ejecucion y/o --actas.')

        if opciones['emparejar_distrito']:
            self._emparejar_por_distrito(opciones['emparejar_distrito'])
        if opciones['purgar_sin_categoria']:
            self._purgar_sin_categoria()

        self.stdout.write(self.style.SUCCESS(
            f'Catálogo maestro: {ProyectoCatalogo.objects.count()} filas, '
            f'{ProyectoCatalogo.objects.exclude(categoria_programatica="").count()}'
            ' con categoría programática.'))

    # --- Lectura -----------------------------------------------------------

    def _archivos(self, ruta):
        """Los `.xls` del SIGEP son BIFF viejo: openpyxl no los abre."""
        patron = os.path.join(ruta, '*.xls') if os.path.isdir(ruta) else ruta
        return sorted(glob.glob(patron))

    def _leer_reportes(self, ruta, registros, lector, rotulo, solo_obra):
        import xlrd

        archivos = self._archivos(ruta)
        if not archivos:
            self.stdout.write(self.style.WARNING(f'Sin .xls en {ruta}'))
            return

        leidas = omitidas = ilegibles = 0
        for archivo in archivos:
            hoja = xlrd.open_workbook(archivo).sheet_by_index(0)
            categorias = lector(hoja)
            if categorias is None:
                ilegibles += 1
                self.stdout.write(self.style.WARNING(
                    f'{os.path.basename(archivo)}: cabecera inesperada, '
                    'se omite.'))
                continue
            for cat in categorias:
                if not cat.es_proyecto or (solo_obra and not cat.es_obra):
                    omitidas += 1
                    continue
                leidas += 1
                clave = (normalizar(cat.nombre), cat.sisin)
                registros.setdefault(clave, {
                    'nombre': cat.nombre,
                    'sisin': cat.sisin,
                    'categoria': cat.codigo,
                    'denominacion': cat.nombre,
                    'origen': OrigenProyecto.SIGEP,
                    'veces': 0,
                })

        detalle = f', {omitidas} omitidas' if omitidas else ''
        detalle += f', {ilegibles} archivo(s) ilegibles' if ilegibles else ''
        self.stdout.write(
            f'{rotulo}: {leidas} categorías en {len(archivos)} archivo(s)'
            f'{detalle}.')

    def _leer_actas(self, archivo, hoja, registros):
        """Cuenta priorizaciones. No da de alta: un nombre de acta sin
        categoría programática no entra al catálogo."""
        import openpyxl

        ws = openpyxl.load_workbook(archivo, data_only=True)[hoja]
        contados = huerfanos = 0
        for fila in range(2, ws.max_row + 1):
            for columna in COL_ACTAS:
                nombre = ' '.join(str(ws.cell(fila, columna).value or '').split())
                if not nombre:
                    continue
                busqueda = normalizar(nombre)
                destino = next(
                    (v for (n, _), v in registros.items() if n == busqueda), None)
                if destino is None:
                    huerfanos += 1
                    continue
                destino['veces'] += 1
                contados += 1
        self.stdout.write(
            f'Actas: {contados} priorizaciones contadas, {huerfanos} sobre '
            'nombres que no están en ningún reporte.')

    # --- Escritura ---------------------------------------------------------

    @transaction.atomic
    def _guardar(self, registros, solo_altas=False):
        """Da de alta lo que falta y, salvo `solo_altas`, completa lo que ya
        estaba incompleto.

        Una categoría cuyo nombre ya existe en el catálogo no se puede insertar:
        la clave es (`nombre_busqueda`, `sisin`) y `proyecto_catalogo_unico` la
        rechaza. Para esas filas la única forma de completar la categoría es
        actualizar la que ya está, y `--solo-altas` renuncia a hacerlo.
        """
        conteo = {'creados': 0, 'completados': 0, 'sin_cambio': 0,
                  'intactos': 0, 'sin_categoria': 0}
        for (busqueda, sisin), datos in registros.items():
            if not datos['categoria']:
                conteo['sin_categoria'] += 1
                continue
            previo = ProyectoCatalogo.objects.filter(
                nombre_busqueda=busqueda, sisin=sisin).first()
            if previo and solo_altas:
                conteo['intactos'] += 1
                continue
            if previo:
                conteo['sin_cambio' if previo.categoria_programatica
                       else 'completados'] += 1
            else:
                conteo['creados'] += 1
            if self.seco:
                continue
            obj, _ = ProyectoCatalogo.objects.update_or_create(
                nombre_busqueda=busqueda, sisin=sisin,
                defaults={
                    'nombre': datos['nombre'],
                    'categoria_programatica': datos['categoria'],
                    'denominacion_categoria': datos['denominacion'],
                    'origen': datos['origen'],
                })
            # `veces_priorizado` se conserva: lo llenan las actas, y un reporte
            # del SIGEP que no las conoce lo pondría en cero.
            if datos['veces'] > obj.veces_priorizado:
                obj.veces_priorizado = datos['veces']
                obj.save(update_fields=['veces_priorizado'])

        prefijo = '[dry-run] ' if self.seco else ''
        detalle = [f"{conteo['creados']} creados"]
        if solo_altas:
            detalle.append(f"{conteo['intactos']} ya existían, sin tocar")
        else:
            detalle.append(f"{conteo['completados']} completados")
            detalle.append(f"{conteo['sin_cambio']} ya estaban completos")
        if conteo['sin_categoria']:
            detalle.append(f"{conteo['sin_categoria']} omitidos sin categoría")
        self.stdout.write(self.style.SUCCESS(
            prefijo + ', '.join(detalle) + '.'))

    @transaction.atomic
    def _emparejar_por_distrito(self, umbral):
        """Completa la categoría por parecido de nombre dentro del distrito.

        El parecido solo no alcanza: `... DISTRITO 4 (OTB ESMERALDA NORTE)` se
        parece 0.80 a `... DISTRITO 6`, y aceptarlo le pondría al acta la
        categoría del distrito equivocado. El distrito es la llave dura; la
        similitud solo desempata dentro de él.
        """
        con_categoria = list(
            ProyectoCatalogo.objects.exclude(categoria_programatica=''))
        candidatos = {}
        for p in con_categoria:
            candidatos.setdefault(distrito_de(p.nombre), []).append(p)

        emparejados = sin_distrito = sin_par = 0
        for huerfano in ProyectoCatalogo.objects.filter(
                categoria_programatica=''):
            distrito = distrito_de(huerfano.nombre)
            if not distrito:
                sin_distrito += 1
                continue
            pares = candidatos.get(distrito, [])
            claves = [p.nombre_busqueda for p in pares]
            mejor = difflib.get_close_matches(
                huerfano.nombre_busqueda, claves, n=1, cutoff=umbral)
            if not mejor:
                sin_par += 1
                continue
            fuente = pares[claves.index(mejor[0])]
            emparejados += 1
            if self.seco:
                continue
            huerfano.categoria_programatica = fuente.categoria_programatica
            huerfano.denominacion_categoria = fuente.nombre
            huerfano.save(update_fields=['categoria_programatica',
                                         'denominacion_categoria'])

        prefijo = '[dry-run] ' if self.seco else ''
        self.stdout.write(self.style.SUCCESS(
            f'{prefijo}Emparejado por distrito (umbral {umbral}): '
            f'{emparejados} completados, {sin_par} sin par en su distrito, '
            f'{sin_distrito} sin distrito en el nombre.'))

    @transaction.atomic
    def _purgar_sin_categoria(self):
        """Saca del catálogo lo que quedó sin categoría programática.

        Las actas ya emitidas no cambian: `ProyectoPriorizado` copia nombre y
        categoría en vez de leerlos del catálogo, y su vínculo es SET_NULL.
        """
        sobrantes = ProyectoCatalogo.objects.filter(categoria_programatica='')
        cuantas = sobrantes.count()
        priorizaciones = sum(p.veces_priorizado for p in sobrantes)
        if self.seco:
            self.stdout.write(self.style.WARNING(
                f'[dry-run] Se borrarían {cuantas} filas sin categoría '
                f'({priorizaciones} priorizaciones históricas).'))
            return
        sobrantes.delete()
        self.stdout.write(self.style.WARNING(
            f'Purgadas {cuantas} filas sin categoría '
            f'({priorizaciones} priorizaciones históricas).'))
