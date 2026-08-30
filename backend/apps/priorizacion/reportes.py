"""Reporte de proyectos programados: una fila por proyecto priorizado.

El listado de actas muestra una fila por acta; este reporte baja un nivel y
emite una fila por proyecto, que es la unidad con la que se trabaja el POA.
Un acta sin proyectos no aporta filas: no hay nada programado que informar.

Las dos salidas —Excel y PDF— se arman sobre las mismas filas para que no
puedan divergir: si el PDF y la planilla mostraran totales distintos, ninguno
de los dos serviría para respaldar nada.
"""
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

# Carta apaisada: nueve columnas no entran en vertical sin partir los nombres
# de proyecto en tiras ilegibles.
CARTA_APAISADA = landscape((216 * mm, 279 * mm))

# Verde institucional de la plataforma, el mismo que `styles.scss` expone como
# `--pip-green-*`. El reporte impreso y la pantalla tienen que verse de la misma
# entidad: un archivo que sale con otra paleta no parece emitido por el sistema.
VERDE_700 = '007229'   # --primary, encabezado
VERDE_800 = '034E1F'   # tinta de la fila de total
VERDE_100 = 'E3F2E9'   # fondo suave, fila de total
VERDE_CEBRA = 'F1F9F4'  # el mismo suave a media intensidad, para las rayas
VERDE_BORDE = 'CBE3D5'  # líneas de la grilla, verde apagado

# El ancho en mm es del PDF; el de Excel va en caracteres, que es la unidad de
# `column_dimensions`. Se declaran juntos para que agregar una columna obligue
# a decidir las dos medidas de una sola vez.
COLUMNAS = [
    ('gestion', 'GESTIÓN POA', 16, 12),
    ('distrito', 'DISTRITO', 26, 22),
    ('otb', 'OTB / JUNTA VECINAL', 42, 38),
    ('categoria_programatica', 'CATEGORÍA PROGRAMÁTICA', 30, 26),
    ('sisin', 'SISIN', 21, 20),
    ('numero_proyecto', 'N° PROYECTO', 17, 12),
    ('nombre', 'NOMBRE DEL PROYECTO', 52, 52),
    ('monto', 'MONTO PROYECTO', 24, 18),
    ('responsable_registro', 'RESPONSABLE DEL REGISTRO', 30, 28),
]

CLAVES = [c[0] for c in COLUMNAS]
TITULOS = [c[1] for c in COLUMNAS]

# `monto` es la única columna numérica con decimales: se alinea a la derecha y
# se suma. `numero_proyecto` y `gestion` son números que se leen como rótulos.
COL_MONTO = CLAVES.index('monto')


def filas_reporte(actas):
    """Aplana actas y proyectos en las filas del reporte.

    Recibe el queryset YA filtrado por la vista. No lo vuelve a filtrar: el
    reporte tiene que decir exactamente lo que la pantalla venía mostrando.
    """
    filas = []
    for acta in actas:
        # `proyectos.all()` aprovecha el prefetch del viewset; ordenar acá con
        # un `order_by` dispararía una consulta nueva por cada acta.
        for proyecto in sorted(acta.proyectos.all(), key=lambda p: p.orden):
            filas.append({
                'gestion': acta.gestion,
                'distrito': acta.distrito.nombre,
                'otb': acta.otb,
                'categoria_programatica': proyecto.categoria_programatica or '',
                'sisin': proyecto.sisin or '',
                'numero_proyecto': proyecto.orden,
                'nombre': proyecto.nombre,
                'monto': proyecto.monto or Decimal('0'),
                'responsable_registro': acta.responsable_registro or '',
            })
    return filas


def total_montos(filas):
    return sum((f['monto'] for f in filas), Decimal('0'))


# --- Excel -----------------------------------------------------------------

_RELLENO_TITULO = PatternFill('solid', fgColor=VERDE_700)
_RELLENO_CEBRA = PatternFill('solid', fgColor=VERDE_CEBRA)
_RELLENO_TOTAL = PatternFill('solid', fgColor=VERDE_100)
_BORDE = Border(*[Side(style='thin', color=VERDE_BORDE)] * 4)


def generar_reporte_excel(filas, subtitulo=''):
    """La planilla, con encabezado congelado y autofiltro.

    Se entrega como planilla de trabajo y no como imagen del PDF: quien la
    recibe filtra y suma por su cuenta, que es para lo que pide el Excel.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Proyectos programados'

    hoja.append(['PROYECTOS PROGRAMADOS'])
    hoja.merge_cells(start_row=1, start_column=1,
                     end_row=1, end_column=len(COLUMNAS))
    hoja['A1'].font = Font(bold=True, size=13, color=VERDE_800)
    hoja['A1'].alignment = Alignment(horizontal='center')

    hoja.append([subtitulo])
    hoja.merge_cells(start_row=2, start_column=1,
                     end_row=2, end_column=len(COLUMNAS))
    hoja['A2'].font = Font(italic=True, size=9, color=VERDE_700)
    hoja['A2'].alignment = Alignment(horizontal='center')

    hoja.append(TITULOS)
    for col in range(1, len(COLUMNAS) + 1):
        celda = hoja.cell(row=3, column=col)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = _RELLENO_TITULO
        celda.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
        celda.border = _BORDE

    for fila in filas:
        hoja.append([fila[c] for c in CLAVES])

    primera = 4
    ultima = primera + len(filas) - 1
    for indice in range(len(filas)):
        for col in range(1, len(COLUMNAS) + 1):
            celda = hoja.cell(row=primera + indice, column=col)
            celda.border = _BORDE
            celda.alignment = Alignment(
                vertical='top',
                wrap_text=col - 1 in (CLAVES.index('otb'),
                                      CLAVES.index('nombre')),
                horizontal='right' if col - 1 == COL_MONTO else 'left')
            # Rayado cebra: con nueve columnas, seguir un renglón sin la guía
            # del fondo alternado es donde se cruzan los datos al leer.
            if indice % 2:
                celda.fill = _RELLENO_CEBRA
            if col - 1 == COL_MONTO:
                celda.number_format = '#,##0.00'

    if filas:
        total = hoja.max_row + 1
        # openpyxl numera las columnas desde 1 y `COL_MONTO` es un índice desde
        # 0: la columna del monto es `COL_MONTO + 1` y el rótulo va en la de al
        # lado, que es justamente `COL_MONTO`.
        rotulo = hoja.cell(row=total, column=COL_MONTO, value='TOTAL')
        rotulo.font = Font(bold=True, color=VERDE_800)
        rotulo.alignment = Alignment(horizontal='right')
        celda = hoja.cell(row=total, column=COL_MONTO + 1,
                          value=total_montos(filas))
        celda.font = Font(bold=True, color=VERDE_800)
        celda.number_format = '#,##0.00'
        celda.alignment = Alignment(horizontal='right')
        for col in range(1, len(COLUMNAS) + 1):
            hoja.cell(row=total, column=col).fill = _RELLENO_TOTAL
        # El autofiltro se declara sobre los datos y NO sobre la fila de total:
        # incluida, al filtrar cualquier columna el total se esconde con ella.
        hoja.auto_filter.ref = (
            f'A3:{get_column_letter(len(COLUMNAS))}{ultima}')

    for indice, (_, _, _, ancho) in enumerate(COLUMNAS, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.freeze_panes = 'A4'

    flujo = io.BytesIO()
    libro.save(flujo)
    return flujo.getvalue()


# --- PDF -------------------------------------------------------------------

_CELDA = ParagraphStyle('celda', fontName='Helvetica', fontSize=6.5,
                        leading=8)
_CELDA_TITULO = ParagraphStyle('celdaTitulo', fontName='Helvetica-Bold',
                               fontSize=6.5, leading=8,
                               textColor=colors.white)
_CELDA_NUM = ParagraphStyle('celdaNum', parent=_CELDA, alignment=TA_RIGHT)
_CELDA_TOTAL = ParagraphStyle('celdaTotal', fontName='Helvetica-Bold',
                              fontSize=7, leading=9, alignment=TA_RIGHT,
                              textColor=colors.HexColor(f'#{VERDE_800}'))
_TITULO = ParagraphStyle('titulo', fontName='Helvetica-Bold', fontSize=13,
                         leading=16, alignment=1,
                         textColor=colors.HexColor(f'#{VERDE_800}'))
_SUBTITULO = ParagraphStyle('subtitulo', fontName='Helvetica', fontSize=8.5,
                            leading=11, alignment=1,
                            textColor=colors.HexColor(f'#{VERDE_700}'))


def _monto(valor):
    return f'{valor:,.2f}'.replace(',', '@').replace('.', ',').replace('@', '.')


def generar_reporte_pdf(filas, subtitulo=''):
    """El reporte en PDF carta apaisado, armado en el servidor.

    Mismo criterio que el acta oficial: la medida la fija el servidor y no el
    diálogo de impresión del navegador.
    """
    flujo = io.BytesIO()
    documento = SimpleDocTemplate(
        flujo, pagesize=CARTA_APAISADA,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title='Proyectos programados')

    cuerpo = [[Paragraph(t, _CELDA_TITULO) for t in TITULOS]]
    for fila in filas:
        cuerpo.append([
            Paragraph(str(fila['gestion']), _CELDA),
            Paragraph(fila['distrito'], _CELDA),
            Paragraph(fila['otb'], _CELDA),
            Paragraph(fila['categoria_programatica'], _CELDA),
            Paragraph(fila['sisin'], _CELDA),
            Paragraph(str(fila['numero_proyecto']), _CELDA),
            Paragraph(fila['nombre'], _CELDA),
            Paragraph(_monto(fila['monto']), _CELDA_NUM),
            Paragraph(fila['responsable_registro'], _CELDA),
        ])

    estilo = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{VERDE_700}')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor(f'#{VERDE_BORDE}')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor(f'#{VERDE_CEBRA}')]),
    ]

    if filas:
        # El rótulo va en la PRIMERA celda del bloque combinado: ReportLab
        # toma el contenido de esa y descarta el de las que absorbe.
        cuerpo.append(
            [Paragraph('TOTAL', _CELDA_TOTAL)]
            + [Paragraph('', _CELDA_TOTAL)] * (COL_MONTO - 1)
            + [Paragraph(_monto(total_montos(filas)), _CELDA_TOTAL),
               Paragraph('', _CELDA_TOTAL)])
        estilo += [
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{VERDE_100}')),
            # TOTAL ocupa todo lo que va antes del monto: suelto en su celda
            # queda separado del número que rotula por el ancho del nombre.
            ('SPAN', (0, -1), (COL_MONTO - 1, -1)),
        ]

    tabla = Table(cuerpo, colWidths=[c[2] * mm for c in COLUMNAS],
                  repeatRows=1)
    tabla.setStyle(TableStyle(estilo))

    documento.build([
        Paragraph('PROYECTOS PROGRAMADOS', _TITULO),
        Paragraph(subtitulo, _SUBTITULO),
        Spacer(1, 5 * mm),
        tabla,
    ])
    return flujo.getvalue()
