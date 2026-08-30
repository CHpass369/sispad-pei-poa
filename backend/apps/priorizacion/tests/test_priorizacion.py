"""Módulo Priorización POA: buscador, acta oficial y circuito de revisión."""
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from apps.accounts.models import Rol
from apps.gestion.models import GestionFiscal
from apps.priorizacion.models import (
    ActaPriorizacion, PlantillaActa, ProyectoCatalogo, ProyectoPriorizado,
    normalizar,
)
from apps.priorizacion.views import MESES, anio_en_letras
from apps.territorio.models import Distrito

User = get_user_model()
API = '/api/v1/priorizacion'

PLANTILLA = {
    'titulo': 'ACTA DE PRIORIZACIÓN DE PROYECTOS Y ACTIVIDADES',
    'subtitulo': 'POA {gestion}',
    'encabezado': ('El Sr. {presidente} presidente de la {otb} del {distrito}, '
                   'en fecha {dia} de {mes} del año {anio_letras}, realizo la '
                   'priorización del proyecto para el POA {gestion}, mismo se '
                   'detalla a continuación:'),
    'nota': 'Nota:  Se aclara que, una vez priorizado el proyecto, no se podrá '
            'realizar ninguna modificación ni cambio de proyecto.',
    'cierre': 'En constancia de conformidad firman al pie del presente '
              'documento los siguientes:',
    'firmas': [{'rol': 'Presidente de la OTB', 'campo': 'presidente'},
               {'rol': 'Responsable del registro', 'campo': 'responsable'}],
}


class NormalizacionTests(TestCase):
    def test_ignora_tildes_y_puntuacion(self):
        # "ADQ." y "ADQ" tienen que encontrar lo mismo.
        self.assertEqual(normalizar('ADQ. LUMINARIAS, BRAZOS'),
                         'ADQ LUMINARIAS BRAZOS')
        self.assertEqual(normalizar('Construcción'), 'CONSTRUCCION')
        self.assertEqual(normalizar('  doble   espacio '), 'DOBLE ESPACIO')
        self.assertEqual(normalizar(None), '')


class AnioEnLetrasTests(TestCase):
    def test_escribe_el_anio_como_el_acta(self):
        self.assertEqual(anio_en_letras(2025), 'dos mil veinticinco')
        self.assertEqual(anio_en_letras(2026), 'dos mil veintiséis')
        self.assertEqual(anio_en_letras(2031), 'dos mil treinta y uno')
        self.assertEqual(anio_en_letras(2040), 'dos mil cuarenta')
        self.assertEqual(anio_en_letras(2042), 'dos mil cuarenta y dos')
        self.assertEqual(anio_en_letras(2000), 'dos mil')


class BuscadorProyectosTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(
            user=User.objects.create_user(email='t@test.com', password='x12345678'))
        for nombre, sisin, cat, veces in [
            ('ADQ. LUMINARIAS, BRAZOS Y ACCESORIOS DISTRITO 4', '', '', 6),
            ('ADQ. LUMINARIAS, BRAZOS Y ACCESORIOS DISTRITO 1', '', '', 8),
            ('PAVIMENTO FLEXIBLE DISTRITO 4', '', '', 2),
            ('CONST. SISTEMA DE MICRORIEGO', '02874735200000',
             '100 02874735200000 000', 0),
        ]:
            ProyectoCatalogo.objects.create(
                nombre=nombre, sisin=sisin, categoria_programatica=cat,
                veces_priorizado=veces)

    def buscar(self, q, extra=''):
        return self.client.get(f'{API}/catalogo-proyectos/?q={q}{extra}').json()

    def test_cada_palabra_acota_mas(self):
        self.assertEqual(self.buscar('lumin')['total'], 2)
        self.assertEqual(self.buscar('lumin distrito 4')['total'], 1)
        self.assertEqual(self.buscar('distrito 4')['total'], 2)

    def test_no_importan_tildes_ni_puntuacion_ni_mayusculas(self):
        self.assertEqual(self.buscar('adq.')['total'], 2)
        self.assertEqual(self.buscar('ADQ')['total'], 2)

    def test_ordena_por_lo_mas_priorizado(self):
        primero = self.buscar('lumin')['resultados'][0]
        self.assertEqual(primero['veces_priorizado'], 8)

    def test_filtra_los_que_tienen_sisin(self):
        d = self.buscar('const', '&con_sisin=1')
        self.assertEqual(d['total'], 1)
        self.assertEqual(d['resultados'][0]['categoria_programatica'],
                         '100 02874735200000 000')

    def test_una_palabra_que_no_esta_no_devuelve_nada(self):
        self.assertEqual(self.buscar('helipuerto')['total'], 0)


class ActaTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.tecnico = User.objects.create_user(email='t@test.com', password='x12345678')
        self.jefatura = User.objects.create_user(email='j@test.com', password='x12345678')
        rol, _ = Rol.objects.get_or_create(codigo='jefe_poa',
                                           defaults={'nombre': 'Jefatura POA'})
        self.jefatura.roles.add(rol)
        self.client.force_authenticate(user=self.tecnico)

        # Sin gestión habilitada SIS-POA no opera: el candado (ADR-007) es
        # la precondición de todo el circuito de priorización.
        GestionFiscal.objects.update(activa=False)
        self.gestion = GestionFiscal.objects.update_or_create(
            anio=2027, defaults={'estado': 'HABILITADA', 'activa': True},
        )[0]

        self.distrito = Distrito.objects.create(codigo='D2', nombre='DISTRITO 2')
        PlantillaActa.objects.create(nombre='Acta', **PLANTILLA)

    def crear_acta(self, **extra):
        # `fecha` no viaja en el cuerpo: es de solo lectura y la pone el
        # servidor al registrar. Mandarla no hace nada.
        datos = {
            'gestion': 2027, 'distrito': str(self.distrito.id),
            'otb': 'OTB SAN JOSE DE KORIPILA',
            'presidente': 'LIZETTE SHIRLEY CUBA ALDUNATE',
            'responsable_registro': 'LILIANA AYALA',
            'proyectos': [
                {'nombre': 'CONST. PAVIMENTO ZONA SUDOESTE', 'monto': '220000',
                 'sisin': '', 'categoria_programatica': '170 0 001'},
                {'nombre': 'ADQ. LUMINARIAS', 'monto': '10000',
                 'sisin': '', 'categoria_programatica': ''},
            ],
        }
        datos.update(extra)
        return self.client.post(f'{API}/actas/', datos, format='json')

    def acta_sin_fecha(self):
        """Un acta sin fecha, que por API ya no se puede crear.

        Las hay igual: son las 18 que se importaron de planillas donde la
        fecha venía en blanco. `esta_completa` las sigue frenando.
        """
        acta_id = self.crear_acta().json()['id']
        ActaPriorizacion.objects.filter(pk=acta_id).update(fecha=None)
        return acta_id

    def fecha_del_acta_en_letras(self, acta_id):
        """El encabezado, escrito con la fecha que el acta tiene guardada."""
        fecha = ActaPriorizacion.objects.get(pk=acta_id).fecha
        return (f'en fecha {fecha.day:02d} de {MESES[fecha.month - 1]} '
                f'del año {anio_en_letras(fecha.year)}')

    # --- Registro ----------------------------------------------------------

    def test_la_gestion_arranca_sin_actas(self):
        self.assertEqual(
            self.client.get(f'{API}/actas/?gestion=2027').json()['count'], 0)

    def test_registra_el_acta_con_sus_proyectos_numerados(self):
        r = self.crear_acta()
        self.assertEqual(r.status_code, status.HTTP_201_CREATED)
        acta = ActaPriorizacion.objects.get()
        self.assertEqual([p.orden for p in acta.proyectos.all()], [1, 2])
        self.assertEqual(float(acta.monto_total), 230000.0)

    def test_una_otb_no_prioriza_dos_veces_en_la_misma_gestion(self):
        self.crear_acta()
        self.assertEqual(self.crear_acta().status_code,
                         status.HTTP_400_BAD_REQUEST)

    def test_no_se_prioriza_fuera_de_la_gestion_habilitada(self):
        # Antes la misma OTB podía priorizar en otra gestión. Con el candado
        # duro no: SIS-POA opera sobre la gestión habilitada y sobre ninguna
        # otra, y el rechazo llega con su propio código (ADR-007).
        self.crear_acta()
        respuesta = self.crear_acta(gestion=2028)
        self.assertEqual(respuesta.status_code, status.HTTP_409_CONFLICT)
        self.assertEqual(respuesta.json()['error']['code'],
                         'fuera_de_gestion_habilitada')

    def test_editar_reemplaza_la_lista_completa_de_proyectos(self):
        acta_id = self.crear_acta().json()['id']
        r = self.client.put(f'{API}/actas/{acta_id}/', {
            'gestion': 2027, 'distrito': str(self.distrito.id),
            'otb': 'OTB SAN JOSE DE KORIPILA', 'presidente': 'OTRO',
            'responsable_registro': '',
            'proyectos': [{'nombre': 'UNICO', 'monto': '500', 'sisin': '',
                           'categoria_programatica': ''}],
        }, format='json')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(ProyectoPriorizado.objects.count(), 1)

    def test_la_fecha_la_pone_el_servidor_y_no_el_cliente(self):
        # El formulario ni siquiera la pide: la muestra ya asignada. Por eso el
        # acta oficial imprime el día del registro y no lo que mande el cuerpo.
        r = self.crear_acta(fecha='2020-01-01')
        self.assertEqual(r.json()['fecha'], str(timezone.localdate()))

    # --- Orden del listado -------------------------------------------------

    def _dos_actas_en_distritos_distintos(self):
        """La primera va al distrito que el orden viejo ponía adelante.

        Así el orden por código de distrito y el orden por registro dan
        resultados opuestos, y el test distingue cuál de los dos se aplicó.
        """
        primero = Distrito.objects.create(codigo='D1', nombre='DISTRITO 1')
        self.crear_acta(distrito=str(primero.id), otb='OTB ANTIGUA')
        self.crear_acta(otb='OTB RECIENTE', proyectos=[
            {'nombre': 'ADQ. CONTENEDORES', 'monto': '900000', 'sisin': '',
             'categoria_programatica': ''},
        ])

    def _listar(self, consulta=''):
        return [a['otb'] for a in
                self.client.get(f'{API}/actas/{consulta}').json()['results']]

    def test_el_listado_arranca_por_lo_ultimo_registrado(self):
        self._dos_actas_en_distritos_distintos()
        self.assertEqual(self._listar(), ['OTB RECIENTE', 'OTB ANTIGUA'])

    def test_el_encabezado_ordena_por_la_columna_pedida(self):
        self._dos_actas_en_distritos_distintos()
        self.assertEqual(self._listar('?ordering=otb'),
                         ['OTB ANTIGUA', 'OTB RECIENTE'])
        self.assertEqual(self._listar('?ordering=-otb'),
                         ['OTB RECIENTE', 'OTB ANTIGUA'])

    def test_ordena_por_las_columnas_calculadas_del_encabezado(self):
        # `proyectos` y `monto Bs` son propiedades del modelo: sin anotarlas
        # el encabezado no podría ordenar por ellas.
        self._dos_actas_en_distritos_distintos()
        self.assertEqual(self._listar('?ordering=-cuenta_proyectos'),
                         ['OTB ANTIGUA', 'OTB RECIENTE'])
        self.assertEqual(self._listar('?ordering=-suma_monto'),
                         ['OTB RECIENTE', 'OTB ANTIGUA'])

    # --- Paginación --------------------------------------------------------

    def _cargar_actas(self, cuantas):
        """`cuantas` actas, cada una de una OTB distinta."""
        for i in range(cuantas):
            self.crear_acta(otb=f'OTB {i:03d}')

    def test_el_listado_viene_paginado_y_dice_de_a_cuanto(self):
        self._cargar_actas(27)
        d = self.client.get(f'{API}/actas/').json()
        self.assertEqual(d['count'], 27)
        self.assertEqual(d['page_size'], 25)
        self.assertEqual(len(d['results']), 25)
        segunda = self.client.get(f'{API}/actas/?page=2').json()
        self.assertEqual(len(segunda['results']), 2)

    def test_el_resumen_cuenta_todas_las_actas_y_no_la_pagina(self):
        # Las tarjetas de la pantalla se arman con esto. Si contaran `results`
        # dirían 25 actas y 50 proyectos habiendo 27 y 54.
        self._cargar_actas(27)
        resumen = self.client.get(f'{API}/actas/').json()['resumen']
        self.assertEqual(resumen['actas'], 27)
        self.assertEqual(resumen['proyectos'], 54)
        self.assertEqual(float(resumen['monto']), 27 * 230000.0)

    def test_el_resumen_respeta_el_filtro_de_busqueda(self):
        self._cargar_actas(3)
        resumen = self.client.get(f'{API}/actas/?q=OTB 001').json()['resumen']
        self.assertEqual(resumen['actas'], 1)
        self.assertEqual(resumen['proyectos'], 2)

    def test_sin_actas_el_resumen_va_en_cero_y_no_en_nulo(self):
        d = self.client.get(f'{API}/actas/').json()
        self.assertEqual(d['resumen'],
                         {'actas': 0, 'proyectos': 0, 'monto': 0})

    def test_la_cuenta_de_proyectos_no_se_infla_al_sumar_montos(self):
        # Contar la relación y sumar sobre ella en el mismo `annotate` es donde
        # Django infla la cuenta por el join.
        self.crear_acta()
        acta = self.client.get(f'{API}/actas/').json()['results'][0]
        self.assertEqual(len(acta['proyectos']), 2)
        self.assertEqual(float(acta['monto_total']), 230000.0)

    # --- Acta oficial ------------------------------------------------------

    def test_el_acta_reproduce_el_formato_oficial(self):
        acta_id = self.crear_acta().json()['id']
        d = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/').json()
        self.assertEqual(d['titulo'],
                         'ACTA DE PRIORIZACIÓN DE PROYECTOS Y ACTIVIDADES')
        self.assertEqual(d['subtitulo'], 'POA 2027')
        self.assertEqual(d['distrito'], 'DISTRITO 2')
        self.assertIn('El Sr. LIZETTE SHIRLEY CUBA ALDUNATE presidente de la '
                      'OTB SAN JOSE DE KORIPILA del DISTRITO 2', d['encabezado'])
        self.assertIn(self.fecha_del_acta_en_letras(acta_id), d['encabezado'])
        self.assertEqual(d['total'], 230000.0)
        # Firma el presidente y nadie más: el responsable del registro es un
        # dato interno y no se imprime, aunque la plantilla declare su casilla.
        self.assertEqual([f['nombre'] for f in d['firmas']],
                         ['LIZETTE SHIRLEY CUBA ALDUNATE'])
        self.assertNotIn('LILIANA AYALA', str(d['firmas']))

    def test_no_antepone_otb_al_nombre_de_la_organizacion(self):
        # Hay juntas vecinales y sindicatos que no son OTB.
        acta_id = self.crear_acta(otb='J.V. DIN LA PAZ').json()['id']
        d = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/').json()
        self.assertIn('presidente de la J.V. DIN LA PAZ', d['encabezado'])
        self.assertNotIn('de la OTB J.V.', d['encabezado'])

    def test_sin_fecha_no_se_emite(self):
        acta_id = self.acta_sin_fecha()
        r = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)

    def test_sin_proyectos_no_se_emite(self):
        acta_id = self.crear_acta(proyectos=[]).json()['id']
        r = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)

    def test_sin_plantilla_lo_dice_en_vez_de_reventar(self):
        acta_id = self.crear_acta().json()['id']
        PlantillaActa.objects.all().delete()
        r = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('plantilla', r.json()['error'])

    def test_una_plantilla_de_la_gestion_le_gana_a_la_general(self):
        PlantillaActa.objects.create(nombre='2027', gestion=2027,
                                     **{**PLANTILLA, 'titulo': 'ACTA 2027'})
        acta_id = self.crear_acta().json()['id']
        d = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/').json()
        self.assertEqual(d['titulo'], 'ACTA 2027')

    def test_una_llave_suelta_en_la_plantilla_no_rompe_la_emision(self):
        # El texto lo escribe un usuario: format() explotaría con un KeyError.
        PlantillaActa.objects.all().update(nota='Presupuesto {no_existe} y {')
        acta_id = self.crear_acta().json()['id']
        r = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)

    # --- Circuito ----------------------------------------------------------

    def test_validar_exige_que_el_acta_este_completa(self):
        acta_id = self.acta_sin_fecha()
        r = self.client.post(f'{API}/actas/{acta_id}/validar/')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)

    def test_aprobar_es_de_la_jefatura_y_solo_sobre_lo_validado(self):
        acta_id = self.crear_acta().json()['id']
        self.client.force_authenticate(user=self.jefatura)
        self.assertEqual(
            self.client.post(f'{API}/actas/{acta_id}/aprobar/').status_code,
            status.HTTP_400_BAD_REQUEST)

        self.client.force_authenticate(user=self.tecnico)
        self.client.post(f'{API}/actas/{acta_id}/validar/')
        self.assertEqual(
            self.client.post(f'{API}/actas/{acta_id}/aprobar/').status_code,
            status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.jefatura)
        r = self.client.post(f'{API}/actas/{acta_id}/aprobar/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r.json()['estado'], 'APROBADO')

    def test_observar_exige_motivo_y_lo_guarda(self):
        acta_id = self.crear_acta().json()['id']
        self.client.force_authenticate(user=self.jefatura)
        self.assertEqual(
            self.client.post(f'{API}/actas/{acta_id}/observar/',
                             {'comentario': '  '}).status_code,
            status.HTTP_400_BAD_REQUEST)
        r = self.client.post(f'{API}/actas/{acta_id}/observar/',
                             {'comentario': 'Falta el monto del proyecto 2'})
        self.assertEqual(r.json()['observacion'], 'Falta el monto del proyecto 2')

    def test_un_acta_aprobada_no_se_elimina(self):
        acta_id = self.crear_acta().json()['id']
        self.client.post(f'{API}/actas/{acta_id}/validar/')
        self.client.force_authenticate(user=self.jefatura)
        self.client.post(f'{API}/actas/{acta_id}/aprobar/')
        self.client.force_authenticate(user=self.tecnico)
        r = self.client.delete(f'{API}/actas/{acta_id}/')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(ActaPriorizacion.objects.filter(id=acta_id).exists())

    # --- Matrices ----------------------------------------------------------

    def test_la_matriz_consolida_por_distrito(self):
        self.crear_acta()
        d = self.client.get(f'{API}/matrices/?gestion=2027').json()
        self.assertEqual(d['total_filas'], 2)
        self.assertEqual(d['total_monto'], 230000.0)
        self.assertEqual(len(d['resumen']), 1)
        self.assertEqual(d['resumen'][0]['actas'], 1)
        # El conteo no puede inflarse por el join con proyectos.
        self.assertEqual(d['resumen'][0]['proyectos'], 2)

    def test_la_matriz_sin_actas_viene_vacia(self):
        d = self.client.get(f'{API}/matrices/').json()
        self.assertEqual(d['total_filas'], 0)
        self.assertEqual(d['resumen'], [])

    def test_la_matriz_de_otra_gestion_se_rechaza(self):
        respuesta = self.client.get(f'{API}/matrices/?gestion=2029')
        self.assertEqual(respuesta.status_code, status.HTTP_409_CONFLICT)


class ActaPDFTests(ActaTests):
    """El PDF lo arma el servidor: la medida no puede quedar a criterio del
    diálogo de impresión del navegador."""

    def test_sale_en_carta_exacta(self):
        # El acta se emitía en oficio (216 x 330) hasta que `0f3ac99` la pasó a
        # carta. La medida sigue clavada en el servidor: lo que no puede pasar
        # es que la decida el diálogo de impresión del navegador.
        import re
        acta_id = self.crear_acta().json()['id']
        r = self.client.get(f'{API}/actas/{acta_id}/pdf/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        caja = re.search(rb'/MediaBox \[([^\]]+)\]', r.content)
        ancho, alto = [float(v) for v in caja.group(1).split()[2:]]
        # 216 x 279 mm en puntos, con tolerancia de redondeo.
        self.assertAlmostEqual(ancho, 216 * 72 / 25.4, places=1)
        self.assertAlmostEqual(alto, 279 * 72 / 25.4, places=1)

    def test_se_descarga_con_nombre_propio(self):
        acta_id = self.crear_acta().json()['id']
        r = self.client.get(f'{API}/actas/{acta_id}/pdf/')
        self.assertIn('attachment', r['Content-Disposition'])
        self.assertIn('.pdf', r['Content-Disposition'])

    def test_la_huella_del_contenido_es_estable_y_viaja_en_la_cabecera(self):
        acta_id = self.crear_acta().json()['id']
        primera = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        segunda = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        self.assertEqual(primera, segunda)
        self.assertEqual(len(primera), 64)
        # El JSON del acta declara la misma huella que el PDF.
        d = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/').json()
        self.assertEqual(d['huella'], primera)

    def test_cambiar_un_monto_cambia_la_huella(self):
        acta_id = self.crear_acta().json()['id']
        antes = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        p = ProyectoPriorizado.objects.filter(acta_id=acta_id).first()
        p.monto = 999
        p.save()
        despues = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        self.assertNotEqual(antes, despues)

    def test_la_huella_no_depende_de_la_redaccion_de_la_plantilla(self):
        # Lo que se verifica es qué se priorizó y por cuánto, no cómo se
        # redactó el acta: cambiar la plantilla no invalida lo firmado.
        acta_id = self.crear_acta().json()['id']
        antes = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        PlantillaActa.objects.all().update(nota='Otra nota distinta')
        despues = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        self.assertEqual(antes, despues)

    def test_el_pdf_incluye_la_aclaracion_de_recursos(self):
        PlantillaActa.objects.all().update(
            aclaracion='Aclarar que las transferencias del TGN del POA {gestion}')
        acta_id = self.crear_acta().json()['id']
        d = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/').json()
        self.assertIn('POA 2027', d['aclaracion'])
        self.assertEqual(
            self.client.get(f'{API}/actas/{acta_id}/pdf/').status_code,
            status.HTTP_200_OK)

    def test_sin_fecha_no_hay_pdf(self):
        acta_id = self.acta_sin_fecha()
        r = self.client.get(f'{API}/actas/{acta_id}/pdf/')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)



class ReporteProyectosTests(ActaTests):
    """El reporte de proyectos programados, por filtros y en los dos formatos."""

    XLSX = ('application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet')

    def filas_del_xlsx(self, contenido):
        from openpyxl import load_workbook
        hoja = load_workbook(io.BytesIO(contenido)).active
        # Fila 1 título, 2 subtítulo, 3 encabezados: los datos arrancan en 4.
        return [[c.value for c in fila] for fila in hoja.iter_rows(min_row=4)]

    def test_emite_una_fila_por_proyecto_y_no_por_acta(self):
        self.crear_acta()
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r['Content-Type'], self.XLSX)
        # Un acta con dos proyectos son dos filas, más la del total.
        self.assertEqual(r['X-Reporte-Filas'], '2')
        filas = self.filas_del_xlsx(r.content)
        self.assertEqual(len(filas), 3)

    def test_las_columnas_son_las_nueve_pedidas_y_en_orden(self):
        self.crear_acta()
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        from openpyxl import load_workbook
        hoja = load_workbook(io.BytesIO(r.content)).active
        self.assertEqual(
            [c.value for c in hoja[3]],
            ['GESTIÓN POA', 'DISTRITO', 'OTB / JUNTA VECINAL',
             'CATEGORÍA PROGRAMÁTICA', 'SISIN', 'N° PROYECTO',
             'NOMBRE DEL PROYECTO', 'MONTO PROYECTO',
             'RESPONSABLE DEL REGISTRO'])

    def test_cada_fila_trae_los_datos_del_acta_y_del_proyecto(self):
        self.crear_acta()
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        primera = self.filas_del_xlsx(r.content)[0]
        # openpyxl guarda la celda vacía como ausente: el SISIN en blanco
        # vuelve a leerse como None, no como la cadena que se escribió.
        self.assertEqual(primera[:8], [
            2027, 'DISTRITO 2', 'OTB SAN JOSE DE KORIPILA', '170 0 001', None,
            1, 'CONST. PAVIMENTO ZONA SUDOESTE', 220000])
        self.assertEqual(primera[8], 'LILIANA AYALA')

    def test_respeta_el_filtro_de_distrito_aplicado(self):
        self.crear_acta()
        otro = Distrito.objects.create(codigo='D5', nombre='DISTRITO 5')
        self.crear_acta(distrito=str(otro.id), otb='OTB VILLA OBRAJES')

        completo = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        self.assertEqual(completo['X-Reporte-Filas'], '4')

        filtrado = self.client.get(
            f'{API}/actas/reporte/?formato=xlsx&distrito={otro.id}')
        self.assertEqual(filtrado['X-Reporte-Filas'], '2')
        distritos = {f[1] for f in self.filas_del_xlsx(filtrado.content)[:2]}
        self.assertEqual(distritos, {'DISTRITO 5'})

    def test_respeta_la_busqueda_por_otb(self):
        self.crear_acta()
        self.crear_acta(otb='OTB VILLA OBRAJES')
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx&q=OBRAJES')
        self.assertEqual(r['X-Reporte-Filas'], '2')

    def test_exporta_todo_lo_filtrado_y_no_solo_la_pagina(self):
        """La paginación es de la pantalla; el reporte no la hereda."""
        for n in range(30):
            self.crear_acta(otb=f'OTB NUMERO {n:02d}')
        listado = self.client.get(f'{API}/actas/').json()
        self.assertEqual(len(listado['results']), listado['page_size'])
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        self.assertEqual(r['X-Reporte-Filas'], '60')

    def test_el_pdf_sale_en_carta_apaisada(self):
        self.crear_acta()
        r = self.client.get(f'{API}/actas/reporte/?formato=pdf')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        self.assertTrue(r.content.startswith(b'%PDF'))
        # Carta apaisada en puntos: 279 mm x 216 mm. Que el ancho sea mayor
        # que el alto es justamente lo que se está afirmando.
        self.assertIn(b'/MediaBox [ 0 0 790.8661 612.2835 ]', r.content)

    def test_un_acta_sin_proyectos_no_aporta_filas(self):
        acta_id = self.crear_acta().json()['id']
        ProyectoPriorizado.objects.filter(acta_id=acta_id).delete()
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        self.assertEqual(r['X-Reporte-Filas'], '0')
        self.assertEqual(self.filas_del_xlsx(r.content), [])

    def test_rechaza_un_formato_que_no_existe(self):
        r = self.client.get(f'{API}/actas/reporte/?formato=docx')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)

    def test_el_subtitulo_dice_que_recorte_se_exporto(self):
        self.crear_acta()
        from openpyxl import load_workbook
        r = self.client.get(
            f'{API}/actas/reporte/?formato=xlsx&distrito={self.distrito.id}')
        hoja = load_workbook(io.BytesIO(r.content)).active
        self.assertIn('Gestión POA 2027', hoja['A2'].value)
        self.assertIn('DISTRITO 2', hoja['A2'].value)

    def test_sin_filtros_el_subtitulo_lo_dice(self):
        self.crear_acta()
        from openpyxl import load_workbook
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        hoja = load_workbook(io.BytesIO(r.content)).active
        self.assertIn('todas las actas de la gestión', hoja['A2'].value)

    def test_pide_autenticacion(self):
        self.client.force_authenticate(user=None)
        r = self.client.get(f'{API}/actas/reporte/?formato=xlsx')
        self.assertEqual(r.status_code, status.HTTP_401_UNAUTHORIZED)


class ContenidoQRTests(TestCase):
    """Lo que se lee al escanear el QR del acta."""

    DATOS = {
        'acta_id': 'abc', 'gestion': 2027, 'distrito': 'DISTRITO 2',
        'otb': 'OTB SAN JOSE DE KORIPILA', 'presidente': 'LIZETTE CUBA',
        'fecha': '2026-09-03',
        'proyectos': [{'nro': 1, 'descripcion': 'X', 'monto': 1.0}],
        'total': 1.0,
        'firmas': [{'rol': 'Presidente de la OTB', 'nombre': 'LIZETTE CUBA'},
                   {'rol': 'Responsable del registro', 'nombre': 'LILIANA AYALA'}],
    }

    def contenido(self, datos=None):
        from datetime import datetime

        from apps.priorizacion.pdf import contenido_qr, hash_acta
        datos = datos or self.DATOS
        return contenido_qr(datos, hash_acta(datos), datetime(2026, 8, 19, 18, 42))

    def test_lleva_los_nombres_de_los_firmantes(self):
        texto = self.contenido()
        self.assertIn('Firmantes: LIZETTE CUBA, LILIANA AYALA', texto)

    def test_lleva_la_fecha_y_hora_de_generacion(self):
        self.assertIn('Generada: 19/08/2026 18:42', self.contenido())

    def test_cierra_con_la_entidad_y_la_gestion(self):
        self.assertTrue(self.contenido().rstrip().endswith(
            'Gobierno Autonomo Municipal de Sacaba - POA 2027'))

    def test_incluye_la_huella_del_contenido(self):
        from apps.priorizacion.pdf import hash_acta
        self.assertIn(hash_acta(self.DATOS), self.contenido())

    def test_va_sin_tildes(self):
        # Los lectores de ventanilla devuelven símbolos rotos con UTF-8.
        texto = self.contenido()
        self.assertEqual(texto, texto.encode('ascii', 'ignore').decode())

    def test_un_acta_sin_firmantes_lo_dice_en_vez_de_dejar_el_campo_colgando(self):
        datos = {**self.DATOS, 'firmas': []}
        self.assertIn('Firmantes: sin registrar', self.contenido(datos))

    def test_ignora_los_firmantes_sin_nombre_cargado(self):
        datos = {**self.DATOS, 'firmas': [
            {'rol': 'Presidente', 'nombre': 'JUAN'},
            {'rol': 'Responsable', 'nombre': ''}]}
        self.assertIn('Firmantes: JUAN\n', self.contenido(datos))


class TextoDelPDFTests(ActaTests):
    """El PDF se lee de verdad, no se supone."""

    def texto_pdf(self, acta_id):
        import io

        from pypdf import PdfReader
        contenido = self.client.get(f'{API}/actas/{acta_id}/pdf/').content
        return '\n'.join(p.extract_text() for p in
                         PdfReader(io.BytesIO(contenido)).pages)

    def test_imprime_el_acta_completa(self):
        texto = self.texto_pdf(self.crear_acta().json()['id'])
        self.assertIn('ACTA DE PRIORIZACIÓN DE PROYECTOS Y ACTIVIDADES', texto)
        self.assertIn('DISTRITO 2', texto)
        self.assertIn('CONST. PAVIMENTO ZONA SUDOESTE', texto)
        self.assertIn('LIZETTE SHIRLEY CUBA ALDUNATE', texto)
        self.assertIn('230.000,00', texto)

    def test_la_aclaracion_de_la_plantilla_ya_no_se_imprime(self):
        # El hotfix `6b9cd73` reemplazó ese párrafo por el texto condicional de
        # pavimento. La aclaración sigue viva en la plantilla y en el acta que
        # se ve en pantalla; en el PDF no sale.
        PlantillaActa.objects.all().update(
            aclaracion='Aclarar que las transferencias del TGN del POA {gestion}')
        acta_id = self.crear_acta().json()['id']
        d = self.client.get(f'{API}/actas/{acta_id}/acta-oficial/').json()
        self.assertIn('transferencias del TGN', d['aclaracion'])
        self.assertNotIn('transferencias del TGN', self.texto_pdf(acta_id))

    def test_no_lleva_el_rotulo_de_verificacion(self):
        texto = self.texto_pdf(self.crear_acta().json()['id'])
        self.assertNotIn('SHA-256', texto)
        self.assertNotIn('Verificación del contenido', texto)

    def test_cierra_con_la_entidad_y_la_gestion(self):
        # Con tilde: el pie del PDF se imprime, no se escanea. El que va sin
        # tildes es el contenido del QR, y eso lo cubre `ContenidoQRTests`.
        texto = self.texto_pdf(self.crear_acta().json()['id'])
        self.assertIn('Gobierno Autónomo Municipal de Sacaba', texto)
        self.assertIn('POA 2027', texto)

    def test_la_huella_queda_impresa_bajo_el_qr(self):
        acta_id = self.crear_acta().json()['id']
        huella = self.client.get(f'{API}/actas/{acta_id}/pdf/')['X-Acta-Huella']
        self.assertIn(huella, self.texto_pdf(acta_id))
