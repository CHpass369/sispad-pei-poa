"""Importa el CATÁLOGO NACIONAL MAESTRO PGDES/PDES/PAD desde el XLSX oficial.

Uso:
    python manage.py importar_catalogo_nacional [--archivo RUTA]
        [--gestion 2026] [--dry-run]

El command detecta el marco por las hojas del XLSX:

MARCO 2026-2035 (por defecto) — ``SISPE_Catalogo_Nacional_Maestro_PAD_v0_1.xlsx``
(base upstream SIS-PE, control de calidad con incidencias registradas):
- ``01_FUENTES`` (F01..F09) → ``planificacion.Plan`` + ``VersionCatalogoPlan``
  vigente por plan (PGDESA-2026-2050, PDESA-2026-2030, PDS-PES-2026).
- ``02_PGDESA_EJES`` (7) → ``EjePGDESA`` (código 01..07).
- ``03_PDESA_COMPONENTES`` (38) → ``ComponentePDESA`` (correlativo por eje).
- ``04_PDESA_LINEAMIENTOS`` (170) → ``LineamientoPAD`` (correlativo por
  componente, CGEO municipal 031001 Sacaba, FK componente).
- ``05_PDS_SECTORES`` (24) → ``SectorEconomico`` (clasificador 2026, F03).
- ``08_ODS`` (17) / ``11_NDT_METAS`` (17) / ``12_KMGBF_30x30`` (23) →
  ``articulacion.AcuerdoInternacional`` (tipos ODS / NDT / COMPROMISO_3030).
- ``06_PDS_RESULTADOS`` (0 registros): NO se carga.

MARCO 2021-2025 (compat, vía ``--archivo``) —
``catalogo_nacional_maestro_ptdi_sis_pe_pgdes_pdes_2021_2025.xlsx``:
- ``00_INSTRUMENTOS`` → ``planificacion.Plan`` + ``VersionCatalogoPlan``.
- ``01_PGDES_PILARES`` (13) → ``EjePGDESA`` (código oficial 01..13).
- ``02_PDES_EJES`` (10) → ``ComponentePDESA`` (padre: pilar articulado).
- ``03_PDES_METAS`` (44) / ``04_PDES_RESULTADOS`` (156) /
  ``05_PDES_ACCIONES`` (227) → ``NodoPlanificacion``.
- ``06_PILAR_EJE`` (19 ARTICULA) → ``ArticulacionPlanificacion`` pilar→eje.
- ``08_FUENTES`` (5) → ``normativa.VersionNormativa``.

Idempotente: ``update_or_create`` por clave natural (codigo+versión).
No borra registros demo: los datos oficiales viven en versiones nuevas y el
demo queda intacto en sus propias versiones.
"""
import os
import re
from datetime import date
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.articulacion.models import AcuerdoInternacional
from apps.codificacion.models import (
    ComponentePDESA,
    EjePGDESA,
    EntidadTerritorialCGEO,
    LineamientoPAD,
    SectorEconomico,
    VersionCatalogoPlan,
)
from apps.normativa.models import VersionNormativa
from apps.planificacion.models import (
    ArticulacionPlanificacion,
    NodoPlanificacion,
    Plan,
)

RUTA_XLSX_DEFECTO = os.environ.get(
    'CATALOGO_NACIONAL_XLSX',
    'C:\\Users\\Metatron\\Desktop\\Documentos\\SIS POA\\BASE DE DATOS\\'
    'SISPE_Catalogo_Nacional_Maestro_PAD_v0_1.xlsx',
)

# XLSX del marco 2021-2025 (carga de NodoPlanificacion y fuentes): sigue
# soportado vía --archivo; el default ahora es el marco 2026-2035.
RUTA_XLSX_2021_2025 = os.environ.get(
    'CATALOGO_NACIONAL_XLSX_2021_2025',
    'C:\\Users\\Metatron\\Desktop\\Documentos\\SIS POA\\BASE DE DATOS\\'
    'catalogo_nacional_maestro_ptdi_sis_pe_pgdes_pdes_2021_2025.xlsx',
)

HOJAS = {
    'instrumentos': '00_INSTRUMENTOS',
    'pilares': '01_PGDES_PILARES',
    'ejes': '02_PDES_EJES',
    'metas': '03_PDES_METAS',
    'resultados': '04_PDES_RESULTADOS',
    'acciones': '05_PDES_ACCIONES',
    'pilar_eje': '06_PILAR_EJE',
    'relaciones': '07_RELACIONES',
    'fuentes': '08_FUENTES',
}

HOJAS_MARCO_2026 = {
    'fuentes': '01_FUENTES',
    'ejes': '02_PGDESA_EJES',
    'componentes': '03_PDESA_COMPONENTES',
    'lineamientos': '04_PDESA_LINEAMIENTOS',
    'sectores': '05_PDS_SECTORES',
    'ods': '08_ODS',
    'ndt': '11_NDT_METAS',
    'kmgbf': '12_KMGBF_30x30',
}

# Primera columna de la fila de cabecera en las hojas del XLSX marco 2026
# (las hojas llevan 1-2 filas de título/nota antes de la cabecera real).
CLAVES_CABECERA_MARCO_2026 = {
    'codigo', 'codigo_tecnico', 'fuente_id', 'id_resultado', 'sector_codigo',
}

TIPO_NORMATIVA = {
    'LEY': VersionNormativa.tipo.field.choices[0][0],  # 'ley'
    'PLAN': 'otro',
}

GESTION_PGDES_DEFECTO = 2025
GESTION_PDES_DEFECTO = 2021
GESTION_MARCO_2026 = 2026

# Códigos oficiales de los planes del marco 2026-2035 (convención del seed
# del repo: scripts/seed.py y catalogos/importer/marco_superior.py).
PLANES_MARCO_2026 = (
    {
        'codigo': 'PGDESA-2026-2050',
        'tipo': 'pgdesa',
        'nombre': 'Plan General de Desarrollo Sostenible del Estado 2026-2050',
        'gestion_inicio': 2026,
        'gestion_fin': 2050,
        'fuente_id': 'F01',
    },
    {
        'codigo': 'PDESA-2026-2030',
        'tipo': 'pdesa',
        'nombre': 'Plan de Desarrollo Económico y Social 2026-2030',
        'gestion_inicio': 2026,
        'gestion_fin': 2030,
        'fuente_id': 'F02',
    },
    {
        'codigo': 'PDS-PES-2026',
        'tipo': 'sectorial',
        'nombre': 'Planes de Desarrollo Sectorial / Planes Estratégicos Sectoriales 2026',
        'gestion_inicio': 2026,
        'gestion_fin': 2026,
        'fuente_id': 'F03',
    },
)

CGEO_MUNICIPIO_PAD = '031001'  # Sacaba (municipio único de la plataforma PIP)


def _filas(ws):
    """Convierte una hoja en listas de dicts (cabecera → valor)."""
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    cabeceras = [str(c).strip() if c is not None else '' for c in filas[0]]
    return [
        dict(zip(cabeceras, fila))
        for fila in filas[1:]
        if any(v is not None for v in fila)
    ]


def _filas_marco_2026(ws):
    """Hoja del XLSX marco 2026: detecta la fila de cabecera real.

    Las hojas llevan 1-2 filas de título/nota antes de la cabecera; la
    cabecera se reconoce por la primera columna (clave de catálogo).
    """
    filas = list(ws.iter_rows(values_only=True))
    for i, fila in enumerate(filas):
        if not fila or not any(v is not None for v in fila):
            continue
        if str(fila[0]).strip() in CLAVES_CABECERA_MARCO_2026:
            cabeceras = [str(c).strip() if c is not None else '' for c in fila]
            return [
                dict(zip(cabeceras, fila))
                for fila in filas[i + 1:]
                if any(v is not None for v in fila)
            ]
    return []


def _texto(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def _entero(valor):
    texto = _texto(valor)
    if not texto:
        return None
    try:
        return int(texto)
    except (TypeError, ValueError):
        return None


def _codigo_2_digitos(valor):
    numero = _entero(valor)
    return str(numero).zfill(2) if numero is not None else ''


def _fecha(valor):
    texto = _texto(valor)
    if not texto:
        return None
    coincidencia = re.match(r'^(\d{4})-(\d{2})-(\d{2})', texto)
    if coincidencia:
        return date(*(int(g) for g in coincidencia.groups()))
    return None


class Command(BaseCommand):
    help = (
        'Importa el catálogo nacional maestro PGDES/PDES/PAD desde el XLSX '
        'oficial. Detecta el marco por las hojas: por defecto el marco '
        '2026-2035 (ejes, componentes, lineamientos, sectores y acuerdos '
        'internacionales); con --archivo del XLSX 2021-2025 carga además el '
        'árbol de planificación. Idempotente por codigo+version.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            default=RUTA_XLSX_DEFECTO,
            help='Ruta del XLSX del catálogo nacional maestro.',
        )
        parser.add_argument(
            '--gestion',
            type=int,
            default=None,
            help=(
                'Gestión de las versiones de catálogo. Por defecto: 2026 '
                'para el marco 2026-2035; 2025 para PGDES y 2021 para PDES '
                'en el marco 2021-2025.'
            ),
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Planifica y cuenta sin persistir (rollback al final).',
        )

    def handle(self, *args, **options):
        ruta = Path(options['archivo']).expanduser().resolve()
        if not ruta.exists():
            raise CommandError(f'No existe el XLSX: {ruta}')

        if options['dry_run']:
            with transaction.atomic():
                resumen = self._importar(ruta, options['gestion'])
                transaction.set_rollback(True)
        else:
            resumen = self._importar(ruta, options['gestion'])

        self._volcar_resumen(resumen, dry_run=options['dry_run'])

    # ------------------------------------------------------------------
    # Importación
    # ------------------------------------------------------------------
    def _importar(self, ruta, gestion_override):
        libro = load_workbook(ruta, read_only=True, data_only=True)
        try:
            es_marco_2026 = '02_PGDESA_EJES' in libro.sheetnames
        finally:
            libro.close()
        if es_marco_2026:
            return self._importar_marco_2026(ruta, gestion_override)
        return self._importar_2021_2025(ruta, gestion_override)

    def _importar_2021_2025(self, ruta, gestion_override):
        """Marco 2021-2025: catálogo + árbol de nodos (comportamiento original)."""
        resumen = {'modelos': {}, 'omitidos': [], 'plan': {}, 'gestion': {}}

        libro = load_workbook(ruta, read_only=True, data_only=True)
        try:
            hojas = {
                nombre: _filas(libro[hoja])
                for nombre, hoja in HOJAS.items()
            }
        finally:
            libro.close()

        instrumentos = {
            fila['instrumento_id']: fila for fila in hojas['instrumentos']
        }
        inst_pgdes = next(
            f for f in instrumentos.values() if f['tipo'] == 'PGDES'
        )
        inst_pdes = next(
            f for f in instrumentos.values() if f['tipo'] == 'PDES'
        )

        gestion_pgdes = gestion_override or GESTION_PGDES_DEFECTO
        gestion_pdes = gestion_override or GESTION_PDES_DEFECTO
        resumen['gestion'] = {
            'pgdes': gestion_pgdes,
            'pdes': gestion_pdes,
        }

        plan_pgdes = self._plan_para(inst_pgdes, 'pgdesa')
        plan_pdes = self._plan_para(inst_pdes, 'pdesa')
        version_pgdes = self._version_vigente(plan_pgdes, gestion_pgdes, inst_pgdes)
        version_pdes = self._version_vigente(plan_pdes, gestion_pdes, inst_pdes)
        resumen['plan'] = {
            'pgdes': plan_pgdes.codigo,
            'pdes': plan_pdes.codigo,
        }

        pilares = self._importar_ejes_pgdesa(hojas['pilares'], version_pgdes)
        componentes = self._importar_componentes_pdesa(
            hojas['ejes'], hojas['pilar_eje'], version_pdes, pilares,
        )

        nodos = self._importar_nodos_planificacion(
            hojas, plan_pgdes, plan_pdes,
            gestion_pgdes, gestion_pdes,
        )

        self._importar_articulaciones(
            hojas['pilar_eje'], nodos, gestion_pdes,
        )

        self._importar_fuentes(hojas['fuentes'])

        self._conteos(resumen)
        resumen['omitidos'] = (
            self._decisiones_omision(hojas, componentes)
        )
        return resumen

    def _plan_para(self, instrumento, tipo_plan, procedencia='XLSX oficial 2021-2025'):
        """Reutiliza el Plan por (codigo, tipo) o lo crea para el instrumento."""
        codigo = instrumento['codigo']
        gestion_inicio = _entero(instrumento['gestion_inicio'])
        gestion_fin = _entero(instrumento['gestion_fin']) or gestion_inicio
        plan, _ = Plan.objects.get_or_create(
            codigo=codigo,
            tipo=tipo_plan,
            defaults={
                'nombre': instrumento['nombre'][:500],
                'gestion_inicio': gestion_inicio,
                'gestion_fin': gestion_fin,
                'fecha_vigencia_desde': date(gestion_inicio, 1, 1),
                'fecha_vigencia_hasta': date(gestion_fin, 12, 31),
                'descripcion': (
                    'Importado del catálogo nacional maestro SIS-PE '
                    f'({procedencia}).'
                ),
                'activo': True,
            },
        )
        return plan

    def _version_vigente(self, plan, gestion, instrumento, procedencia='XLSX oficial 2021-2025'):
        version, creada = VersionCatalogoPlan.objects.get_or_create(
            plan=plan,
            gestion=gestion,
            defaults={
                'estado': VersionCatalogoPlan.ESTADO_VIGENTE,
                'norma_aprobacion': _texto(
                    instrumento.get('norma_aprobacion')
                ),
                'clasificacion_fuente': VersionCatalogoPlan.FUENTE_OFICIAL,
                'procedencia_fuente': (
                    'Importado del catálogo nacional maestro SIS-PE '
                    f'({procedencia}).'
                ),
            },
        )
        if not creada and version.estado != VersionCatalogoPlan.ESTADO_VIGENTE:
            hay_otra_vigente = VersionCatalogoPlan.objects.filter(
                plan=plan,
                estado=VersionCatalogoPlan.ESTADO_VIGENTE,
            ).exclude(pk=version.pk).exists()
            if not hay_otra_vigente:
                version.estado = VersionCatalogoPlan.ESTADO_VIGENTE
                version.save(update_fields=['estado', 'updated_at'])
        return version

    def _importar_ejes_pgdesa(self, filas, version):
        """01_PGDES_PILARES → EjePGDESA (código oficial 01..13)."""
        por_sistema = {}
        for fila in filas:
            codigo = _codigo_2_digitos(fila['codigo_oficial'])
            if not codigo:
                continue
            obj, _ = EjePGDESA.objects.update_or_create(
                codigo=codigo,
                version_catalogo=version,
                defaults={
                    'denominacion': _texto(fila['denominacion'])[:500],
                    'activo': True,
                },
            )
            por_sistema[fila['codigo_sistema']] = obj
        return por_sistema

    def _importar_componentes_pdesa(
        self, filas, pilar_eje, version, pilares_por_sistema,
    ):
        """02_PDES_EJES → ComponentePDESA, padre = pilar articulado.

        La FK ``eje`` es obligatoria; el pilar padre se resuelve con la
        primera relación ARTICULA de la hoja 06_PILAR_EJE (pilar→eje).
        """
        primer_pilar = {}
        for fila in sorted(
            pilar_eje, key=lambda f: _entero(f['relacion_id']) or 0,
        ):
            primer_pilar.setdefault(fila['eje_codigo'], fila['pilar_codigo'])

        componentes = {}
        for fila in filas:
            codigo = _codigo_2_digitos(fila['codigo_oficial'])
            if not codigo:
                continue
            eje = pilares_por_sistema.get(
                primer_pilar.get(fila['codigo_sistema'])
            )
            if eje is None:
                continue
            obj, _ = ComponentePDESA.objects.update_or_create(
                eje=eje,
                codigo=codigo,
                version_catalogo=version,
                defaults={
                    'denominacion': _texto(fila['denominacion'])[:500],
                    'activo': True,
                },
            )
            componentes[fila['codigo_sistema']] = obj
        return componentes

    def _importar_nodos_planificacion(
        self, hojas, plan_pgdes, plan_pdes,
        gestion_pgdes, gestion_pdes,
    ):
        """Pilares, ejes, metas, resultados y acciones → NodoPlanificacion.

        Jerarquía padre por código oficial: eje (E01) → meta (1.1) →
        resultado (1.1.1) → acción (1.1.1.1). Las acciones del marco
        nacional usan el nivel ``accion_nacional``.
        """
        nodos = {'pilar': {}, 'eje': {}, 'meta': {}, 'resultado': {}, 'accion': {}}

        def _nodo(plan, nivel, codigo, nombre, gestion, orden=0, padre=None):
            nodo, _ = NodoPlanificacion.objects.update_or_create(
                plan=plan,
                codigo=codigo,
                nivel=nivel,
                defaults={
                    'nombre': nombre,
                    'gestion': gestion,
                    'orden': orden,
                    'padre': padre,
                    'activo': True,
                },
            )
            return nodo

        for fila in hojas['pilares']:
            nodos['pilar'][fila['codigo_sistema']] = _nodo(
                plan_pgdes, 'pilar', _texto(fila['codigo_oficial']),
                _texto(fila['denominacion']), gestion_pgdes,
                orden=_entero(fila['numero']) or 0,
            )
        for fila in hojas['ejes']:
            nodos['eje'][fila['codigo_sistema']] = _nodo(
                plan_pdes, 'eje', _texto(fila['codigo_oficial']),
                _texto(fila['denominacion']), gestion_pdes,
                orden=_entero(fila['numero']) or 0,
            )
        for fila in hojas['metas']:
            eje = nodos['eje'].get(_codigo_sistema_eje(fila['eje_id'], hojas))
            nodos['meta'][fila['codigo_sistema']] = _nodo(
                plan_pdes, 'meta', _texto(fila['codigo_oficial']),
                _texto(fila['denominacion']), gestion_pdes,
                orden=_entero(fila['orden']) or 0,
                padre=eje,
            )
        for fila in hojas['resultados']:
            meta = nodos['meta'].get(_codigo_sistema_meta(fila['meta_id'], hojas))
            nodos['resultado'][fila['codigo_sistema']] = _nodo(
                plan_pdes, 'resultado', _texto(fila['codigo_oficial']),
                _texto(fila['denominacion']), gestion_pdes,
                orden=_entero(fila['orden']) or 0,
                padre=meta,
            )
        for fila in hojas['acciones']:
            resultado = nodos['resultado'].get(
                _codigo_sistema_resultado(fila['resultado_id'], hojas)
            )
            nodos['accion'][fila['codigo_sistema']] = _nodo(
                plan_pdes, 'accion_nacional', _texto(fila['codigo_oficial']),
                _texto(fila['denominacion']), gestion_pdes,
                orden=_entero(fila['orden']) or 0,
                padre=resultado,
            )
        return nodos

    def _importar_articulaciones(self, pilar_eje, nodos, gestion):
        """06_PILAR_EJE (ARTICULA) → ArticulacionPlanificacion pilar→eje."""
        for fila in pilar_eje:
            origen = nodos['pilar'].get(fila['pilar_codigo'])
            destino = nodos['eje'].get(fila['eje_codigo'])
            if origen is None or destino is None:
                continue
            ArticulacionPlanificacion.objects.update_or_create(
                nodo_origen=origen,
                nodo_destino=destino,
                gestion=gestion,
                defaults={'es_principal': True},
            )

    def _importar_fuentes(self, filas):
        """08_FUENTES → normativa.VersionNormativa."""
        for fila in filas:
            titulo = _texto(fila['titulo'])
            if not titulo:
                continue
            coincidencia = re.search(
                r'Ley\s*N[º°]?\s*(\d+)', titulo, re.IGNORECASE,
            )
            numero = coincidencia.group(1) if coincidencia else ''
            fecha = _fecha(fila['fecha'])
            gestion = _entero(_texto(fila['fecha'])[:4]) or (
                fecha.year if fecha else GESTION_PDES_DEFECTO
            )
            VersionNormativa.objects.get_or_create(
                titulo=titulo,
                defaults={
                    'tipo': TIPO_NORMATIVA.get(_texto(fila['tipo']), 'otro'),
                    'numero': numero,
                    'fecha_emision': fecha,
                    'gestion': gestion,
                    'resumen': _texto(fila['uso']),
                    'activo': True,
                },
            )

    # ------------------------------------------------------------------
    # Marco 2026-2035 (catálogos PGDESA/PDESA/PDS + acuerdos)
    # ------------------------------------------------------------------
    def _importar_marco_2026(self, ruta, gestion_override):
        resumen = {'modelos': {}, 'omitidos': [], 'plan': {}, 'gestion': {},
                   'lotes': {}}

        gestion = gestion_override or GESTION_MARCO_2026
        resumen['gestion'] = {
            'pgdesa': gestion, 'pdesa': gestion, 'sectorial': gestion,
        }

        libro = load_workbook(ruta, read_only=True, data_only=True)
        try:
            hojas = {
                nombre: _filas_marco_2026(libro[hoja])
                for nombre, hoja in HOJAS_MARCO_2026.items()
            }
        finally:
            libro.close()

        fuentes = {
            _texto(fila['fuente_id']): fila for fila in hojas['fuentes']
        }
        procedencia = 'XLSX oficial SIS-PE marco 2026-2035 (base upstream)'

        planes = {}
        versiones = {}
        for spec in PLANES_MARCO_2026:
            fuente = fuentes.get(spec['fuente_id'], {})
            instrumento = {
                **spec,
                'norma_aprobacion': _texto(fuente.get('archivo_url')),
            }
            plan = self._plan_para(instrumento, spec['tipo'], procedencia)
            version = self._version_vigente(
                plan, gestion, instrumento, procedencia,
            )
            planes[spec['tipo']] = plan
            versiones[spec['tipo']] = version
        resumen['plan'] = {
            'pgdesa': planes['pgdesa'].codigo,
            'pdesa': planes['pdesa'].codigo,
            'sectorial': planes['sectorial'].codigo,
        }

        ejes = self._importar_ejes_pgdesa_marco(
            hojas['ejes'], versiones['pgdesa'],
        )
        componentes = self._importar_componentes_pdesa_marco(
            hojas['componentes'], versiones['pdesa'], ejes,
        )
        sectores = self._importar_sectores_economicos_marco(
            hojas['sectores'], versiones['sectorial'], componentes,
        )
        lineamientos = self._importar_lineamientos_pad_marco(
            hojas['lineamientos'], versiones['pdesa'], componentes,
        )
        acuerdos = self._importar_acuerdos_internacionales_marco(hojas)

        resumen['lotes'] = {
            'EjePGDESA (7 ejes PGDESA 2026-2035)': len(ejes),
            'ComponentePDESA (38 componentes PDESA 2026-2030)': len(componentes),
            'SectorEconomico (24 sectores PDS/PES 2026)': len(sectores),
            'LineamientoPAD (170 lineamientos PAD)': len(lineamientos),
            'AcuerdoInternacional (ODS 17 + NDT 17 + KMGBF 23)': sum(
                acuerdos.values()
            ),
        }
        resumen['lotes'].update({
            f'  AcuerdoInternacional tipo={tipo}': n
            for tipo, n in sorted(acuerdos.items())
        })
        resumen['omitidos'] = self._decisiones_marco_2026()
        return resumen

    def _importar_ejes_pgdesa_marco(self, filas, version):
        """02_PGDESA_EJES → EjePGDESA (código 01..07).

        El modelo EjePGDESA no tiene campo ``objetivo_impacto``; se conserva
        concatenado a la denominación ('<eje> — <objetivo_impacto>'), dentro
        del límite de 500 caracteres del campo.
        """
        ejes = {}
        for fila in filas:
            codigo = _codigo_2_digitos(fila['codigo'])
            if not codigo:
                continue
            eje = _texto(fila['eje'])
            objetivo = _texto(fila['objetivo_impacto'])
            denominacion = (
                f'{eje} — {objetivo}' if objetivo and objetivo != eje else eje
            )
            obj, _ = EjePGDESA.objects.update_or_create(
                codigo=codigo,
                version_catalogo=version,
                defaults={
                    'denominacion': denominacion[:500],
                    'activo': True,
                },
            )
            ejes[codigo] = obj
        return ejes

    def _importar_componentes_pdesa_marco(self, filas, version, ejes):
        """03_PDESA_COMPONENTES → ComponentePDESA (correlativo por eje).

        El código oficial del XLSX es compuesto ('1.1'); el catálogo exige
        correlativo por eje, así que se renumera en el orden de la fuente
        reiniciando por eje (01..99).
        """
        correlativos = {}
        componentes = {}
        for fila in filas:
            codigo_fuente = _texto(fila['codigo'])
            eje = ejes.get(_codigo_2_digitos(fila['eje_codigo']))
            if not codigo_fuente or eje is None:
                continue
            correlativos[eje.pk] = correlativos.get(eje.pk, 0) + 1
            codigo = str(correlativos[eje.pk]).zfill(2)
            nombre = _texto(fila['componente'])
            objetivo = _texto(fila['objetivo_efecto'])
            denominacion = (
                f'{nombre} — {objetivo}'
                if objetivo and objetivo != nombre else nombre
            )
            obj, _ = ComponentePDESA.objects.update_or_create(
                eje=eje,
                codigo=codigo,
                version_catalogo=version,
                defaults={
                    'denominacion': denominacion[:500],
                    'activo': True,
                },
            )
            componentes[codigo_fuente] = obj
        return componentes

    def _importar_sectores_economicos_marco(self, filas, version, componentes):
        """05_PDS_SECTORES → SectorEconomico (código 01..24).

        El clasificador nacional (F03) no declara dependencia de un
        componente PDESA; como la FK es obligatoria, los 24 sectores se
        cuelgan del primer componente del primer eje (codigo '01') —
        convención documentada: el clasificador es NACIONAL, no sectorial.
        """
        componente_base = next(
            (c for c in componentes.values() if c.codigo == '01'), None,
        )
        sectores = {}
        for fila in filas:
            codigo = _codigo_2_digitos(fila['codigo'])
            if not codigo or componente_base is None:
                continue
            obj, _ = SectorEconomico.objects.update_or_create(
                componente=componente_base,
                codigo=codigo,
                version_catalogo=version,
                defaults={
                    'denominacion': _texto(fila['denominacion'])[:500],
                    'activo': True,
                },
            )
            sectores[codigo] = obj
        return sectores

    def _importar_lineamientos_pad_marco(self, filas, version, componentes):
        """04_PDESA_LINEAMIENTOS → LineamientoPAD (170).

        ``codigo_tecnico`` ('1.1.L01') es identificador técnico SIS-PE; el
        código de catálogo es el correlativo por componente (campo ``orden``,
        reinicia por componente). La FK ``componente`` (migración 0011)
        resuelve el componente del PDESA y el CGEO es el municipio único.
        """
        try:
            entidad = EntidadTerritorialCGEO.objects.get(
                codigo=CGEO_MUNICIPIO_PAD,
            )
        except EntidadTerritorialCGEO.DoesNotExist:
            return {}

        lineamientos = {}
        for fila in filas:
            codigo_fuente = _texto(fila['codigo_tecnico'])
            componente = componentes.get(_texto(fila['componente_codigo']))
            correlativo = _entero(fila['orden'])
            if (
                not codigo_fuente
                or componente is None
                or correlativo is None
                or not 1 <= correlativo <= 99
            ):
                continue
            codigo = str(correlativo).zfill(2)
            obj, _ = LineamientoPAD.objects.update_or_create(
                entidad_territorial=entidad,
                componente=componente,
                codigo=codigo,
                version_catalogo=version,
                defaults={
                    'denominacion': _texto(fila['lineamiento'])[:500],
                    'activo': True,
                },
            )
            lineamientos[codigo_fuente] = obj
        return lineamientos

    def _importar_acuerdos_internacionales_marco(self, hojas):
        """08_ODS / 11_NDT_METAS / 12_KMGBF_30x30 → AcuerdoInternacional.

        Los tres catálogos tienen tipo compatible en el modelo (ODS, NDT,
        COMPROMISO_3030 — alias PAD '30/30'); los códigos se conservan tal
        cual publica la fuente (''1''..''17'', ''a.1.1'', ''1''..''23'').
        """
        por_tipo = (
            ('ods', 'ODS', 'objetivo'),
            ('ndt', 'NDT', 'meta'),
            ('kmgbf', 'COMPROMISO_3030', 'meta'),
        )
        conteo = {}
        for hoja, tipo, campo_texto in por_tipo:
            n = 0
            for fila in hojas[hoja]:
                codigo = _texto(fila['codigo'])
                if not codigo:
                    continue
                AcuerdoInternacional.objects.update_or_create(
                    tipo_acuerdo=tipo,
                    codigo=codigo[:10],
                    defaults={
                        'denominacion': _texto(fila[campo_texto]),
                        'activo': True,
                    },
                )
                n += 1
            conteo[tipo] = n
        return conteo

    def _decisiones_marco_2026(self):
        return [
            (
                'EjePGDESA/ComponentePDESA: el modelo no tiene campos '
                'objetivo_impacto/objetivo_efecto; se conservan concatenados '
                'en denominacion ("<nombre> — <objetivo>"), dentro del '
                'límite de 500 caracteres.'
            ),
            (
                'SectorEconomico: el clasificador nacional (F03) no declara '
                'dependencia sectorial; los 24 sectores se cuelgan del '
                'componente 01 del primer eje PGDESA (FK obligatoria).'
            ),
            (
                'ODS/NDT/KMGBF: cargados en articulacion.AcuerdoInternacional '
                '(tipos ODS, NDT, COMPROMISO_3030) con los códigos tal cual '
                'publica la fuente.'
            ),
            (
                '06_PDS_RESULTADOS: la hoja no trae registros (0); no se '
                'inventan resultados sectoriales.'
            ),
            (
                'NodoPlanificacion: el marco 2026-2035 no toca el árbol de '
                'nodos 2021-2025 (ya cargado, 450 nodos).'
            ),
        ]

    # ------------------------------------------------------------------
    # Reporte
    # ------------------------------------------------------------------
    def _conteos(self, resumen):
        for etiqueta, queryset in (
            ('EjePGDESA (pilares PGDES)', EjePGDESA.objects.all()),
            ('ComponentePDESA (ejes PDES)', ComponentePDESA.objects.all()),
            ('NodoPlanificacion', NodoPlanificacion.objects.all()),
            ('ArticulacionPlanificacion', ArticulacionPlanificacion.objects.all()),
            ('VersionNormativa (fuentes)', VersionNormativa.objects.all()),
        ):
            resumen['modelos'][etiqueta] = queryset.count()

    def _decisiones_omision(self, hojas, componentes):
        omisiones = [
            (
                'SectorEconomico y ResultadoSectorial: el XLSX oficial no '
                'trae sectores del clasificador ni resultados sectoriales; '
                'no se inventan datos y las tablas quedan como están.'
            ),
            (
                'LineamientoPAD: el XLSX no trae lineamientos PAD '
                '(catálogo municipal, no nacional); no se cargan.'
            ),
            (
                '07_RELACIONES CONTIENE/DESAGREGA/OPERATIVIZA: quedan '
                'implícitas en la FK padre de NodoPlanificacion; solo se '
                'materializan las 19 ARTICULA pilar->eje.'
            ),
        ]
        if len(componentes) < len(hojas['ejes']):
            omisiones.append(
                'ComponentePDESA: algunos ejes PDES no tienen pilar '
                'articulado en 06_PILAR_EJE y se omitieron.'
            )
        return omisiones

    def _volcar_resumen(self, resumen, dry_run):
        if 'lotes' in resumen:
            return self._volcar_resumen_marco_2026(resumen, dry_run)
        modo = 'DRY-RUN (nada persistido)' if dry_run else 'COMMIT (persistido)'
        self.stdout.write(
            self.style.MIGRATE_HEADING(f'\nCatálogo nacional maestro ({modo})')
        )
        self.stdout.write(
            f"Planes: PGDES={resumen['plan']['pgdes']} "
            f"(versión vigente {resumen['gestion']['pgdes']}) | "
            f"PDES={resumen['plan']['pdes']} "
            f"(versión vigente {resumen['gestion']['pdes']})"
        )
        for etiqueta, conteo in sorted(resumen['modelos'].items()):
            self.stdout.write(f'  {etiqueta}: {conteo}')
        if resumen['omitidos']:
            self.stdout.write(self.style.WARNING('  Decisiones (no cargado):'))
            for omision in resumen['omitidos']:
                self.stdout.write(self.style.WARNING(f'    - {omision}'))

    def _volcar_resumen_marco_2026(self, resumen, dry_run):
        modo = 'DRY-RUN (nada persistido)' if dry_run else 'COMMIT (persistido)'
        self.stdout.write(
            self.style.MIGRATE_HEADING(f'\nCatálogo nacional maestro ({modo})')
        )
        self.stdout.write(
            f"Planes: PGDESA={resumen['plan']['pgdesa']} "
            f"(versión vigente {resumen['gestion']['pgdesa']}) | "
            f"PDESA={resumen['plan']['pdesa']} "
            f"(versión vigente {resumen['gestion']['pdesa']}) | "
            f"SECTORIAL={resumen['plan']['sectorial']} "
            f"(versión vigente {resumen['gestion']['sectorial']})"
        )
        self.stdout.write('  Lotes cargados (versiones de catálogo vigentes):')
        for etiqueta, conteo in resumen['lotes'].items():
            self.stdout.write(f'    {etiqueta}: {conteo}')
        if resumen['omitidos']:
            self.stdout.write(self.style.WARNING('  Decisiones:'))
            for omision in resumen['omitidos']:
                self.stdout.write(self.style.WARNING(f'    - {omision}'))


def _codigo_sistema_eje(eje_id, hojas):
    for fila in hojas['ejes']:
        if fila['elemento_id'] == eje_id:
            return fila['codigo_sistema']
    return None


def _codigo_sistema_meta(meta_id, hojas):
    for fila in hojas['metas']:
        if fila['elemento_id'] == meta_id:
            return fila['codigo_sistema']
    return None


def _codigo_sistema_resultado(resultado_id, hojas):
    for fila in hojas['resultados']:
        if fila['elemento_id'] == resultado_id:
            return fila['codigo_sistema']
    return None
