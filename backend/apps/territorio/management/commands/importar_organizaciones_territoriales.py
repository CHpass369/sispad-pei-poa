"""Carga el padrón de organizaciones sociales territoriales de un distrito.

Fuente: los libros `LISTA DE LIMITES DE OTB *.xlsx` que entrega cada distrito.
Los doce comparten la misma hoja y las mismas ocho columnas:

    Nº | ORGANIZACIÓN SOCIAL TERRITORIAL | NOMBRE DEL DIRIGENTE | CARGO |
    TELEFONO | DISTRITO | ¿presento? | OBSERVACIONES

El distrito se pasa por parámetro y NO se lee de la columna F: en el conjunto
completo esa columna viene vacía en 111 de 368 filas y con seis grafías
distintas en el resto. El archivo llega dentro de la carpeta de su distrito, y
esa es la única fuente confiable.

Una fila cuenta como organización solo si trae número de orden. Lava Lava
—único de los doce— no es una lista plana: intercala encabezados de sección sin
`Nº` que nombran la subcentral, y debajo van sus miembros numerados desde 1. El
primero de esos encabezados dice `SINDICATOS`, que no es ninguna organización.
"""
import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.candado import anio_habilitado
from apps.territorio.models import (
    DirigenteTerritorial, Distrito, TipoUnidadTerritorial, UnidadTerritorial,
    clave_organizacion,
)

HOJA_POR_DEFECTO = '1. Equipamiento'
FILA_INICIAL = 3  # 1 es el rótulo "DETALLES", 2 son los encabezados.

COL_NUMERO, COL_NOMBRE, COL_DIRIGENTE = 0, 1, 2
COL_CARGO, COL_TELEFONO, COL_OBSERVACION = 3, 4, 7

# El tipo se deduce del prefijo del propio nombre. Lo que no cae en ninguno
# queda como 'otro': inventarle un tipo a una organización sería inventar dato.
PREFIJOS = [
    ('OTB', TipoUnidadTerritorial.OTB),
    ('JUNTA VECINAL', TipoUnidadTerritorial.JUNTA_VECINAL),
    ('JV', TipoUnidadTerritorial.JUNTA_VECINAL),
    ('SINDICATO', TipoUnidadTerritorial.SINDICATO),
    ('SUBCENTRAL', TipoUnidadTerritorial.SUBCENTRAL),
    ('COMUNIDAD', TipoUnidadTerritorial.COMUNIDAD),
    ('URB', TipoUnidadTerritorial.ZONA),
    ('ZONA', TipoUnidadTerritorial.ZONA),
]


def deducir_tipo(nombre):
    """`O.T.B. SAN JOSÉ` y `OTB SAN JOSE` dan las dos `otb`."""
    plano = clave_organizacion(nombre)
    for prefijo, tipo in PREFIJOS:
        if plano == prefijo or plano.startswith(prefijo + ' '):
            return tipo
    return TipoUnidadTerritorial.OTRO


def limpiar_telefono(valor):
    """openpyxl devuelve los números como float: `65376999.0` → `65376999`."""
    texto = str(valor or '').strip()
    if texto.endswith('.0'):
        texto = texto[:-2]
    return texto


class Command(BaseCommand):
    help = ('Importa el padrón de organizaciones territoriales de un distrito '
            'y sus dirigentes desde la planilla LISTA DE LIMITES DE OTB.')

    def add_arguments(self, parser):
        parser.add_argument('--archivo', required=True,
                            help='Ruta del .xlsx del distrito.')
        parser.add_argument('--distrito', required=True,
                            help='Código del distrito destino (D1, DA, DCH…).')
        parser.add_argument('--gestion', type=int,
                            help='Gestión POA a la que sirve el padrón. Por defecto, la habilitada.')
        parser.add_argument('--hoja', default=HOJA_POR_DEFECTO)
        parser.add_argument('--dry-run', action='store_true',
                            help='Informa lo que haría, sin escribir nada.')

    def handle(self, *args, **opciones):
        distrito = self._resolver_distrito(opciones['distrito'])
        gestion = opciones['gestion'] or anio_habilitado()
        if not gestion:
            raise CommandError(
                'No hay gestión habilitada: indique --gestion con el año del padrón.')

        filas, encabezados, ilegibles = self._leer(
            opciones['archivo'], opciones['hoja'])
        if not filas:
            raise CommandError(
                'La planilla no tiene ninguna organización con número de orden.')

        registros, descartes = self._consolidar(filas)
        self._informar(distrito, gestion, registros, descartes, encabezados,
                       ilegibles)

        if opciones['dry_run']:
            self.stdout.write(self.style.WARNING(
                f'[dry-run] {len(registros)} organizaciones, sin escribir.'))
            return

        altas, cambios, dirigentes = self._guardar(distrito, gestion, registros)
        self.stdout.write(self.style.SUCCESS(
            f'{distrito.codigo}: {altas} organizaciones nuevas, {cambios} '
            f'actualizadas, {dirigentes} dirigentes de la gestión {gestion}.'))

    # --- Lectura -----------------------------------------------------------

    def _resolver_distrito(self, codigo):
        distrito = Distrito.objects.filter(codigo__iexact=codigo).first()
        if distrito is None:
            disponibles = ', '.join(
                Distrito.objects.order_by('codigo').values_list('codigo', flat=True))
            raise CommandError(
                f'No existe el distrito «{codigo}». Disponibles: {disponibles}')
        return distrito

    def _leer(self, ruta, hoja):
        import openpyxl

        try:
            libro = openpyxl.load_workbook(ruta, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'No se encontró el archivo {ruta}')
        if hoja not in libro.sheetnames:
            raise CommandError(
                f'La hoja «{hoja}» no está en el libro. Hay: {libro.sheetnames}')

        filas, encabezados, ilegibles = [], [], []
        for numero, celdas in enumerate(
                libro[hoja].iter_rows(min_row=FILA_INICIAL, values_only=True),
                start=FILA_INICIAL):
            def celda(indice):
                if indice >= len(celdas) or celdas[indice] is None:
                    return ''
                return str(celdas[indice]).strip()

            crudo = celdas[COL_NOMBRE] if COL_NOMBRE < len(celdas) else None
            if isinstance(crudo, (datetime.datetime, datetime.date)):
                # Excel convirtió el nombre en fecha y el texto original se
                # perdió: `15 de mayo` quedó como 2026-05-15. Deducir el nombre
                # sería inventarlo; se rechaza la fila y se corrige la planilla.
                ilegibles.append((numero, crudo))
                continue
            if not celda(COL_NOMBRE):
                continue  # Las planillas traen cientos de filas en blanco al pie.
            if not celda(COL_NUMERO):
                # Encabezado de sección, no una organización.
                encabezados.append((numero, celda(COL_NOMBRE)))
                continue
            filas.append({
                'fila': numero,
                'orden': celda(COL_NUMERO),
                'nombre': ' '.join(celda(COL_NOMBRE).upper().split()),
                'dirigente': ' '.join(celda(COL_DIRIGENTE).upper().split()),
                'cargo': ' '.join(celda(COL_CARGO).upper().split()),
                'telefono': limpiar_telefono(celda(COL_TELEFONO)),
                'observacion': celda(COL_OBSERVACION),
            })
        return filas, encabezados, ilegibles

    def _consolidar(self, filas):
        """Colapsa por nombre normalizado.

        Ante dos filas de la misma organización gana la que trae dirigente:
        quedarse con la primera a secas perdería el dato si la vacía viniera
        antes, y eso no se nota hasta que alguien emite el acta.
        """
        registros, descartes = {}, []
        for fila in filas:
            clave = clave_organizacion(fila['nombre'])
            previa = registros.get(clave)
            if previa is None:
                registros[clave] = fila
                continue
            if fila['dirigente'] and not previa['dirigente']:
                registros[clave] = fila
                descartes.append((previa['fila'], previa['nombre']))
            else:
                descartes.append((fila['fila'], fila['nombre']))
        return registros, descartes

    # --- Escritura ---------------------------------------------------------

    @transaction.atomic
    def _guardar(self, distrito, gestion, registros):
        altas = cambios = dirigentes = 0
        correlativo = self._ultimo_correlativo(distrito)
        for clave, fila in registros.items():
            unidad = UnidadTerritorial.objects.filter(
                distrito=distrito, nombre_busqueda=clave).first()
            if unidad is None:
                # El código se asigna una sola vez, al alta. Recalcularlo en
                # cada importación renumeraría a todo el distrito porque
                # alguien intercaló una fila en la planilla.
                correlativo += 1
                unidad = UnidadTerritorial(
                    distrito=distrito, codigo=f'{distrito.codigo}-{correlativo:03d}')
                altas += 1
            else:
                cambios += 1
            unidad.nombre = fila['nombre']
            unidad.tipo = deducir_tipo(fila['nombre'])
            unidad.save()

            if not fila['dirigente']:
                continue
            # El cargo entra en la clave de unicidad: una organización puede
            # declarar presidente y secretario general en la misma gestión.
            DirigenteTerritorial.objects.update_or_create(
                unidad=unidad, gestion=gestion, cargo=fila['cargo'],
                defaults={
                    'nombre': fila['dirigente'],
                    'telefono': fila['telefono'],
                    'observacion': fila['observacion'],
                    'vigente': True,
                },
            )
            dirigentes += 1
        return altas, cambios, dirigentes

    def _ultimo_correlativo(self, distrito):
        """Sigue numerando donde quedó el distrito, no desde uno."""
        codigos = UnidadTerritorial.objects.filter(
            distrito=distrito).values_list('codigo', flat=True)
        sufijos = [int(c.rsplit('-', 1)[-1]) for c in codigos
                   if c.rsplit('-', 1)[-1].isdigit()]
        return max(sufijos, default=0)

    # --- Informe -----------------------------------------------------------

    def _informar(self, distrito, gestion, registros, descartes,
                  encabezados=(), ilegibles=()):
        self.stdout.write(
            f'{distrito} · gestión {gestion} · {len(registros)} organizaciones')

        conteo = {}
        for fila in registros.values():
            tipo = deducir_tipo(fila['nombre'])
            conteo[tipo] = conteo.get(tipo, 0) + 1
        etiquetas = dict(TipoUnidadTerritorial.choices)
        detalle = ', '.join(f'{etiquetas[t]}: {n}'
                            for t, n in sorted(conteo.items(), key=lambda x: -x[1]))
        self.stdout.write(f'  tipos deducidos del nombre → {detalle}')

        sin_dirigente = [f['nombre'] for f in registros.values() if not f['dirigente']]
        if sin_dirigente:
            self.stdout.write(self.style.WARNING(
                f'  {len(sin_dirigente)} sin dirigente: '
                f'{", ".join(sin_dirigente[:5])}'))
        sin_cargo = sum(1 for f in registros.values()
                        if f['dirigente'] and not f['cargo'])
        if sin_cargo:
            self.stdout.write(self.style.WARNING(
                f'  {sin_cargo} dirigentes sin cargo declarado'))
        for fila, nombre in descartes:
            self.stdout.write(self.style.WARNING(
                f'  fila {fila} repetida, se ignora: {nombre}'))
        for fila, nombre in encabezados:
            self.stdout.write(self.style.WARNING(
                f'  fila {fila} sin número de orden, se toma como encabezado '
                f'de sección: {nombre}'))
        for fila, valor in ilegibles:
            self.stdout.write(self.style.ERROR(
                f'  fila {fila} RECHAZADA: Excel guardó el nombre como fecha '
                f'({valor:%d de %B}). Escribirlo como texto en la planilla.'))
