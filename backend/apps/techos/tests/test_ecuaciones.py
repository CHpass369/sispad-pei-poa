"""
Tests de ecuación pin del núcleo presupuestario (slice S1).

Fijan el comportamiento actual de las 6 implementaciones de saldo ANTES del
refactor de delegación al BudgetAllocationService (design D11) y verifican,
después del refactor, que las ecuaciones dan resultados idénticos (no cambia
ningún saldo).

Ecuación 2027: recursos 245.290.497,00 - gastos obligatorios 6.464.396,00
= distribuible 238.826.101,00.

Los pins de validar_techo (apps.sis_poa.migration_v2) viven en
test_validar_techo_sispoa.py (C2): este módulo NO importa apps.sis_poa
para poder coleccionar en HEAD, sin el swap poau→sis_poa del branch
hermano.
"""
import io as _io
import types
from datetime import date
from decimal import Decimal

import pytest
from django.utils import timezone

from apps.accounts.models import Usuario
from apps.catalogos.models import FuenteFinanciamiento
from apps.gestion.models import GestionFiscal
from apps.presupuesto.models import ProgramaPresupuestario
from apps.techos.models import (
    DistribucionTecho,
    GastoObligatorio,
    MovimientoTecho,
    RecursoTecho,
    TechoPresupuestario,
)
from apps.techos.services import obtener_saldo_disponible, resumen_techo
from apps.workflow.consolidacion import (
    _total_distribuido,
    _total_techo,
    consolidar_poa_institucional,
)

MONTO_RECURSOS = Decimal('245290497.00')
MONTO_GASTOS = Decimal('6464396.00')
MONTO_DISTRIBUIBLE = Decimal('238826101.00')
MONTO_DISTRIBUIDO = Decimal('50000000.00')
SALDO_ESPERADO = MONTO_DISTRIBUIBLE - MONTO_DISTRIBUIDO  # 188.826.101,00


def _gestion(anio):
    """Gestión fiscal para el techo 1:1 (R2.1). El techo 2027 (fixture) ya
    ocupa GestionFiscal(2027); los techos auxiliares usan 2028 para no
    colisionar con el OneToOne."""
    gestion, _ = GestionFiscal.objects.get_or_create(
        anio=anio, defaults={'estado': 'preparacion'},
    )
    return gestion


@pytest.fixture
def fuente(db):
    return FuenteFinanciamiento.objects.create(
        codigo='41-113',
        gestion=2027,
        denominacion='Coparticipación Tributaria',
        fecha_vigencia_desde=date(2027, 1, 1),
    )


@pytest.fixture
def techo_2027(db, fuente):
    """Techo 2027 con recursos y gastos obligatorios de la ecuación pin."""
    techo = TechoPresupuestario.objects.create(
        gestion=2027,
        gestion_fiscal=_gestion(2027),
        monto_total=MONTO_RECURSOS,
        fuente=fuente,
        concepto='Techo 2027',
    )
    RecursoTecho.objects.create(
        techo=techo, fuente=fuente,
        concepto='Coparticipación Tributaria', monto=Decimal('181658084.00'),
    )
    RecursoTecho.objects.create(
        techo=techo, fuente=fuente,
        concepto='Recursos Hipotecarios', monto=Decimal('63632413.00'),
    )
    GastoObligatorio.objects.create(
        techo=techo, fuente=fuente,
        denominacion='Renta Dignidad', monto=Decimal('3500000.00'),
    )
    GastoObligatorio.objects.create(
        techo=techo, fuente=fuente,
        denominacion='Seguridad Ciudadana', monto=Decimal('2964396.00'),
    )
    DistribucionTecho.objects.create(
        techo=techo, monto_asignado=MONTO_DISTRIBUIDO,
    )
    return techo


@pytest.fixture
def usuario(db):
    return Usuario.objects.create_user(
        email='ecuaciones_test@gamsacaba.gob.bo', password='test123',
    )


# ---------------------------------------------------------------------------
# 1. TechoPresupuestario.saldo_disponible (property)
# ---------------------------------------------------------------------------

def test_saldo_disponible_property_ecuacion_2027(techo_2027):
    """saldo_disponible = techo - gastos obligatorios - distribuido."""
    assert techo_2027.monto_total - MONTO_GASTOS == MONTO_DISTRIBUIBLE
    assert techo_2027.saldo_disponible == SALDO_ESPERADO


def test_saldo_disponible_property_sin_distribuir(techo_2027):
    """Ecuación 2027 con distribución cero: el saldo es el distribuible."""
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=MONTO_RECURSOS, fuente=techo_2027.fuente,
        concepto='Techo 2027 sin distribuir',
    )
    RecursoTecho.objects.create(
        techo=techo, fuente=techo_2027.fuente,
        concepto='Coparticipación Tributaria', monto=MONTO_RECURSOS,
    )
    GastoObligatorio.objects.create(
        techo=techo, fuente=techo_2027.fuente,
        denominacion='Renta Dignidad', monto=MONTO_GASTOS,
    )
    assert techo.saldo_disponible == MONTO_DISTRIBUIBLE


# ---------------------------------------------------------------------------
# 2. services.obtener_saldo_disponible (semántica legacy por movimientos)
# ---------------------------------------------------------------------------

def test_obtener_saldo_disponible_sin_movimientos(techo_2027):
    """Sin movimientos aprobados el saldo legacy es el monto_total."""
    assert obtener_saldo_disponible(techo_2027) == MONTO_RECURSOS


def test_obtener_saldo_disponible_con_reduccion_aprobada(techo_2027, usuario):
    """Una reducción aprobada descuenta del saldo legacy."""
    MovimientoTecho.objects.create(
        techo=techo_2027, movement_type='reduccion',
        amount=Decimal('500000.00'), justification='Reducción pin',
        requested_by=usuario, approved_by=usuario, date=timezone.now(),
    )
    assert obtener_saldo_disponible(techo_2027) == MONTO_RECURSOS - Decimal('500000.00')


# ---------------------------------------------------------------------------
# 3. services.resumen_techo
# ---------------------------------------------------------------------------

def test_resumen_techo_claves_y_valores(techo_2027):
    resumen = resumen_techo(techo_2027)
    assert resumen['techo_id'] == str(techo_2027.id)
    assert resumen['gestion'] == 2027
    assert resumen['monto_total'] == MONTO_RECURSOS
    assert resumen['total_recursos'] == MONTO_RECURSOS
    assert resumen['total_gastos_obligatorios'] == MONTO_GASTOS
    assert resumen['monto_distribuido'] == MONTO_DISTRIBUIDO
    assert resumen['saldo_disponible'] == SALDO_ESPERADO
    assert resumen['excede'] is False


# ---------------------------------------------------------------------------
# 4. workflow/consolidacion: _total_techo / _total_distribuido / saldo_por_distribuir
# ---------------------------------------------------------------------------

def test_workflow_total_techo_por_gestion(techo_2027):
    assert _total_techo(2027) == MONTO_RECURSOS


def test_workflow_total_distribuido_por_gestion(techo_2027):
    assert _total_distribuido(2027) == MONTO_DISTRIBUIDO


def test_workflow_saldo_por_distribuir(techo_2027):
    totales = consolidar_poa_institucional(2027)['totales']
    assert totales['techo'] == MONTO_RECURSOS
    assert totales['techo_distribuido'] == MONTO_DISTRIBUIDO
    assert totales['saldo_por_distribuir'] == MONTO_RECURSOS - MONTO_DISTRIBUIDO


# ---------------------------------------------------------------------------
# 5. reportes/services.py: techo por programa en el consolidado
# ---------------------------------------------------------------------------

def test_reportes_consolidado_techo_por_programa(monkeypatch, techo_2027):
    """El consolidado muestra el techo asignado por programa (sin float).

    W6: además de las celdas, se fija el contrato (output, filename) que
    B1 restaura — los llamadores (_responder_descarga y
    generar_reporte_presupuestario_async) desempaquetan esa tupla.
    """
    import openpyxl

    import apps.reportes.services as reportes_mod

    prog = ProgramaPresupuestario.objects.create(
        codigo='001', nombre='Programa 1', gestion=2027,
    )
    DistribucionTecho.objects.create(
        techo=techo_2027, programa=prog, monto_asignado=Decimal('30000000.00'),
    )

    captured = _io.BytesIO()
    monkeypatch.setattr(
        reportes_mod, 'io', types.SimpleNamespace(BytesIO=lambda: captured),
    )
    output, filename = reportes_mod.generar_poa_consolidado_xlsx(2027)

    # Contrato del generador (B1): (BytesIO, nombre de archivo) — el output
    # es el mismo objeto que el BytesIO capturado por el monkeypatch.
    assert output is captured
    assert filename.startswith('poa_consolidado_2027_')
    assert filename.endswith('.xlsx')

    ws = openpyxl.load_workbook(captured).active
    FILA_PRIMER_PROGRAMA = 4  # encabezados en fila 3
    COL_TECHO = 4  # Techo (Bs)
    COL_SALDO = 5  # Saldo (Bs)
    COL_ESTADO = 7  # Estado
    assert ws.cell(row=FILA_PRIMER_PROGRAMA, column=COL_TECHO).value == 30000000
    assert ws.cell(row=FILA_PRIMER_PROGRAMA, column=COL_SALDO).value == 30000000
    assert ws.cell(row=FILA_PRIMER_PROGRAMA, column=COL_ESTADO).value == 'Completo'


# ---------------------------------------------------------------------------
# 6. validar_techo — movido a test_validar_techo_sispoa.py (C2/C3):
#    importa apps.sis_poa y se salta en HEAD (sin el swap poau→sis_poa).
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 7. BudgetAllocationService (motor único): las ecuaciones también vía servicio
# ---------------------------------------------------------------------------

from apps.techos.services import BudgetAllocationService, budget_service  # noqa: E402


def test_service_ecuacion_2027_resumen(techo_2027):
    """get_techo_resumen reproduce el resumen legacy y suma claves nuevas."""
    resumen = budget_service.get_techo_resumen(techo_2027)
    assert resumen['techo_id'] == str(techo_2027.id)
    assert resumen['gestion'] == 2027
    assert resumen['monto_total'] == MONTO_RECURSOS
    assert resumen['total_recursos'] == MONTO_RECURSOS
    assert resumen['total_gastos_obligatorios'] == MONTO_GASTOS
    assert resumen['monto_distribuido'] == MONTO_DISTRIBUIDO
    assert resumen['saldo_disponible'] == SALDO_ESPERADO
    assert resumen['excede'] is False
    assert resumen['techo_distribuible'] == MONTO_DISTRIBUIBLE
    assert resumen['monto_reservado'] == Decimal('0.00')


def test_service_get_available_ecuacion_2027(techo_2027):
    assert budget_service.get_available(techo_2027) == SALDO_ESPERADO


def test_service_totales_techo(techo_2027):
    assert budget_service.get_total_recursos(techo_2027) == MONTO_RECURSOS
    assert budget_service.get_total_gastos_obligatorios(techo_2027) == MONTO_GASTOS
    assert budget_service.get_techo_distribuible(techo_2027) == MONTO_DISTRIBUIBLE


def test_service_get_distributed_nodo_hoja(techo_2027):
    dist = techo_2027.distribuciones.get()
    assert budget_service.get_distributed_nodo(dist) == MONTO_DISTRIBUIDO


def test_service_agregados_por_gestion(techo_2027):
    assert budget_service.get_techo_agregado_gestion(2027) == MONTO_RECURSOS
    assert budget_service.get_distribuido_agregado_gestion(2027) == MONTO_DISTRIBUIDO
    assert budget_service.get_saldo_por_distribuir_gestion(2027) == (
        MONTO_RECURSOS - MONTO_DISTRIBUIDO
    )


def test_service_get_distribuido_por_programa(techo_2027):
    prog = ProgramaPresupuestario.objects.create(
        codigo='002', nombre='Programa 2', gestion=2027,
    )
    DistribucionTecho.objects.create(
        techo=techo_2027, programa=prog, monto_asignado=Decimal('30000000.00'),
    )
    assert budget_service.get_distribuido_por_programa(prog) == Decimal('30000000.00')


def test_service_estado_techo_parcial(techo_2027):
    assert budget_service.estado_techo(techo_2027) == 'DISTRIBUCION_PARCIAL'


def test_service_estado_techo_sin_configurar(techo_2027):
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=MONTO_RECURSOS, fuente=techo_2027.fuente,
        concepto='Sin distribución',
    )
    assert budget_service.estado_techo(techo) == 'SIN_CONFIGURAR'


def test_service_estado_techo_inconsistente(techo_2027):
    """Σ distribuido > distribuible => INCONSISTENTE (fail-loud, C3)."""
    DistribucionTecho.objects.bulk_create([
        DistribucionTecho(
            techo=techo_2027, monto_asignado=MONTO_DISTRIBUIBLE + Decimal('1.00'),
        ),
    ])
    assert budget_service.estado_techo(techo_2027) == 'INCONSISTENTE'


def test_service_validate_allocation_dentro(techo_2027):
    resultado = budget_service.validate_allocation(techo_2027, Decimal('10000000.00'))
    assert resultado['valido'] is True
    assert resultado['excede'] is False
    assert resultado['monto_solicitado'] == Decimal('10000000.00')
    assert resultado['saldo_disponible'] == SALDO_ESPERADO


def test_service_validate_allocation_excede(techo_2027):
    resultado = budget_service.validate_allocation(techo_2027, Decimal('200000000.00'))
    assert resultado['valido'] is False
    assert resultado['excede'] is True
    assert resultado['saldo_disponible'] == SALDO_ESPERADO


def test_service_validate_allocation_respeto_reserva(techo_2027):
    """La guardia a nivel techo resta reservas (C3): la capacidad efectiva
    es min(saldo_bolsa, techo_distribuible - Σ hojas - reservado_total)."""
    DistribucionTecho.objects.create(
        techo=techo_2027, monto_asignado=Decimal('0.00'),
        monto_reserva=Decimal('10000000.00'),
    )
    resultado = budget_service.validate_allocation(techo_2027, Decimal('180000000.00'))
    assert resultado['valido'] is False
    assert resultado['excede'] is True
    assert resultado['saldo_disponible'] == SALDO_ESPERADO - Decimal('10000000.00')
    assert budget_service.validate_allocation(
        techo_2027, Decimal('178826101.00'),
    )['valido'] is True


def test_service_validate_allocation_monto_negativo(techo_2027):
    resultado = budget_service.validate_allocation(techo_2027, Decimal('-1.00'))
    assert resultado['valido'] is False
    assert 'negativo' in resultado['mensaje']


def test_service_can_allocate(techo_2027):
    assert budget_service.can_allocate(techo_2027, SALDO_ESPERADO) is True
    assert budget_service.can_allocate(
        techo_2027, SALDO_ESPERADO + Decimal('0.01'),
    ) is False


def test_service_get_saldo_por_movimientos(techo_2027, usuario):
    """Compatibilidad legacy V1: saldo por movimientos aprobados."""
    MovimientoTecho.objects.create(
        techo=techo_2027, movement_type='reduccion',
        amount=Decimal('500000.00'), justification='Reducción pin',
        requested_by=usuario, approved_by=usuario, date=timezone.now(),
    )
    assert budget_service.get_saldo_por_movimientos(techo_2027) == (
        MONTO_RECURSOS - Decimal('500000.00')
    )


# ---------------------------------------------------------------------------
# 8. Pins de corrección (W2/W3/W4): edición con exclude_id y guardia de modelo
# ---------------------------------------------------------------------------

def test_service_validate_allocation_edicion_exclude_id(techo_2027):
    """W2: al editar una fila (600→550 sobre techo 1000) su monto viejo
    vuelve a la capacidad. Sin exclude_id el doble conteo (saldo de bolsa
    que aún incluye la fila) colapsa el min() y rechaza toda edición
    positiva."""
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo W2',
    )
    fila = DistribucionTecho.objects.create(
        techo=techo, monto_asignado=Decimal('600.00'),
    )

    # Sin exclude_id: saldo de bolsa 400 (1000 - 600) → 550 rechazado.
    assert budget_service.validate_allocation(
        techo, Decimal('550.00'),
    )['valido'] is False

    # Con exclude_id: los 600 vuelven → capacidad 1000 → 550 válido.
    resultado = budget_service.validate_allocation(
        techo, Decimal('550.00'), exclude_id=fila.pk,
    )
    assert resultado['valido'] is True
    assert resultado['saldo_disponible'] == Decimal('1000.00')

    # Sobre la misma capacidad corregida, 1000.01 sigue excediendo.
    assert budget_service.validate_allocation(
        techo, Decimal('1000.01'), exclude_id=fila.pk,
    )['valido'] is False


def test_distribucion_clean_edicion_fila_cero_respeta_guardia(techo_2027):
    """W3: editar una fila cuyo monto viejo es Decimal('0.00') (falsy) no
    debe tomar el monto nuevo como 'viejo': el exceso se rechaza."""
    from django.core.exceptions import ValidationError

    fila = DistribucionTecho.objects.create(
        techo=techo_2027, monto_asignado=Decimal('0.00'),
    )
    # 200.000.000 > saldo disponible (238.826.101 - 50.000.000 = 188.826.101)
    fila.monto_asignado = Decimal('200000000.00')
    with pytest.raises(ValidationError):
        fila.full_clean()

    # Control positivo: una edición dentro del saldo sí es válida.
    fila.monto_asignado = Decimal('100000000.00')
    fila.full_clean()  # no levanta


def test_distribucion_clean_resta_reserva_del_techo(techo_2027):
    """W4: la guardia del modelo resta el reservado total, igual que
    validate_allocation (C3): una fila que excede la capacidad efectiva
    (techo_distribuible - Σ hojas - reservado) se rechaza."""
    from django.core.exceptions import ValidationError

    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo W4',
    )
    DistribucionTecho.objects.create(
        techo=techo, monto_asignado=Decimal('600.00'),
        monto_reserva=Decimal('200.00'),
    )
    # Capacidad efectiva = 1000 - 600 - 200 = 200 → 300 se rechaza.
    fila = DistribucionTecho(
        techo=techo, monto_asignado=Decimal('300.00'),
    )
    with pytest.raises(ValidationError):
        fila.full_clean()


# ---------------------------------------------------------------------------
# 9. Pins C4 (W1): guardia de movimientos restaurada
#    (validar_movimiento / aplicar_movimiento, apps.techos.services)
# ---------------------------------------------------------------------------

from apps.techos.services import aplicar_movimiento, validar_movimiento  # noqa: E402


def test_validar_movimiento_transferencia_excede_saldo_origen(techo_2027, usuario):
    """C4 (W1): una transferencia cuyo monto excede el saldo por movimientos
    del techo origen se rechaza."""
    destino = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo destino C4',
    )
    movimiento = MovimientoTecho(
        techo=techo_2027, movement_type='transferencia',
        source_ceiling=techo_2027, destination_ceiling=destino,
        amount=MONTO_RECURSOS + Decimal('1.00'),
        justification='Transferencia que excede el saldo',
        requested_by=usuario, date=timezone.now(),
    )
    errores = validar_movimiento(movimiento)
    assert any('excede el saldo disponible' in e for e in errores)


def test_validar_movimiento_transferencia_dentro_del_saldo(techo_2027, usuario):
    """C4 (W1): control positivo — transferencia dentro del saldo sin errores."""
    destino = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo destino C4',
    )
    movimiento = MovimientoTecho(
        techo=techo_2027, movement_type='transferencia',
        source_ceiling=techo_2027, destination_ceiling=destino,
        amount=Decimal('1000.00'),
        justification='Transferencia dentro del saldo',
        requested_by=usuario, date=timezone.now(),
    )
    assert validar_movimiento(movimiento) == []


def test_aplicar_movimiento_reduccion_excede_saldo(techo_2027, usuario):
    """C4 (W1): una reducción que dejaría el saldo por movimientos en
    negativo se rechaza; no persiste nada."""
    movimiento = MovimientoTecho(
        techo=techo_2027, movement_type='reduccion',
        amount=MONTO_RECURSOS + Decimal('1.00'),
        justification='Reducción que deja saldo negativo',
        requested_by=usuario, date=timezone.now(),
    )
    with pytest.raises(ValueError):
        aplicar_movimiento(movimiento)
    assert MovimientoTecho.objects.filter(techo=techo_2027).count() == 0


def test_aplicar_movimiento_pk_none_nuevo_registro_permitido(techo_2027, usuario):
    """C4 (W1): con pk=None (registro nuevo) el excluir_id no excluye nada
    y la validación usa el saldo vigente: una reducción dentro del saldo se
    aplica y persiste. (El display V1 solo cuenta movimientos APROBADOS,
    A11: el registro recién creado no altera obtener_saldo_disponible.)"""
    movimiento = MovimientoTecho(
        techo=techo_2027, movement_type='reduccion',
        amount=Decimal('500000.00'),
        justification='Reducción pin C4',
        requested_by=usuario, date=timezone.now(),
    )
    aplicado = aplicar_movimiento(movimiento)
    assert aplicado.pk is not None
    assert MovimientoTecho.objects.filter(pk=aplicado.pk).exists()
    assert obtener_saldo_disponible(techo_2027) == MONTO_RECURSOS


def test_aplicar_movimiento_misma_ecuacion_que_display_v1(techo_2027, usuario):
    """C4 (W1): aplicar_movimiento valida contra la misma ecuación que
    muestra obtener_saldo_disponible (saldo por movimientos aprobados,
    A11)."""
    MovimientoTecho.objects.create(
        techo=techo_2027, movement_type='reduccion',
        amount=Decimal('500000.00'), justification='Reducción aprobada',
        requested_by=usuario, approved_by=usuario, date=timezone.now(),
    )
    saldo_display = obtener_saldo_disponible(techo_2027)
    assert saldo_display == MONTO_RECURSOS - Decimal('500000.00')

    # Exactamente el saldo mostrado → permitido.
    aplicar_movimiento(MovimientoTecho(
        techo=techo_2027, movement_type='reduccion',
        amount=saldo_display, justification='Reducción al límite V1',
        requested_by=usuario, date=timezone.now(),
    ))
    # Un centavo más que el saldo mostrado → rechazado.
    with pytest.raises(ValueError):
        aplicar_movimiento(MovimientoTecho(
            techo=techo_2027, movement_type='reduccion',
            amount=saldo_display + Decimal('0.01'), justification='Excede V1',
            requested_by=usuario, date=timezone.now(),
        ))


# ---------------------------------------------------------------------------
# 10. Pins C5: TechoPresupuestario.total_gastos_obligatorios (activo=True)
# ---------------------------------------------------------------------------

def test_total_gastos_obligatorios_respeta_gastos_inactivos(techo_2027):
    """C5: la property delega en el motor (activo=True); desactivar un
    GastoObligatorio la actualiza de forma consistente con el motor y con
    el saldo disponible."""
    go = techo_2027.gastos_obligatorios.first()
    go.activo = False
    go.save(update_fields=['activo'])
    esperado = MONTO_GASTOS - go.monto
    assert techo_2027.total_gastos_obligatorios == esperado
    assert techo_2027.total_gastos_obligatorios == (
        budget_service.get_total_gastos_obligatorios(techo_2027)
    )
    # El saldo disponible se libera por el monto desactivado.
    assert techo_2027.saldo_disponible == SALDO_ESPERADO + go.monto


# ---------------------------------------------------------------------------
# 11. Pins D1/D2 (4R): el saldo disponible a nivel techo resta reservado_total
#     y la guardia C3 valida el monto_reserva propio de la fila
# ---------------------------------------------------------------------------

def test_techo_saldo_disponible_resta_reserva(techo_2027):
    """D1: las tres rutas de lectura (get_available, get_techo_resumen,
    property saldo_disponible) restan reservado_total y coinciden con la
    capacidad de validate_allocation: techo 1000, asignado 100, reserva 200
    → saldo disponible 700.00."""
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo D1',
    )
    DistribucionTecho.objects.create(
        techo=techo, monto_asignado=Decimal('100.00'),
        monto_reserva=Decimal('200.00'),
    )
    esperado = Decimal('700.00')
    assert budget_service.get_available(techo) == esperado
    assert budget_service.get_techo_resumen(techo)['saldo_disponible'] == esperado
    assert techo.saldo_disponible == esperado
    # La capacidad de validate_allocation coincide: 700 entra, 700.01 no.
    assert budget_service.validate_allocation(techo, esperado)['valido'] is True
    assert budget_service.validate_allocation(
        techo, esperado + Decimal('0.01'),
    )['valido'] is False


def test_techo_saldo_disponible_cero_con_reserva_total(techo_2027):
    """D1: techo 1000, asignado 100, reserva 900 → saldo disponible 0.00
    en las tres rutas de lectura (reserva consume todo el saldo)."""
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo D1 reserva total',
    )
    DistribucionTecho.objects.create(
        techo=techo, monto_asignado=Decimal('100.00'),
        monto_reserva=Decimal('900.00'),
    )
    assert budget_service.get_available(techo) == Decimal('0.00')
    assert budget_service.get_techo_resumen(techo)['saldo_disponible'] == Decimal('0.00')
    assert techo.saldo_disponible == Decimal('0.00')
    assert budget_service.validate_allocation(
        techo, Decimal('0.00'),
    )['valido'] is True
    assert budget_service.validate_allocation(
        techo, Decimal('0.01'),
    )['valido'] is False


def test_distribucion_clean_reserva_nueva_excede_capacidad(techo_2027):
    """D2 (a): una fila nueva cuyo monto_reserva (en memoria) excede la
    capacidad efectiva se rechaza. Antes solo se validaba monto_asignado:
    una reserva gigante con monto_asignado 0 persistía en silencio."""
    from django.core.exceptions import ValidationError

    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo D2',
    )
    fila = DistribucionTecho(
        techo=techo, monto_asignado=Decimal('0.00'),
        monto_reserva=Decimal('999999999.00'),
    )
    with pytest.raises(ValidationError):
        fila.full_clean()


def test_distribucion_clean_edicion_reserva_excede_capacidad(techo_2027):
    """D2 (b): editar una fila para subir monto_reserva más allá de la
    capacidad efectiva se rechaza: el monto viejo de la fila vuelve a la
    capacidad (exclude_id) y el monto_reserva nuevo se valida."""
    from django.core.exceptions import ValidationError

    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo D2 edición',
    )
    fila = DistribucionTecho.objects.create(
        techo=techo, monto_asignado=Decimal('100.00'),
        monto_reserva=Decimal('100.00'),
    )
    fila.monto_reserva = Decimal('1000.00')
    with pytest.raises(ValidationError):
        fila.full_clean()

    # Control: dentro de la capacidad sí se puede subir la reserva
    # (100 asignado + 800 reserva = 900 ≤ 1000).
    fila.monto_reserva = Decimal('800.00')
    fila.full_clean()  # no levanta


def test_distribucion_clean_reserva_dentro_capacidad_permitida(techo_2027):
    """D2 (c): control positivo — una fila con monto_asignado + monto_reserva
    dentro de la capacidad efectiva se persiste sin error."""
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo D2 válido',
    )
    fila = DistribucionTecho.objects.create(
        techo=techo, monto_asignado=Decimal('100.00'),
        monto_reserva=Decimal('200.00'),
    )
    assert fila.pk is not None


def test_service_estado_techo_inconsistente_reserva_only(techo_2027):
    """D2 (d): sobre-compromiso solo por reservas (monto_asignado 0,
    monto_reserva > distribuible) → INCONSISTENTE (fail-loud), igual que
    el sobre-compromiso por monto_asignado. El estado solo puede existir
    por fuera del ORM (bulk_create/migraciones): DistribucionTecho.save()
    ya rechaza la fila (D2 a)."""
    techo = TechoPresupuestario.objects.create(
        gestion=2028, gestion_fiscal=_gestion(2028),
        monto_total=Decimal('1000.00'), fuente=techo_2027.fuente,
        concepto='Techo D2 estado',
    )
    DistribucionTecho.objects.bulk_create([
        DistribucionTecho(
            techo=techo, monto_asignado=Decimal('0.00'),
            monto_reserva=Decimal('1500.00'),
        ),
    ])
    assert budget_service.estado_techo(techo) == 'INCONSISTENTE'
